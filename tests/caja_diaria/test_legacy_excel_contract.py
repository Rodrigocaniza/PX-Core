"""Characterization tests: OBSERVED_LEGACY behavior, not approved policy."""

from __future__ import annotations

import json
import unittest
from pathlib import Path

from openpyxl import Workbook

import CajaDiaria


FIXTURES = Path(__file__).parent / "fixtures"


def load_cases() -> dict:
    return json.loads((FIXTURES / "legacy_cases.json").read_text(encoding="utf-8"))


def build_sheet(case: dict, *, title: str = "Día"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.cell(1, 1, case["date"])
    headers = load_cases()["headers"]
    for column, header in enumerate(headers, 1):
        sheet.cell(3, column, header)
    sheet.cell(4, 1, "CAJA INICIAL")
    sheet.cell(4, 9, case["opening_cash"])
    for row_number, row in enumerate(case["rows"], 5):
        for column, value in enumerate(row, 1):
            sheet.cell(row_number, column, value)
    return sheet


class LegacyExcelContractTests(unittest.TestCase):
    def test_reported_column_order_matches_legacy_positions(self):
        data = load_cases()
        self.assertEqual(len(data["headers"]), 14)
        self.assertEqual(data["headers"][7:], [
            "TOTAL", "Efectivo", "Tarj./Cheq.", "Ordenes", "Cuotas", "Saldo", "Gastos"
        ])

    def test_representative_day_preserves_rows_text_balance_and_totals(self):
        case = load_cases()["cases"]["representative_day"]
        records, opening, totals, errors = CajaDiaria.analizar_hoja(build_sheet(case), "PC")

        self.assertEqual(errors, [])
        self.assertEqual(opening, case["opening_cash"])
        self.assertEqual(len(records), case["expected"]["entries"])
        self.assertEqual(records[1]["saldo"], "cancelado")
        self.assertEqual(records[2]["saldo"], "50000")
        self.assertEqual(records[2]["ordenes"], "ORD-10")
        self.assertEqual(records[2]["cuotas"], "2x50000")
        self.assertEqual(totals, {
            "total": case["expected"]["total"],
            "efectivo": case["expected"]["cash"],
            "tarjeta_cheque": case["expected"]["card_check"],
            "gastos": case["expected"]["expenses"],
            "efectivo_final": case["expected"]["final_cash"],
        })
        self.assertEqual([r["descripcion"] for r in records].count("CLIENTE MULTIFILA"), 2)

    def test_empty_template_with_opening_is_not_an_operational_day(self):
        case = load_cases()["cases"]["empty_day"]
        records, opening, totals, errors = CajaDiaria.analizar_hoja(build_sheet(case), "PC")
        self.assertEqual((records, opening, totals, errors), ([], None, None, []))

    def test_optional_cells_become_empty_strings(self):
        case = load_cases()["cases"]["blank_optional_values"]
        records, _, _, _ = CajaDiaria.analizar_hoja(build_sheet(case), "PC")
        self.assertEqual(records[0]["sobre"], "")
        self.assertEqual(records[0]["saldo"], "")
        self.assertEqual(records[0]["tarjeta_cheque"], 0)

    def test_missing_headers_omits_sheet(self):
        sheet = Workbook().active
        sheet["A1"] = "01-08-2026"
        result = CajaDiaria.analizar_hoja(sheet, "PC")
        self.assertEqual(result[:3], ([], None, None))
        self.assertIn("no se encontró la fila de encabezados", result[3][0])

    def test_invalid_money_becomes_zero_with_non_blocking_error(self):
        case = load_cases()["cases"]["blank_optional_values"]
        sheet = build_sheet(case)
        sheet.cell(5, 8, "NO-ES-MONTO")
        records, _, totals, errors = CajaDiaria.analizar_hoja(sheet, "PC")
        self.assertEqual(records[0]["total"], 0)
        self.assertEqual(totals["total"], 0)
        self.assertTrue(any("TOTAL inválido" in error for error in errors))

    def test_duplicate_opening_uses_last_value_with_warning(self):
        case = load_cases()["cases"]["blank_optional_values"]
        sheet = build_sheet(case)
        sheet.insert_rows(5)
        sheet.cell(5, 1, "CAJA INICIAL")
        sheet.cell(5, 9, 250000)
        _, opening, _, errors = CajaDiaria.analizar_hoja(sheet, "PC")
        self.assertEqual(opening, 250000)
        self.assertTrue(any("más de una" in error for error in errors))

    def test_blank_description_with_amount_is_discarded(self):
        case = load_cases()["cases"]["empty_day"]
        sheet = build_sheet(case)
        sheet.cell(5, 8, 900000)
        records, _, totals, _ = CajaDiaria.analizar_hoja(sheet, "PC")
        self.assertEqual(records, [])
        self.assertIsNone(totals)

    def test_real_like_values_are_preserved_and_calculation_rows_are_not_entries(self):
        case = load_cases()["cases"]["blank_optional_values"]
        sheet = build_sheet(case)
        sheet.cell(5, 4, ".000037")
        sheet.cell(5, 11, "Caja Muni")
        sheet.cell(5, 12, 5)
        sheet.cell(5, 13, "cancelado")
        sheet.cell(6, 1, "TOTALES")
        sheet.cell(6, 8, case["rows"][0][7] or 0)
        sheet.cell(6, 9, case["opening_cash"] + (case["rows"][0][8] or 0))
        sheet.cell(6, 10, case["rows"][0][9] or 0)
        sheet.cell(6, 14, case["rows"][0][13] or 0)
        sheet.cell(8, 9, sheet.cell(6, 9).value - sheet.cell(6, 14).value)

        records, _, _, errors = CajaDiaria.analizar_hoja(sheet, "PC")

        self.assertEqual(errors, [])
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["cod"], ".000037")
        self.assertEqual(records[0]["ordenes"], "Caja Muni")
        self.assertEqual(records[0]["cuotas"], "5")
        self.assertEqual(records[0]["saldo"], "cancelado")

    def test_declared_totals_mismatch_is_reported(self):
        case = load_cases()["cases"]["blank_optional_values"]
        sheet = build_sheet(case)
        sheet.cell(6, 1, "TOTALES")
        sheet.cell(6, 8, 999)
        sheet.cell(6, 9, case["opening_cash"])
        sheet.cell(6, 10, 0)
        sheet.cell(6, 14, 0)
        _, _, _, errors = CajaDiaria.analizar_hoja(sheet, "PC")
        self.assertTrue(any("no coinciden" in error for error in errors))


if __name__ == "__main__":
    unittest.main()
