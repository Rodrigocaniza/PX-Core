"""Informes mensuales de BC Gestión.

Este módulo reutiliza las funciones de cierre de ``Movimientos.py``. De esa
forma, la vista de Informes y el Cierre mensual siempre parten de los mismos
totales.
"""

from calendar import monthrange
from datetime import datetime
from pathlib import Path
import re

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import customtkinter as ctk

import Movimientos
from datos import leer_datos


COLOR_FONDO = ("#F7F9FC", "#0B1220")
COLOR_PANEL = ("#FFFFFF", "#131D2E")
COLOR_PANEL_SECUNDARIO = ("#F1F5F9", "#1A2639")
COLOR_BORDE = ("#D9E2EC", "#26354A")
COLOR_TEXTO = ("#182230", "#F4F7FB")
COLOR_TEXTO_SUAVE = ("#617084", "#9CAFC5")
COLOR_PRIMARIO = "#2F6FED"
COLOR_PRIMARIO_HOVER = "#2458BF"
COLOR_VERDE = "#12A67A"
COLOR_ROJO = "#E05A67"
COLOR_NARANJA = "#F09A3E"
COLOR_VIOLETA = "#7A67E8"
COLOR_CELESTE = "#27A7D8"
COLOR_AZUL_NOCHE = ("#102A43", "#0D2033")

MESES_ESPANOL = (
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
)


def periodo_en_palabras(periodo):
    """Convierte MM-AAAA en un título más amable para el informe."""
    try:
        fecha = datetime.strptime(periodo, "%m-%Y")
    except ValueError:
        return periodo
    return f"{MESES_ESPANOL[fecha.month - 1]} {fecha.year}"


def nombre_archivo_seguro(informe, extension):
    """Crea un nombre breve y válido para Windows."""
    unidad = re.sub(
        r"[^A-Za-z0-9_-]+",
        "_",
        informe["unidad"],
    ).strip("_")
    unidad = unidad or "Empresa"
    return (
        f"Informe_BC_{informe['periodo'].replace('-', '_')}_"
        f"{unidad}.{extension}"
    )


def nombre_archivo_comparacion(comparacion, extension):
    """Crea el nombre del archivo de una comparación entre meses."""
    unidad = re.sub(
        r"[^A-Za-z0-9_-]+",
        "_",
        comparacion["unidad"],
    ).strip("_")
    unidad = unidad or "Empresa"
    desde = comparacion["periodo_desde"].replace("-", "_")
    hasta = comparacion["periodo_hasta"].replace("-", "_")
    return f"Comparacion_BC_{desde}_a_{hasta}_{unidad}.{extension}"


def nombre_archivo_anual(informe, extension):
    """Crea el nombre del archivo de un informe anual."""
    unidad = re.sub(
        r"[^A-Za-z0-9_-]+",
        "_",
        informe["unidad"],
    ).strip("_")
    unidad = unidad or "Empresa"
    return f"Informe_Anual_BC_{informe['anio']}_{unidad}.{extension}"


def resumir_detalles_por_dia(detalles):
    """Agrupa cada fecha por unidad/categoría y muestra movimientos internos."""
    agrupados = {}

    for registro in detalles:
        categoria = str(registro.get("categoria", "")).strip()
        unidad = str(registro.get("unidad", "")).strip()
        es_interno = registro["tipo"] in [
            "Transferencia interna",
            "Deposito interno",
        ]
        ruta_interna = (
            str(registro.get("detalle", "")).strip()
            if es_interno
            else ""
        )
        if es_interno:
            grupo = categoria or "Movimiento interno"
        elif categoria == "Operación diaria" and unidad not in ["", "General"]:
            grupo = unidad
        else:
            grupo = categoria or unidad or "Otros"

        clave = (
            registro["orden"],
            registro["fecha"],
            grupo,
            ruta_interna,
        )
        if clave not in agrupados:
            agrupados[clave] = {
                "orden": registro["orden"],
                "fecha": registro["fecha"],
                "grupo": grupo,
                "entradas": 0,
                "salidas": 0,
                "resultado": 0,
                "cantidad": 0,
                "ruta_interna": ruta_interna,
                "monto_interno": 0,
            }

        fila = agrupados[clave]
        if registro["tipo"] == "Ingreso":
            fila["entradas"] += registro["monto"]
        elif registro["tipo"] == "Egreso":
            fila["salidas"] += registro["monto"]
        elif es_interno:
            fila["monto_interno"] += registro["monto"]
        fila["cantidad"] += 1
        fila["resultado"] = fila["entradas"] - fila["salidas"]

    orden_unidades = {
        Movimientos.normalizar_texto(unidad): indice
        for indice, unidad in enumerate(Movimientos.UNIDADES)
    }

    def clave_orden(fila):
        grupo_normalizado = Movimientos.normalizar_texto(fila["grupo"])
        posicion = orden_unidades.get(
            grupo_normalizado,
            len(orden_unidades) + 1,
        )
        return (
            -fila["orden"].toordinal(),
            posicion,
            grupo_normalizado,
        )

    return sorted(agrupados.values(), key=clave_orden)


def exportar_excel(informe, ruta):
    """Exporta un informe visual, filtrable y verificable en Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ModuleNotFoundError as error:
        raise RuntimeError(
            "Falta instalar openpyxl. Cerrá el sistema y volvé a abrir "
            "iniciar_interfaz.bat para completar la instalación."
        ) from error

    ruta = Path(ruta)
    libro = Workbook()
    resumen = libro.active
    resumen.title = "Resumen"

    azul = "2F6FED"
    azul_oscuro = "102A43"
    verde = "12A67A"
    rojo = "E05A67"
    naranja = "F09A3E"
    gris = "E9EFF5"
    gris_claro = "F7F9FC"
    gris_texto = "617084"
    blanco = "FFFFFF"
    borde = Side(style="thin", color="DCE4EE")
    borde_fuerte = Side(style="medium", color="C8D4E3")
    formato_gs = '"Gs. "#,##0;[Red]-"Gs. "#,##0;"-"'

    def titulo_hoja(hoja, texto, ultima_columna):
        hoja.sheet_view.showGridLines = False
        hoja.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=ultima_columna,
        )
        celda = hoja.cell(1, 1, texto)
        celda.font = Font(
            name="Aptos Display",
            size=16,
            bold=True,
            color=blanco,
        )
        celda.fill = PatternFill("solid", fgColor=azul_oscuro)
        celda.alignment = Alignment(vertical="center")
        hoja.row_dimensions[1].height = 30

    def encabezado(hoja, fila, columnas):
        for columna, texto in enumerate(columnas, 1):
            celda = hoja.cell(fila, columna, texto)
            celda.font = Font(bold=True, color=blanco)
            celda.fill = PatternFill("solid", fgColor=azul)
            celda.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            celda.border = Border(bottom=borde)

    def ajustar_columnas(hoja, limites=None):
        limites = limites or {}
        for columna in range(1, hoja.max_column + 1):
            letra = get_column_letter(columna)
            mayor = 0
            for celda in hoja[letra]:
                valor = "" if celda.value is None else str(celda.value)
                mayor = max(mayor, len(valor))
            minimo, maximo = limites.get(columna, (11, 42))
            hoja.column_dimensions[letra].width = min(
                max(mayor + 2, minimo),
                maximo,
            )

    def preparar_impresion(hoja, orientacion="landscape"):
        hoja.page_setup.orientation = orientacion
        hoja.page_setup.paperSize = hoja.PAPERSIZE_A4
        hoja.page_setup.fitToWidth = 1
        hoja.page_setup.fitToHeight = 0
        hoja.sheet_properties.pageSetUpPr.fitToPage = True
        hoja.oddFooter.center.text = (
            f"BC Gestión - {informe['periodo']} - Página &P de &N"
        )
        hoja.oddFooter.center.size = 8
        hoja.oddFooter.center.color = gris_texto
        hoja.sheet_properties.outlinePr.summaryBelow = True

    def crear_tarjeta_excel(
        hoja,
        rango_titulo,
        rango_valor,
        titulo,
        valor,
        color,
        formato=formato_gs,
    ):
        fondos_tarjeta = {
            verde: "EAF8F3",
            rojo: "FFF0F2",
            azul: "EDF3FF",
            naranja: "FFF5E8",
            "7A67E8": "F2EFFF",
            "27A7D8": "EAF8FC",
        }
        fondo_tarjeta = fondos_tarjeta.get(color, blanco)
        borde_color = Side(style="medium", color=color)
        hoja.merge_cells(rango_titulo)
        hoja.merge_cells(rango_valor)
        celda_titulo = hoja[rango_titulo.split(":")[0]]
        celda_valor = hoja[rango_valor.split(":")[0]]
        celda_titulo.value = titulo.upper()
        celda_valor.value = valor
        for fila in hoja[rango_titulo]:
            for celda in fila:
                celda.fill = PatternFill("solid", fgColor=fondo_tarjeta)
                celda.border = Border(
                    top=borde,
                    left=borde_color,
                    right=borde,
                )
        for fila in hoja[rango_valor]:
            for celda in fila:
                celda.fill = PatternFill("solid", fgColor=fondo_tarjeta)
                celda.border = Border(
                    bottom=borde_fuerte,
                    left=borde_color,
                    right=borde,
                )
        celda_titulo.font = Font(
            name="Aptos",
            size=9,
            bold=True,
            color=gris_texto,
        )
        celda_titulo.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )
        celda_valor.font = Font(
            name="Aptos Display",
            size=15,
            bold=True,
            color=color,
        )
        celda_valor.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )
        celda_valor.number_format = formato

    def agregar_tabla(hoja, referencia, nombre):
        tabla = Table(displayName=nombre, ref=referencia)
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        hoja.add_table(tabla)

    # El resumen se diseña como un tablero de una sola página.
    titulo_hoja(
        resumen,
        f"BC Inversiones EAS  |  Resumen ejecutivo - "
        f"{periodo_en_palabras(informe['periodo'])}",
        12,
    )
    resumen.merge_cells("A2:C2")
    resumen["A2"] = f"VISTA: {informe['unidad']}"
    resumen.merge_cells("D2:F2")
    resumen["D2"] = f"PERÍODO: {informe['periodo']}"
    resumen.merge_cells("G2:L2")
    resumen["G2"] = f"GENERADO: {datetime.now():%d/%m/%Y %H:%M}"
    for celda in ("A2", "D2", "G2"):
        resumen[celda].font = Font(size=9, bold=True, color=gris_texto)

    crear_tarjeta_excel(
        resumen, "A4:C4", "A5:C6", "Ingresos",
        informe["ingresos"], verde,
    )
    crear_tarjeta_excel(
        resumen, "D4:F4", "D5:F6", "Utilidad / resultado",
        "=A5-J5", verde if informe["utilidad"] >= 0 else rojo,
    )
    crear_tarjeta_excel(
        resumen, "G4:I4", "G5:I6", "Margen",
        "=IFERROR(D5/A5,0)", azul, "0.00%",
    )
    crear_tarjeta_excel(
        resumen, "J4:L4", "J5:L6", "Egresos",
        informe["egresos"], rojo,
    )
    crear_tarjeta_excel(
        resumen, "A8:C8", "A9:C10", "Fondo",
        informe["fondo"] if informe["fondo"] is not None else 0,
        naranja,
    )
    crear_tarjeta_excel(
        resumen, "D8:F8", "D9:F10", "Inversiones",
        informe["inversiones"], "7A67E8",
    )
    crear_tarjeta_excel(
        resumen, "G8:I8", "G9:I10", "Salida real de dinero",
        informe["salida_caja"], "27A7D8",
    )
    crear_tarjeta_excel(
        resumen, "J8:L8", "J9:L10",
        "Retenciones y otros descuentos",
        informe["retenciones"], naranja,
    )
    resumen.merge_cells("A11:L11")
    resumen["A11"] = (
        f"{'CIERRE POSITIVO' if informe['utilidad'] >= 0 else 'CIERRE A REVISAR'}"
        f"  |  El resultado del mes fue "
        f"Gs. {Movimientos.formatear_monto(informe['utilidad'])} "
        f"con un margen de "
        f"{Movimientos.formatear_porcentaje(informe['margen'])}%."
    )
    resumen["A11"].font = Font(
        name="Aptos",
        size=10,
        bold=True,
        color=blanco,
    )
    resumen["A11"].fill = PatternFill(
        "solid",
        fgColor=verde if informe["utilidad"] >= 0 else rojo,
    )
    resumen["A11"].alignment = Alignment(
        horizontal="left",
        vertical="center",
    )
    resumen.row_dimensions[11].height = 24

    def agregar_composicion_resumen(
        fila_titulo,
        titulo,
        categorias_informe,
        nombre_hoja,
        celda_total,
        color,
    ):
        resumen.merge_cells(
            start_row=fila_titulo,
            start_column=1,
            end_row=fila_titulo,
            end_column=12,
        )
        celda_titulo = resumen.cell(fila_titulo, 1, titulo.upper())
        celda_titulo.font = Font(
            name="Aptos Display",
            size=12,
            bold=True,
            color=azul_oscuro,
        )

        fila_encabezado = fila_titulo + 1
        fila_inicio = fila_titulo + 2
        encabezado(resumen, fila_encabezado, ["Categoría", "Monto", "%"])
        categorias_principales = min(5, len(categorias_informe))
        for indice in range(categorias_principales):
            fila = fila_inicio + indice
            fila_origen = 3 + indice
            resumen.cell(
                fila,
                1,
                (
                    f'=IF(LEN(\'{nombre_hoja}\'!A{fila_origen})>26,'
                    f'LEFT(\'{nombre_hoja}\'!A{fila_origen},23)&"...",'
                    f'\'{nombre_hoja}\'!A{fila_origen})'
                ),
            )
            resumen.cell(
                fila,
                2,
                f"='{nombre_hoja}'!B{fila_origen}",
            )
            resumen.cell(
                fila,
                3,
                f"=IFERROR(B{fila}/{celda_total},0)",
            )

        if len(categorias_informe) > categorias_principales:
            fila_otros = fila_inicio + categorias_principales
            resumen.cell(fila_otros, 1, "Otros")
            resumen.cell(
                fila_otros,
                2,
                (
                    f"=SUM('{nombre_hoja}'!B{3 + categorias_principales}:"
                    f"B{2 + len(categorias_informe)})"
                ),
            )
            resumen.cell(
                fila_otros,
                3,
                f"=IFERROR(B{fila_otros}/{celda_total},0)",
            )
            cantidad_grafico = categorias_principales + 1
        else:
            cantidad_grafico = categorias_principales

        ultima_fila = fila_inicio + max(cantidad_grafico, 1) - 1
        for fila in range(fila_inicio, ultima_fila + 1):
            resumen.cell(fila, 2).number_format = formato_gs
            resumen.cell(fila, 3).number_format = "0.0%"
            if fila % 2 == 0:
                for columna in range(1, 4):
                    resumen.cell(fila, columna).fill = PatternFill(
                        "solid",
                        fgColor=gris_claro,
                    )

        if cantidad_grafico:
            grafico = BarChart()
            datos = Reference(
                resumen,
                min_col=2,
                min_row=fila_encabezado,
                max_row=ultima_fila,
            )
            categorias = Reference(
                resumen,
                min_col=1,
                min_row=fila_inicio,
                max_row=ultima_fila,
            )
            grafico.add_data(datos, titles_from_data=True)
            grafico.set_categories(categorias)
            grafico.type = "bar"
            grafico.title = titulo.title()
            grafico.style = 10
            grafico.legend = None
            grafico.x_axis.title = "Guaraníes"
            grafico.y_axis.title = ""
            grafico.height = 4.7
            grafico.width = 13.2
            if grafico.series:
                grafico.series[0].graphicalProperties.solidFill = color
                grafico.series[0].graphicalProperties.line.solidFill = color
            resumen.add_chart(grafico, f"E{fila_encabezado}")

    agregar_composicion_resumen(
        12,
        "Ingresos por categoría",
        informe["categorias_ingresos"],
        "Ingresos",
        "$A$5",
        verde,
    )
    agregar_composicion_resumen(
        22,
        "Egresos por categoría",
        informe["categorias_egresos"],
        "Egresos",
        "$J$5",
        rojo,
    )

    resumen.merge_cells("A32:C32")
    resumen["A32"] = "CONTROL DEL INFORME"
    resumen["A32"].font = Font(bold=True, color=blanco)
    resumen["A32"].fill = PatternFill("solid", fgColor=azul)
    resumen["A33"] = "Total de egresos"
    resumen["B33"] = "=J5"
    resumen["A34"] = "Suma del desglose"
    if informe["categorias_egresos"]:
        ultima_fila_egresos = len(informe["categorias_egresos"]) + 2
        resumen["B34"] = (
            f"=SUM('Egresos'!B3:B{ultima_fila_egresos})"
        )
    else:
        resumen["B34"] = 0
    resumen["A35"] = "Estado"
    resumen["B35"] = '=IF(B33=B34,"OK","REVISAR")'
    for fila in (33, 34):
        resumen.cell(fila, 2).number_format = formato_gs
    resumen["B35"].font = Font(bold=True, color=verde)
    resumen.freeze_panes = "A4"
    for columna in range(1, 13):
        resumen.column_dimensions[get_column_letter(columna)].width = 12.5
    resumen.column_dimensions["A"].width = 22
    resumen.column_dimensions["B"].width = 17
    resumen.column_dimensions["C"].width = 9
    resumen.row_dimensions[1].height = 34
    resumen.row_dimensions[4].height = 21
    resumen.row_dimensions[5].height = 25
    resumen.row_dimensions[6].height = 18
    resumen.row_dimensions[8].height = 21
    resumen.row_dimensions[9].height = 25
    resumen.row_dimensions[10].height = 18
    resumen.print_area = "A1:L36"
    preparar_impresion(resumen)

    def crear_hoja_categorias(nombre, categorias, color):
        hoja = libro.create_sheet(nombre)
        titulo_hoja(
            hoja,
            f"{nombre} por categoría - {informe['periodo']}",
            3,
        )
        encabezado(hoja, 2, ["Categoría", "Monto", "% del total"])
        fila_total = max(len(categorias) + 3, 3)
        for fila, (concepto, monto) in enumerate(categorias, 3):
            hoja.cell(fila, 1, concepto)
            hoja.cell(fila, 2, monto)
            hoja.cell(fila, 3, f"=IFERROR(B{fila}/$B${fila_total},0)")
            hoja.cell(fila, 2).number_format = formato_gs
            hoja.cell(fila, 3).number_format = "0.0%"
            hoja.cell(fila, 1).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )
            lineas = max(1, (len(str(concepto)) + 64) // 65)
            hoja.row_dimensions[fila].height = min(45, 15 * lineas)
            if fila % 2:
                for columna in range(1, 3):
                    hoja.cell(fila, columna).fill = PatternFill(
                        "solid",
                        fgColor="F6F8FB",
                    )

        hoja.cell(fila_total, 1, "TOTAL")
        if categorias:
            hoja.cell(
                fila_total,
                2,
                f"=SUM(B3:B{fila_total - 1})",
            )
        else:
            hoja.cell(fila_total, 2, 0)
        hoja.cell(fila_total, 2).number_format = formato_gs
        hoja.cell(fila_total, 3, 1 if categorias else 0)
        hoja.cell(fila_total, 3).number_format = "0.0%"
        for columna in range(1, 4):
            hoja.cell(fila_total, columna).font = Font(
                bold=True,
                color=blanco,
            )
            hoja.cell(fila_total, columna).fill = PatternFill(
                "solid",
                fgColor=color,
            )
        if categorias:
            agregar_tabla(
                hoja,
                f"A2:C{fila_total - 1}",
                f"Tabla{nombre}",
            )
        hoja.freeze_panes = "A3"
        ajustar_columnas(
            hoja,
            {1: (32, 58), 2: (18, 23), 3: (15, 18)},
        )
        hoja.print_title_rows = "1:2"
        preparar_impresion(hoja, "portrait")
        return hoja

    crear_hoja_categorias(
        "Ingresos",
        informe["categorias_ingresos"],
        verde,
    )
    crear_hoja_categorias(
        "Egresos",
        informe["categorias_egresos"],
        rojo,
    )

    comparacion = libro.create_sheet("Comparación")
    titulo_hoja(
        comparacion,
        f"Comparación por unidad - {informe['periodo']}",
        6,
    )
    encabezado(
        comparacion,
        2,
        [
            "Unidad",
            "Ingresos",
            "Egresos",
            "Resultado",
            "Transf. recibidas",
            "Transf. enviadas",
        ],
    )
    for fila, unidad in enumerate(informe["comparacion_unidades"], 3):
        valores = [
            unidad["unidad"],
            unidad["ingresos"],
            unidad["egresos"],
            unidad["resultado"],
            unidad["recibidas"],
            unidad["enviadas"],
        ]
        for columna, valor in enumerate(valores, 1):
            comparacion.cell(fila, columna, valor)
            if columna > 1:
                comparacion.cell(fila, columna).number_format = formato_gs
        if fila % 2:
            for columna in range(1, 7):
                comparacion.cell(fila, columna).fill = PatternFill(
                    "solid",
                    fgColor=gris_claro,
                )
    if informe["comparacion_unidades"]:
        agregar_tabla(
            comparacion,
            f"A2:F{len(informe['comparacion_unidades']) + 2}",
            "TablaComparacion",
        )
    comparacion.freeze_panes = "A3"
    ajustar_columnas(
        comparacion,
        {
            1: (16, 24),
            2: (16, 22),
            3: (16, 22),
            4: (16, 22),
            5: (18, 24),
            6: (18, 24),
        },
    )
    comparacion.print_title_rows = "1:2"
    preparar_impresion(comparacion)

    detalle = libro.create_sheet("Detalle diario")
    titulo_hoja(
        detalle,
        f"Detalle diario - {informe['periodo']} - {informe['unidad']}",
        7,
    )
    fila = 3
    fecha_actual = None
    total_entradas_dia = 0
    total_salidas_dia = 0
    total_interno_dia = 0

    def cerrar_dia(fila_destino):
        detalle.cell(fila_destino, 1, "TOTAL DEL DÍA")
        detalle.merge_cells(
            start_row=fila_destino,
            start_column=1,
            end_row=fila_destino,
            end_column=2,
        )
        detalle.cell(fila_destino, 3, total_entradas_dia)
        detalle.cell(fila_destino, 4, total_salidas_dia)
        detalle.cell(
            fila_destino,
            5,
            total_entradas_dia - total_salidas_dia,
        )
        detalle.cell(fila_destino, 6, "Movido internamente")
        detalle.cell(fila_destino, 7, total_interno_dia)
        for columna in range(1, 8):
            celda = detalle.cell(fila_destino, columna)
            celda.font = Font(bold=True, color=azul_oscuro)
            celda.fill = PatternFill("solid", fgColor=gris)
            celda.border = Border(top=borde_fuerte, bottom=borde)
            if columna in [3, 4, 5, 7]:
                celda.number_format = formato_gs
        return fila_destino + 2

    for registro in informe["resumen_diario"]:
        if registro["fecha"] != fecha_actual:
            if fecha_actual is not None:
                fila = cerrar_dia(fila)
            fecha_actual = registro["fecha"]
            total_entradas_dia = 0
            total_salidas_dia = 0
            total_interno_dia = 0
            detalle.merge_cells(
                start_row=fila,
                start_column=1,
                end_row=fila,
                end_column=7,
            )
            celda_fecha = detalle.cell(fila, 1, str(fecha_actual))
            celda_fecha.font = Font(bold=True, color=blanco, size=11)
            celda_fecha.fill = PatternFill("solid", fgColor=azul_oscuro)
            celda_fecha.alignment = Alignment(vertical="center")
            detalle.row_dimensions[fila].height = 23
            fila += 1
            encabezado(
                detalle,
                fila,
                [
                    "Categoría / unidad",
                    "Mov.",
                    "Entró",
                    "Salió",
                    "Resultado",
                    "Ruta interna / depósito",
                    "Monto movido",
                ],
            )
            fila += 1

        valores = [
            registro["grupo"],
            registro["cantidad"],
            registro["entradas"],
            registro["salidas"],
            registro["resultado"],
            registro["ruta_interna"],
            registro["monto_interno"],
        ]
        for columna, valor in enumerate(valores, 1):
            celda = detalle.cell(fila, columna, valor)
            celda.border = Border(bottom=borde)
            if columna in [3, 4, 5, 7]:
                celda.number_format = formato_gs
                celda.alignment = Alignment(horizontal="right")
        if fila % 2:
            for columna in range(1, 8):
                detalle.cell(fila, columna).fill = PatternFill(
                    "solid",
                    fgColor=gris_claro,
                )
        total_entradas_dia += registro["entradas"]
        total_salidas_dia += registro["salidas"]
        total_interno_dia += registro["monto_interno"]
        fila += 1

    if fecha_actual is not None:
        fila = cerrar_dia(fila)

    detalle.freeze_panes = "A3"
    ajustar_columnas(
        detalle,
        {
            1: (24, 34),
            2: (14, 18),
            3: (18, 22),
            4: (18, 22),
            5: (18, 22),
            6: (30, 46),
            7: (18, 22),
        },
    )
    detalle.print_title_rows = "1:1"
    preparar_impresion(detalle)

    movimientos = libro.create_sheet("Movimientos")
    titulo_hoja(
        movimientos,
        f"Movimientos verificables - {informe['periodo']} - {informe['unidad']}",
        6,
    )
    encabezado(
        movimientos,
        2,
        ["Fecha", "Unidad", "Tipo", "Categoría", "Concepto / detalle", "Monto"],
    )
    for fila, registro in enumerate(informe["detalles"], 3):
        valores = [
            registro["fecha"],
            registro["unidad"],
            registro["tipo"],
            registro["categoria"],
            registro["detalle"],
            registro["monto"],
        ]
        for columna, valor in enumerate(valores, 1):
            movimientos.cell(fila, columna, valor)
        movimientos.cell(fila, 6).number_format = formato_gs
        movimientos.cell(fila, 4).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        movimientos.cell(fila, 5).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        lineas_categoria = (
            len(str(registro["categoria"])) + 27
        ) // 28
        lineas_detalle = (len(str(registro["detalle"])) + 59) // 60
        movimientos.row_dimensions[fila].height = min(
            45,
            15 * max(1, lineas_categoria, lineas_detalle),
        )
    if informe["detalles"]:
        agregar_tabla(
            movimientos,
            f"A2:F{len(informe['detalles']) + 2}",
            "TablaMovimientos",
        )
    movimientos.freeze_panes = "A3"
    ajustar_columnas(
        movimientos,
        {
            1: (13, 18),
            2: (13, 20),
            3: (12, 18),
            4: (16, 28),
            5: (32, 58),
            6: (18, 22),
        },
    )
    movimientos.print_title_rows = "1:2"
    preparar_impresion(movimientos)

    try:
        libro.calculation.fullCalcOnLoad = True
        libro.calculation.forceFullCalc = True
    except AttributeError:
        pass

    libro.save(ruta)
    return ruta


def exportar_pdf(informe, ruta):
    """Exporta un informe ejecutivo y un detalle legible en PDF."""
    try:
        from reportlab.lib import colors
        from reportlab.graphics.charts.barcharts import HorizontalBarChart
        from reportlab.graphics.shapes import Drawing, Rect, String
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            KeepTogether,
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
        from xml.sax.saxutils import escape
    except ModuleNotFoundError as error:
        raise RuntimeError(
            "Falta instalar reportlab. Cerrá el sistema y volvé a abrir "
            "iniciar_interfaz.bat para completar la instalación."
        ) from error

    ruta = Path(ruta)
    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        "TituloBC",
        parent=estilos["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#102A43"),
        alignment=TA_LEFT,
        spaceAfter=5,
    )
    subtitulo = ParagraphStyle(
        "SubtituloBC",
        parent=estilos["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        textColor=colors.HexColor("#102A43"),
        spaceBefore=8,
        spaceAfter=6,
    )
    cuerpo = ParagraphStyle(
        "CuerpoBC",
        parent=estilos["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#182230"),
    )
    cuerpo_centrado = ParagraphStyle(
        "CuerpoCentradoBC",
        parent=cuerpo,
        alignment=1,
    )
    cuerpo_derecha = ParagraphStyle(
        "CuerpoDerechaBC",
        parent=cuerpo,
        alignment=TA_RIGHT,
    )
    portada_texto = ParagraphStyle(
        "PortadaTextoBC",
        parent=cuerpo,
        leading=20,
    )
    portada_texto_derecha = ParagraphStyle(
        "PortadaTextoDerechaBC",
        parent=portada_texto,
        alignment=TA_RIGHT,
    )
    fecha_estilo = ParagraphStyle(
        "FechaBC",
        parent=subtitulo,
        fontSize=10,
        leading=12,
        textColor=colors.white,
        backColor=colors.HexColor("#102A43"),
        borderPadding=(4, 6, 4, 6),
        spaceBefore=6,
        spaceAfter=0,
        keepWithNext=True,
    )

    documento = SimpleDocTemplate(
        str(ruta),
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=17 * mm,
        bottomMargin=15 * mm,
        title=f"Informe BC {informe['periodo']}",
        author="BC Inversiones EAS",
    )
    portada_izquierda = Paragraph(
        (
            '<font color="#8EC5FF" size="8"><b>BC INVERSIONES EAS  ·  '
            "RESUMEN MENSUAL</b></font><br/>"
            f'<font color="#FFFFFF" size="21"><b>'
            f"{escape(periodo_en_palabras(informe['periodo']))}"
            "</b></font><br/>"
            f'<font color="#C7D9EA" size="9">'
            f"{escape(informe['unidad'])}</font>"
        ),
        portada_texto,
    )
    color_resultado = "#61D8B2" if informe["utilidad"] >= 0 else "#FF9AA5"
    portada_derecha = Paragraph(
        (
            '<font color="#AFC7DC" size="8"><b>RESULTADO DEL MES</b></font>'
            "<br/>"
            f'<font color="{color_resultado}" size="17"><b>'
            f"{escape(formatear_monto(informe['utilidad']))}</b></font><br/>"
            f'<font color="#D7E5F1" size="9">'
            f"Margen {Movimientos.formatear_porcentaje(informe['margen'])}%"
            "</font>"
        ),
        portada_texto_derecha,
    )
    portada = Table(
        [[portada_izquierda, portada_derecha]],
        colWidths=[170 * mm, 86 * mm],
        rowHeights=[31 * mm],
        hAlign="LEFT",
    )
    portada.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#102A43")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (0, 0), 14),
                ("RIGHTPADDING", (1, 0), (1, 0), 14),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                (
                    "LINEBEFORE",
                    (1, 0),
                    (1, 0),
                    1,
                    colors.HexColor("#315D80"),
                ),
            ]
        )
    )
    elementos = [portada, Spacer(1, 4 * mm)]

    estilo_tabla = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F6FED")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("LEADING", (0, 0), (-1, -1), 10),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
                colors.white,
                colors.HexColor("#F4F7FB"),
            ]),
            (
                "LINEBELOW",
                (0, 0),
                (-1, -1),
                0.35,
                colors.HexColor("#DCE4EE"),
            ),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]
    )

    def tarjeta_pdf(nombre, valor, color):
        fondos = {
            "#12A67A": "#EAF8F3",
            "#E05A67": "#FFF0F2",
            "#2F6FED": "#EDF3FF",
            "#F09A3E": "#FFF5E8",
            "#7A67E8": "#F2EFFF",
            "#27A7D8": "#EAF8FC",
        }
        contenido = Paragraph(
            (
                f'<font color="#617084" size="7"><b>{escape(nombre.upper())}'
                f"</b></font><br/>"
                f'<font color="{color}" size="12"><b>{escape(valor)}</b></font>'
            ),
            cuerpo,
        )
        tabla = Table([[contenido]], colWidths=[62 * mm], rowHeights=[18 * mm])
        tabla.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, -1),
                        colors.HexColor(fondos.get(color, "#FFFFFF")),
                    ),
                    (
                        "BOX",
                        (0, 0),
                        (-1, -1),
                        0.8,
                        colors.HexColor(color),
                    ),
                    ("LEFTPADDING", (0, 0), (-1, -1), 7),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        return tabla

    fondo_texto = (
        formatear_monto(informe["fondo"])
        if informe["fondo"] is not None
        else "No aplica"
    )
    fila_kpi_1 = [
        tarjeta_pdf("Ingresos", formatear_monto(informe["ingresos"]), "#12A67A"),
        tarjeta_pdf(
            "Utilidad",
            formatear_monto(informe["utilidad"]),
            "#12A67A" if informe["utilidad"] >= 0 else "#E05A67",
        ),
        tarjeta_pdf(
            "Margen",
            Movimientos.formatear_porcentaje(informe["margen"]) + "%",
            "#2F6FED",
        ),
        tarjeta_pdf("Egresos", formatear_monto(informe["egresos"]), "#E05A67"),
    ]
    fila_kpi_2 = [
        tarjeta_pdf("Fondo de estabilidad", fondo_texto, "#F09A3E"),
        tarjeta_pdf(
            "Inversiones del negocio",
            formatear_monto(informe["inversiones"]),
            "#7A67E8",
        ),
        tarjeta_pdf(
            "Salida real de dinero",
            formatear_monto(informe["salida_caja"]),
            "#27A7D8",
        ),
        tarjeta_pdf(
            "Retenciones y descuentos",
            formatear_monto(informe["retenciones"]),
            "#F09A3E",
        ),
    ]
    tabla_kpi = Table(
        [fila_kpi_1, fila_kpi_2],
        colWidths=[64 * mm] * 4,
        rowHeights=[20 * mm, 20 * mm],
        hAlign="LEFT",
    )
    tabla_kpi.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 1),
                ("RIGHTPADDING", (0, 0), (-1, -1), 1),
                ("TOPPADDING", (0, 0), (-1, -1), 1),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ]
        )
    )
    def nombre_breve_categoria(nombre):
        texto = str(nombre)
        texto = texto.replace("Ingresos operativos · ", "Operativo ")
        texto = texto.replace("Egresos operativos · ", "Operativo ")
        texto = texto.replace("Adicional · ", "")
        return texto if len(texto) <= 18 else texto[:15] + "..."

    def crear_bloque_composicion_pdf(
        titulo_bloque,
        titulo_grafico,
        categorias,
        total,
        color_barra,
        color_borde,
        texto_sin_datos,
    ):
        categorias_principales = list(categorias[:5])
        if len(categorias) > 5:
            categorias_principales.append(
                (
                    "Otros",
                    sum(monto for _nombre, monto in categorias[5:]),
                )
            )

        dibujo = Drawing(108 * mm, 72 * mm)
        barras = HorizontalBarChart()
        barras.x = 35 * mm
        barras.y = 10 * mm
        barras.width = 68 * mm
        barras.height = 54 * mm
        valores_barras = [
            monto for _nombre, monto in categorias_principales
        ] or [0]
        barras.data = [valores_barras]
        barras.categoryAxis.categoryNames = [
            nombre_breve_categoria(nombre)
            for nombre, _monto in categorias_principales
        ] or [texto_sin_datos]
        barras.categoryAxis.labels.fontSize = 7
        barras.categoryAxis.labels.dx = -3
        barras.valueAxis.labels.fontSize = 6.5
        barras.valueAxis.labelTextFormat = (
            lambda valor: f"{valor / 1_000_000:.0f} M"
        )
        barras.valueAxis.valueMin = 0
        maximo_barra = max(barras.data[0]) if barras.data[0] else 0
        barras.valueAxis.valueMax = max(maximo_barra * 1.12, 1)
        barras.bars[0].fillColor = colors.HexColor(color_barra)
        barras.bars[0].strokeColor = colors.HexColor(color_borde)
        barras.barSpacing = 2
        barras.groupSpacing = 7
        dibujo.add(barras)
        dibujo.add(
            String(
                54 * mm,
                2 * mm,
                titulo_grafico,
                fontName="Helvetica-Bold",
                fontSize=9,
                textAnchor="middle",
                fillColor=colors.HexColor("#102A43"),
            )
        )

        leyenda_datos = [["Categoría", "Monto", "%"]]
        total_porcentaje = total or 1
        for nombre, monto in categorias_principales:
            leyenda_datos.append(
                [
                    Paragraph(escape(str(nombre)), cuerpo),
                    Paragraph(formatear_monto(monto), cuerpo_derecha),
                    Paragraph(
                        f"{monto / total_porcentaje:.1%}",
                        cuerpo_derecha,
                    ),
                ]
            )
        tabla_leyenda = Table(
            leyenda_datos,
            colWidths=[78 * mm, 43 * mm, 18 * mm],
            repeatRows=1,
        )
        tabla_leyenda.setStyle(estilo_tabla)
        tabla_leyenda.setStyle(
            TableStyle(
                [
                    ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        bloque_grafico = Table(
            [[dibujo, tabla_leyenda]],
            colWidths=[112 * mm, 143 * mm],
            hAlign="LEFT",
        )
        bloque_grafico.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ]
            )
        )
        return [
            Paragraph(titulo_bloque, subtitulo),
            bloque_grafico,
        ]

    lectura_mes = Table(
        [[
            Paragraph(
                (
                    f"<b>{'CIERRE POSITIVO' if informe['utilidad'] >= 0 else 'CIERRE A REVISAR'}</b>"
                    f"  ·  El resultado fue "
                    f"<b>{escape(formatear_monto(informe['utilidad']))}</b> "
                    f"con un margen de "
                    f"<b>{Movimientos.formatear_porcentaje(informe['margen'])}%</b>."
                ),
                cuerpo,
            )
        ]],
        colWidths=[256 * mm],
        rowHeights=[12 * mm],
        hAlign="LEFT",
    )
    lectura_mes.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, -1),
                    colors.HexColor(
                        "#EAF8F3" if informe["utilidad"] >= 0 else "#FFF0F2"
                    ),
                ),
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.8,
                    colors.HexColor(
                        "#12A67A" if informe["utilidad"] >= 0 else "#E05A67"
                    ),
                ),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    grafico_balance = Drawing(256 * mm, 40 * mm)
    grafico_balance.add(
        String(
            0,
            35 * mm,
            "Balance visual del mes",
            fontName="Helvetica-Bold",
            fontSize=10,
            fillColor=colors.HexColor("#102A43"),
        )
    )
    maximo_balance = max(
        informe["ingresos"],
        informe["egresos"],
        abs(informe["utilidad"]),
        1,
    )
    ancho_maximo = 185 * mm
    filas_balance = [
        ("Ingresos", informe["ingresos"], "#12A67A"),
        ("Egresos", informe["egresos"], "#E05A67"),
        (
            "Resultado",
            abs(informe["utilidad"]),
            "#2F6FED" if informe["utilidad"] >= 0 else "#E05A67",
        ),
    ]
    for indice, (etiqueta, valor, color) in enumerate(filas_balance):
        y = (25 - indice * 9) * mm
        grafico_balance.add(
            String(
                0,
                y + 1.2 * mm,
                etiqueta,
                fontName="Helvetica-Bold",
                fontSize=8,
                fillColor=colors.HexColor("#617084"),
            )
        )
        grafico_balance.add(
            Rect(
                27 * mm,
                y,
                ancho_maximo,
                5 * mm,
                rx=2 * mm,
                ry=2 * mm,
                fillColor=colors.HexColor("#E9EFF5"),
                strokeColor=None,
            )
        )
        grafico_balance.add(
            Rect(
                27 * mm,
                y,
                ancho_maximo * valor / maximo_balance,
                5 * mm,
                rx=2 * mm,
                ry=2 * mm,
                fillColor=colors.HexColor(color),
                strokeColor=None,
            )
        )
        grafico_balance.add(
            String(
                218 * mm,
                y + 1.2 * mm,
                formatear_monto(
                    -valor
                    if etiqueta == "Resultado" and informe["utilidad"] < 0
                    else valor
                ),
                fontName="Helvetica-Bold",
                fontSize=8,
                fillColor=colors.HexColor(color),
            )
        )

    elementos.extend(
        [
            tabla_kpi,
            Spacer(1, 4 * mm),
            lectura_mes,
            Spacer(1, 6 * mm),
            grafico_balance,
            PageBreak(),
        ]
    )
    elementos.extend(
        crear_bloque_composicion_pdf(
            "Dónde se generaron los ingresos",
            "Principales ingresos",
            informe["categorias_ingresos"],
            informe["ingresos"],
            "#12A67A",
            "#0E7B5A",
            "Sin ingresos",
        )
    )
    elementos.append(PageBreak())
    elementos.extend(
        crear_bloque_composicion_pdf(
            "Cómo se distribuyeron los egresos",
            "Principales egresos",
            informe["categorias_egresos"],
            informe["egresos"],
            "#E05A67",
            "#B83D4A",
            "Sin egresos",
        )
    )
    elementos.append(PageBreak())

    comparacion_datos = [[
        "Unidad",
        "Ingresos",
        "Egresos",
        "Resultado",
        "Transf. recibidas",
        "Transf. enviadas",
    ]]
    for unidad in informe["comparacion_unidades"]:
        comparacion_datos.append(
            [
                unidad["unidad"],
                Movimientos.formatear_monto(unidad["ingresos"]),
                Movimientos.formatear_monto(unidad["egresos"]),
                Movimientos.formatear_monto(unidad["resultado"]),
                Movimientos.formatear_monto(unidad["recibidas"]),
                Movimientos.formatear_monto(unidad["enviadas"]),
            ]
        )
    tabla_comparacion = Table(
        comparacion_datos,
        colWidths=[35 * mm, 38 * mm, 38 * mm, 38 * mm, 43 * mm, 43 * mm],
        repeatRows=1,
    )
    tabla_comparacion.setStyle(estilo_tabla)
    tabla_comparacion.setStyle(
        TableStyle([("ALIGN", (1, 1), (-1, -1), "RIGHT")])
    )
    elementos.extend(
        [
            Paragraph("Resumen por unidad y categorías", titulo),
            Paragraph("Comparación por unidad", subtitulo),
            tabla_comparacion,
            Spacer(1, 4 * mm),
        ]
    )

    def crear_tabla_categorias_pdf(categorias, total, tipo):
        categorias = list(categorias)
        punto_medio = (len(categorias) + 1) // 2
        bloque_izquierdo = categorias[:punto_medio]
        bloque_derecho = categorias[punto_medio:]
        categorias_datos = [[
            f"Categoría de {tipo}",
            "Monto",
            f"Categoría de {tipo}",
            "Monto",
        ]]
        for indice in range(punto_medio):
            izquierda = bloque_izquierdo[indice]
            derecha = (
                bloque_derecho[indice]
                if indice < len(bloque_derecho)
                else ("", 0)
            )
            categorias_datos.append(
                [
                    Paragraph(escape(str(izquierda[0])), cuerpo),
                    Paragraph(
                        formatear_monto(izquierda[1]),
                        cuerpo_derecha,
                    ),
                    (
                        Paragraph(escape(str(derecha[0])), cuerpo)
                        if derecha[0]
                        else ""
                    ),
                    (
                        Paragraph(
                            formatear_monto(derecha[1]),
                            cuerpo_derecha,
                        )
                        if derecha[0]
                        else ""
                    ),
                ]
            )
        categorias_datos.append(
            [
                Paragraph("<b>TOTAL</b>", cuerpo),
                Paragraph(
                    f"<b>{escape(formatear_monto(total))}</b>",
                    cuerpo_derecha,
                ),
                "",
                "",
            ]
        )
        tabla = Table(
            categorias_datos,
            colWidths=[84 * mm, 42 * mm, 84 * mm, 42 * mm],
            repeatRows=1,
        )
        tabla.setStyle(estilo_tabla)
        tabla.setStyle(
            TableStyle(
                [
                    ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                    ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    (
                        "BACKGROUND",
                        (0, -1),
                        (-1, -1),
                        colors.HexColor("#EAF0F7"),
                    ),
                ]
            )
        )
        return tabla

    elementos.extend(
        [
            Paragraph("Ingresos por categoría", subtitulo),
            crear_tabla_categorias_pdf(
                informe["categorias_ingresos"],
                informe["ingresos"],
                "ingreso",
            ),
            PageBreak(),
            Paragraph("Egresos por categoría", subtitulo),
            crear_tabla_categorias_pdf(
                informe["categorias_egresos"],
                informe["egresos"],
                "egreso",
            ),
            PageBreak(),
            Paragraph("Detalle por día", titulo),
            Paragraph(
            (
                "Cada fecha muestra una sola fila por categoría o unidad, "
                "con lo que entró, lo que salió y el resultado del día. "
                "Las transferencias y depósitos propios se muestran aparte "
                "porque no modifican la utilidad."
            ),
            cuerpo,
            ),
            Spacer(1, 2 * mm),
        ]
    )

    fecha_actual = None
    registros_dia = []

    def agregar_bloque_dia(fecha, registros):
        if not registros:
            return
        encabezado_fecha = Paragraph(escape(str(fecha)), fecha_estilo)
        datos = [[
            Paragraph("Categoría / unidad", cuerpo_centrado),
            Paragraph("Mov.", cuerpo_centrado),
            Paragraph("Entró", cuerpo_centrado),
            Paragraph("Salió", cuerpo_centrado),
            Paragraph("Resultado", cuerpo_centrado),
            Paragraph("Ruta interna / depósito", cuerpo_centrado),
            Paragraph("Monto movido", cuerpo_centrado),
        ]]
        total_entradas = 0
        total_salidas = 0
        total_interno = 0
        for registro in registros:
            datos.append(
                [
                    Paragraph(
                        f"<b>{escape(str(registro['grupo']))}</b>",
                        cuerpo,
                    ),
                    Paragraph(str(registro["cantidad"]), cuerpo_centrado),
                    Paragraph(
                        formatear_monto(registro["entradas"]),
                        cuerpo_derecha,
                    ),
                    Paragraph(
                        formatear_monto(registro["salidas"]),
                        cuerpo_derecha,
                    ),
                    Paragraph(
                        formatear_monto(registro["resultado"]),
                        cuerpo_derecha,
                    ),
                    Paragraph(
                        escape(str(registro["ruta_interna"])) or "-",
                        cuerpo,
                    ),
                    Paragraph(
                        (
                            formatear_monto(registro["monto_interno"])
                            if registro["monto_interno"]
                            else "-"
                        ),
                        cuerpo_derecha,
                    ),
                ]
            )
            total_entradas += registro["entradas"]
            total_salidas += registro["salidas"]
            total_interno += registro["monto_interno"]
        datos.append(
            [
                Paragraph("<b>TOTAL DEL DÍA</b>", cuerpo),
                "",
                Paragraph(
                    f"<b>{escape(formatear_monto(total_entradas))}</b>",
                    cuerpo_derecha,
                ),
                Paragraph(
                    f"<b>{escape(formatear_monto(total_salidas))}</b>",
                    cuerpo_derecha,
                ),
                Paragraph(
                    f"<b>{escape(formatear_monto(total_entradas - total_salidas))}</b>",
                    cuerpo_derecha,
                ),
                Paragraph("<b>Movido internamente</b>", cuerpo),
                Paragraph(
                    f"<b>{escape(formatear_monto(total_interno))}</b>",
                    cuerpo_derecha,
                ),
            ]
        )

        tabla = Table(
            datos,
            colWidths=[
                42 * mm,
                14 * mm,
                33 * mm,
                33 * mm,
                33 * mm,
                68 * mm,
                34 * mm,
            ],
            repeatRows=1,
        )
        tabla.setStyle(estilo_tabla)
        tabla.setStyle(
            TableStyle(
                [
                    ("ALIGN", (1, 1), (1, -1), "CENTER"),
                    ("ALIGN", (2, 1), (4, -1), "RIGHT"),
                    ("ALIGN", (6, 1), (6, -1), "RIGHT"),
                    ("TEXTCOLOR", (2, 1), (2, -2), colors.HexColor("#18A874")),
                    ("TEXTCOLOR", (3, 1), (3, -2), colors.HexColor("#E25555")),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    (
                        "BACKGROUND",
                        (0, -1),
                        (-1, -1),
                        colors.HexColor("#EAF0F7"),
                    ),
                ]
            )
        )
        elementos.append(
            KeepTogether(
                [
                    encabezado_fecha,
                    tabla,
                    Spacer(1, 2 * mm),
                ]
            )
        )

    for registro in informe["resumen_diario"]:
        if fecha_actual is None:
            fecha_actual = registro["fecha"]
        if registro["fecha"] != fecha_actual:
            agregar_bloque_dia(fecha_actual, registros_dia)
            fecha_actual = registro["fecha"]
            registros_dia = []
        registros_dia.append(registro)
    agregar_bloque_dia(fecha_actual, registros_dia)

    def encabezado_pie(canvas, documento_actual):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#617084"))
        canvas.drawString(
            documento_actual.leftMargin,
            landscape(A4)[1] - 10 * mm,
            f"BC Gestión - Informe {informe['periodo']}",
        )
        canvas.drawRightString(
            landscape(A4)[0] - documento_actual.rightMargin,
            8 * mm,
            f"Página {documento_actual.page}",
        )
        canvas.restoreState()

    documento.build(
        elementos,
        onFirstPage=encabezado_pie,
        onLaterPages=encabezado_pie,
    )
    return ruta


def limites_periodo(periodo):
    """Convierte MM-AAAA en los límites completos del mes."""
    try:
        fecha = datetime.strptime(periodo.strip(), "%m-%Y")
    except ValueError as error:
        raise ValueError(
            "El período no es válido. Usá el formato MM-AAAA."
        ) from error

    ultimo_dia = monthrange(fecha.year, fecha.month)[1]
    return (
        datetime(fecha.year, fecha.month, 1),
        datetime(fecha.year, fecha.month, ultimo_dia),
    )


def formatear_monto(valor):
    return "Gs. " + Movimientos.formatear_monto(valor)


def fecha_en_rango(fecha_texto, fecha_desde, fecha_hasta):
    try:
        fecha = Movimientos.convertir_fecha(fecha_texto)
    except ValueError:
        return None

    if fecha_desde <= fecha <= fecha_hasta:
        return fecha
    return None


def agregar_categoria(diccionario, concepto, monto):
    if monto:
        diccionario[concepto] = diccionario.get(concepto, 0) + monto


def detalle_movimientos(
    fecha_desde,
    fecha_hasta,
    unidad_seleccionada,
    categorias_egresos,
    categorias_ingresos,
):
    """Obtiene movimientos externos e internos del período."""
    detalles = []
    todas = unidad_seleccionada == "Empresa completa"
    unidad_normalizada = Movimientos.normalizar_texto(
        unidad_seleccionada
    )

    for linea in leer_datos(Movimientos.RUTA_MOVIMIENTOS):
        datos = Movimientos.separar_movimiento(linea)
        if datos is None:
            continue

        fecha = fecha_en_rango(
            datos["fecha"],
            fecha_desde,
            fecha_hasta,
        )
        if fecha is None:
            continue

        tipo = datos["tipo"]
        origen = Movimientos.normalizar_texto(datos["origen"])
        destino = Movimientos.normalizar_texto(datos["destino"])
        monto = datos["monto"]

        if tipo in ["Ingreso", "Cobro externo"]:
            if not todas and destino != unidad_normalizada:
                continue

            unidad = datos["destino"]
            agregar_categoria(
                categorias_ingresos,
                f"Ingresos operativos · {unidad}",
                monto,
            )
            detalle = (
                "Ingreso externo"
                if tipo == "Ingreso"
                else (
                    "Depósito externo recibido en "
                    f"{datos['origen']}"
                )
            )
            detalles.append(
                {
                    "orden": fecha,
                    "fecha": datos["fecha"],
                    "tipo": "Ingreso",
                    "categoria": "Operación diaria",
                    "detalle": detalle,
                    "unidad": unidad,
                    "monto": monto,
                }
            )

        elif tipo == "Egreso":
            if not todas and origen != unidad_normalizada:
                continue

            unidad = datos["origen"]
            agregar_categoria(
                categorias_egresos,
                f"Egresos operativos · {unidad}",
                monto,
            )
            detalles.append(
                {
                    "orden": fecha,
                    "fecha": datos["fecha"],
                    "tipo": "Egreso",
                    "categoria": "Operación diaria",
                    "detalle": "Egreso externo",
                    "unidad": unidad,
                    "monto": monto,
                }
            )

        elif tipo in ["Transferencia interna", "Deposito interno"]:
            if (
                not todas
                and origen != unidad_normalizada
                and destino != unidad_normalizada
            ):
                continue

            es_deposito = tipo == "Deposito interno"
            detalles.append(
                {
                    "orden": fecha,
                    "fecha": datos["fecha"],
                    "tipo": tipo,
                    "categoria": (
                        "Depósito bancario"
                        if es_deposito
                        else "Transferencia interna"
                    ),
                    "detalle": (
                        f"{datos['origen']} → {datos['destino']}"
                    ),
                    "unidad": datos["origen"],
                    "monto": monto,
                }
            )

    return detalles


def detalle_componentes_generales(
    fecha_desde,
    fecha_hasta,
    categorias_egresos,
    categorias_ingresos,
):
    """Detalla conceptos generales que no tienen una unidad asignada."""
    detalles = []

    for linea in leer_datos(Movimientos.RUTA_ADICIONALES):
        datos = Movimientos.separar_adicional(linea)
        if datos is None:
            continue

        fecha = fecha_en_rango(
            datos["fecha"],
            fecha_desde,
            fecha_hasta,
        )
        if fecha is None:
            continue

        categoria = (
            categorias_ingresos
            if datos["tipo"] == "Ingreso"
            else categorias_egresos
        )
        agregar_categoria(
            categoria,
            f"Adicional · {datos['descripcion']}",
            datos["monto"],
        )
        observacion = datos.get("observacion", "").strip()
        detalles.append(
            {
                "orden": fecha,
                "fecha": datos["fecha"],
                "tipo": datos["tipo"],
                "categoria": "Adicional",
                "detalle": (
                    datos["descripcion"]
                    if observacion in ["", "-"]
                    else f"{datos['descripcion']} · {observacion}"
                ),
                "unidad": "General",
                "monto": datos["monto"],
            }
        )

    total_cuotas, detalle_cuotas = Movimientos.resumen_cuotas(
        fecha_desde,
        fecha_hasta,
    )
    agregar_categoria(
        categorias_egresos,
        "Cuotas de préstamos",
        total_cuotas,
    )
    for nombre, cuota in detalle_cuotas:
        fecha = fecha_en_rango(
            cuota["fecha"],
            fecha_desde,
            fecha_hasta,
        )
        if fecha is None:
            continue
        nombre_visible = nombre
        referencia = cuota.get("prestamo_id", "")
        if (
            nombre == "Préstamo"
            and referencia.upper().startswith("HIST-")
        ):
            nombre_visible = (
                referencia[5:].replace("-", " ").title()
            )
        detalles.append(
            {
                "orden": fecha,
                "fecha": cuota["fecha"],
                "tipo": "Egreso",
                "categoria": "Préstamo",
                "detalle": (
                    f"{nombre_visible} · Cuota {cuota['numero']}"
                ),
                "unidad": "General",
                "monto": cuota["monto"],
            }
        )

    nomina, detalle_nomina = Movimientos.resumen_nomina_liquidada(
        fecha_desde,
        fecha_hasta,
    )
    agregar_categoria(
        categorias_egresos,
        "Nómina",
        nomina["egreso_planilla"],
    )
    for liquidacion in detalle_nomina:
        detalles.append(
            {
                "orden": fecha_hasta,
                "fecha": liquidacion["periodo"],
                "tipo": "Egreso",
                "categoria": "Nómina",
                "detalle": liquidacion["nombre"],
                "unidad": "General",
                "monto": liquidacion["egreso_planilla"],
            }
        )

    return detalles


def obtener_comparacion_unidades(fecha_desde, fecha_hasta):
    comparacion = []
    for unidad in Movimientos.UNIDADES:
        (
            ingresos,
            egresos,
            resultado,
            recibidas,
            enviadas,
        ) = Movimientos.calcular_resumen_unidad(
            unidad,
            fecha_desde,
            fecha_hasta,
        )
        comparacion.append(
            {
                "unidad": unidad,
                "ingresos": ingresos,
                "egresos": egresos,
                "resultado": resultado,
                "recibidas": recibidas,
                "enviadas": enviadas,
            }
        )
    return comparacion


def obtener_informe(periodo, unidad_seleccionada="Empresa completa"):
    """Construye un informe mensual verificable y compatible con el cierre."""
    fecha_desde, fecha_hasta = limites_periodo(periodo)
    periodo_normalizado = fecha_desde.strftime("%m-%Y")
    categorias_egresos = {}
    categorias_ingresos = {}

    detalles = detalle_movimientos(
        fecha_desde,
        fecha_hasta,
        unidad_seleccionada,
        categorias_egresos,
        categorias_ingresos,
    )
    empresa_completa = unidad_seleccionada == "Empresa completa"

    if empresa_completa:
        indicadores = Movimientos.calcular_indicadores_cierre(
            fecha_desde,
            fecha_hasta,
        )
        detalles.extend(
            detalle_componentes_generales(
                fecha_desde,
                fecha_hasta,
                categorias_egresos,
                categorias_ingresos,
            )
        )
        ingresos = indicadores["ingresos"]
        egresos = indicadores["egresos"]
        utilidad = indicadores["utilidad_mes"]
        margen = indicadores["margen_porcentual"]
        salida_caja = indicadores["salida_caja_total"]
        retenciones = indicadores["diferencia_egreso_caja"]
        registro_fondo = Movimientos.obtener_registro_fondo(
            periodo_normalizado
        )
        fondo = (
            registro_fondo[1]["monto_aplicado"]
            if registro_fondo is not None
            else indicadores["fondo_calculado"]
        )
        fondo_estado = (
            registro_fondo[1]["modo"]
            if registro_fondo is not None
            else "CALCULADO"
        )
    else:
        (
            ingresos,
            egresos,
            utilidad,
            _recibidas,
            _enviadas,
        ) = Movimientos.calcular_resumen_unidad(
            unidad_seleccionada,
            fecha_desde,
            fecha_hasta,
        )
        margen = utilidad / ingresos * 100 if ingresos else 0
        salida_caja = egresos
        retenciones = 0
        fondo = None
        fondo_estado = "NO APLICA"

    total_categorias_ingreso = sum(categorias_ingresos.values())
    total_categorias_egreso = sum(categorias_egresos.values())
    if total_categorias_ingreso != ingresos:
        raise ValueError(
            "No se pudo conciliar el detalle de ingresos con el total."
        )
    if total_categorias_egreso != egresos:
        raise ValueError(
            "No se pudo conciliar el detalle de egresos con el total."
        )

    detalles.sort(
        key=lambda item: (
            item["orden"],
            item["tipo"],
            item["detalle"],
        ),
        reverse=True,
    )
    resumen_diario = resumir_detalles_por_dia(detalles)
    total_entradas_diarias = sum(
        fila["entradas"] for fila in resumen_diario
    )
    total_salidas_diarias = sum(
        fila["salidas"] for fila in resumen_diario
    )
    if total_entradas_diarias != ingresos:
        raise ValueError(
            "No se pudo conciliar el resumen diario de ingresos con el total."
        )
    if total_salidas_diarias != egresos:
        raise ValueError(
            "No se pudo conciliar el resumen diario de egresos con el total."
        )

    return {
        "periodo": periodo_normalizado,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "unidad": unidad_seleccionada,
        "empresa_completa": empresa_completa,
        "ingresos": ingresos,
        "egresos": egresos,
        "utilidad": utilidad,
        "margen": margen,
        "fondo": fondo,
        "fondo_estado": fondo_estado,
        "salida_caja": salida_caja,
        "retenciones": retenciones,
        "inversiones": Movimientos.resumen_inversiones(
            fecha_desde,
            fecha_hasta,
        ),
        "categorias_egresos": sorted(
            categorias_egresos.items(),
            key=lambda item: item[1],
            reverse=True,
        ),
        "categorias_ingresos": sorted(
            categorias_ingresos.items(),
            key=lambda item: item[1],
            reverse=True,
        ),
        "comparacion_unidades": obtener_comparacion_unidades(
            fecha_desde,
            fecha_hasta,
        ),
        "resumen_diario": resumen_diario,
        "detalles": detalles,
    }


def obtener_periodos_comparacion(periodo_desde, periodo_hasta):
    """Devuelve todos los meses de un rango, con un máximo razonable."""
    fecha_desde, _ = limites_periodo(periodo_desde)
    fecha_hasta, _ = limites_periodo(periodo_hasta)
    fecha_desde = fecha_desde.replace(day=1)
    fecha_hasta = fecha_hasta.replace(day=1)

    if fecha_desde > fecha_hasta:
        raise ValueError(
            "El período inicial no puede ser posterior al período final."
        )

    periodos = []
    actual = fecha_desde
    while actual <= fecha_hasta:
        periodos.append(actual.strftime("%m-%Y"))
        if len(periodos) > 24:
            raise ValueError(
                "La comparación admite como máximo 24 meses."
            )
        if actual.month == 12:
            actual = actual.replace(
                year=actual.year + 1,
                month=1,
            )
        else:
            actual = actual.replace(month=actual.month + 1)
    return periodos


def calcular_variacion_porcentual(actual, anterior):
    if anterior in (None, 0):
        return None
    return (actual - anterior) / abs(anterior) * 100


def obtener_comparacion_mensual(
    periodo_desde,
    periodo_hasta,
    unidad_seleccionada="Empresa completa",
):
    """Construye una comparación conciliada usando cada informe mensual."""
    periodos = obtener_periodos_comparacion(
        periodo_desde,
        periodo_hasta,
    )
    filas = []

    for periodo in periodos:
        informe = obtener_informe(periodo, unidad_seleccionada)
        tiene_datos = bool(informe["detalles"])
        fila = {
            "periodo": periodo,
            "tiene_datos": tiene_datos,
            "estado": "Con datos" if tiene_datos else "Sin datos",
            "ingresos": informe["ingresos"] if tiene_datos else None,
            "egresos": informe["egresos"] if tiene_datos else None,
            "utilidad": informe["utilidad"] if tiene_datos else None,
            "margen": informe["margen"] if tiene_datos else None,
            "fondo": informe["fondo"] if tiene_datos else None,
            "variacion_ingresos": None,
            "variacion_egresos": None,
            "variacion_utilidad": None,
        }
        filas.append(fila)

    for indice, fila in enumerate(filas):
        if indice == 0 or not fila["tiene_datos"]:
            continue
        anterior = filas[indice - 1]
        if not anterior["tiene_datos"]:
            continue
        fila["variacion_ingresos"] = calcular_variacion_porcentual(
            fila["ingresos"],
            anterior["ingresos"],
        )
        fila["variacion_egresos"] = calcular_variacion_porcentual(
            fila["egresos"],
            anterior["egresos"],
        )
        fila["variacion_utilidad"] = calcular_variacion_porcentual(
            fila["utilidad"],
            anterior["utilidad"],
        )

    filas_con_datos = [fila for fila in filas if fila["tiene_datos"]]
    total_ingresos = sum(fila["ingresos"] for fila in filas_con_datos)
    total_egresos = sum(fila["egresos"] for fila in filas_con_datos)
    total_utilidad = total_ingresos - total_egresos
    margen_total = (
        total_utilidad / total_ingresos * 100
        if total_ingresos
        else 0
    )
    fondos = [
        fila["fondo"]
        for fila in filas_con_datos
        if fila["fondo"] is not None
    ]
    mejor_mes = (
        max(filas_con_datos, key=lambda fila: fila["utilidad"])
        if filas_con_datos
        else None
    )
    peor_mes = (
        min(filas_con_datos, key=lambda fila: fila["utilidad"])
        if filas_con_datos
        else None
    )

    return {
        "tipo": "comparacion_mensual",
        "periodo_desde": periodos[0],
        "periodo_hasta": periodos[-1],
        "unidad": unidad_seleccionada,
        "filas": filas,
        "meses_solicitados": len(filas),
        "meses_con_datos": len(filas_con_datos),
        "total_ingresos": total_ingresos,
        "total_egresos": total_egresos,
        "total_utilidad": total_utilidad,
        "margen_total": margen_total,
        "total_fondo": sum(fondos) if fondos else None,
        "mejor_mes": mejor_mes,
        "peor_mes": peor_mes,
    }


def obtener_informe_anual(anio, unidad_seleccionada="Empresa completa"):
    """Construye el cuadro completo de enero a diciembre de un año."""
    texto_anio = str(anio).strip()
    if not re.fullmatch(r"\d{4}", texto_anio):
        raise ValueError("Ingresá el año con cuatro números. Ejemplo: 2026.")

    numero_anio = int(texto_anio)
    if numero_anio < 2000 or numero_anio > 2100:
        raise ValueError("El año debe estar comprendido entre 2000 y 2100.")

    informe = obtener_comparacion_mensual(
        f"01-{numero_anio}",
        f"12-{numero_anio}",
        unidad_seleccionada,
    )
    informe["tipo"] = "informe_anual"
    informe["anio"] = numero_anio

    for indice, fila in enumerate(informe["filas"]):
        fila["mes"] = MESES_ESPANOL[indice]

    return informe


def exportar_excel_anual(informe, ruta):
    """Exporta el cuadro anual en una hoja clara y lista para imprimir."""
    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ModuleNotFoundError as error:
        raise RuntimeError(
            "Falta instalar openpyxl. Cerrá el sistema y volvé a abrir "
            "iniciar_interfaz.bat para completar la instalación."
        ) from error

    ruta = Path(ruta)
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Informe anual"
    datos_grafica = libro.create_sheet("_Grafica anual")
    hoja.sheet_view.showGridLines = False

    azul = "246BFD"
    azul_oscuro = "173B73"
    gris_claro = "F6F8FB"
    blanco = "FFFFFF"
    borde = Side(style="thin", color="DCE4EE")
    formato_gs = '"Gs. "#,##0;[Red]-"Gs. "#,##0;"-"'
    formato_porcentaje = '0.00%;[Red]-0.00%;"-"'

    hoja.merge_cells("A1:E1")
    hoja["A1"] = f"BC Inversiones EAS - Informe anual {informe['anio']}"
    hoja["A1"].font = Font(
        name="Aptos Display",
        size=17,
        bold=True,
        color=blanco,
    )
    hoja["A1"].fill = PatternFill("solid", fgColor=azul_oscuro)
    hoja["A1"].alignment = Alignment(
        horizontal="left",
        vertical="center",
    )
    hoja.row_dimensions[1].height = 32

    hoja.merge_cells("A2:E2")
    hoja["A2"] = f"Vista: {informe['unidad']}"
    hoja["A2"].font = Font(bold=True, color=azul_oscuro)
    hoja["A2"].alignment = Alignment(horizontal="left")

    encabezados = ["Mes", "Ingresos", "Egresos", "Utilidad", "Margen"]
    fila_encabezado = 4
    for columna, titulo in enumerate(encabezados, 1):
        celda = hoja.cell(fila_encabezado, columna, titulo)
        celda.fill = PatternFill("solid", fgColor=azul)
        celda.font = Font(bold=True, color=blanco)
        celda.alignment = Alignment(horizontal="center")
        celda.border = Border(bottom=borde)

    fila_inicial = fila_encabezado + 1
    for indice, fila in enumerate(informe["filas"], fila_inicial):
        hoja.cell(indice, 1, fila["mes"])
        if fila["tiene_datos"]:
            hoja.cell(indice, 2, fila["ingresos"])
            hoja.cell(indice, 3, fila["egresos"])
            hoja.cell(indice, 4, f"=B{indice}-C{indice}")
            hoja.cell(indice, 5, f'=IFERROR(D{indice}/B{indice},0)')
        else:
            hoja.cell(indice, 2, None)
            hoja.cell(indice, 3, None)
            hoja.cell(indice, 4, None)
            hoja.cell(indice, 5, None)

        fondo = gris_claro if indice % 2 else blanco
        for columna in range(1, 6):
            celda = hoja.cell(indice, columna)
            celda.fill = PatternFill("solid", fgColor=fondo)
            celda.border = Border(bottom=borde)
        hoja.cell(indice, 1).alignment = Alignment(horizontal="left")
        for columna in range(2, 6):
            hoja.cell(indice, columna).alignment = Alignment(
                horizontal="right"
            )

    fila_total = fila_inicial + 12
    hoja.cell(fila_total, 1, "TOTAL DEL AÑO")
    hoja.cell(
        fila_total,
        2,
        f"=SUM(B{fila_inicial}:B{fila_total - 1})",
    )
    hoja.cell(
        fila_total,
        3,
        f"=SUM(C{fila_inicial}:C{fila_total - 1})",
    )
    hoja.cell(fila_total, 4, f"=B{fila_total}-C{fila_total}")
    hoja.cell(
        fila_total,
        5,
        f'=IFERROR(D{fila_total}/B{fila_total},0)',
    )
    for columna in range(1, 6):
        celda = hoja.cell(fila_total, columna)
        celda.fill = PatternFill("solid", fgColor="DCE8FF")
        celda.font = Font(bold=True, color=azul_oscuro)
        celda.border = Border(
            top=Side(style="medium", color=azul),
            bottom=Side(style="medium", color=azul),
        )

    for fila in range(fila_inicial, fila_total + 1):
        for columna in range(2, 5):
            hoja.cell(fila, columna).number_format = formato_gs
        hoja.cell(fila, 5).number_format = formato_porcentaje

    hoja.column_dimensions["A"].width = 20
    for columna in ("B", "C", "D"):
        hoja.column_dimensions[columna].width = 22
    hoja.column_dimensions["E"].width = 15
    hoja.freeze_panes = "A5"

    grafica = LineChart()
    grafica.title = f"Evolución mensual {informe['anio']} (Gs.)"
    grafica.y_axis.title = "Guaraníes"
    grafica.x_axis.title = "Mes"
    grafica.height = 7
    grafica.width = 11
    grafica.style = 13

    datos_grafica.append(["Mes", "Ingresos", "Egresos", "Utilidad"])
    for fila in informe["filas"]:
        if not fila["tiene_datos"]:
            continue
        datos_grafica.append(
            [
                fila["mes"],
                fila["ingresos"],
                fila["egresos"],
                fila["utilidad"],
            ]
        )
    ultima_fila_grafica = max(datos_grafica.max_row, 2)
    categorias = Reference(
        datos_grafica,
        min_col=1,
        min_row=2,
        max_row=ultima_fila_grafica,
    )
    valores = Reference(
        datos_grafica,
        min_col=2,
        max_col=4,
        min_row=1,
        max_row=ultima_fila_grafica,
    )
    grafica.add_data(valores, titles_from_data=True)
    grafica.set_categories(categorias)
    grafica.display_blanks = "gap"
    for serie, nombre in zip(
        grafica.series,
        ("Ingresos", "Egresos", "Utilidad"),
    ):
        serie.tx = SeriesLabel(v=nombre)
    hoja.add_chart(grafica, "G4")
    datos_grafica.sheet_state = "hidden"

    hoja.page_setup.orientation = "landscape"
    hoja.page_setup.paperSize = hoja.PAPERSIZE_A4
    hoja.page_setup.fitToWidth = 1
    hoja.page_setup.fitToHeight = 1
    hoja.sheet_properties.pageSetUpPr.fitToPage = True
    hoja.print_area = f"A1:N{fila_total}"

    try:
        libro.calculation.fullCalcOnLoad = True
        libro.calculation.forceFullCalc = True
    except AttributeError:
        pass
    libro.save(ruta)
    return ruta


def exportar_pdf_anual(informe, ruta):
    """Exporta el cuadro anual completo en una sola página horizontal."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
        from xml.sax.saxutils import escape
    except ModuleNotFoundError as error:
        raise RuntimeError(
            "Falta instalar reportlab. Cerrá el sistema y volvé a abrir "
            "iniciar_interfaz.bat para completar la instalación."
        ) from error

    ruta = Path(ruta)
    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        "TituloAnual",
        parent=estilos["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#173B73"),
        alignment=TA_LEFT,
    )
    cuerpo = ParagraphStyle(
        "CuerpoAnual",
        parent=estilos["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#182230"),
    )
    documento = SimpleDocTemplate(
        str(ruta),
        pagesize=landscape(A4),
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=17 * mm,
        bottomMargin=15 * mm,
        title=f"Informe anual BC {informe['anio']}",
        author="BC Inversiones EAS",
    )

    elementos = [
        Paragraph(
            f"BC Inversiones EAS - Informe anual {informe['anio']}",
            titulo,
        ),
        Paragraph(
            f"Vista: {escape(informe['unidad'])}",
            cuerpo,
        ),
        Spacer(1, 4 * mm),
    ]

    tarjetas = [
        ("INGRESOS", formatear_monto(informe["total_ingresos"]), "#18A874"),
        ("EGRESOS", formatear_monto(informe["total_egresos"]), "#E25555"),
        ("UTILIDAD", formatear_monto(informe["total_utilidad"]), "#173B73"),
        (
            "MARGEN",
            Movimientos.formatear_porcentaje(informe["margen_total"]) + "%",
            "#246BFD",
        ),
    ]
    celdas = []
    for nombre, valor, color in tarjetas:
        celdas.append(
            Paragraph(
                (
                    f'<font color="#617084" size="7"><b>{nombre}</b></font>'
                    f'<br/><font color="{color}" size="12"><b>'
                    f"{escape(valor)}</b></font>"
                ),
                cuerpo,
            )
        )
    tabla_tarjetas = Table(
        [celdas],
        colWidths=[64 * mm] * 4,
        rowHeights=[18 * mm],
    )
    tabla_tarjetas.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#DCE4EE")),
                ("INNERGRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#DCE4EE")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    elementos.append(tabla_tarjetas)
    elementos.append(Spacer(1, 5 * mm))

    filas_tabla = [["Mes", "Ingresos", "Egresos", "Utilidad", "Margen"]]
    for fila in informe["filas"]:
        if fila["tiene_datos"]:
            filas_tabla.append(
                [
                    fila["mes"],
                    formatear_monto(fila["ingresos"]),
                    formatear_monto(fila["egresos"]),
                    formatear_monto(fila["utilidad"]),
                    (
                        Movimientos.formatear_porcentaje(
                            fila["margen"]
                        ) + "%"
                    ),
                ]
            )
        else:
            filas_tabla.append([fila["mes"], "-", "-", "-", "-"])

    filas_tabla.append(
        [
            "TOTAL DEL AÑO",
            formatear_monto(informe["total_ingresos"]),
            formatear_monto(informe["total_egresos"]),
            formatear_monto(informe["total_utilidad"]),
            (
                Movimientos.formatear_porcentaje(
                    informe["margen_total"]
                ) + "%"
            ),
        ]
    )
    tabla = Table(
        filas_tabla,
        repeatRows=1,
        colWidths=[44 * mm, 53 * mm, 53 * mm, 53 * mm, 32 * mm],
    )
    tabla.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#246BFD")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#DCE8FF")),
                ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor("#173B73")),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [
                    colors.white,
                    colors.HexColor("#F6F8FB"),
                ]),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#DCE4EE")),
                ("LINEABOVE", (0, -1), (-1, -1), 1.2, colors.HexColor("#246BFD")),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    elementos.append(tabla)
    elementos.append(Spacer(1, 3 * mm))
    elementos.append(
        Paragraph(
            (
                "Los meses sin registros se muestran con guiones. El margen "
                "anual se calcula sobre los totales del año."
            ),
            cuerpo,
        )
    )

    def encabezado_pie(canvas, documento_actual):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#617084"))
        canvas.drawString(
            documento_actual.leftMargin,
            landscape(A4)[1] - 10 * mm,
            f"BC Gestión - Informe anual {informe['anio']}",
        )
        canvas.drawRightString(
            landscape(A4)[0] - documento_actual.rightMargin,
            8 * mm,
            f"Página {documento_actual.page}",
        )
        canvas.restoreState()

    documento.build(
        elementos,
        onFirstPage=encabezado_pie,
        onLaterPages=encabezado_pie,
    )
    return ruta


def exportar_excel_comparacion(comparacion, ruta):
    """Exporta la comparación con tablero, gráfica y datos filtrables."""
    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ModuleNotFoundError as error:
        raise RuntimeError(
            "Falta instalar openpyxl. Cerrá el sistema y volvé a abrir "
            "iniciar_interfaz.bat para completar la instalación."
        ) from error

    ruta = Path(ruta)
    libro = Workbook()
    resumen = libro.active
    resumen.title = "Resumen"
    datos = libro.create_sheet("Datos mensuales")
    datos_grafica = libro.create_sheet("_Grafica")

    azul = "246BFD"
    azul_oscuro = "173B73"
    verde = "18A874"
    rojo = "E25555"
    naranja = "E99A35"
    gris = "EAF0F7"
    gris_claro = "F6F8FB"
    blanco = "FFFFFF"
    borde = Side(style="thin", color="DCE4EE")
    formato_gs = '"Gs. "#,##0;[Red]-"Gs. "#,##0;"-"'
    formato_porcentaje = '0.00%;[Red]-0.00%;"-"'

    resumen.sheet_view.showGridLines = False
    resumen.merge_cells("A1:J1")
    resumen["A1"] = "BC Inversiones EAS - Comparación mensual"
    resumen["A1"].font = Font(
        name="Aptos Display",
        size=17,
        bold=True,
        color=blanco,
    )
    resumen["A1"].fill = PatternFill("solid", fgColor=azul_oscuro)
    resumen["A1"].alignment = Alignment(vertical="center")
    resumen.row_dimensions[1].height = 32
    resumen.merge_cells("A2:J2")
    resumen["A2"] = (
        f"{comparacion['periodo_desde']} a "
        f"{comparacion['periodo_hasta']} · {comparacion['unidad']}"
    )
    resumen["A2"].font = Font(bold=True, color=azul_oscuro)

    ultima_fila_datos = len(comparacion["filas"]) + 1
    tarjetas = [
        (
            "Ingresos acumulados",
            f"=SUM('Datos mensuales'!C2:C{ultima_fila_datos})",
            verde,
        ),
        (
            "Egresos acumulados",
            f"=SUM('Datos mensuales'!D2:D{ultima_fila_datos})",
            rojo,
        ),
        (
            "Utilidad acumulada",
            f"=SUM('Datos mensuales'!E2:E{ultima_fila_datos})",
            verde,
        ),
        ("Margen total", "=IFERROR(E5/A5,0)", azul),
        (
            "Fondo acumulado",
            f"=SUM('Datos mensuales'!G2:G{ultima_fila_datos})",
            naranja,
        ),
    ]
    for indice, (titulo, formula, color) in enumerate(tarjetas):
        columna = indice * 2 + 1
        resumen.merge_cells(
            start_row=4,
            start_column=columna,
            end_row=4,
            end_column=columna + 1,
        )
        resumen.merge_cells(
            start_row=5,
            start_column=columna,
            end_row=6,
            end_column=columna + 1,
        )
        celda_titulo = resumen.cell(4, columna, titulo)
        celda_valor = resumen.cell(5, columna)
        celda_valor.value = formula
        for fila in (4, 5, 6):
            for col in (columna, columna + 1):
                resumen.cell(fila, col).fill = PatternFill(
                    "solid",
                    fgColor=gris_claro,
                )
                resumen.cell(fila, col).border = Border(
                    left=borde,
                    right=borde,
                    top=borde,
                    bottom=borde,
                )
        celda_titulo.font = Font(bold=True, color=color)
        celda_valor.font = Font(size=14, bold=True, color=azul_oscuro)
        celda_valor.alignment = Alignment(
            vertical="center",
            horizontal="center",
        )
        celda_titulo.alignment = Alignment(horizontal="center")
        if titulo == "Margen total":
            celda_valor.number_format = formato_porcentaje
        else:
            celda_valor.number_format = formato_gs

    resumen["A8"] = "Meses con datos"
    resumen["B8"] = (
        f'=COUNTIF(\'Datos mensuales\'!B2:B{ultima_fila_datos},'
        f'"Con datos")&" de {comparacion["meses_solicitados"]}"'
    )
    resumen["A9"] = "Mejor mes"
    resumen["B9"] = (
        f'=IFERROR(INDEX(\'Datos mensuales\'!A2:A{ultima_fila_datos},'
        f'MATCH(MAX(\'Datos mensuales\'!E2:E{ultima_fila_datos}),'
        f'\'Datos mensuales\'!E2:E{ultima_fila_datos},0)),"Sin datos")'
    )
    resumen["A10"] = "Menor resultado"
    resumen["B10"] = (
        f'=IFERROR(INDEX(\'Datos mensuales\'!A2:A{ultima_fila_datos},'
        f'MATCH(MIN(\'Datos mensuales\'!E2:E{ultima_fila_datos}),'
        f'\'Datos mensuales\'!E2:E{ultima_fila_datos},0)),"Sin datos")'
    )
    for fila in range(8, 11):
        resumen.cell(fila, 1).font = Font(bold=True, color=azul_oscuro)
        resumen.cell(fila, 1).fill = PatternFill("solid", fgColor=gris)
        resumen.cell(fila, 2).fill = PatternFill("solid", fgColor=gris_claro)

    encabezados = [
        "Período",
        "Estado",
        "Ingresos",
        "Egresos",
        "Utilidad",
        "Margen",
        "Fondo",
        "Var. ingresos",
        "Var. egresos",
        "Var. utilidad",
    ]
    datos.append(encabezados)
    for celda in datos[1]:
        celda.fill = PatternFill("solid", fgColor=azul)
        celda.font = Font(bold=True, color=blanco)
        celda.alignment = Alignment(horizontal="center")

    for numero_fila, fila in enumerate(comparacion["filas"], 2):
        if fila["tiene_datos"]:
            valores = [
                fila["periodo"],
                fila["estado"],
                fila["ingresos"],
                fila["egresos"],
                None,
                None,
                fila["fondo"],
                None,
                None,
                None,
            ]
        else:
            valores = [fila["periodo"], fila["estado"]] + [None] * 8
        datos.append(valores)
        if fila["tiene_datos"]:
            datos.cell(numero_fila, 5).value = (
                f"=C{numero_fila}-D{numero_fila}"
            )
            datos.cell(numero_fila, 6).value = (
                f'=IFERROR(E{numero_fila}/C{numero_fila},0)'
            )
            if numero_fila > 2:
                fila_anterior = numero_fila - 1
                formulas_variacion = {
                    8: "C",
                    9: "D",
                    10: "E",
                }
                for columna_destino, letra_origen in (
                    formulas_variacion.items()
                ):
                    datos.cell(
                        numero_fila,
                        columna_destino,
                    ).value = (
                        f'=IF(AND($B{numero_fila}="Con datos",'
                        f'$B{fila_anterior}="Con datos",'
                        f'{letra_origen}{fila_anterior}<>0),'
                        f'{letra_origen}{numero_fila}/'
                        f'{letra_origen}{fila_anterior}-1,"")'
                    )

    for fila in range(2, datos.max_row + 1):
        for columna in range(3, 6):
            datos.cell(fila, columna).number_format = formato_gs
        datos.cell(fila, 7).number_format = formato_gs
        for columna in (6, 8, 9, 10):
            datos.cell(fila, columna).number_format = formato_porcentaje
        color_fondo = (
            gris_claro
            if fila % 2 == 0
            else blanco
        )
        if datos.cell(fila, 2).value == "Sin datos":
            color_fondo = "FFF4E5"
        for columna in range(1, 11):
            datos.cell(fila, columna).fill = PatternFill(
                "solid",
                fgColor=color_fondo,
            )
            datos.cell(fila, columna).border = Border(bottom=borde)

    tabla = Table(
        displayName="ComparacionMensualBC",
        ref=f"A1:J{datos.max_row}",
    )
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showFirstColumn=False,
        showLastColumn=False,
    )
    datos.add_table(tabla)
    datos.freeze_panes = "A2"
    anchos = [13, 13, 18, 18, 18, 13, 18, 16, 16, 16]
    for indice, ancho in enumerate(anchos, 1):
        datos.column_dimensions[
            datos.cell(1, indice).column_letter
        ].width = ancho

    datos_grafica.append(["Período", "Ingresos", "Egresos", "Utilidad"])
    fila_grafica = 2
    for fila_datos, fila in enumerate(comparacion["filas"], 2):
        if not fila["tiene_datos"]:
            continue
        for columna, letra_origen in enumerate(("A", "C", "D", "E"), 1):
            datos_grafica.cell(
                fila_grafica,
                columna,
                f"='Datos mensuales'!{letra_origen}{fila_datos}",
            )
        fila_grafica += 1

    grafica = LineChart()
    grafica.title = "Evolución mensual (Gs.)"
    grafica.y_axis.title = "Guaraníes"
    grafica.x_axis.title = "Período"
    grafica.height = 8
    grafica.width = 19
    grafica.style = 13
    categorias = Reference(
        datos_grafica,
        min_col=1,
        min_row=2,
        max_row=max(fila_grafica - 1, 2),
    )
    valores = Reference(
        datos_grafica,
        min_col=2,
        max_col=4,
        min_row=1,
        max_row=max(fila_grafica - 1, 2),
    )
    grafica.add_data(valores, titles_from_data=True)
    grafica.set_categories(categorias)
    grafica.display_blanks = "gap"
    for serie, nombre in zip(
        grafica.series,
        ("Ingresos", "Egresos", "Utilidad"),
    ):
        serie.tx = SeriesLabel(v=nombre)
    resumen.add_chart(grafica, "D8")

    for columna in range(1, 11):
        resumen.column_dimensions[get_column_letter(columna)].width = 14
    resumen.page_setup.orientation = "landscape"
    resumen.page_setup.paperSize = resumen.PAPERSIZE_A4
    resumen.page_setup.fitToWidth = 1
    resumen.sheet_properties.pageSetUpPr.fitToPage = True
    datos.page_setup.orientation = "landscape"
    datos.page_setup.paperSize = datos.PAPERSIZE_A4
    datos.page_setup.fitToWidth = 1
    datos.sheet_properties.pageSetUpPr.fitToPage = True
    datos_grafica.sheet_state = "hidden"

    try:
        libro.calculation.fullCalcOnLoad = True
        libro.calculation.forceFullCalc = True
    except AttributeError:
        pass
    libro.save(ruta)
    return ruta


def exportar_pdf_comparacion(comparacion, ruta):
    """Exporta el tablero comparativo y su tabla completa en PDF."""
    try:
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.graphics.shapes import Drawing, String
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            KeepTogether,
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
        from xml.sax.saxutils import escape
    except ModuleNotFoundError as error:
        raise RuntimeError(
            "Falta instalar reportlab. Cerrá el sistema y volvé a abrir "
            "iniciar_interfaz.bat para completar la instalación."
        ) from error

    ruta = Path(ruta)
    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        "TituloComparacion",
        parent=estilos["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#173B73"),
        alignment=TA_LEFT,
    )
    cuerpo = ParagraphStyle(
        "CuerpoComparacion",
        parent=estilos["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#182230"),
    )
    cuerpo_derecha = ParagraphStyle(
        "CuerpoDerechaComparacion",
        parent=cuerpo,
        alignment=TA_RIGHT,
    )
    subtitulo = ParagraphStyle(
        "SubtituloComparacion",
        parent=cuerpo,
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=15,
        textColor=colors.HexColor("#173B73"),
        spaceAfter=5,
    )
    documento = SimpleDocTemplate(
        str(ruta),
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=17 * mm,
        bottomMargin=15 * mm,
        title=(
            f"Comparación BC {comparacion['periodo_desde']} a "
            f"{comparacion['periodo_hasta']}"
        ),
        author="BC Inversiones EAS",
    )
    elementos = [
        Paragraph("BC Inversiones EAS", titulo),
        Paragraph(
            (
                "Comparación mensual "
                f"{comparacion['periodo_desde']} a "
                f"{comparacion['periodo_hasta']} · "
                f"Vista: {escape(comparacion['unidad'])}"
            ),
            cuerpo,
        ),
        Spacer(1, 4 * mm),
    ]

    tarjetas = [
        ("Ingresos", formatear_monto(comparacion["total_ingresos"]), "#18A874"),
        ("Egresos", formatear_monto(comparacion["total_egresos"]), "#E25555"),
        ("Utilidad", formatear_monto(comparacion["total_utilidad"]), "#173B73"),
        (
            "Margen",
            Movimientos.formatear_porcentaje(
                comparacion["margen_total"]
            ) + "%",
            "#246BFD",
        ),
    ]
    celdas = []
    for nombre, valor, color in tarjetas:
        celdas.append(
            Paragraph(
                (
                    f'<font color="#617084" size="7"><b>{nombre.upper()}'
                    f"</b></font><br/>"
                    f'<font color="{color}" size="12"><b>'
                    f"{escape(valor)}</b></font>"
                ),
                cuerpo,
            )
        )
    tabla_tarjetas = Table(
        [celdas],
        colWidths=[64 * mm] * 4,
        rowHeights=[18 * mm],
    )
    tabla_tarjetas.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#DCE4EE")),
                ("INNERGRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#DCE4EE")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    elementos.append(tabla_tarjetas)
    elementos.append(Spacer(1, 4 * mm))

    filas_con_datos = [
        fila
        for fila in comparacion["filas"]
        if fila["tiene_datos"]
    ]
    if filas_con_datos:
        dibujo = Drawing(735, 215)
        dibujo.add(
            String(
                10,
                198,
                "Evolución mensual de ingresos, egresos y utilidad",
                fontName="Helvetica-Bold",
                fontSize=11,
                fillColor=colors.HexColor("#173B73"),
            )
        )
        grafica = VerticalBarChart()
        grafica.x = 48
        grafica.y = 38
        grafica.height = 145
        grafica.width = 650
        grafica.data = [
            [fila["ingresos"] for fila in filas_con_datos],
            [fila["egresos"] for fila in filas_con_datos],
            [fila["utilidad"] for fila in filas_con_datos],
        ]
        grafica.categoryAxis.categoryNames = [
            fila["periodo"] for fila in filas_con_datos
        ]
        grafica.categoryAxis.labels.fontSize = 7
        grafica.categoryAxis.labels.angle = 30
        grafica.categoryAxis.labels.dy = -8
        valores = [
            valor
            for serie in grafica.data
            for valor in serie
        ]
        grafica.valueAxis.valueMin = min(0, min(valores))
        grafica.valueAxis.valueMax = max(valores) * 1.12 or 1
        grafica.valueAxis.labels.fontSize = 7
        grafica.bars[0].fillColor = colors.HexColor("#18A874")
        grafica.bars[1].fillColor = colors.HexColor("#E25555")
        grafica.bars[2].fillColor = colors.HexColor("#246BFD")
        grafica.barSpacing = 1
        grafica.groupSpacing = 8
        dibujo.add(grafica)
        dibujo.add(
            String(
                500,
                198,
                "Ingresos",
                fontSize=8,
                fillColor=colors.HexColor("#18A874"),
            )
        )
        dibujo.add(
            String(
                560,
                198,
                "Egresos",
                fontSize=8,
                fillColor=colors.HexColor("#E25555"),
            )
        )
        dibujo.add(
            String(
                620,
                198,
                "Utilidad",
                fontSize=8,
                fillColor=colors.HexColor("#246BFD"),
            )
        )
        elementos.append(dibujo)

    def porcentaje_pdf(valor):
        if valor is None:
            return "-"
        return f"{valor:+.2f}%".replace(".", ",")

    filas_tabla = [[
        "Período",
        "Estado",
        "Ingresos",
        "Egresos",
        "Utilidad",
        "Margen",
        "Fondo",
        "Var. ingresos",
        "Var. egresos",
        "Var. utilidad",
    ]]
    for fila in comparacion["filas"]:
        if fila["tiene_datos"]:
            filas_tabla.append(
                [
                    fila["periodo"],
                    fila["estado"],
                    formatear_monto(fila["ingresos"]),
                    formatear_monto(fila["egresos"]),
                    formatear_monto(fila["utilidad"]),
                    (
                        Movimientos.formatear_porcentaje(
                            fila["margen"]
                        ) + "%"
                    ),
                    (
                        formatear_monto(fila["fondo"])
                        if fila["fondo"] is not None
                        else "-"
                    ),
                    porcentaje_pdf(fila["variacion_ingresos"]),
                    porcentaje_pdf(fila["variacion_egresos"]),
                    porcentaje_pdf(fila["variacion_utilidad"]),
                ]
            )
        else:
            filas_tabla.append(
                [fila["periodo"], "Sin datos"] + ["-"] * 8
            )

    tabla = Table(
        filas_tabla,
        repeatRows=1,
        colWidths=[
            19 * mm,
            19 * mm,
            27 * mm,
            27 * mm,
            27 * mm,
            18 * mm,
            27 * mm,
            23 * mm,
            23 * mm,
            23 * mm,
        ],
    )
    tabla.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#246BFD")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 6.5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
                    colors.white,
                    colors.HexColor("#F4F7FB"),
                ]),
                ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
                ("ALIGN", (0, 0), (1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#DCE4EE")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elementos.append(PageBreak())
    elementos.append(KeepTogether([
        Paragraph("Detalle por mes", subtitulo),
        Paragraph(
            (
                "Los meses sin registros se muestran como Sin datos. "
                "Las variaciones se calculan contra el mes inmediato anterior."
            ),
            cuerpo,
        ),
        Spacer(1, 3 * mm),
    ]))
    elementos.append(tabla)

    def encabezado_pie(canvas, documento_actual):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#617084"))
        canvas.drawString(
            documento_actual.leftMargin,
            landscape(A4)[1] - 10 * mm,
            (
                "BC Gestión - Comparación "
                f"{comparacion['periodo_desde']} a "
                f"{comparacion['periodo_hasta']}"
            ),
        )
        canvas.drawRightString(
            landscape(A4)[0] - documento_actual.rightMargin,
            8 * mm,
            f"Página {documento_actual.page}",
        )
        canvas.restoreState()

    documento.build(
        elementos,
        onFirstPage=encabezado_pie,
        onLaterPages=encabezado_pie,
    )
    return ruta


def crear_tarjeta(master, titulo, valor, color, descripcion, fila, columna):
    fondos = {
        COLOR_VERDE: ("#EAF8F3", "#12342C"),
        COLOR_ROJO: ("#FFF0F2", "#3A2027"),
        COLOR_PRIMARIO: ("#EDF3FF", "#182C4D"),
        COLOR_NARANJA: ("#FFF5E8", "#3A2B18"),
        COLOR_VIOLETA: ("#F2EFFF", "#2B2547"),
        COLOR_CELESTE: ("#EAF8FC", "#173440"),
    }
    tarjeta = ctk.CTkFrame(
        master,
        fg_color=fondos.get(color, COLOR_PANEL),
        corner_radius=18,
        border_width=1,
        border_color=color,
    )
    tarjeta.grid(
        row=fila,
        column=columna,
        sticky="nsew",
        padx=5,
        pady=5,
    )
    ctk.CTkFrame(
        tarjeta,
        height=5,
        corner_radius=3,
        fg_color=color,
    ).pack(fill="x", padx=16, pady=(14, 8))
    ctk.CTkLabel(
        tarjeta,
        text=titulo,
        font=ctk.CTkFont(size=11, weight="bold"),
        text_color=COLOR_TEXTO_SUAVE,
        anchor="w",
    ).pack(fill="x", padx=16, pady=(0, 2))
    ctk.CTkLabel(
        tarjeta,
        text=valor,
        font=ctk.CTkFont(size=21, weight="bold"),
        text_color=color,
        anchor="w",
    ).pack(fill="x", padx=16)
    ctk.CTkLabel(
        tarjeta,
        text=descripcion,
        font=ctk.CTkFont(size=10),
        text_color=COLOR_TEXTO_SUAVE,
        anchor="w",
        justify="left",
        wraplength=210,
    ).pack(fill="x", padx=16, pady=(4, 16))


def dibujar_portada_mensual(master, informe):
    """Encabezado visual con el resultado más importante del mes."""
    positivo = informe["utilidad"] >= 0
    estado = "Cierre positivo" if positivo else "Cierre a revisar"
    color_resultado = "#61D8B2" if positivo else "#FF9AA5"

    portada = ctk.CTkFrame(
        master,
        fg_color=COLOR_AZUL_NOCHE,
        corner_radius=20,
        border_width=0,
    )
    portada.pack(fill="x", pady=(2, 12))
    portada.grid_columnconfigure(0, weight=1)

    ctk.CTkLabel(
        portada,
        text="BC INVERSIONES EAS  ·  RESUMEN MENSUAL",
        font=ctk.CTkFont(size=11, weight="bold"),
        text_color="#8EC5FF",
        anchor="w",
    ).grid(row=0, column=0, sticky="ew", padx=24, pady=(20, 3))
    ctk.CTkLabel(
        portada,
        text=periodo_en_palabras(informe["periodo"]),
        font=ctk.CTkFont(size=27, weight="bold"),
        text_color="#FFFFFF",
        anchor="w",
    ).grid(row=1, column=0, sticky="ew", padx=24)
    ctk.CTkLabel(
        portada,
        text=f"{informe['unidad']}  ·  {estado}",
        font=ctk.CTkFont(size=12),
        text_color="#C7D9EA",
        anchor="w",
    ).grid(row=2, column=0, sticky="ew", padx=24, pady=(4, 22))

    resultado = ctk.CTkFrame(
        portada,
        fg_color=("#173B5C", "#142E46"),
        corner_radius=16,
        border_width=1,
        border_color=("#315D80", "#315D80"),
    )
    resultado.grid(
        row=0,
        column=1,
        rowspan=3,
        sticky="nsew",
        padx=(10, 20),
        pady=18,
    )
    ctk.CTkLabel(
        resultado,
        text="RESULTADO DEL MES",
        font=ctk.CTkFont(size=10, weight="bold"),
        text_color="#AFC7DC",
        anchor="w",
    ).pack(fill="x", padx=20, pady=(15, 2))
    ctk.CTkLabel(
        resultado,
        text=formatear_monto(informe["utilidad"]),
        font=ctk.CTkFont(size=22, weight="bold"),
        text_color=color_resultado,
        anchor="w",
    ).pack(fill="x", padx=20)
    ctk.CTkLabel(
        resultado,
        text=(
            f"Margen {Movimientos.formatear_porcentaje(informe['margen'])}%"
        ),
        font=ctk.CTkFont(size=11),
        text_color="#D7E5F1",
        anchor="w",
    ).pack(fill="x", padx=20, pady=(3, 15))


def dibujar_tarjetas(master, informe):
    zona = ctk.CTkFrame(master, fg_color="transparent")
    zona.pack(fill="x", pady=(0, 8))
    for columna in range(4):
        zona.grid_columnconfigure(columna, weight=1, uniform="informe")

    datos = [
        (
            "Ingresos",
            formatear_monto(informe["ingresos"]),
            COLOR_VERDE,
            "Todo lo que generó el mes",
        ),
        (
            "Utilidad",
            formatear_monto(informe["utilidad"]),
            COLOR_VERDE if informe["utilidad"] >= 0 else COLOR_ROJO,
            "Ingresos menos egresos",
        ),
        (
            "Margen",
            Movimientos.formatear_porcentaje(informe["margen"]) + "%",
            COLOR_PRIMARIO,
            "Resultado sobre ingresos",
        ),
        (
            "Egresos",
            formatear_monto(informe["egresos"]),
            COLOR_ROJO,
            "Uso total de recursos",
        ),
    ]
    for indice, datos_tarjeta in enumerate(datos):
        crear_tarjeta(
            zona,
            *datos_tarjeta,
            0,
            indice,
        )


def dibujar_indicadores_secundarios(master, informe):
    """Muestra datos de control sin competir con los KPI principales."""
    panel = ctk.CTkFrame(
        master,
        fg_color=COLOR_PANEL,
        corner_radius=16,
        border_width=1,
        border_color=COLOR_BORDE,
    )
    panel.pack(fill="x", pady=(0, 8))
    for columna in range(4):
        panel.grid_columnconfigure(columna, weight=1, uniform="secundarios")

    fondo_valor = (
        formatear_monto(informe["fondo"])
        if informe["fondo"] is not None
        else "No aplica"
    )
    datos = [
        ("Fondo de estabilidad", fondo_valor, COLOR_NARANJA),
        (
            "Inversiones del negocio",
            formatear_monto(informe["inversiones"]),
            COLOR_VIOLETA,
        ),
        (
            "Salida real de dinero",
            formatear_monto(informe["salida_caja"]),
            COLOR_CELESTE,
        ),
        (
            "Retenciones y descuentos",
            formatear_monto(informe["retenciones"]),
            COLOR_NARANJA,
        ),
    ]
    for columna, (titulo, valor, color) in enumerate(datos):
        bloque = ctk.CTkFrame(panel, fg_color="transparent")
        bloque.grid(row=0, column=columna, sticky="nsew", padx=8, pady=13)
        ctk.CTkLabel(
            bloque,
            text="●  " + titulo,
            font=ctk.CTkFont(size=10, weight="bold"),
            text_color=color,
            anchor="w",
        ).pack(fill="x", padx=8)
        ctk.CTkLabel(
            bloque,
            text=valor,
            font=ctk.CTkFont(size=15, weight="bold"),
            text_color=COLOR_TEXTO,
            anchor="w",
        ).pack(fill="x", padx=8, pady=(3, 0))


def dibujar_comparacion(master, informe):
    panel = ctk.CTkFrame(
        master,
        fg_color=COLOR_PANEL,
        corner_radius=16,
        border_width=1,
        border_color=COLOR_BORDE,
    )
    panel.pack(fill="x", pady=7)
    ctk.CTkLabel(
        panel,
        text="Comparación por unidad",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color=COLOR_TEXTO,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(16, 4))
    ctk.CTkLabel(
        panel,
        text=(
            "Compara solamente movimientos asignados directamente a cada "
            "unidad. Nómina, adicionales y cuotas aparecen como generales."
        ),
        font=ctk.CTkFont(size=11),
        text_color=COLOR_TEXTO_SUAVE,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(0, 10))

    columnas = (
        "unidad",
        "ingresos",
        "egresos",
        "resultado",
        "recibidas",
        "enviadas",
    )
    tabla = ttk.Treeview(
        panel,
        columns=columnas,
        show="headings",
        height=max(len(informe["comparacion_unidades"]), 1),
    )
    titulos = {
        "unidad": "Unidad",
        "ingresos": "Ingresos",
        "egresos": "Egresos",
        "resultado": "Resultado",
        "recibidas": "Transf. recibidas",
        "enviadas": "Transf. enviadas",
    }
    for columna in columnas:
        tabla.heading(columna, text=titulos[columna])
        tabla.column(
            columna,
            width=140,
            minwidth=95,
            anchor="center",
            stretch=True,
        )
    for unidad in informe["comparacion_unidades"]:
        tabla.insert(
            "",
            "end",
            values=(
                unidad["unidad"],
                Movimientos.formatear_monto(unidad["ingresos"]),
                Movimientos.formatear_monto(unidad["egresos"]),
                Movimientos.formatear_monto(unidad["resultado"]),
                Movimientos.formatear_monto(unidad["recibidas"]),
                Movimientos.formatear_monto(unidad["enviadas"]),
            ),
        )
    tabla.pack(fill="x", padx=18, pady=(0, 18))


def dibujar_barras_categorias(master, informe, tipo):
    es_ingreso = tipo == "ingresos"
    titulo = (
        "Dónde se generaron los ingresos"
        if es_ingreso
        else "Cómo se distribuyeron los egresos"
    )
    categorias = (
        informe["categorias_ingresos"]
        if es_ingreso
        else informe["categorias_egresos"]
    )
    total = informe["ingresos"] if es_ingreso else informe["egresos"]
    color_barra = COLOR_VERDE if es_ingreso else COLOR_ROJO
    texto_vacio = (
        "No hay ingresos registrados para este filtro."
        if es_ingreso
        else "No hay egresos registrados para este filtro."
    )

    panel = ctk.CTkFrame(
        master,
        fg_color=COLOR_PANEL,
        corner_radius=16,
        border_width=1,
        border_color=COLOR_BORDE,
    )
    panel.pack(fill="x", pady=7)
    ctk.CTkLabel(
        panel,
        text=titulo,
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color=COLOR_TEXTO,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(16, 4))
    ctk.CTkLabel(
        panel,
        text=(
            f"Las categorías suman exactamente "
            f"{formatear_monto(total)}."
        ),
        font=ctk.CTkFont(size=11),
        text_color=COLOR_TEXTO_SUAVE,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(0, 12))

    if not categorias:
        ctk.CTkLabel(
            panel,
            text=texto_vacio,
            font=ctk.CTkFont(size=12),
            text_color=COLOR_TEXTO_SUAVE,
        ).pack(fill="x", padx=18, pady=(0, 18))
        return

    maximo = max(valor for _, valor in categorias)
    for indice, (concepto, valor) in enumerate(categorias):
        fila = ctk.CTkFrame(
            panel,
            fg_color=(
                COLOR_PANEL_SECUNDARIO
                if indice % 2 == 0
                else "transparent"
            ),
            corner_radius=8,
        )
        fila.pack(fill="x", padx=18, pady=2)
        fila.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(
            fila,
            text=concepto,
            width=250,
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=COLOR_TEXTO,
            anchor="w",
        ).grid(row=0, column=0, sticky="w", padx=10, pady=9)
        barra = ctk.CTkProgressBar(
            fila,
            height=12,
            progress_color=color_barra,
            fg_color=COLOR_BORDE,
        )
        barra.grid(row=0, column=1, sticky="ew", padx=10, pady=9)
        barra.set(valor / maximo if maximo else 0)
        ctk.CTkLabel(
            fila,
            text=formatear_monto(valor),
            width=155,
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=COLOR_TEXTO,
            anchor="e",
        ).grid(row=0, column=2, sticky="e", padx=10, pady=9)
    ctk.CTkFrame(panel, height=10, fg_color="transparent").pack()


def dibujar_conciliacion(master, informe):
    panel = ctk.CTkFrame(
        master,
        fg_color=COLOR_PANEL,
        corner_radius=16,
        border_width=1,
        border_color=COLOR_BORDE,
    )
    panel.pack(fill="x", pady=7)
    ctk.CTkLabel(
        panel,
        text="Control del total",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color=COLOR_TEXTO,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(16, 10))

    filas = [
        ("Total del informe", informe["egresos"]),
        ("Suma del desglose", sum(
            valor for _, valor in informe["categorias_egresos"]
        )),
    ]
    if informe["empresa_completa"]:
        filas.extend(
            [
                ("Salida real de dinero", informe["salida_caja"]),
                (
                    "Retenciones y otros descuentos",
                    informe["retenciones"],
                ),
            ]
        )

    for indice, (texto, valor) in enumerate(filas):
        fila = ctk.CTkFrame(
            panel,
            fg_color=(
                COLOR_PANEL_SECUNDARIO
                if indice % 2 == 0
                else "transparent"
            ),
        )
        fila.pack(fill="x", padx=18, pady=2)
        fila.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(
            fila,
            text=texto,
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=COLOR_TEXTO,
            anchor="w",
        ).grid(row=0, column=0, sticky="w", padx=12, pady=9)
        ctk.CTkLabel(
            fila,
            text=formatear_monto(valor),
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=COLOR_TEXTO,
            anchor="e",
        ).grid(row=0, column=1, sticky="e", padx=12, pady=9)
    ctk.CTkFrame(panel, height=12, fg_color="transparent").pack()


def dibujar_detalle(master, informe):
    panel = ctk.CTkFrame(
        master,
        fg_color=COLOR_PANEL,
        corner_radius=16,
        border_width=1,
        border_color=COLOR_BORDE,
    )
    panel.pack(fill="x", pady=7)
    ctk.CTkLabel(
        panel,
        text="Detalle diario por categoría",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color=COLOR_TEXTO,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(16, 4))
    ctk.CTkLabel(
        panel,
        text=(
            f"{len(informe['resumen_diario'])} grupos diarios, formados por "
            f"{len(informe['detalles'])} movimientos verificables. "
            "Las transferencias y depósitos propios se informan sin alterar "
            "el resultado."
        ),
        font=ctk.CTkFont(size=11),
        text_color=COLOR_TEXTO_SUAVE,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(0, 10))

    zona_tabla = ctk.CTkFrame(panel, fg_color="transparent")
    zona_tabla.pack(fill="x", padx=18, pady=(0, 18))
    zona_tabla.grid_columnconfigure(0, weight=1)

    columnas = (
        "fecha",
        "grupo",
        "cantidad",
        "entradas",
        "salidas",
        "resultado",
        "ruta_interna",
        "monto_interno",
    )
    tabla = ttk.Treeview(
        zona_tabla,
        columns=columnas,
        show="headings",
        height=min(max(len(informe["resumen_diario"]), 3), 15),
    )
    configuracion = [
        ("fecha", "Fecha", 95, "center"),
        ("grupo", "Categoría / unidad", 210, "w"),
        ("cantidad", "Mov.", 70, "center"),
        ("entradas", "Entró", 145, "e"),
        ("salidas", "Salió", 145, "e"),
        ("resultado", "Resultado", 145, "e"),
        ("ruta_interna", "Ruta interna / depósito", 255, "w"),
        ("monto_interno", "Monto movido", 145, "e"),
    ]
    for columna, titulo, ancho, ancla in configuracion:
        tabla.heading(columna, text=titulo)
        tabla.column(
            columna,
            width=ancho,
            minwidth=75,
            anchor=ancla,
            stretch=columna == "grupo",
        )
    for detalle in informe["resumen_diario"]:
        tabla.insert(
            "",
            "end",
            values=(
                detalle["fecha"],
                detalle["grupo"],
                detalle["cantidad"],
                Movimientos.formatear_monto(detalle["entradas"]),
                Movimientos.formatear_monto(detalle["salidas"]),
                Movimientos.formatear_monto(detalle["resultado"]),
                detalle["ruta_interna"] or "-",
                (
                    Movimientos.formatear_monto(detalle["monto_interno"])
                    if detalle["monto_interno"]
                    else "-"
                ),
            ),
        )

    desplazamiento_vertical = ttk.Scrollbar(
        zona_tabla,
        orient="vertical",
        command=tabla.yview,
    )
    desplazamiento_horizontal = ttk.Scrollbar(
        zona_tabla,
        orient="horizontal",
        command=tabla.xview,
    )
    tabla.configure(
        yscrollcommand=desplazamiento_vertical.set,
        xscrollcommand=desplazamiento_horizontal.set,
    )
    tabla.grid(row=0, column=0, sticky="nsew")
    desplazamiento_vertical.grid(row=0, column=1, sticky="ns")
    desplazamiento_horizontal.grid(row=1, column=0, sticky="ew")


def dibujar_informe(zona_resultado, informe):
    for widget in zona_resultado.winfo_children():
        widget.destroy()

    dibujar_portada_mensual(zona_resultado, informe)

    if not informe["empresa_completa"]:
        aviso = ctk.CTkFrame(
            zona_resultado,
            fg_color=("#FFF7E8", "#3B2C16"),
            corner_radius=12,
            border_width=1,
            border_color=("#F2D49B", "#684B20"),
        )
        aviso.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(
            aviso,
            text=(
                "Esta vista incluye solo movimientos asignados directamente "
                f"a {informe['unidad']}. La nómina, los adicionales y las "
                "cuotas no tienen unidad asignada y aparecen únicamente en "
                "Empresa completa."
            ),
            font=ctk.CTkFont(size=11),
            text_color=COLOR_TEXTO,
            anchor="w",
            justify="left",
            wraplength=900,
        ).pack(fill="x", padx=16, pady=12)

    dibujar_tarjetas(zona_resultado, informe)
    dibujar_indicadores_secundarios(zona_resultado, informe)
    dibujar_barras_categorias(zona_resultado, informe, "ingresos")
    dibujar_comparacion(zona_resultado, informe)
    dibujar_barras_categorias(zona_resultado, informe, "egresos")
    dibujar_conciliacion(zona_resultado, informe)
    dibujar_detalle(zona_resultado, informe)


def texto_variacion(valor):
    if valor is None:
        return "-"
    return f"{valor:+.2f}%".replace(".", ",")


def abreviar_monto_grafica(valor):
    absoluto = abs(valor)
    signo = "-" if valor < 0 else ""
    if absoluto >= 1_000_000_000:
        return f"{signo}{absoluto / 1_000_000_000:.1f} mil M"
    if absoluto >= 1_000_000:
        return f"{signo}{absoluto / 1_000_000:.0f} M"
    if absoluto >= 1_000:
        return f"{signo}{absoluto / 1_000:.0f} mil"
    return f"{valor:.0f}"


def dibujar_grafica_tendencia(master, comparacion):
    panel = ctk.CTkFrame(
        master,
        fg_color=COLOR_PANEL,
        corner_radius=16,
        border_width=1,
        border_color=COLOR_BORDE,
    )
    panel.pack(fill="x", pady=7)
    ctk.CTkLabel(
        panel,
        text="Evolución mensual",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color=COLOR_TEXTO,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(16, 2))
    ctk.CTkLabel(
        panel,
        text=(
            "Los meses sin registros quedan como espacios vacíos y no como "
            "meses con resultado cero."
        ),
        font=ctk.CTkFont(size=11),
        text_color=COLOR_TEXTO_SUAVE,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(0, 8))

    modo_oscuro = ctk.get_appearance_mode() == "Dark"
    fondo = "#131D2E" if modo_oscuro else "#FFFFFF"
    rejilla = "#26354A" if modo_oscuro else "#DCE4EE"
    texto = "#F4F7FB" if modo_oscuro else "#617084"
    lienzo = tk.Canvas(
        panel,
        height=300,
        background=fondo,
        highlightthickness=0,
    )
    lienzo.pack(fill="x", padx=18, pady=(0, 18))

    def redibujar(_evento=None):
        lienzo.delete("all")
        filas = comparacion["filas"]
        ancho = max(lienzo.winfo_width(), 760)
        alto = 300
        izquierda = 72
        derecha = ancho - 24
        arriba = 24
        abajo = 238
        valores = [
            fila[clave]
            for fila in filas
            if fila["tiene_datos"]
            for clave in ("ingresos", "egresos", "utilidad")
        ]
        if not valores:
            lienzo.create_text(
                ancho / 2,
                alto / 2,
                text="No hay meses con datos en este rango.",
                fill=texto,
                font=("Segoe UI", 11),
            )
            return

        minimo = min(0, min(valores))
        maximo = max(valores)
        if maximo == minimo:
            maximo = minimo + 1
        amplitud = maximo - minimo
        maximo += amplitud * 0.08
        minimo -= amplitud * 0.05

        def posicion_y(valor):
            proporcion = (valor - minimo) / (maximo - minimo)
            return abajo - proporcion * (abajo - arriba)

        for indice in range(5):
            valor = minimo + (maximo - minimo) * indice / 4
            y = posicion_y(valor)
            lienzo.create_line(
                izquierda,
                y,
                derecha,
                y,
                fill=rejilla,
                dash=(3, 4),
            )
            lienzo.create_text(
                izquierda - 8,
                y,
                text=abreviar_monto_grafica(valor),
                fill=texto,
                anchor="e",
                font=("Segoe UI", 8),
            )

        cantidad = max(len(filas), 1)
        paso_x = (
            (derecha - izquierda) / (cantidad - 1)
            if cantidad > 1
            else 0
        )
        salto_etiqueta = 2 if cantidad > 12 else 1
        posiciones_x = []
        for indice, fila in enumerate(filas):
            x = (
                izquierda + paso_x * indice
                if cantidad > 1
                else (izquierda + derecha) / 2
            )
            posiciones_x.append(x)
            if indice % salto_etiqueta == 0 or indice == cantidad - 1:
                lienzo.create_text(
                    x,
                    abajo + 16,
                    text=fila["periodo"],
                    fill=texto,
                    anchor="n",
                    font=("Segoe UI", 8),
                )

        series = [
            ("ingresos", "#18A874", "Ingresos"),
            ("egresos", "#E25555", "Egresos"),
            ("utilidad", "#246BFD", "Utilidad"),
        ]
        for clave, color, _nombre in series:
            punto_anterior = None
            for indice, fila in enumerate(filas):
                if not fila["tiene_datos"]:
                    punto_anterior = None
                    continue
                punto = (
                    posiciones_x[indice],
                    posicion_y(fila[clave]),
                )
                if punto_anterior is not None:
                    lienzo.create_line(
                        *punto_anterior,
                        *punto,
                        fill=color,
                        width=3,
                    )
                lienzo.create_oval(
                    punto[0] - 4,
                    punto[1] - 4,
                    punto[0] + 4,
                    punto[1] + 4,
                    fill=color,
                    outline=fondo,
                    width=1,
                )
                punto_anterior = punto

        inicio_leyenda = max(izquierda, derecha - 270)
        for indice, (_clave, color, nombre) in enumerate(series):
            x = inicio_leyenda + indice * 90
            lienzo.create_line(
                x,
                278,
                x + 18,
                278,
                fill=color,
                width=4,
            )
            lienzo.create_text(
                x + 24,
                278,
                text=nombre,
                fill=texto,
                anchor="w",
                font=("Segoe UI", 8, "bold"),
            )

    lienzo.bind("<Configure>", redibujar)
    lienzo.after_idle(redibujar)


def dibujar_tabla_comparacion(master, comparacion):
    panel = ctk.CTkFrame(
        master,
        fg_color=COLOR_PANEL,
        corner_radius=16,
        border_width=1,
        border_color=COLOR_BORDE,
    )
    panel.pack(fill="x", pady=7)
    ctk.CTkLabel(
        panel,
        text="Detalle comparativo por mes",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color=COLOR_TEXTO,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(16, 4))
    ctk.CTkLabel(
        panel,
        text=(
            "Las variaciones comparan contra el mes inmediatamente anterior. "
            "Usá ← y → para recorrer todas las columnas."
        ),
        font=ctk.CTkFont(size=11),
        text_color=COLOR_TEXTO_SUAVE,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(0, 10))

    marco_tabla = ctk.CTkFrame(panel, fg_color="transparent")
    marco_tabla.pack(fill="x", padx=18, pady=(0, 18))
    columnas = (
        "periodo",
        "estado",
        "ingresos",
        "egresos",
        "utilidad",
        "margen",
        "fondo",
        "var_ingresos",
        "var_egresos",
        "var_utilidad",
    )
    tabla = ttk.Treeview(
        marco_tabla,
        columns=columnas,
        show="headings",
        height=min(max(len(comparacion["filas"]), 2), 12),
    )
    titulos = {
        "periodo": "Período",
        "estado": "Estado",
        "ingresos": "Ingresos",
        "egresos": "Egresos",
        "utilidad": "Utilidad",
        "margen": "Margen",
        "fondo": "Fondo",
        "var_ingresos": "Var. ingresos",
        "var_egresos": "Var. egresos",
        "var_utilidad": "Var. utilidad",
    }
    anchos = {
        "periodo": 90,
        "estado": 90,
        "ingresos": 145,
        "egresos": 145,
        "utilidad": 145,
        "margen": 90,
        "fondo": 145,
        "var_ingresos": 105,
        "var_egresos": 105,
        "var_utilidad": 105,
    }
    for columna in columnas:
        tabla.heading(columna, text=titulos[columna])
        tabla.column(
            columna,
            width=anchos[columna],
            minwidth=anchos[columna],
            anchor="center",
            stretch=False,
        )

    for fila in comparacion["filas"]:
        if fila["tiene_datos"]:
            valores = (
                fila["periodo"],
                fila["estado"],
                formatear_monto(fila["ingresos"]),
                formatear_monto(fila["egresos"]),
                formatear_monto(fila["utilidad"]),
                Movimientos.formatear_porcentaje(fila["margen"]) + "%",
                (
                    formatear_monto(fila["fondo"])
                    if fila["fondo"] is not None
                    else "-"
                ),
                texto_variacion(fila["variacion_ingresos"]),
                texto_variacion(fila["variacion_egresos"]),
                texto_variacion(fila["variacion_utilidad"]),
            )
            etiqueta = (
                "negativo"
                if fila["utilidad"] < 0
                else "normal"
            )
        else:
            valores = (
                fila["periodo"],
                "Sin datos",
                "-",
                "-",
                "-",
                "-",
                "-",
                "-",
                "-",
                "-",
            )
            etiqueta = "sin_datos"
        tabla.insert("", "end", values=valores, tags=(etiqueta,))

    tabla.tag_configure("sin_datos", foreground="#9B6B22")
    tabla.tag_configure("negativo", foreground="#C43D3D")
    barra_x = ttk.Scrollbar(
        marco_tabla,
        orient="horizontal",
        command=tabla.xview,
    )
    tabla.configure(xscrollcommand=barra_x.set)
    tabla.pack(fill="x")
    barra_x.pack(fill="x")
    tabla.bind(
        "<Left>",
        lambda _evento: (
            tabla.xview_scroll(-1, "units"),
            "break",
        )[1],
    )
    tabla.bind(
        "<Right>",
        lambda _evento: (
            tabla.xview_scroll(1, "units"),
            "break",
        )[1],
    )
    hijos = tabla.get_children()
    if hijos:
        tabla.selection_set(hijos[0])


def dibujar_comparacion_mensual(zona_resultado, comparacion):
    for widget in zona_resultado.winfo_children():
        widget.destroy()

    ctk.CTkLabel(
        zona_resultado,
        text=(
            f"Comparación · {comparacion['periodo_desde']} a "
            f"{comparacion['periodo_hasta']} · {comparacion['unidad']}"
        ),
        font=ctk.CTkFont(size=20, weight="bold"),
        text_color=COLOR_TEXTO,
        anchor="w",
    ).pack(fill="x", pady=(2, 8))

    if comparacion["meses_con_datos"] == 0:
        aviso = ctk.CTkFrame(
            zona_resultado,
            fg_color=("#FFF7E8", "#3B2C16"),
            corner_radius=12,
            border_width=1,
            border_color=("#F2D49B", "#684B20"),
        )
        aviso.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(
            aviso,
            text="No se encontraron registros dentro del rango seleccionado.",
            font=ctk.CTkFont(size=11),
            text_color=COLOR_TEXTO,
        ).pack(fill="x", padx=16, pady=12)

    tarjetas = ctk.CTkFrame(zona_resultado, fg_color="transparent")
    tarjetas.pack(fill="x", pady=(0, 8))
    for columna in range(3):
        tarjetas.grid_columnconfigure(columna, weight=1, uniform="comparacion")
    fondo_texto = (
        formatear_monto(comparacion["total_fondo"])
        if comparacion["total_fondo"] is not None
        else "No aplica"
    )
    datos_tarjetas = [
        (
            "Ingresos acumulados",
            formatear_monto(comparacion["total_ingresos"]),
            COLOR_VERDE,
            f"{comparacion['meses_con_datos']} meses con datos",
        ),
        (
            "Egresos acumulados",
            formatear_monto(comparacion["total_egresos"]),
            COLOR_ROJO,
            "Suma de los meses con registros",
        ),
        (
            "Utilidad acumulada",
            formatear_monto(comparacion["total_utilidad"]),
            (
                COLOR_VERDE
                if comparacion["total_utilidad"] >= 0
                else COLOR_ROJO
            ),
            "Ingresos menos egresos",
        ),
        (
            "Margen total",
            (
                Movimientos.formatear_porcentaje(
                    comparacion["margen_total"]
                ) + "%"
            ),
            COLOR_PRIMARIO,
            "Resultado acumulado sobre ingresos",
        ),
        (
            "Fondo acumulado",
            fondo_texto,
            COLOR_NARANJA,
            "Suma de fondos de los meses con datos",
        ),
        (
            "Mejor mes",
            (
                comparacion["mejor_mes"]["periodo"]
                if comparacion["mejor_mes"]
                else "Sin datos"
            ),
            "#7B61FF",
            (
                formatear_monto(
                    comparacion["mejor_mes"]["utilidad"]
                )
                if comparacion["mejor_mes"]
                else "No hay resultado para comparar"
            ),
        ),
    ]
    for indice, datos_tarjeta in enumerate(datos_tarjetas):
        crear_tarjeta(
            tarjetas,
            *datos_tarjeta,
            indice // 3,
            indice % 3,
        )

    dibujar_grafica_tendencia(zona_resultado, comparacion)
    dibujar_tabla_comparacion(zona_resultado, comparacion)


def dibujar_informe_anual(zona_resultado, informe):
    """Muestra el cuadro anual compacto con una fila de total al final."""
    for widget in zona_resultado.winfo_children():
        widget.destroy()

    ctk.CTkLabel(
        zona_resultado,
        text=f"Informe anual {informe['anio']} · {informe['unidad']}",
        font=ctk.CTkFont(size=20, weight="bold"),
        text_color=COLOR_TEXTO,
        anchor="w",
    ).pack(fill="x", pady=(2, 8))

    panel = ctk.CTkFrame(
        zona_resultado,
        fg_color=COLOR_PANEL,
        corner_radius=16,
        border_width=1,
        border_color=COLOR_BORDE,
    )
    panel.pack(fill="x", pady=7)
    ctk.CTkLabel(
        panel,
        text="Resumen de enero a diciembre",
        font=ctk.CTkFont(size=16, weight="bold"),
        text_color=COLOR_TEXTO,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(16, 4))
    ctk.CTkLabel(
        panel,
        text=(
            "Los meses sin registros aparecen con guiones. El margen de la "
            "última fila se calcula sobre los totales del año."
        ),
        font=ctk.CTkFont(size=11),
        text_color=COLOR_TEXTO_SUAVE,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(0, 10))

    marco_tabla = ctk.CTkFrame(panel, fg_color="transparent")
    marco_tabla.pack(fill="x", padx=18, pady=(0, 18))
    columnas = ("mes", "ingresos", "egresos", "utilidad", "margen")
    tabla = ttk.Treeview(
        marco_tabla,
        columns=columnas,
        show="headings",
        height=13,
    )
    titulos = {
        "mes": "Mes",
        "ingresos": "Ingresos",
        "egresos": "Egresos",
        "utilidad": "Utilidad",
        "margen": "Margen",
    }
    anchos = {
        "mes": 175,
        "ingresos": 210,
        "egresos": 210,
        "utilidad": 210,
        "margen": 130,
    }
    for columna in columnas:
        tabla.heading(columna, text=titulos[columna])
        tabla.column(
            columna,
            width=anchos[columna],
            minwidth=anchos[columna],
            anchor="center" if columna == "margen" else "e",
            stretch=True,
        )
    tabla.column("mes", anchor="w")

    for fila in informe["filas"]:
        if fila["tiene_datos"]:
            valores = (
                fila["mes"],
                formatear_monto(fila["ingresos"]),
                formatear_monto(fila["egresos"]),
                formatear_monto(fila["utilidad"]),
                Movimientos.formatear_porcentaje(fila["margen"]) + "%",
            )
            etiqueta = "negativo" if fila["utilidad"] < 0 else "normal"
        else:
            valores = (fila["mes"], "-", "-", "-", "-")
            etiqueta = "sin_datos"
        tabla.insert("", "end", values=valores, tags=(etiqueta,))

    tabla.insert(
        "",
        "end",
        values=(
            "TOTAL DEL AÑO",
            formatear_monto(informe["total_ingresos"]),
            formatear_monto(informe["total_egresos"]),
            formatear_monto(informe["total_utilidad"]),
            (
                Movimientos.formatear_porcentaje(
                    informe["margen_total"]
                ) + "%"
            ),
        ),
        tags=("total",),
    )
    tabla.tag_configure("sin_datos", foreground="#9B6B22")
    tabla.tag_configure("negativo", foreground="#C43D3D")
    tabla.tag_configure(
        "total",
        background="#DCE8FF",
        foreground="#173B73",
        font=("Segoe UI", 10, "bold"),
    )
    tabla.pack(fill="x")
    hijos = tabla.get_children()
    if hijos:
        tabla.selection_set(hijos[0])


def mostrar_informes(aplicacion):
    """Abre la pantalla principal del módulo Informes."""
    aplicacion.limpiar_contenedor()
    aplicacion.marcar_seleccion("Informes")
    aplicacion.bind(
        "<Escape>",
        lambda _evento: aplicacion.mostrar_inicio(),
    )
    estado_informe = {
        "actual": None,
        "tipo": None,
    }

    pagina = ctk.CTkScrollableFrame(
        aplicacion.contenedor,
        fg_color="transparent",
        corner_radius=0,
    )
    pagina.grid(row=0, column=0, sticky="nsew")

    aplicacion.crear_encabezado(
        pagina,
        "Informes",
        (
            "Analizá un mes, compará períodos o revisá el cuadro completo "
            "de un año usando los mismos totales conciliados del cierre."
        ),
    )

    controles = ctk.CTkFrame(
        pagina,
        fg_color=COLOR_PANEL,
        corner_radius=16,
        border_width=1,
        border_color=COLOR_BORDE,
    )
    controles.pack(fill="x", padx=34, pady=(0, 14))
    controles.grid_columnconfigure(4, weight=1)

    ctk.CTkLabel(
        controles,
        text="Período",
        font=ctk.CTkFont(size=12, weight="bold"),
        text_color=COLOR_TEXTO,
    ).grid(row=0, column=0, padx=(20, 8), pady=18)
    entrada_periodo = ctk.CTkEntry(
        controles,
        width=130,
        height=40,
        placeholder_text="MM-AAAA",
        border_color=COLOR_BORDE,
        fg_color=COLOR_PANEL_SECUNDARIO,
        text_color=COLOR_TEXTO,
    )
    entrada_periodo.insert(0, datetime.now().strftime("%m-%Y"))
    entrada_periodo.grid(row=0, column=1, padx=(0, 18), pady=18)

    ctk.CTkLabel(
        controles,
        text="Vista",
        font=ctk.CTkFont(size=12, weight="bold"),
        text_color=COLOR_TEXTO,
    ).grid(row=0, column=2, padx=(0, 8), pady=18)
    selector_unidad = ctk.CTkComboBox(
        controles,
        values=["Empresa completa", *Movimientos.UNIDADES],
        width=190,
        height=40,
        state="readonly",
        border_color=COLOR_BORDE,
        fg_color=COLOR_PANEL_SECUNDARIO,
        button_color=COLOR_PRIMARIO,
        button_hover_color=COLOR_PRIMARIO_HOVER,
        dropdown_fg_color=COLOR_PANEL,
        dropdown_text_color=COLOR_TEXTO,
        text_color=COLOR_TEXTO,
    )
    selector_unidad.set("Empresa completa")
    selector_unidad.grid(row=0, column=3, padx=(0, 18), pady=18)

    comparador = ctk.CTkFrame(
        controles,
        fg_color=COLOR_PANEL_SECUNDARIO,
        corner_radius=12,
    )
    comparador.grid(
        row=1,
        column=0,
        columnspan=6,
        sticky="ew",
        padx=20,
        pady=(0, 12),
    )
    comparador.grid_columnconfigure(6, weight=1)
    ctk.CTkLabel(
        comparador,
        text="Comparar meses",
        font=ctk.CTkFont(size=12, weight="bold"),
        text_color=COLOR_TEXTO,
    ).grid(row=0, column=0, padx=(14, 10), pady=12)
    ctk.CTkLabel(
        comparador,
        text="Desde",
        font=ctk.CTkFont(size=11),
        text_color=COLOR_TEXTO_SUAVE,
    ).grid(row=0, column=1, padx=(0, 6), pady=12)
    entrada_desde = ctk.CTkEntry(
        comparador,
        width=110,
        height=36,
        placeholder_text="MM-AAAA",
        border_color=COLOR_BORDE,
        fg_color=COLOR_PANEL,
        text_color=COLOR_TEXTO,
    )
    entrada_desde.insert(0, f"01-{datetime.now().year}")
    entrada_desde.grid(row=0, column=2, padx=(0, 12), pady=12)
    ctk.CTkLabel(
        comparador,
        text="Hasta",
        font=ctk.CTkFont(size=11),
        text_color=COLOR_TEXTO_SUAVE,
    ).grid(row=0, column=3, padx=(0, 6), pady=12)
    entrada_hasta = ctk.CTkEntry(
        comparador,
        width=110,
        height=36,
        placeholder_text="MM-AAAA",
        border_color=COLOR_BORDE,
        fg_color=COLOR_PANEL,
        text_color=COLOR_TEXTO,
    )
    entrada_hasta.insert(0, datetime.now().strftime("%m-%Y"))
    entrada_hasta.grid(row=0, column=4, padx=(0, 12), pady=12)

    anual = ctk.CTkFrame(
        controles,
        fg_color=COLOR_PANEL_SECUNDARIO,
        corner_radius=12,
    )
    anual.grid(
        row=2,
        column=0,
        columnspan=6,
        sticky="ew",
        padx=20,
        pady=(0, 12),
    )
    anual.grid_columnconfigure(3, weight=1)
    ctk.CTkLabel(
        anual,
        text="Informe anual",
        font=ctk.CTkFont(size=12, weight="bold"),
        text_color=COLOR_TEXTO,
    ).grid(row=0, column=0, padx=(14, 10), pady=12)
    ctk.CTkLabel(
        anual,
        text="Año",
        font=ctk.CTkFont(size=11),
        text_color=COLOR_TEXTO_SUAVE,
    ).grid(row=0, column=1, padx=(0, 6), pady=12)
    entrada_anio = ctk.CTkEntry(
        anual,
        width=100,
        height=36,
        placeholder_text="AAAA",
        border_color=COLOR_BORDE,
        fg_color=COLOR_PANEL,
        text_color=COLOR_TEXTO,
    )
    entrada_anio.insert(0, str(datetime.now().year))
    entrada_anio.grid(row=0, column=2, padx=(0, 12), pady=12)

    zona_resultado = ctk.CTkFrame(
        pagina,
        fg_color="transparent",
    )
    zona_resultado.pack(
        fill="both",
        expand=True,
        padx=34,
        pady=(0, 30),
    )

    boton_excel = None
    boton_pdf = None

    def generar():
        try:
            informe = obtener_informe(
                entrada_periodo.get(),
                selector_unidad.get(),
            )
        except (ValueError, OSError, IndexError, TypeError) as error:
            from tkinter import messagebox

            messagebox.showerror(
                "No se pudo generar el informe",
                str(error),
                parent=aplicacion,
            )
            return

        entrada_periodo.delete(0, "end")
        entrada_periodo.insert(0, informe["periodo"])
        estado_informe["actual"] = informe
        estado_informe["tipo"] = "mensual"
        boton_excel.configure(state="normal")
        boton_pdf.configure(state="normal")
        dibujar_informe(zona_resultado, informe)

    def generar_comparacion():
        try:
            comparacion = obtener_comparacion_mensual(
                entrada_desde.get(),
                entrada_hasta.get(),
                selector_unidad.get(),
            )
        except (ValueError, OSError, IndexError, TypeError) as error:
            messagebox.showerror(
                "No se pudo comparar los meses",
                str(error),
                parent=aplicacion,
            )
            return

        entrada_desde.delete(0, "end")
        entrada_desde.insert(0, comparacion["periodo_desde"])
        entrada_hasta.delete(0, "end")
        entrada_hasta.insert(0, comparacion["periodo_hasta"])
        estado_informe["actual"] = comparacion
        estado_informe["tipo"] = "comparacion"
        boton_excel.configure(state="normal")
        boton_pdf.configure(state="normal")
        dibujar_comparacion_mensual(
            zona_resultado,
            comparacion,
        )

    def generar_anual():
        try:
            informe_anual = obtener_informe_anual(
                entrada_anio.get(),
                selector_unidad.get(),
            )
        except (ValueError, OSError, IndexError, TypeError) as error:
            messagebox.showerror(
                "No se pudo generar el informe anual",
                str(error),
                parent=aplicacion,
            )
            return

        entrada_anio.delete(0, "end")
        entrada_anio.insert(0, str(informe_anual["anio"]))
        estado_informe["actual"] = informe_anual
        estado_informe["tipo"] = "anual"
        boton_excel.configure(state="normal")
        boton_pdf.configure(state="normal")
        dibujar_informe_anual(zona_resultado, informe_anual)

    def guardar_exportacion(tipo):
        informe = estado_informe["actual"]
        if informe is None:
            generar()
            informe = estado_informe["actual"]
        if informe is None:
            return

        extension = "xlsx" if tipo == "excel" else "pdf"
        tipo_archivo = (
            [("Libro de Excel", "*.xlsx")]
            if tipo == "excel"
            else [("Documento PDF", "*.pdf")]
        )
        es_comparacion = estado_informe["tipo"] == "comparacion"
        es_anual = estado_informe["tipo"] == "anual"
        nombre_inicial = (
            nombre_archivo_anual(informe, extension)
            if es_anual
            else
            nombre_archivo_comparacion(informe, extension)
            if es_comparacion
            else nombre_archivo_seguro(informe, extension)
        )
        ruta = filedialog.asksaveasfilename(
            parent=aplicacion,
            title=(
                "Guardar informe anual en Excel"
                if es_anual and tipo == "excel"
                else "Guardar informe anual en PDF"
                if es_anual
                else
                "Guardar comparación en Excel"
                if es_comparacion and tipo == "excel"
                else "Guardar comparación en PDF"
                if es_comparacion
                else "Guardar informe en Excel"
                if tipo == "excel"
                else "Guardar informe en PDF"
            ),
            defaultextension=f".{extension}",
            initialfile=nombre_inicial,
            filetypes=tipo_archivo,
        )
        if not ruta:
            return

        try:
            if es_anual and tipo == "excel":
                exportar_excel_anual(informe, ruta)
            elif es_anual:
                exportar_pdf_anual(informe, ruta)
            elif es_comparacion and tipo == "excel":
                exportar_excel_comparacion(informe, ruta)
            elif es_comparacion:
                exportar_pdf_comparacion(informe, ruta)
            elif tipo == "excel":
                exportar_excel(informe, ruta)
            else:
                exportar_pdf(informe, ruta)
        except (OSError, RuntimeError, ValueError, TypeError) as error:
            messagebox.showerror(
                "No se pudo guardar el informe",
                str(error),
                parent=aplicacion,
            )
            return

        messagebox.showinfo(
            "Informe guardado",
            f"El archivo se guardó correctamente en:\n{ruta}",
            parent=aplicacion,
        )

    ctk.CTkButton(
        controles,
        text="Generar informe",
        command=generar,
        width=150,
        height=40,
        corner_radius=9,
        fg_color=COLOR_PRIMARIO,
        hover_color=COLOR_PRIMARIO_HOVER,
        font=ctk.CTkFont(size=12, weight="bold"),
    ).grid(row=0, column=5, padx=(0, 20), pady=18)

    ctk.CTkButton(
        comparador,
        text="Generar comparación",
        command=generar_comparacion,
        width=165,
        height=36,
        corner_radius=9,
        fg_color="#7B61FF",
        hover_color="#654BCF",
        font=ctk.CTkFont(size=11, weight="bold"),
    ).grid(row=0, column=7, padx=(12, 14), pady=12)

    ctk.CTkButton(
        anual,
        text="Generar informe anual",
        command=generar_anual,
        width=175,
        height=36,
        corner_radius=9,
        fg_color=COLOR_NARANJA,
        hover_color="#C77C25",
        font=ctk.CTkFont(size=11, weight="bold"),
    ).grid(row=0, column=4, padx=(12, 14), pady=12)

    acciones_exportacion = ctk.CTkFrame(
        controles,
        fg_color="transparent",
    )
    acciones_exportacion.grid(
        row=3,
        column=0,
        columnspan=6,
        sticky="e",
        padx=20,
        pady=(0, 18),
    )

    ctk.CTkLabel(
        acciones_exportacion,
        text=(
            "Exporta la vista actual · "
            "Atajos: Ctrl+E Excel · Ctrl+P PDF"
        ),
        font=ctk.CTkFont(size=11),
        text_color=COLOR_TEXTO_SUAVE,
    ).pack(side="left", padx=(0, 14))

    boton_excel = ctk.CTkButton(
        acciones_exportacion,
        text="Descargar Excel",
        command=lambda: guardar_exportacion("excel"),
        width=145,
        height=38,
        corner_radius=9,
        fg_color=COLOR_VERDE,
        hover_color="#12855C",
        font=ctk.CTkFont(size=12, weight="bold"),
        state="disabled",
    )
    boton_excel.pack(side="left", padx=(0, 8))

    boton_pdf = ctk.CTkButton(
        acciones_exportacion,
        text="Descargar PDF",
        command=lambda: guardar_exportacion("pdf"),
        width=145,
        height=38,
        corner_radius=9,
        fg_color=COLOR_ROJO,
        hover_color="#BD4040",
        font=ctk.CTkFont(size=12, weight="bold"),
        state="disabled",
    )
    boton_pdf.pack(side="left")

    entrada_periodo.bind("<Return>", lambda _evento: generar())
    entrada_desde.bind(
        "<Return>",
        lambda _evento: generar_comparacion(),
    )
    entrada_hasta.bind(
        "<Return>",
        lambda _evento: generar_comparacion(),
    )
    entrada_anio.bind(
        "<Return>",
        lambda _evento: generar_anual(),
    )

    def actualizar_vista(_valor=None):
        if estado_informe["tipo"] == "comparacion":
            generar_comparacion()
        elif estado_informe["tipo"] == "anual":
            generar_anual()
        else:
            generar()

    selector_unidad.configure(command=actualizar_vista)
    aplicacion.bind(
        "<Control-e>",
        lambda _evento: guardar_exportacion("excel"),
    )
    aplicacion.bind(
        "<Control-E>",
        lambda _evento: guardar_exportacion("excel"),
    )
    aplicacion.bind(
        "<Control-p>",
        lambda _evento: guardar_exportacion("pdf"),
    )
    aplicacion.bind(
        "<Control-P>",
        lambda _evento: guardar_exportacion("pdf"),
    )
    generar()
