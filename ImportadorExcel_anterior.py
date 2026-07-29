"""Importación mensual validada desde Excel para BC Gestión.

La importación trabaja exclusivamente con la plantilla oficial. Primero
construye una vista previa y recién después, con confirmación del usuario,
agrega los registros nuevos. Los registros idénticos ya existentes se omiten.
"""

from __future__ import annotations

from datetime import date, datetime
import hashlib
import os
from pathlib import Path
import re
import subprocess
import unicodedata
from uuid import uuid4
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook

import Liquidaciones
import Movimientos
import Socios
from datos import guardar_datos, leer_datos


BASE_DIR = Path(__file__).resolve().parent
PLANTILLA = BASE_DIR / "Plantilla_Carga_Mensual_BC_Gestion.xlsx"
MARCADOR_PLANTILLA = "BC_GESTION_IMPORT_V1"

COLOR_PANEL = ("#FFFFFF", "#131D2E")
COLOR_PANEL_SECUNDARIO = ("#EAF0F7", "#1A2639")
COLOR_BORDE = ("#DCE4EE", "#26354A")
COLOR_TEXTO = ("#182230", "#F4F7FB")
COLOR_TEXTO_SUAVE = ("#617084", "#9CAFC5")
COLOR_PRIMARIO = "#246BFD"
COLOR_PRIMARIO_HOVER = "#1855D6"
COLOR_VERDE = "#18A874"
COLOR_ROJO = "#E25555"

RUTAS = {
    "Movimientos": Path(Movimientos.RUTA_MOVIMIENTOS),
    "Adicionales": Path(Movimientos.RUTA_ADICIONALES),
    "Conceptos": Path(Movimientos.RUTA_CONCEPTOS_ADICIONALES),
    "Inversiones": Path(Movimientos.RUTA_INVERSIONES),
    "Prestamos": Path(Movimientos.RUTA_PRESTAMOS),
    "Cuotas": Path(Movimientos.RUTA_CUOTAS_PRESTAMOS),
    "Retiros": Path(Socios.RUTA_RETIROS_SOCIOS),
    "Liquidaciones": Path(Liquidaciones.RUTA_LIQUIDACIONES),
}

NOMBRES_RESUMEN = {
    "Movimientos": "Movimientos",
    "Adicionales": "Adicionales",
    "Inversiones": "Inversiones",
    "Prestamos": "Préstamos",
    "Cuotas": "Cuotas",
    "Retiros": "Retiros de socios",
    "Liquidaciones": "Liquidaciones",
}


def normalizar(valor):
    texto = "" if valor is None else str(valor)
    texto = texto.strip().lower()
    return "".join(
        caracter
        for caracter in unicodedata.normalize("NFD", texto)
        if unicodedata.category(caracter) != "Mn"
    )


def texto_seguro(valor):
    texto = "" if valor is None else str(valor)
    return " ".join(
        texto.replace("|", "/").replace("\r", " ").replace("\n", " ").split()
    )


def celda_vacia(valor):
    return valor is None or str(valor).strip() == ""


def parsear_monto(valor, permitir_cero=False):
    if isinstance(valor, bool) or celda_vacia(valor):
        raise ValueError("el monto está vacío")

    if isinstance(valor, (int, float)):
        numero = float(valor)
        if not numero.is_integer():
            raise ValueError("el monto debe ser un número entero")
        monto = int(numero)
    else:
        texto = str(valor).strip()
        texto = re.sub(r"(?i)g(?:s)?\.?", "", texto)
        texto = texto.replace("₲", "").replace(" ", "")

        if re.fullmatch(r"-?\d+[.,]00", texto):
            texto = texto[:-3]
        else:
            texto = texto.replace(".", "").replace(",", "")

        if not re.fullmatch(r"-?\d+", texto):
            raise ValueError("el monto contiene caracteres no válidos")
        monto = int(texto)

    minimo = 0 if permitir_cero else 1
    if monto < minimo:
        if permitir_cero:
            raise ValueError("el monto no puede ser negativo")
        raise ValueError("el monto debe ser mayor que cero")
    return monto


def parsear_entero(valor, nombre, minimo=0, predeterminado=None):
    if celda_vacia(valor):
        if predeterminado is not None:
            return predeterminado
        raise ValueError(f"{nombre} está vacío")

    if isinstance(valor, bool):
        raise ValueError(f"{nombre} no es válido")

    try:
        numero = float(valor)
    except (TypeError, ValueError):
        raise ValueError(f"{nombre} debe ser un número entero") from None

    if not numero.is_integer() or int(numero) < minimo:
        raise ValueError(f"{nombre} debe ser un entero desde {minimo}")
    return int(numero)


def parsear_fecha(valor):
    if isinstance(valor, datetime):
        fecha = valor.date()
    elif isinstance(valor, date):
        fecha = valor
    else:
        texto = texto_seguro(valor)
        formatos = (
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%Y-%m-%d",
            "%Y/%m/%d",
        )
        fecha = None
        for formato in formatos:
            try:
                fecha = datetime.strptime(texto, formato).date()
                break
            except ValueError:
                continue
        if fecha is None:
            raise ValueError("fecha inválida; usá DD-MM-AAAA")
    return fecha.strftime("%d-%m-%Y"), fecha.strftime("%m-%Y")


def parsear_periodo(valor):
    if isinstance(valor, (date, datetime)):
        return valor.strftime("%m-%Y")

    texto = texto_seguro(valor)
    for formato in ("%m-%Y", "%m/%Y", "%Y-%m", "%Y/%m"):
        try:
            return datetime.strptime(texto, formato).strftime("%m-%Y")
        except ValueError:
            continue
    raise ValueError("período inválido; usá MM-AAAA")


def obtener_hoja(libro, nombre):
    objetivo = normalizar(nombre)
    for hoja in libro.worksheets:
        if normalizar(hoja.title) == objetivo:
            return hoja
    return None


def buscar_encabezados(hoja, requeridos):
    requeridos_normalizados = {normalizar(valor) for valor in requeridos}

    for numero_fila in range(1, min(hoja.max_row, 15) + 1):
        columnas = {}
        for numero_columna, celda in enumerate(
            hoja[numero_fila],
            start=1,
        ):
            encabezado = normalizar(celda.value)
            if encabezado:
                columnas[encabezado] = numero_columna

        if requeridos_normalizados.issubset(columnas):
            return numero_fila, columnas
    return None, {}


def valor_fila(fila, columnas, nombre, opcionales=()):
    candidatos = (nombre, *opcionales)
    for candidato in candidatos:
        indice = columnas.get(normalizar(candidato))
        if indice is not None:
            return fila[indice - 1].value
    return None


def filas_de_hoja(libro, nombre, encabezados_requeridos):
    hoja = obtener_hoja(libro, nombre)
    if hoja is None:
        return None, None, []

    fila_encabezado, columnas = buscar_encabezados(
        hoja,
        encabezados_requeridos,
    )
    if fila_encabezado is None:
        return hoja, None, []

    filas = []
    for numero in range(fila_encabezado + 1, hoja.max_row + 1):
        fila = hoja[numero]
        if all(celda_vacia(celda.value) for celda in fila):
            continue
        filas.append((numero, fila))
    return hoja, columnas, filas


def clave_movimiento(datos):
    return (
        normalizar(datos["tipo"]),
        datos["fecha"],
        normalizar(datos["origen"]),
        normalizar(datos["destino"]),
        datos["monto"],
    )


def clave_adicional(datos):
    return (
        normalizar(datos["tipo"]),
        datos["fecha"],
        normalizar(datos["descripcion"]),
        datos["monto"],
        normalizar(datos["observacion"]),
    )


def clave_inversion(datos):
    return (
        datos["fecha"],
        normalizar(datos["descripcion"]),
        datos["monto"],
    )


def clave_prestamo(datos):
    return (
        normalizar(datos["descripcion"]),
        normalizar(datos["banco"]),
        datos["fecha"],
        datos["monto_recibido"],
        datos["costo_total"],
        datos["cantidad_cuotas"],
    )


def clave_retiro(datos):
    return (
        datos["fecha"],
        datos["socio_id"],
        datos["monto"],
        normalizar(
            datos.get(
                "tipo",
                Socios.TIPO_MOVIMIENTO_PREDETERMINADO,
            )
        ),
        normalizar(datos["observacion"]),
    )


def crear_resultado(ruta):
    return {
        "ruta": str(ruta),
        "periodos": set(),
        "errores": [],
        "advertencias": [],
        "duplicados": [],
        "correcciones": [],
        "lineas": {nombre: [] for nombre in RUTAS},
        "conteos": {nombre: 0 for nombre in NOMBRES_RESUMEN},
        "totales": {nombre: 0 for nombre in NOMBRES_RESUMEN},
        "resumen_filas": [],
        "totales_retiros_tipo": {
            tipo: 0
            for tipo in Socios.TIPOS_MOVIMIENTO_PERSONAL
        },
        "huella": None,
    }


def error(resultado, hoja, fila, mensaje):
    resultado["errores"].append(f"{hoja}, fila {fila}: {mensaje}")


def duplicado(resultado, hoja, fila, mensaje):
    resultado["duplicados"].append(f"{hoja}, fila {fila}: {mensaje}")


def agregar_periodo(resultado, periodo):
    resultado["periodos"].add(periodo)


def catalogo_unidades():
    return {
        normalizar(unidad): unidad
        for unidad in Movimientos.UNIDADES
    }


def catalogo_socios():
    resultado = {}
    for socio in Socios.SOCIOS:
        resultado[normalizar(socio["id"])] = socio["id"]
        resultado[normalizar(socio["nombre"])] = socio["id"]
    return resultado


def analizar_movimientos(libro, resultado):
    hoja, columnas, filas = filas_de_hoja(
        libro,
        "Movimientos",
        ["Fecha", "Tipo", "Origen", "Destino", "Monto"],
    )
    if hoja is None:
        resultado["advertencias"].append(
            "No se encontró la hoja Movimientos."
        )
        return
    if columnas is None:
        resultado["errores"].append(
            "La hoja Movimientos no tiene los encabezados esperados."
        )
        return

    existentes = set()
    movimientos_existentes = []
    for linea in leer_datos(Movimientos.RUTA_MOVIMIENTOS):
        datos = Movimientos.separar_movimiento(linea)
        if datos is not None:
            existentes.add(clave_movimiento(datos))
            movimientos_existentes.append((linea, datos))

    nuevas = set()
    unidades = catalogo_unidades()
    tipos = {
        "ingreso": "Ingreso",
        "egreso": "Egreso",
        "transferencia interna": "Transferencia interna",
        "deposito interno": "Deposito interno",
        "deposito bancario": "Deposito interno",
        "deposito externo": "Cobro externo",
        "cobro externo": "Cobro externo",
    }

    for numero, fila in filas:
        try:
            fecha, periodo = parsear_fecha(
                valor_fila(fila, columnas, "Fecha")
            )
            agregar_periodo(resultado, periodo)
            tipo_texto = normalizar(
                valor_fila(fila, columnas, "Tipo")
            )
            if tipo_texto not in tipos:
                raise ValueError(
                    "tipo inválido; usá Ingreso, Egreso, "
                    "Transferencia interna, Depósito interno, "
                    "Depósito externo o Cobro externo"
                )

            tipo = tipos[tipo_texto]
            origen = texto_seguro(
                valor_fila(fila, columnas, "Origen")
            )
            destino = texto_seguro(
                valor_fila(fila, columnas, "Destino")
            )
            monto = parsear_monto(
                valor_fila(fila, columnas, "Monto")
            )

            if tipo == "Ingreso":
                origen = origen or "Externo"
                destino = unidades.get(normalizar(destino), "")
                if not destino:
                    raise ValueError("Destino debe ser una unidad válida")
                linea = f"Ingreso|{fecha}|{origen}|{destino}|{monto}|Si"
            elif tipo == "Egreso":
                origen = unidades.get(normalizar(origen), "")
                destino = destino or "Externo"
                if not origen:
                    raise ValueError("Origen debe ser una unidad válida")
                linea = f"Egreso|{fecha}|{origen}|{destino}|{monto}|Si"
            elif tipo == "Transferencia interna":
                origen = unidades.get(normalizar(origen), "")
                destino = unidades.get(normalizar(destino), "")
                if not origen or not destino:
                    raise ValueError(
                        "Origen y Destino deben ser unidades válidas"
                    )
                if origen == destino:
                    raise ValueError(
                        "Origen y Destino no pueden ser iguales"
                    )
                linea = (
                    f"Transferencia interna|{fecha}|{origen}|"
                    f"{destino}|{monto}|No"
                )
            elif tipo == "Deposito interno":
                if not origen or not destino:
                    raise ValueError(
                        "completá quién deposita y el banco de destino"
                    )
                origen = unidades.get(normalizar(origen), origen)
                linea = (
                    f"Deposito interno|{fecha}|{origen}|"
                    f"{destino}|{monto}|No"
                )
            else:
                if not origen:
                    raise ValueError("completá el banco que recibió")
                destino = unidades.get(normalizar(destino), "")
                if not destino:
                    raise ValueError("Destino debe ser una unidad válida")
                linea = (
                    f"Cobro externo|{fecha}|{origen}|"
                    f"{destino}|{monto}|Si"
                )

            datos = Movimientos.separar_movimiento(linea)
            clave = clave_movimiento(datos)
            if clave in existentes or clave in nuevas:
                duplicado(
                    resultado,
                    hoja.title,
                    numero,
                    "el movimiento ya existe y se omitirá",
                )
                continue

            if tipo == "Cobro externo":
                candidatos = []
                for (
                    linea_existente,
                    datos_existentes,
                ) in movimientos_existentes:
                    tipo_existente = datos_existentes["tipo"]
                    if tipo_existente not in (
                        "Cobro externo",
                        "Deposito interno",
                    ):
                        continue
                    misma_fecha = (
                        datos_existentes["fecha"] == datos["fecha"]
                    )
                    mismo_monto = (
                        datos_existentes["monto"] == datos["monto"]
                    )
                    if tipo_existente == "Cobro externo":
                        misma_unidad = (
                            normalizar(datos_existentes["destino"])
                            == normalizar(datos["destino"])
                        )
                        banco_anterior = normalizar(
                            datos_existentes["origen"]
                        )
                    else:
                        misma_unidad = (
                            normalizar(datos_existentes["origen"])
                            == normalizar(datos["destino"])
                        )
                        banco_anterior = normalizar(
                            datos_existentes["destino"]
                        )

                    banco_sin_especificar = (
                        "no especificado" in banco_anterior
                    )
                    mismo_banco = (
                        banco_anterior
                        == normalizar(datos["origen"])
                    )
                    if (
                        misma_fecha
                        and mismo_monto
                        and misma_unidad
                        and (mismo_banco or banco_sin_especificar)
                        and linea_existente != linea
                    ):
                        candidatos.append(
                            (linea_existente, tipo_existente)
                        )

                if len(candidatos) > 1:
                    raise ValueError(
                        "hay más de un movimiento que podría ser este "
                        "depósito externo; corregilo manualmente para evitar "
                        "reemplazar el movimiento equivocado"
                    )
                if len(candidatos) == 1:
                    linea_anterior, tipo_anterior = candidatos[0]
                    resultado["correcciones"].append(
                        {
                            "hoja": hoja.title,
                            "fila": numero,
                            "ruta": "Movimientos",
                            "anterior": linea_anterior,
                            "nuevo": linea,
                            "detalle": (
                                f"{fecha}: {tipo_anterior} → "
                                f"Cobro externo en {origen} "
                                f"(Gs. {Movimientos.formatear_monto(monto)})"
                            ),
                        }
                    )
                    movimientos_existentes = [
                        (linea_actual, datos_actuales)
                        for linea_actual, datos_actuales
                        in movimientos_existentes
                        if linea_actual != linea_anterior
                    ]
                    movimientos_existentes.append((linea, datos))
                    existentes.discard(
                        clave_movimiento(
                            Movimientos.separar_movimiento(linea_anterior)
                        )
                    )
                    existentes.add(clave)
                    agregar_periodo(resultado, periodo)
                    continue

            nuevas.add(clave)
            resultado["lineas"]["Movimientos"].append(linea)
            resultado["conteos"]["Movimientos"] += 1
            resultado["totales"]["Movimientos"] += monto
            agregar_periodo(resultado, periodo)
        except ValueError as exc:
            error(resultado, hoja.title, numero, str(exc))


def analizar_adicionales(libro, resultado):
    hoja, columnas, filas = filas_de_hoja(
        libro,
        "Adicionales",
        ["Fecha", "Tipo", "Concepto", "Monto"],
    )
    if hoja is None:
        resultado["advertencias"].append(
            "No se encontró la hoja Adicionales."
        )
        return
    if columnas is None:
        resultado["errores"].append(
            "La hoja Adicionales no tiene los encabezados esperados."
        )
        return

    existentes = set()
    for linea in leer_datos(Movimientos.RUTA_ADICIONALES):
        datos = Movimientos.separar_adicional(linea)
        if datos is not None:
            existentes.add(clave_adicional(datos))
    nuevas = set()

    catalogo_lineas = leer_datos(
        Movimientos.RUTA_CONCEPTOS_ADICIONALES
    )
    conceptos = {
        (normalizar(partes[0]), normalizar(partes[1]))
        for linea in catalogo_lineas
        if len(partes := linea.split("|", 1)) == 2
    }
    conceptos_nuevos = set()

    for numero, fila in filas:
        try:
            fecha, periodo = parsear_fecha(
                valor_fila(fila, columnas, "Fecha")
            )
            agregar_periodo(resultado, periodo)
            tipo_normalizado = normalizar(
                valor_fila(fila, columnas, "Tipo")
            )
            if tipo_normalizado not in ("ingreso", "egreso"):
                raise ValueError("Tipo debe ser Ingreso o Egreso")
            tipo = "Ingreso" if tipo_normalizado == "ingreso" else "Egreso"
            concepto = texto_seguro(
                valor_fila(fila, columnas, "Concepto")
            )
            if not concepto:
                raise ValueError("Concepto está vacío")
            monto = parsear_monto(
                valor_fila(fila, columnas, "Monto")
            )
            observacion = texto_seguro(
                valor_fila(
                    fila,
                    columnas,
                    "Observación",
                    ("Observacion",),
                )
            )
            linea = (
                f"{tipo}|{fecha}|{concepto}|{monto}|{observacion}"
            )
            datos = Movimientos.separar_adicional(linea)
            clave = clave_adicional(datos)
            if clave in existentes or clave in nuevas:
                duplicado(
                    resultado,
                    hoja.title,
                    numero,
                    "el adicional ya existe y se omitirá",
                )
                continue

            nuevas.add(clave)
            resultado["lineas"]["Adicionales"].append(linea)
            resultado["conteos"]["Adicionales"] += 1
            resultado["totales"]["Adicionales"] += monto
            agregar_periodo(resultado, periodo)

            clave_concepto = (
                normalizar(tipo),
                normalizar(concepto),
            )
            if (
                clave_concepto not in conceptos
                and clave_concepto not in conceptos_nuevos
            ):
                conceptos_nuevos.add(clave_concepto)
                resultado["lineas"]["Conceptos"].append(
                    f"{tipo}|{concepto}"
                )
        except ValueError as exc:
            error(resultado, hoja.title, numero, str(exc))


def analizar_inversiones(libro, resultado):
    hoja, columnas, filas = filas_de_hoja(
        libro,
        "Inversiones",
        ["Fecha", "Descripción", "Monto"],
    )
    if hoja is None:
        resultado["advertencias"].append(
            "No se encontró la hoja Inversiones."
        )
        return
    if columnas is None:
        resultado["errores"].append(
            "La hoja Inversiones no tiene los encabezados esperados."
        )
        return

    existentes = set()
    for linea in leer_datos(Movimientos.RUTA_INVERSIONES):
        datos = Movimientos.separar_inversion(linea)
        if datos is not None:
            existentes.add(clave_inversion(datos))
    nuevas = set()

    for numero, fila in filas:
        try:
            fecha, periodo = parsear_fecha(
                valor_fila(fila, columnas, "Fecha")
            )
            agregar_periodo(resultado, periodo)
            descripcion = texto_seguro(
                valor_fila(
                    fila,
                    columnas,
                    "Descripción",
                    ("Descripcion",),
                )
            )
            if not descripcion:
                raise ValueError("Descripción está vacía")
            monto = parsear_monto(
                valor_fila(fila, columnas, "Monto")
            )
            datos = {
                "fecha": fecha,
                "descripcion": descripcion,
                "monto": monto,
            }
            clave = clave_inversion(datos)
            if clave in existentes or clave in nuevas:
                duplicado(
                    resultado,
                    hoja.title,
                    numero,
                    "la inversión ya existe y se omitirá",
                )
                continue

            nuevas.add(clave)
            resultado["lineas"]["Inversiones"].append(
                f"{fecha}|{descripcion}|{monto}"
            )
            resultado["conteos"]["Inversiones"] += 1
            resultado["totales"]["Inversiones"] += monto
            agregar_periodo(resultado, periodo)
        except ValueError as exc:
            error(resultado, hoja.title, numero, str(exc))


def analizar_prestamos(libro, resultado):
    existentes = {}
    claves_existentes = {}
    mayor_id = 0
    for linea in leer_datos(Movimientos.RUTA_PRESTAMOS):
        datos = Movimientos.separar_prestamo(linea)
        if datos is None:
            continue
        existentes[datos["id"]] = datos
        claves_existentes[clave_prestamo(datos)] = datos["id"]
        if datos["id"].isdigit():
            mayor_id = max(mayor_id, int(datos["id"]))

    hoja, columnas, filas = filas_de_hoja(
        libro,
        "Prestamos",
        [
            "Referencia",
            "Descripción",
            "Banco",
            "Fecha recepción",
            "Monto recibido",
            "Total a devolver",
            "Cantidad cuotas",
        ],
    )
    referencias = {}
    prestamos_nuevos = {}
    claves_nuevas = {}

    if hoja is None:
        resultado["advertencias"].append(
            "No se encontró la hoja Prestamos."
        )
    elif columnas is None:
        resultado["errores"].append(
            "La hoja Prestamos no tiene los encabezados esperados."
        )
    else:
        for numero, fila in filas:
            try:
                referencia = texto_seguro(
                    valor_fila(fila, columnas, "Referencia")
                )
                if not referencia:
                    raise ValueError(
                        "Referencia está vacía; usá por ejemplo PRESTAMO1"
                    )
                ref_normalizada = normalizar(referencia)
                if (
                    ref_normalizada in referencias
                    or referencia in existentes
                ):
                    raise ValueError("Referencia repetida o ya utilizada")

                descripcion = texto_seguro(
                    valor_fila(
                        fila,
                        columnas,
                        "Descripción",
                        ("Descripcion",),
                    )
                )
                banco = texto_seguro(
                    valor_fila(fila, columnas, "Banco")
                )
                if not descripcion or not banco:
                    raise ValueError("completá Descripción y Banco")
                fecha, periodo = parsear_fecha(
                    valor_fila(
                        fila,
                        columnas,
                        "Fecha recepción",
                        ("Fecha recepcion",),
                    )
                )
                agregar_periodo(resultado, periodo)
                monto_recibido = parsear_monto(
                    valor_fila(fila, columnas, "Monto recibido")
                )
                costo_total = parsear_monto(
                    valor_fila(fila, columnas, "Total a devolver")
                )
                cantidad_cuotas = parsear_entero(
                    valor_fila(fila, columnas, "Cantidad cuotas"),
                    "Cantidad cuotas",
                    minimo=1,
                )
                if costo_total < monto_recibido:
                    raise ValueError(
                        "Total a devolver no puede ser menor al monto recibido"
                    )

                datos_sin_id = {
                    "descripcion": descripcion,
                    "banco": banco,
                    "fecha": fecha,
                    "monto_recibido": monto_recibido,
                    "costo_total": costo_total,
                    "cantidad_cuotas": cantidad_cuotas,
                }
                clave = clave_prestamo(datos_sin_id)
                if clave in claves_existentes:
                    prestamo_id = claves_existentes[clave]
                    referencias[ref_normalizada] = prestamo_id
                    duplicado(
                        resultado,
                        hoja.title,
                        numero,
                        f"el préstamo ya existe con ID {prestamo_id}",
                    )
                    continue
                if clave in claves_nuevas:
                    referencias[ref_normalizada] = claves_nuevas[clave]
                    duplicado(
                        resultado,
                        hoja.title,
                        numero,
                        "el préstamo está repetido en el Excel",
                    )
                    continue

                mayor_id += 1
                prestamo_id = str(mayor_id)
                datos = {"id": prestamo_id, **datos_sin_id}
                referencias[ref_normalizada] = prestamo_id
                claves_nuevas[clave] = prestamo_id
                prestamos_nuevos[prestamo_id] = datos
                existentes[prestamo_id] = datos
                resultado["lineas"]["Prestamos"].append(
                    f"{prestamo_id}|{descripcion}|{banco}|{fecha}|"
                    f"{monto_recibido}|{costo_total}|{cantidad_cuotas}"
                )
                resultado["conteos"]["Prestamos"] += 1
                resultado["totales"]["Prestamos"] += monto_recibido
                agregar_periodo(resultado, periodo)
            except ValueError as exc:
                error(resultado, hoja.title, numero, str(exc))

    return existentes, referencias, prestamos_nuevos


def analizar_cuotas(
    libro,
    resultado,
    prestamos,
    referencias,
):
    hoja, columnas, filas = filas_de_hoja(
        libro,
        "Cuotas",
        ["ID o referencia", "Número cuota", "Fecha pago", "Monto"],
    )
    if hoja is None:
        resultado["advertencias"].append(
            "No se encontró la hoja Cuotas."
        )
        return
    if columnas is None:
        resultado["errores"].append(
            "La hoja Cuotas no tiene los encabezados esperados."
        )
        return

    numeros_pagados = {}
    total_pagado = {}
    adicionales_existentes = [
        (linea, Movimientos.separar_adicional(linea))
        for linea in leer_datos(Movimientos.RUTA_ADICIONALES)
    ]
    for linea in leer_datos(Movimientos.RUTA_CUOTAS_PRESTAMOS):
        cuota = Movimientos.separar_cuota(linea)
        if cuota is None:
            continue
        numeros_pagados.setdefault(cuota["prestamo_id"], set()).add(
            cuota["numero"]
        )
        total_pagado[cuota["prestamo_id"]] = (
            total_pagado.get(cuota["prestamo_id"], 0)
            + cuota["monto"]
        )

    for numero_fila, fila in filas:
        try:
            referencia = texto_seguro(
                valor_fila(fila, columnas, "ID o referencia")
            )
            prestamo_id = referencias.get(
                normalizar(referencia),
                referencia,
            )
            es_historica = normalizar(referencia).startswith("hist-")
            if prestamo_id not in prestamos and not es_historica:
                raise ValueError(
                    "el ID o referencia no corresponde a un préstamo; "
                    "para una cuota histórica sin datos originales usá "
                    "una referencia que empiece con HIST-"
                )
            prestamo = prestamos.get(prestamo_id)
            numero_cuota = parsear_entero(
                valor_fila(
                    fila,
                    columnas,
                    "Número cuota",
                    ("Numero cuota",),
                ),
                "Número cuota",
                minimo=1,
            )
            if (
                prestamo is not None
                and numero_cuota > prestamo["cantidad_cuotas"]
            ):
                raise ValueError(
                    "el número supera la cantidad total de cuotas"
                )
            if numero_cuota in numeros_pagados.setdefault(
                prestamo_id,
                set(),
            ):
                duplicado(
                    resultado,
                    hoja.title,
                    numero_fila,
                    "esa cuota ya está registrada y se omitirá",
                )
                continue

            fecha, periodo = parsear_fecha(
                valor_fila(fila, columnas, "Fecha pago")
            )
            agregar_periodo(resultado, periodo)
            monto = parsear_monto(
                valor_fila(fila, columnas, "Monto")
            )
            if prestamo is not None:
                saldo = (
                    prestamo["costo_total"]
                    - total_pagado.get(prestamo_id, 0)
                )
                if monto > saldo:
                    raise ValueError(
                        f"el monto supera el saldo disponible de Gs. "
                        f"{Movimientos.formatear_monto(max(saldo, 0))}"
                    )

            if es_historica:
                candidatos_adicional = []
                for linea_adicional, datos_adicional in (
                    adicionales_existentes
                ):
                    if datos_adicional is None:
                        continue
                    texto_adicional = normalizar(
                        " ".join(
                            [
                                datos_adicional["descripcion"],
                                datos_adicional.get("observacion", ""),
                            ]
                        )
                    )
                    if (
                        datos_adicional["tipo"] == "Egreso"
                        and datos_adicional["fecha"] == fecha
                        and datos_adicional["monto"] == monto
                        and "cuota" in texto_adicional
                        and "prestamo" in texto_adicional
                    ):
                        candidatos_adicional.append(linea_adicional)

                if len(candidatos_adicional) > 1:
                    raise ValueError(
                        "hay más de un egreso adicional que podría ser esta "
                        "cuota histórica; corregilo manualmente"
                    )
                if len(candidatos_adicional) == 1:
                    linea_adicional = candidatos_adicional[0]
                    resultado["correcciones"].append(
                        {
                            "hoja": hoja.title,
                            "fila": numero_fila,
                            "ruta": "Adicionales",
                            "anterior": linea_adicional,
                            "nuevo": "",
                            "detalle": (
                                f"{fecha}: Egreso adicional → "
                                f"Cuota histórica {referencia} "
                                f"(Gs. "
                                f"{Movimientos.formatear_monto(monto)})"
                            ),
                        }
                    )
                    adicionales_existentes = [
                        (linea_actual, datos_actuales)
                        for linea_actual, datos_actuales
                        in adicionales_existentes
                        if linea_actual != linea_adicional
                    ]

            numeros_pagados[prestamo_id].add(numero_cuota)
            total_pagado[prestamo_id] = (
                total_pagado.get(prestamo_id, 0) + monto
            )
            resultado["lineas"]["Cuotas"].append(
                f"{prestamo_id}|{numero_cuota}|{fecha}|{monto}"
            )
            resultado["conteos"]["Cuotas"] += 1
            resultado["totales"]["Cuotas"] += monto
            agregar_periodo(resultado, periodo)
        except ValueError as exc:
            error(resultado, hoja.title, numero_fila, str(exc))


def analizar_retiros(libro, resultado):
    hoja, columnas, filas = filas_de_hoja(
        libro,
        "Retiros socios",
        ["Fecha", "Socio", "Monto"],
    )
    if hoja is None:
        resultado["advertencias"].append(
            "No se encontró la hoja Retiros socios."
        )
        return
    if columnas is None:
        resultado["errores"].append(
            "La hoja Retiros socios no tiene los encabezados esperados."
        )
        return

    existentes = set()
    for linea in leer_datos(Socios.RUTA_RETIROS_SOCIOS):
        datos = Socios.separar_retiro(linea)
        if datos is not None:
            existentes.add(clave_retiro(datos))
    nuevas = set()
    socios = catalogo_socios()

    for numero, fila in filas:
        try:
            fecha, periodo = parsear_fecha(
                valor_fila(fila, columnas, "Fecha")
            )
            agregar_periodo(resultado, periodo)
            socio = socios.get(
                normalizar(valor_fila(fila, columnas, "Socio"))
            )
            if socio is None:
                raise ValueError("Socio debe ser Sol o Rodrigo")
            valor_tipo = valor_fila(
                fila,
                columnas,
                "Tipo",
                ("Tipo de retiro",),
            )
            if celda_vacia(valor_tipo):
                tipo = Socios.TIPO_MOVIMIENTO_PREDETERMINADO
            else:
                tipo = Socios.normalizar_tipo_movimiento_personal(
                    valor_tipo
                )
                if tipo is None:
                    opciones = " o ".join(
                        Socios.TIPOS_MOVIMIENTO_PERSONAL
                    )
                    raise ValueError(f"Tipo debe ser {opciones}")
            monto = parsear_monto(
                valor_fila(fila, columnas, "Monto")
            )
            observacion = texto_seguro(
                valor_fila(
                    fila,
                    columnas,
                    "Observación",
                    ("Observacion",),
                )
            ) or "-"
            datos = {
                "id": uuid4().hex[:12].upper(),
                "fecha": fecha,
                "socio_id": socio,
                "monto": monto,
                "tipo": tipo,
                "observacion": observacion,
            }
            clave = clave_retiro(datos)
            if clave in existentes or clave in nuevas:
                duplicado(
                    resultado,
                    hoja.title,
                    numero,
                    "el retiro ya existe y se omitirá",
                )
                continue

            nuevas.add(clave)
            resultado["lineas"]["Retiros"].append(
                Socios.crear_linea_retiro(datos)
            )
            resultado["conteos"]["Retiros"] += 1
            resultado["totales"]["Retiros"] += monto
            resultado["totales_retiros_tipo"][tipo] += monto
            agregar_periodo(resultado, periodo)
        except ValueError as exc:
            error(resultado, hoja.title, numero, str(exc))


def entero_opcional(fila, columnas, nombre, predeterminado=0):
    valor = valor_fila(fila, columnas, nombre)
    return parsear_entero(
        valor,
        nombre,
        minimo=0,
        predeterminado=predeterminado,
    )


def monto_opcional(fila, columnas, nombre, predeterminado=0):
    valor = valor_fila(fila, columnas, nombre)
    if celda_vacia(valor):
        return predeterminado
    return parsear_monto(valor, permitir_cero=True)


def analizar_liquidaciones(libro, resultado):
    hoja, columnas, filas = filas_de_hoja(
        libro,
        "Liquidaciones",
        ["Periodo", "Cédula", "Nombre", "Sueldo bruto"],
    )
    if hoja is None:
        resultado["advertencias"].append(
            "No se encontró la hoja Liquidaciones."
        )
        return
    if columnas is None:
        resultado["errores"].append(
            "La hoja Liquidaciones no tiene los encabezados esperados."
        )
        return

    existentes = {}
    for linea in leer_datos(Liquidaciones.RUTA_LIQUIDACIONES):
        datos = Liquidaciones.separar_liquidacion(linea)
        if datos is not None:
            clave = (normalizar(datos["cedula"]), datos["periodo"])
            existentes.setdefault(clave, []).append((linea, datos))
    nuevas = set()

    for numero, fila in filas:
        try:
            periodo = parsear_periodo(
                valor_fila(fila, columnas, "Periodo")
            )
            agregar_periodo(resultado, periodo)
            cedula = texto_seguro(
                valor_fila(
                    fila,
                    columnas,
                    "Cédula",
                    ("Cedula",),
                )
            )
            nombre = texto_seguro(
                valor_fila(fila, columnas, "Nombre")
            )
            if not cedula or not nombre:
                raise ValueError("completá Cédula y Nombre")

            sueldo_bruto = parsear_monto(
                valor_fila(fila, columnas, "Sueldo bruto")
            )
            comisiones = monto_opcional(
                fila,
                columnas,
                "Comisiones",
            )
            descuento_ips = monto_opcional(
                fila,
                columnas,
                "Descuento IPS",
            )
            adelantos = monto_opcional(
                fila,
                columnas,
                "Adelantos",
            )
            otros_descuentos = monto_opcional(
                fila,
                columnas,
                "Otros descuentos",
            )
            descuento_ausencias = monto_opcional(
                fila,
                columnas,
                "Descuento ausencias",
            )
            descuento_reposos = monto_opcional(
                fila,
                columnas,
                "Descuento reposos",
            )
            dias_liquidados = entero_opcional(
                fila,
                columnas,
                "Días liquidados",
                30,
            )
            sueldo_referencia = monto_opcional(
                fila,
                columnas,
                "Sueldo referencia",
                sueldo_bruto,
            )
            dias_ausencia = entero_opcional(
                fila,
                columnas,
                "Días ausencia",
                0,
            )
            dias_reposo = entero_opcional(
                fila,
                columnas,
                "Días reposo",
                0,
            )
            tipo = texto_seguro(
                valor_fila(fila, columnas, "Tipo liquidación")
            ) or "Importada desde Excel"

            remuneracion = max(
                0,
                sueldo_bruto
                - descuento_ausencias
                - descuento_reposos,
            ) + comisiones
            neto_valor = valor_fila(
                fila,
                columnas,
                "Neto a cobrar",
            )
            if celda_vacia(neto_valor):
                neto_cobrar = max(
                    0,
                    remuneracion
                    - descuento_ips
                    - adelantos
                    - otros_descuentos,
                )
            else:
                neto_cobrar = parsear_monto(
                    neto_valor,
                    permitir_cero=True,
                )

            clave = (normalizar(cedula), periodo)
            if clave in nuevas:
                duplicado(
                    resultado,
                    hoja.title,
                    numero,
                    "la liquidación está repetida dentro del Excel",
                )
                continue

            datos = {
                "cedula": cedula,
                "nombre": nombre,
                "periodo": periodo,
                "dias_liquidados": dias_liquidados,
                "sueldo_referencia": sueldo_referencia,
                "sueldo_bruto": sueldo_bruto,
                "descuento_ips": descuento_ips,
                "neto_cobrar": neto_cobrar,
                "tipo_liquidacion": tipo,
                "dias_ausencia": dias_ausencia,
                "dias_reposo": dias_reposo,
                "descuento_ausencias": descuento_ausencias,
                "descuento_reposos": descuento_reposos,
                "adelantos": adelantos,
                "otros_descuentos": otros_descuentos,
                "comisiones": comisiones,
            }
            linea = Liquidaciones.crear_linea_liquidacion(datos)

            if clave in existentes:
                candidatos = existentes[clave]
                if len(candidatos) != 1:
                    raise ValueError(
                        "hay más de una liquidación cargada para esa cédula "
                        "y período; corregilas manualmente antes de importar"
                    )

                linea_anterior, datos_anteriores = candidatos[0]
                if linea_anterior == linea:
                    duplicado(
                        resultado,
                        hoja.title,
                        numero,
                        "la liquidación ya está correcta y se omitirá",
                    )
                    nuevas.add(clave)
                    continue

                resultado["correcciones"].append(
                    {
                        "hoja": hoja.title,
                        "fila": numero,
                        "ruta": "Liquidaciones",
                        "anterior": linea_anterior,
                        "nuevo": linea,
                        "detalle": (
                            f"{periodo}: liquidación de {nombre} actualizada "
                            f"de Gs. {Movimientos.formatear_monto(
                                datos_anteriores['sueldo_bruto']
                                + datos_anteriores.get('comisiones', 0)
                            )} a Gs. {Movimientos.formatear_monto(
                                sueldo_bruto + comisiones
                            )}"
                        ),
                    }
                )
                existentes[clave] = [(linea, datos)]
                nuevas.add(clave)
                continue

            nuevas.add(clave)
            resultado["lineas"]["Liquidaciones"].append(linea)
            resultado["conteos"]["Liquidaciones"] += 1
            resultado["totales"]["Liquidaciones"] += remuneracion
            agregar_periodo(resultado, periodo)
        except ValueError as exc:
            error(resultado, hoja.title, numero, str(exc))


def analizar_archivo(ruta):
    ruta = Path(ruta)
    resultado = crear_resultado(ruta)

    if not ruta.exists():
        resultado["errores"].append("El archivo seleccionado no existe.")
        return resultado
    if ruta.suffix.lower() != ".xlsx":
        resultado["errores"].append(
            "Seleccioná un archivo Excel con extensión .xlsx."
        )
        return resultado

    try:
        libro = load_workbook(
            ruta,
            data_only=True,
            read_only=False,
        )
    except Exception as exc:
        resultado["errores"].append(
            f"No se pudo abrir el Excel: {exc}"
        )
        return resultado

    hoja_leeme = obtener_hoja(libro, "LEEME")
    if (
        hoja_leeme is None
        or texto_seguro(hoja_leeme["H1"].value) != MARCADOR_PLANTILLA
    ):
        resultado["errores"].append(
            "El archivo no corresponde a la plantilla oficial de "
            "BC Gestión. Usá el botón Abrir plantilla."
        )
        libro.close()
        return resultado

    analizar_movimientos(libro, resultado)
    analizar_adicionales(libro, resultado)
    analizar_inversiones(libro, resultado)
    prestamos, referencias, _ = analizar_prestamos(
        libro,
        resultado,
    )
    analizar_cuotas(
        libro,
        resultado,
        prestamos,
        referencias,
    )
    analizar_retiros(libro, resultado)
    analizar_liquidaciones(libro, resultado)
    libro.close()

    if len(resultado["periodos"]) > 1:
        periodos = ", ".join(sorted(resultado["periodos"]))
        resultado["errores"].append(
            "El archivo contiene más de un mes "
            f"({periodos}). Usá un Excel separado por mes."
        )
    elif not resultado["periodos"]:
        resultado["errores"].append(
            "No se encontraron filas con datos para importar."
        )

    total_nuevos = (
        sum(resultado["conteos"].values())
        + len(resultado["correcciones"])
    )
    if total_nuevos == 0 and not resultado["errores"]:
        resultado["advertencias"].append(
            "No hay registros nuevos: todo ya estaba cargado."
        )

    resultado["huella"] = hashlib.sha256(ruta.read_bytes()).hexdigest()
    return resultado


def crear_respaldo_importacion(periodo):
    carpeta = BASE_DIR / "Respaldos"
    carpeta.mkdir(parents=True, exist_ok=True)
    marca = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    ruta = carpeta / f"Antes_importacion_{periodo}_{marca}.zip"

    with ZipFile(ruta, "w", compression=ZIP_DEFLATED) as archivo:
        for ruta_datos in dict.fromkeys(RUTAS.values()):
            ruta_absoluta = (
                ruta_datos
                if ruta_datos.is_absolute()
                else BASE_DIR / ruta_datos
            )
            if ruta_absoluta.exists():
                archivo.write(
                    ruta_absoluta,
                    ruta_absoluta.relative_to(BASE_DIR),
                )
    return ruta


def aplicar_resultado(resultado):
    if resultado["errores"]:
        raise ValueError("La vista previa todavía contiene errores.")

    ruta_excel = Path(resultado["ruta"])
    huella_actual = hashlib.sha256(ruta_excel.read_bytes()).hexdigest()
    if huella_actual != resultado["huella"]:
        raise ValueError(
            "El Excel cambió después de la vista previa. Analizalo otra vez."
        )

    periodo = next(iter(resultado["periodos"]))
    respaldo = crear_respaldo_importacion(periodo)
    originales = {}

    try:
        for nombre, ruta in RUTAS.items():
            originales[nombre] = leer_datos(ruta)

        contenidos_actualizados = {
            nombre: list(contenido)
            for nombre, contenido in originales.items()
        }

        for correccion in resultado["correcciones"]:
            nombre = correccion["ruta"]
            contenido = contenidos_actualizados[nombre]
            coincidencias = [
                indice
                for indice, linea in enumerate(contenido)
                if linea == correccion["anterior"]
            ]
            if len(coincidencias) != 1:
                raise ValueError(
                    "El movimiento que debía corregirse cambió después de "
                    "la vista previa. Analizá el Excel nuevamente."
                )
            contenido[coincidencias[0]] = correccion["nuevo"]

        for nombre, lineas_nuevas in resultado["lineas"].items():
            if not lineas_nuevas:
                continue
            contenidos_actualizados[nombre].extend(lineas_nuevas)

        for nombre, contenido in contenidos_actualizados.items():
            if contenido == originales[nombre]:
                continue
            guardar_datos(
                RUTAS[nombre],
                contenido,
            )
    except Exception:
        for nombre, contenido in originales.items():
            guardar_datos(RUTAS[nombre], contenido)
        raise

    return respaldo


def texto_resumen(resultado):
    lineas = []
    if len(resultado["periodos"]) == 1:
        lineas.append(
            f"PERÍODO DETECTADO: {next(iter(resultado['periodos']))}"
        )
    else:
        lineas.append("PERÍODO DETECTADO: pendiente")
    lineas.append("")
    lineas.append("REGISTROS NUEVOS")

    total_registros = 0
    for clave, etiqueta in NOMBRES_RESUMEN.items():
        cantidad = resultado["conteos"][clave]
        total = resultado["totales"][clave]
        total_registros += cantidad
        lineas.append(
            f"  {etiqueta}: {cantidad} "
            f"(Gs. {Movimientos.formatear_monto(total)})"
        )
    lineas.append(f"  TOTAL DE FILAS NUEVAS: {total_registros}")
    if resultado["conteos"]["Retiros"]:
        lineas.append("")
        lineas.append("MOVIMIENTOS PERSONALES DE SOCIOS")
        for tipo in Socios.TIPOS_MOVIMIENTO_PERSONAL:
            lineas.append(
                f"  {tipo}: Gs. "
                f"{Movimientos.formatear_monto(
                    resultado['totales_retiros_tipo'][tipo]
                )}"
            )

    lineas.append("")
    lineas.append(
        f"CORRECCIONES DE CLASIFICACIÓN: "
        f"{len(resultado['correcciones'])}"
    )
    lineas.append(
        f"DUPLICADOS OMITIDOS: {len(resultado['duplicados'])}"
    )
    lineas.append(f"ADVERTENCIAS: {len(resultado['advertencias'])}")
    lineas.append(f"ERRORES: {len(resultado['errores'])}")

    if resultado["errores"]:
        lineas.append("")
        lineas.append("ERRORES A CORREGIR")
        lineas.extend(f"  • {item}" for item in resultado["errores"])

    if resultado["correcciones"]:
        lineas.append("")
        lineas.append("CORRECCIONES QUE SE APLICARÁN")
        lineas.extend(
            f"  • {item['detalle']}"
            for item in resultado["correcciones"]
        )

    if resultado["duplicados"]:
        lineas.append("")
        lineas.append("DUPLICADOS (NO SE IMPORTARÁN)")
        lineas.extend(f"  • {item}" for item in resultado["duplicados"][:100])
        if len(resultado["duplicados"]) > 100:
            lineas.append("  • ... y más duplicados.")

    if resultado["advertencias"]:
        lineas.append("")
        lineas.append("ADVERTENCIAS")
        lineas.extend(
            f"  • {item}"
            for item in resultado["advertencias"]
        )

    if not resultado["errores"]:
        lineas.append("")
        lineas.append(
            "VALIDACIÓN CORRECTA. Podés importar con seguridad."
        )
    return "\n".join(lineas)


def abrir_archivo(ruta):
    ruta = str(ruta)
    if os.name == "nt":
        os.startfile(ruta)
        return
    try:
        subprocess.Popen(["xdg-open", ruta])
    except OSError:
        subprocess.Popen(["open", ruta])


def abrir_importador(ventana_padre):
    import customtkinter as ctk
    from tkinter import filedialog, messagebox

    ventana = ctk.CTkToplevel(ventana_padre)
    ventana.title("Importar mes desde Excel")
    ventana.geometry("980x720")
    ventana.minsize(820, 620)
    ventana.transient(ventana_padre)
    ventana.grab_set()
    ventana.grid_columnconfigure(0, weight=1)
    ventana.grid_rowconfigure(2, weight=1)

    estado = {
        "resultado": None,
        "ruta": ctk.StringVar(value=""),
    }

    encabezado = ctk.CTkFrame(
        ventana,
        fg_color=COLOR_PANEL,
        corner_radius=0,
        border_width=0,
    )
    encabezado.grid(row=0, column=0, sticky="ew")
    encabezado.grid_columnconfigure(0, weight=1)

    ctk.CTkLabel(
        encabezado,
        text="Importar mes desde Excel",
        font=ctk.CTkFont(size=24, weight="bold"),
        text_color=COLOR_TEXTO,
        anchor="w",
    ).grid(row=0, column=0, sticky="ew", padx=28, pady=(22, 4))
    ctk.CTkLabel(
        encabezado,
        text=(
            "Primero se valida todo. Nada se guarda hasta que confirmes "
            "la importación."
        ),
        font=ctk.CTkFont(size=13),
        text_color=COLOR_TEXTO_SUAVE,
        anchor="w",
    ).grid(row=1, column=0, sticky="ew", padx=28, pady=(0, 20))

    seleccion = ctk.CTkFrame(
        ventana,
        fg_color=COLOR_PANEL_SECUNDARIO,
        corner_radius=14,
        border_width=1,
        border_color=COLOR_BORDE,
    )
    seleccion.grid(row=1, column=0, sticky="ew", padx=24, pady=18)
    seleccion.grid_columnconfigure(0, weight=1)

    entrada = ctk.CTkEntry(
        seleccion,
        textvariable=estado["ruta"],
        placeholder_text="Seleccioná el Excel mensual...",
        height=40,
    )
    entrada.grid(
        row=0,
        column=0,
        sticky="ew",
        padx=(16, 8),
        pady=16,
    )

    cuadro = ctk.CTkTextbox(
        ventana,
        font=ctk.CTkFont(family="Consolas", size=13),
        fg_color=COLOR_PANEL,
        border_width=1,
        border_color=COLOR_BORDE,
        text_color=COLOR_TEXTO,
        wrap="word",
    )
    cuadro.grid(row=2, column=0, sticky="nsew", padx=24, pady=(0, 12))
    cuadro.insert(
        "1.0",
        "Seleccioná la plantilla completada y presioná Analizar Excel.",
    )
    cuadro.configure(state="disabled")

    acciones = ctk.CTkFrame(
        ventana,
        fg_color="transparent",
    )
    acciones.grid(row=3, column=0, sticky="ew", padx=24, pady=(0, 22))
    for columna in range(4):
        acciones.grid_columnconfigure(columna, weight=1)

    def mostrar_texto(texto):
        cuadro.configure(state="normal")
        cuadro.delete("1.0", "end")
        cuadro.insert("1.0", texto)
        cuadro.configure(state="disabled")

    def seleccionar():
        ruta = filedialog.askopenfilename(
            parent=ventana,
            title="Seleccionar carga mensual",
            filetypes=[("Excel", "*.xlsx")],
        )
        if ruta:
            estado["ruta"].set(ruta)
            estado["resultado"] = None
            boton_importar.configure(state="disabled")
            analizar()

    def analizar(evento=None):
        ruta = estado["ruta"].get().strip()
        if not ruta:
            messagebox.showwarning(
                "Falta el archivo",
                "Seleccioná primero el Excel mensual.",
                parent=ventana,
            )
            return "break"
        ventana.configure(cursor="watch")
        ventana.update_idletasks()
        resultado = analizar_archivo(ruta)
        ventana.configure(cursor="")
        estado["resultado"] = resultado
        mostrar_texto(texto_resumen(resultado))
        total = (
            sum(resultado["conteos"].values())
            + len(resultado["correcciones"])
        )
        boton_importar.configure(
            state=(
                "normal"
                if not resultado["errores"] and total > 0
                else "disabled"
            )
        )
        return "break"

    def importar():
        resultado = estado["resultado"]
        if resultado is None or resultado["errores"]:
            return
        total_nuevos = sum(resultado["conteos"].values())
        total_correcciones = len(resultado["correcciones"])
        total = total_nuevos + total_correcciones
        periodo = next(iter(resultado["periodos"]))
        if not messagebox.askyesno(
            "Confirmar importación",
            (
                f"Período: {periodo}\n"
                f"Registros nuevos: {total_nuevos}\n"
                f"Correcciones: {total_correcciones}\n\n"
                "Antes de guardar se creará un respaldo automático.\n\n"
                "¿Continuar?"
            ),
            parent=ventana,
        ):
            return

        try:
            respaldo = aplicar_resultado(resultado)
        except Exception as exc:
            messagebox.showerror(
                "No se pudo importar",
                str(exc),
                parent=ventana,
            )
            return

        boton_importar.configure(state="disabled")
        messagebox.showinfo(
            "Importación terminada",
            (
                f"Se procesaron {total} registros del período {periodo}.\n"
                f"Nuevos: {total_nuevos} · "
                f"Correcciones: {total_correcciones}\n\n"
                f"Respaldo creado:\n{respaldo}"
            ),
            parent=ventana,
        )
        mostrar_texto(
            texto_resumen(resultado)
            + "\n\nIMPORTACIÓN COMPLETADA CORRECTAMENTE."
        )

    def abrir_plantilla():
        if not PLANTILLA.exists():
            messagebox.showerror(
                "Plantilla no encontrada",
                (
                    "Copiá Plantilla_Carga_Mensual_BC_Gestion.xlsx "
                    "dentro de PX-Core."
                ),
                parent=ventana,
            )
            return
        abrir_archivo(PLANTILLA)

    ctk.CTkButton(
        seleccion,
        text="Elegir Excel",
        command=seleccionar,
        height=40,
        width=125,
        fg_color=COLOR_PRIMARIO,
        hover_color=COLOR_PRIMARIO_HOVER,
    ).grid(row=0, column=1, padx=(0, 16), pady=16)

    ctk.CTkButton(
        acciones,
        text="Abrir plantilla",
        command=abrir_plantilla,
        height=40,
        fg_color=COLOR_PANEL_SECUNDARIO,
        hover_color=COLOR_BORDE,
        text_color=COLOR_TEXTO,
    ).grid(row=0, column=0, sticky="ew", padx=(0, 6))
    ctk.CTkButton(
        acciones,
        text="Analizar Excel",
        command=analizar,
        height=40,
        fg_color=COLOR_PRIMARIO,
        hover_color=COLOR_PRIMARIO_HOVER,
    ).grid(row=0, column=1, sticky="ew", padx=6)
    boton_importar = ctk.CTkButton(
        acciones,
        text="Importar mes",
        command=importar,
        height=40,
        fg_color=COLOR_VERDE,
        hover_color="#12835B",
        state="disabled",
    )
    boton_importar.grid(row=0, column=2, sticky="ew", padx=6)
    ctk.CTkButton(
        acciones,
        text="Cerrar",
        command=ventana.destroy,
        height=40,
        fg_color=COLOR_ROJO,
        hover_color="#B93E3E",
    ).grid(row=0, column=3, sticky="ew", padx=(6, 0))

    entrada.bind("<Return>", analizar)
    ventana.bind("<Escape>", lambda _evento: ventana.destroy())
    ventana.after(100, entrada.focus_set)
