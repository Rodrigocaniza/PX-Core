"""Registro documental de facturas. No afecta movimientos ni resultados."""

import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from uuid import uuid4

import customtkinter as ctk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from datos import guardar_datos, leer_datos


BASE_DIR = Path(__file__).resolve().parent
RUTA_FACTURAS = BASE_DIR / "Datos" / "facturas.txt"
RUTA_PROVEEDORES = BASE_DIR / "Datos" / "proveedores.txt"

ORIGENES = ["PC", "P2", "Administración", "General EAS", "Sol", "Cta"]
PLAZOS = ["30", "60", "90", "Personalizado"]

COLOR_FONDO = ("#F4F7FB", "#0B1220")
COLOR_PANEL = ("#FFFFFF", "#131D2E")
COLOR_PANEL_SECUNDARIO = ("#EAF0F7", "#1A2639")
COLOR_BORDE = ("#DCE4EE", "#26354A")
COLOR_TEXTO = ("#182230", "#F4F7FB")
COLOR_TEXTO_SUAVE = ("#617084", "#9CAFC5")
COLOR_PRIMARIO = "#246BFD"
COLOR_PRIMARIO_HOVER = "#1855D6"
COLOR_VERDE = "#18A874"
COLOR_ROJO = "#E25555"


def _texto(valor, nombre, obligatorio=True):
    valor = str(valor or "").strip()
    if obligatorio and not valor:
        raise ValueError(f"Completá el campo {nombre}.")
    if any(caracter in valor for caracter in ("\n", "\r")):
        raise ValueError(f"El campo {nombre} contiene un salto de línea.")
    return valor


def _fecha(texto):
    if isinstance(texto, datetime):
        return texto
    if isinstance(texto, date):
        return datetime.combine(texto, datetime.min.time())
    try:
        return datetime.strptime(str(texto).strip(), "%d-%m-%Y")
    except ValueError as error:
        raise ValueError("La fecha debe tener formato DD-MM-AAAA.") from error


def _fecha_excel(valor):
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, date):
        return datetime.combine(valor, datetime.min.time())
    if isinstance(valor, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=float(valor))
    texto = str(valor or "").strip()
    for formato in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, formato)
        except ValueError:
            continue
    raise ValueError("Fecha de Excel inválida.")


def _monto(texto):
    if isinstance(texto, (int, float)):
        valor = int(round(texto))
    else:
        limpio = (
            str(texto or "")
            .strip()
            .replace("Gs.", "")
            .replace(".", "")
            .replace(",", "")
            .replace(" ", "")
        )
        try:
            valor = int(limpio)
        except ValueError as error:
            raise ValueError(
                "Ingresá un monto válido usando solo números."
            ) from error
    if valor <= 0:
        raise ValueError("El monto debe ser mayor a cero.")
    return valor


def _formatear_monto(valor):
    return f"{int(valor):,}".replace(",", ".")


def _normalizar(valor):
    return re.sub(r"\s+", " ", str(valor or "").strip()).casefold()


def _leer_json(ruta):
    registros = []
    for posicion, linea in enumerate(leer_datos(ruta)):
        try:
            dato = json.loads(linea)
        except (json.JSONDecodeError, TypeError):
            continue
        if isinstance(dato, dict):
            registros.append((posicion, dato))
    return registros


def _guardar_json(ruta, registros):
    lineas = [
        json.dumps(registro, ensure_ascii=False, separators=(",", ":"))
        for registro in registros
    ]
    guardar_datos(ruta, lineas)


def obtener_facturas():
    return _leer_json(RUTA_FACTURAS)


def obtener_proveedores():
    return _leer_json(RUTA_PROVEEDORES)


def guardar_factura(registro):
    facturas = [dato for _, dato in obtener_facturas()]
    facturas.append(registro)
    _guardar_json(RUTA_FACTURAS, facturas)


def guardar_facturas(registros):
    facturas = [dato for _, dato in obtener_facturas()]
    facturas.extend(registros)
    _guardar_json(RUTA_FACTURAS, facturas)


def guardar_proveedor(registro):
    proveedores = [dato for _, dato in obtener_proveedores()]
    nombre_normalizado = _normalizar(registro["nombre"])
    if any(
        _normalizar(proveedor.get("nombre")) == nombre_normalizado
        for proveedor in proveedores
    ):
        raise ValueError("Ese proveedor ya está registrado.")
    proveedores.append(registro)
    _guardar_json(RUTA_PROVEEDORES, proveedores)


def eliminar_factura_por_id(factura_id):
    facturas = [
        dato
        for _, dato in obtener_facturas()
        if dato.get("id") != factura_id
    ]
    _guardar_json(RUTA_FACTURAS, facturas)


def eliminar_facturas_por_periodo(periodo):
    try:
        mes = datetime.strptime(periodo.strip(), "%m-%Y")
    except ValueError as error:
        raise ValueError(
            "El período debe tener formato MM-AAAA."
        ) from error

    conservadas = []
    eliminadas = []

    for _, factura in obtener_facturas():
        try:
            fecha = _fecha(factura.get("fecha", ""))
        except ValueError:
            conservadas.append(factura)
            continue

        if (fecha.month, fecha.year) == (mes.month, mes.year):
            eliminadas.append(factura)
        else:
            conservadas.append(factura)

    if eliminadas:
        _guardar_json(RUTA_FACTURAS, conservadas)

    return eliminadas


def _clave_factura(registro):
    referencia = registro.get("ruc") or registro.get("proveedor")
    return (
        registro.get("fecha", ""),
        _normalizar(registro.get("numero")),
        _normalizar(referencia),
        int(registro.get("monto", 0)),
    )


def _separar_proveedor_ruc(texto):
    valor = re.sub(r"\s+", " ", str(texto or "").strip())
    coincidencia = re.search(r"(\d{5,}-\d)$", valor)
    if not coincidencia:
        return valor, ""
    ruc = coincidencia.group(1)
    nombre = valor[: coincidencia.start()].strip(" -")
    return nombre or valor, ruc


def _registro_importado(
    tipo,
    fecha,
    numero,
    monto,
    proveedor,
    origen="",
    descripcion="",
    ruc="",
    vencimiento="",
    plazo_dias=0,
    hoja="",
):
    fecha_dt = _fecha_excel(fecha)
    return {
        "id": uuid4().hex[:12].upper(),
        "tipo": tipo,
        "fecha": fecha_dt.strftime("%d-%m-%Y"),
        "numero": str(numero or "").strip(),
        "monto": _monto(monto),
        "origen": str(origen or "").strip(),
        "proveedor": str(proveedor or "").strip(),
        "descripcion": str(descripcion or "").strip(),
        "ruc": str(ruc or "").strip(),
        "plazo_dias": int(plazo_dias or 0),
        "vencimiento": vencimiento,
        "hoja_origen": hoja,
        "registrado_en": datetime.now().isoformat(timespec="seconds"),
    }


def leer_facturas_excel(ruta):
    libro = load_workbook(ruta, data_only=True, read_only=False)
    registros = []
    errores = []

    if "Contado" in libro.sheetnames:
        hoja = libro["Contado"]

        for fila in range(4, hoja.max_row + 1):
            fecha = hoja.cell(fila, 1).value
            numero = hoja.cell(fila, 2).value
            ruc = hoja.cell(fila, 3).value
            montos = [
                hoja.cell(fila, columna).value
                for columna in range(4, 8)
            ]
            concepto = hoja.cell(fila, 8).value
            proveedor = hoja.cell(fila, 9).value
            sucursal = hoja.cell(fila, 10).value

            if not isinstance(fecha, (datetime, date)):
                continue
            if numero in (None, ""):
                continue

            montos_validos = [
                float(valor)
                for valor in montos
                if isinstance(valor, (int, float)) and valor != 0
            ]
            if not montos_validos:
                continue

            categorias = [
                nombre
                for nombre, valor in zip(
                    (
                        "Caja Óptica",
                        "Cuenta",
                        "Máxima Visión",
                        "Gastos personales",
                    ),
                    montos,
                )
                if isinstance(valor, (int, float)) and valor != 0
            ]

            try:
                registro = _registro_importado(
                    "CONTADO",
                    fecha,
                    numero,
                    sum(montos_validos),
                    proveedor or "",
                    origen=sucursal or "",
                    descripcion=concepto or "",
                    ruc=ruc or "",
                    hoja="Contado",
                )
                registro["categoria_cuenta"] = " / ".join(categorias)
                registros.append(registro)
            except (ValueError, TypeError) as error:
                errores.append(f"Contado fila {fila}: {error}")

    if "Crédito" in libro.sheetnames:
        hoja = libro["Crédito"]

        bloques = (
            {
                "fecha": 1,
                "factura": 2,
                "monto": 3,
                "sucursal": 4,
                "vencimiento": None,
            },
            {
                "fecha": 5,
                "factura": 6,
                "monto": 7,
                "vencimiento": 8,
                "sucursal": 9,
            },
            {
                "fecha": 10,
                "factura": 11,
                "monto": 12,
                "sucursal": 13,
                "vencimiento": None,
            },
        )

        palabras_ignoradas = {
            "",
            "fecha",
            "factura",
            "monto",
            "suc",
            "vto.",
            "credito",
            "crédito",
            "enero",
            "febrero",
            "marzo",
            "abril",
            "mayo",
            "junio",
            "julio",
            "agosto",
            "septiembre",
            "octubre",
            "noviembre",
            "diciembre",
            "total",
        }

        for bloque in bloques:
            proveedor = ""
            ruc = ""

            for fila in range(1, hoja.max_row + 1):
                valor_fecha = hoja.cell(fila, bloque["fecha"]).value
                numero = hoja.cell(fila, bloque["factura"]).value
                monto = hoja.cell(fila, bloque["monto"]).value

                if isinstance(valor_fecha, str):
                    texto = re.sub(r"\s+", " ", valor_fecha.strip())
                    normalizado = _normalizar(texto)

                    if (
                        normalizado not in palabras_ignoradas
                        and numero in (None, "")
                        and monto in (None, "")
                    ):
                        proveedor, ruc = _separar_proveedor_ruc(texto)
                    continue

                # Solo una fila completa constituye una factura.
                # Textos como "2 facturas" o números sueltos se ignoran.
                if not isinstance(valor_fecha, (datetime, date)):
                    continue
                if numero in (None, ""):
                    continue
                if not isinstance(monto, (int, float)) or monto <= 0:
                    continue
                if not proveedor:
                    errores.append(
                        f"Crédito fila {fila}: proveedor no identificado."
                    )
                    continue

                try:
                    fecha_dt = _fecha_excel(valor_fecha)
                    vencimiento = ""
                    plazo = 0

                    if bloque["vencimiento"]:
                        valor_vto = hoja.cell(
                            fila,
                            bloque["vencimiento"],
                        ).value
                        if isinstance(valor_vto, (datetime, date)):
                            fecha_vto = _fecha_excel(valor_vto)
                            vencimiento = fecha_vto.strftime("%d-%m-%Y")
                            plazo = max((fecha_vto - fecha_dt).days, 0)

                    sucursal = hoja.cell(
                        fila,
                        bloque["sucursal"],
                    ).value

                    registros.append(
                        _registro_importado(
                            "CREDITO",
                            fecha_dt,
                            numero,
                            monto,
                            proveedor,
                            origen=sucursal or "",
                            descripcion="",
                            ruc=ruc,
                            vencimiento=vencimiento,
                            plazo_dias=plazo,
                            hoja="Crédito",
                        )
                    )
                except (ValueError, TypeError) as error:
                    errores.append(
                        f"Crédito fila {fila}, {proveedor}: {error}"
                    )

    hoja_especial = next(
        (
            libro[nombre]
            for nombre in libro.sheetnames
            if "IPS" in nombre.upper()
            or "MUNICIPALIDAD" in nombre.upper()
            or "SET" in nombre.upper()
        ),
        None,
    )

    if hoja_especial is not None:
        for fila in range(3, hoja_especial.max_row + 1):
            fecha = hoja_especial.cell(fila, 2).value
            numero = hoja_especial.cell(fila, 3).value
            monto = hoja_especial.cell(fila, 4).value
            concepto = hoja_especial.cell(fila, 5).value
            proveedor = hoja_especial.cell(fila, 6).value

            if not isinstance(fecha, (datetime, date)):
                continue
            if numero in (None, ""):
                continue
            if not isinstance(monto, (int, float)) or monto <= 0:
                continue

            try:
                registros.append(
                    _registro_importado(
                        "CONTADO",
                        fecha,
                        numero,
                        monto,
                        proveedor or "",
                        origen="Administración",
                        descripcion=concepto or "",
                        hoja=hoja_especial.title,
                    )
                )
            except (ValueError, TypeError) as error:
                errores.append(
                    f"{hoja_especial.title} fila {fila}: {error}"
                )

    libro.close()
    return registros, errores


class PanelFacturas(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color=COLOR_FONDO, corner_radius=0)
        self.factura_seleccionada_id = None
        self.registros_importacion = []
        self.ruta_importacion = None
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self._crear_encabezado()
        self.pestanas = ctk.CTkTabview(
            self,
            fg_color=COLOR_PANEL,
            segmented_button_selected_color=COLOR_PRIMARIO,
            segmented_button_selected_hover_color=COLOR_PRIMARIO_HOVER,
        )
        self.pestanas.grid(
            row=1,
            column=0,
            sticky="nsew",
            padx=24,
            pady=(0, 24),
        )

        for nombre in (
            "Factura al contado",
            "Factura a crédito",
            "Proveedores",
            "Informes",
            "Importar Excel",
        ):
            self.pestanas.add(nombre)

        self._crear_contado()
        self._crear_credito()
        self._crear_proveedores()
        self._crear_informes()
        self._crear_importador()
        self._actualizar_proveedores()
        self._actualizar_informe()

    def _crear_encabezado(self):
        encabezado = ctk.CTkFrame(self, fg_color="transparent")
        encabezado.grid(row=0, column=0, sticky="ew", padx=28, pady=(20, 14))
        encabezado.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            encabezado,
            text="Registro documental de facturas",
            font=ctk.CTkFont(size=25, weight="bold"),
            text_color=COLOR_TEXTO,
        ).grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(
            encabezado,
            text="Carga, consulta, importación y exportación documental.",
            font=ctk.CTkFont(size=12),
            text_color=COLOR_TEXTO_SUAVE,
        ).grid(row=1, column=0, sticky="w", pady=(3, 0))

        self.etiqueta_reloj = ctk.CTkLabel(
            encabezado,
            text="",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLOR_TEXTO,
        )
        self.etiqueta_reloj.grid(row=0, column=1, rowspan=2, padx=(20, 0))
        self._actualizar_reloj()

    def _actualizar_reloj(self):
        if not self.winfo_exists():
            return
        self.etiqueta_reloj.configure(
            text=datetime.now().strftime("%d-%m-%Y  %H:%M:%S")
        )
        self.after(1000, self._actualizar_reloj)

    def _campo(self, master, fila, etiqueta, placeholder=""):
        ctk.CTkLabel(
            master,
            text=etiqueta,
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLOR_TEXTO,
            anchor="w",
        ).grid(row=fila, column=0, sticky="w", padx=18, pady=8)
        entrada = ctk.CTkEntry(
            master,
            placeholder_text=placeholder,
            fg_color=COLOR_PANEL_SECUNDARIO,
            border_color=COLOR_BORDE,
            text_color=COLOR_TEXTO,
            height=38,
        )
        entrada.grid(
            row=fila,
            column=1,
            sticky="ew",
            padx=(0, 18),
            pady=8,
        )
        return entrada

    def _panel_formulario(self, tab):
        tab.grid_rowconfigure(0, weight=1)
        tab.grid_columnconfigure(0, weight=1)
        panel = ctk.CTkFrame(tab, fg_color="transparent")
        panel.grid(row=0, column=0, sticky="nsew", padx=18, pady=18)
        panel.grid_columnconfigure(1, weight=1)
        return panel

    def _crear_contado(self):
        tab = self.pestanas.tab("Factura al contado")
        panel = self._panel_formulario(tab)

        self.contado_fecha = self._campo(panel, 0, "Fecha", "DD-MM-AAAA")
        self.contado_fecha.insert(0, datetime.now().strftime("%d-%m-%Y"))
        self.contado_numero = self._campo(panel, 1, "Número de factura")
        self.contado_monto = self._campo(panel, 2, "Monto", "Ej.: 150000")

        ctk.CTkLabel(
            panel,
            text="Origen / sucursal",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLOR_TEXTO,
        ).grid(row=3, column=0, sticky="w", padx=18, pady=8)
        self.contado_origen = ctk.CTkComboBox(
            panel,
            values=ORIGENES,
            state="normal",
            height=38,
            fg_color=COLOR_PANEL_SECUNDARIO,
            border_color=COLOR_BORDE,
            button_color=COLOR_PRIMARIO,
            text_color=COLOR_TEXTO,
        )
        self.contado_origen.set(ORIGENES[0])
        self.contado_origen.grid(
            row=3,
            column=1,
            sticky="ew",
            padx=(0, 18),
            pady=8,
        )

        self.contado_local = self._campo(panel, 4, "Local / negocio")
        self.contado_descripcion = self._campo(
            panel,
            5,
            "Descripción de la compra",
        )

        ctk.CTkButton(
            panel,
            text="Guardar factura al contado",
            command=self._guardar_contado,
            fg_color=COLOR_VERDE,
            hover_color="#12835B",
            height=42,
            font=ctk.CTkFont(size=13, weight="bold"),
        ).grid(
            row=6,
            column=0,
            columnspan=2,
            sticky="ew",
            padx=18,
            pady=20,
        )

    def _guardar_contado(self):
        try:
            fecha = _fecha(self.contado_fecha.get())
            registro = {
                "id": uuid4().hex[:12].upper(),
                "tipo": "CONTADO",
                "fecha": fecha.strftime("%d-%m-%Y"),
                "numero": _texto(self.contado_numero.get(), "Número de factura"),
                "monto": _monto(self.contado_monto.get()),
                "origen": _texto(self.contado_origen.get(), "Origen"),
                "proveedor": _texto(self.contado_local.get(), "Local / negocio"),
                "descripcion": _texto(
                    self.contado_descripcion.get(),
                    "Descripción",
                ),
                "ruc": "",
                "plazo_dias": 0,
                "vencimiento": "",
                "registrado_en": datetime.now().isoformat(timespec="seconds"),
            }
            guardar_factura(registro)
        except (ValueError, OSError) as error:
            messagebox.showerror("No se pudo guardar", str(error), parent=self)
            return

        for entrada in (
            self.contado_numero,
            self.contado_monto,
            self.contado_local,
            self.contado_descripcion,
        ):
            entrada.delete(0, "end")
        self._actualizar_informe()
        messagebox.showinfo(
            "Factura guardada",
            "La factura al contado quedó registrada.",
            parent=self,
        )

    def _crear_credito(self):
        tab = self.pestanas.tab("Factura a crédito")
        panel = self._panel_formulario(tab)

        self.credito_fecha = self._campo(panel, 0, "Fecha", "DD-MM-AAAA")
        self.credito_fecha.insert(0, datetime.now().strftime("%d-%m-%Y"))
        self.credito_numero = self._campo(panel, 1, "Número de factura")
        self.credito_monto = self._campo(panel, 2, "Monto")

        ctk.CTkLabel(
            panel,
            text="Proveedor",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLOR_TEXTO,
        ).grid(row=3, column=0, sticky="w", padx=18, pady=8)
        self.credito_proveedor = ctk.CTkComboBox(
            panel,
            values=[],
            state="normal",
            height=38,
            fg_color=COLOR_PANEL_SECUNDARIO,
            border_color=COLOR_BORDE,
            button_color=COLOR_PRIMARIO,
            text_color=COLOR_TEXTO,
        )
        self.credito_proveedor.grid(
            row=3,
            column=1,
            sticky="ew",
            padx=(0, 18),
            pady=8,
        )

        self.credito_descripcion = self._campo(
            panel,
            4,
            "Descripción",
            "Opcional",
        )

        ctk.CTkLabel(
            panel,
            text="Plazo",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLOR_TEXTO,
        ).grid(row=5, column=0, sticky="w", padx=18, pady=8)

        zona_plazo = ctk.CTkFrame(panel, fg_color="transparent")
        zona_plazo.grid(
            row=5,
            column=1,
            sticky="ew",
            padx=(0, 18),
            pady=8,
        )
        zona_plazo.grid_columnconfigure((0, 1), weight=1)

        self.credito_plazo = ctk.CTkComboBox(
            zona_plazo,
            values=PLAZOS,
            state="readonly",
            command=self._cambiar_plazo,
            fg_color=COLOR_PANEL_SECUNDARIO,
            border_color=COLOR_BORDE,
            button_color=COLOR_PRIMARIO,
            text_color=COLOR_TEXTO,
        )
        self.credito_plazo.set("30")
        self.credito_plazo.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        self.credito_dias = ctk.CTkEntry(
            zona_plazo,
            placeholder_text="Días personalizados",
            state="disabled",
            fg_color=COLOR_PANEL_SECUNDARIO,
            border_color=COLOR_BORDE,
            text_color=COLOR_TEXTO,
        )
        self.credito_dias.grid(row=0, column=1, sticky="ew", padx=(8, 0))

        self.etiqueta_vencimiento = ctk.CTkLabel(
            panel,
            text="Vencimiento estimado: -",
            text_color=COLOR_TEXTO_SUAVE,
            font=ctk.CTkFont(size=13, weight="bold"),
        )
        self.etiqueta_vencimiento.grid(
            row=6,
            column=0,
            columnspan=2,
            sticky="w",
            padx=18,
            pady=8,
        )

        self.credito_fecha.bind(
            "<KeyRelease>",
            lambda _evento: self._mostrar_vencimiento(),
        )
        self.credito_dias.bind(
            "<KeyRelease>",
            lambda _evento: self._mostrar_vencimiento(),
        )
        self._mostrar_vencimiento()

        ctk.CTkButton(
            panel,
            text="Guardar factura a crédito",
            command=self._guardar_credito,
            fg_color=COLOR_PRIMARIO,
            hover_color=COLOR_PRIMARIO_HOVER,
            height=42,
            font=ctk.CTkFont(size=13, weight="bold"),
        ).grid(
            row=7,
            column=0,
            columnspan=2,
            sticky="ew",
            padx=18,
            pady=20,
        )

    def _cambiar_plazo(self, valor):
        self.credito_dias.configure(
            state="normal" if valor == "Personalizado" else "disabled"
        )
        self._mostrar_vencimiento()

    def _dias_credito(self):
        valor = self.credito_plazo.get()
        if valor == "Personalizado":
            try:
                dias = int(self.credito_dias.get().strip())
            except ValueError as error:
                raise ValueError(
                    "Ingresá una cantidad válida de días."
                ) from error
            if dias <= 0:
                raise ValueError("El plazo debe ser mayor a cero.")
            return dias
        return int(valor)

    def _mostrar_vencimiento(self):
        try:
            vencimiento = _fecha(self.credito_fecha.get()) + timedelta(
                days=self._dias_credito()
            )
            texto = (
                "Vencimiento estimado: "
                + vencimiento.strftime("%d-%m-%Y")
            )
        except ValueError:
            texto = "Vencimiento estimado: -"
        self.etiqueta_vencimiento.configure(text=texto)

    def _guardar_credito(self):
        try:
            fecha = _fecha(self.credito_fecha.get())
            dias = self._dias_credito()
            registro = {
                "id": uuid4().hex[:12].upper(),
                "tipo": "CREDITO",
                "fecha": fecha.strftime("%d-%m-%Y"),
                "numero": _texto(self.credito_numero.get(), "Número de factura"),
                "monto": _monto(self.credito_monto.get()),
                "origen": "",
                "proveedor": _texto(
                    self.credito_proveedor.get(),
                    "Proveedor",
                ),
                "descripcion": _texto(
                    self.credito_descripcion.get(),
                    "Descripción",
                    obligatorio=False,
                ),
                "ruc": "",
                "plazo_dias": dias,
                "vencimiento": (
                    fecha + timedelta(days=dias)
                ).strftime("%d-%m-%Y"),
                "registrado_en": datetime.now().isoformat(timespec="seconds"),
            }
            guardar_factura(registro)
        except (ValueError, OSError) as error:
            messagebox.showerror("No se pudo guardar", str(error), parent=self)
            return

        self.credito_numero.delete(0, "end")
        self.credito_monto.delete(0, "end")
        self.credito_proveedor.set("")
        self.credito_descripcion.delete(0, "end")
        self._actualizar_informe()
        messagebox.showinfo(
            "Factura guardada",
            "La factura a crédito quedó registrada.",
            parent=self,
        )

    def _crear_proveedores(self):
        tab = self.pestanas.tab("Proveedores")
        panel = self._panel_formulario(tab)

        self.proveedor_nombre = self._campo(
            panel,
            0,
            "Nombre / razón social",
        )
        self.proveedor_ruc = self._campo(panel, 1, "RUC", "Opcional")
        self.proveedor_telefono = self._campo(
            panel,
            2,
            "Teléfono",
            "Opcional",
        )
        self.proveedor_direccion = self._campo(
            panel,
            3,
            "Dirección",
            "Opcional",
        )
        self.proveedor_observacion = self._campo(
            panel,
            4,
            "Observación",
            "Opcional",
        )

        ctk.CTkButton(
            panel,
            text="Crear proveedor",
            command=self._guardar_proveedor,
            fg_color=COLOR_PRIMARIO,
            hover_color=COLOR_PRIMARIO_HOVER,
            height=42,
            font=ctk.CTkFont(size=13, weight="bold"),
        ).grid(
            row=5,
            column=0,
            columnspan=2,
            sticky="ew",
            padx=18,
            pady=20,
        )

        self.lista_proveedores = ctk.CTkTextbox(
            panel,
            height=220,
            fg_color=COLOR_PANEL_SECUNDARIO,
            border_width=1,
            border_color=COLOR_BORDE,
            text_color=COLOR_TEXTO,
        )
        self.lista_proveedores.grid(
            row=6,
            column=0,
            columnspan=2,
            sticky="nsew",
            padx=18,
            pady=(0, 18),
        )

    def _guardar_proveedor(self):
        try:
            registro = {
                "id": uuid4().hex[:12].upper(),
                "nombre": _texto(
                    self.proveedor_nombre.get(),
                    "Nombre / razón social",
                ),
                "ruc": _texto(
                    self.proveedor_ruc.get(),
                    "RUC",
                    obligatorio=False,
                ),
                "telefono": _texto(
                    self.proveedor_telefono.get(),
                    "Teléfono",
                    obligatorio=False,
                ),
                "direccion": _texto(
                    self.proveedor_direccion.get(),
                    "Dirección",
                    obligatorio=False,
                ),
                "observacion": _texto(
                    self.proveedor_observacion.get(),
                    "Observación",
                    obligatorio=False,
                ),
                "registrado_en": datetime.now().isoformat(timespec="seconds"),
            }
            guardar_proveedor(registro)
        except (ValueError, OSError) as error:
            messagebox.showerror("No se pudo guardar", str(error), parent=self)
            return

        for entrada in (
            self.proveedor_nombre,
            self.proveedor_ruc,
            self.proveedor_telefono,
            self.proveedor_direccion,
            self.proveedor_observacion,
        ):
            entrada.delete(0, "end")
        self._actualizar_proveedores()
        messagebox.showinfo(
            "Proveedor creado",
            "El proveedor quedó disponible para futuras facturas.",
            parent=self,
        )

    def _actualizar_proveedores(self):
        proveedores = [dato for _, dato in obtener_proveedores()]
        nombres = sorted(
            (
                proveedor.get("nombre", "")
                for proveedor in proveedores
                if proveedor.get("nombre")
            ),
            key=str.casefold,
        )
        self.credito_proveedor.configure(values=nombres)

        self.lista_proveedores.configure(state="normal")
        self.lista_proveedores.delete("1.0", "end")
        if not proveedores:
            self.lista_proveedores.insert(
                "end",
                "Todavía no hay proveedores registrados.",
            )
        else:
            for proveedor in sorted(
                proveedores,
                key=lambda item: item.get("nombre", "").casefold(),
            ):
                detalle = proveedor.get("nombre", "")
                if proveedor.get("ruc"):
                    detalle += f" · RUC {proveedor['ruc']}"
                if proveedor.get("telefono"):
                    detalle += f" · {proveedor['telefono']}"
                self.lista_proveedores.insert("end", detalle + "\n")
        self.lista_proveedores.configure(state="disabled")

    def _crear_informes(self):
        tab = self.pestanas.tab("Informes")
        tab.grid_rowconfigure(2, weight=1)
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_columnconfigure(1, weight=0)

        filtros = ctk.CTkFrame(tab, fg_color="transparent")
        filtros.grid(row=0, column=0, sticky="ew", padx=12, pady=12)
        for columna in range(6):
            filtros.grid_columnconfigure(columna, weight=1)

        self.filtro_periodo = ctk.CTkEntry(
            filtros,
            placeholder_text="MM-AAAA",
            fg_color=COLOR_PANEL_SECUNDARIO,
            border_color=COLOR_BORDE,
            text_color=COLOR_TEXTO,
        )
        self.filtro_periodo.insert(0, datetime.now().strftime("%m-%Y"))
        self.filtro_periodo.grid(row=0, column=0, sticky="ew", padx=5)

        self.filtro_tipo = ctk.CTkComboBox(
            filtros,
            values=["Todos", "Contado", "Crédito"],
            state="readonly",
            fg_color=COLOR_PANEL_SECUNDARIO,
            border_color=COLOR_BORDE,
            button_color=COLOR_PRIMARIO,
            text_color=COLOR_TEXTO,
        )
        self.filtro_tipo.set("Todos")
        self.filtro_tipo.grid(row=0, column=1, sticky="ew", padx=5)

        self.filtro_proveedor = ctk.CTkEntry(
            filtros,
            placeholder_text="Proveedor / local",
            fg_color=COLOR_PANEL_SECUNDARIO,
            border_color=COLOR_BORDE,
            text_color=COLOR_TEXTO,
        )
        self.filtro_proveedor.grid(row=0, column=2, sticky="ew", padx=5)

        ctk.CTkButton(
            filtros,
            text="Aplicar filtros",
            command=self._actualizar_informe,
            fg_color=COLOR_PRIMARIO,
            hover_color=COLOR_PRIMARIO_HOVER,
        ).grid(row=0, column=3, sticky="ew", padx=5)

        ctk.CTkButton(
            filtros,
            text="Exportar Excel",
            command=self._exportar_informe_excel,
            fg_color=COLOR_VERDE,
            hover_color="#12835B",
        ).grid(row=0, column=4, sticky="ew", padx=5)

        ctk.CTkButton(
            filtros,
            text="Exportar PDF",
            command=self._exportar_informe_pdf,
            fg_color=COLOR_ROJO,
            hover_color="#BE3F3F",
        ).grid(row=0, column=5, sticky="ew", padx=5)

        self.resumen = ctk.CTkLabel(
            tab,
            text="",
            justify="left",
            anchor="w",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COLOR_TEXTO,
            fg_color=COLOR_PANEL_SECUNDARIO,
            corner_radius=10,
        )
        self.resumen.grid(
            row=1,
            column=0,
            sticky="ew",
            padx=17,
            pady=(0, 12),
        )

        columnas = (
            "fecha",
            "tipo",
            "numero",
            "proveedor",
            "monto",
            "origen",
            "plazo",
            "vencimiento",
        )
        self.tabla = ttk.Treeview(
            tab,
            columns=columnas,
            show="headings",
            height=15,
        )
        titulos = {
            "fecha": "Fecha",
            "tipo": "Tipo",
            "numero": "N.º factura",
            "proveedor": "Proveedor / local",
            "monto": "Monto",
            "origen": "Origen",
            "plazo": "Plazo",
            "vencimiento": "Vencimiento",
        }
        anchos = {
            "fecha": 90,
            "tipo": 80,
            "numero": 120,
            "proveedor": 220,
            "monto": 110,
            "origen": 110,
            "plazo": 70,
            "vencimiento": 100,
        }
        for columna in columnas:
            self.tabla.heading(
                columna,
                text=titulos[columna],
                command=lambda c=columna: self._ordenar_treeview(
                    self.tabla,
                    c,
                ),
            )
            self.tabla.column(
                columna,
                width=anchos[columna],
                anchor="center",
            )
        self.tabla.column("proveedor", anchor="w")
        self.tabla.grid(
            row=2,
            column=0,
            sticky="nsew",
            padx=(17, 0),
            pady=(0, 10),
        )
        barra_tabla = ttk.Scrollbar(
            tab,
            orient="vertical",
            command=self.tabla.yview,
        )
        barra_tabla.grid(
            row=2,
            column=1,
            sticky="ns",
            padx=(0, 17),
            pady=(0, 10),
        )
        self.tabla.configure(yscrollcommand=barra_tabla.set)
        self.tabla.bind(
            "<<TreeviewSelect>>",
            self._seleccionar_factura,
        )

        ctk.CTkButton(
            tab,
            text="Eliminar factura seleccionada",
            command=self._eliminar_factura,
            fg_color=COLOR_ROJO,
            hover_color="#BE3F3F",
            height=38,
        ).grid(
            row=3,
            column=0,
            sticky="e",
            padx=17,
            pady=(0, 15),
        )

    def _facturas_filtradas(self):
        periodo = self.filtro_periodo.get().strip()
        try:
            mes = datetime.strptime(periodo, "%m-%Y")
        except ValueError as error:
            raise ValueError(
                "El período debe tener formato MM-AAAA."
            ) from error

        tipo = {
            "Contado": "CONTADO",
            "Crédito": "CREDITO",
        }.get(self.filtro_tipo.get())
        proveedor = _normalizar(self.filtro_proveedor.get())

        resultado = []
        for _, factura in obtener_facturas():
            try:
                fecha = _fecha(factura.get("fecha", ""))
            except ValueError:
                continue
            if (fecha.month, fecha.year) != (mes.month, mes.year):
                continue
            if tipo and factura.get("tipo") != tipo:
                continue
            if proveedor and proveedor not in _normalizar(
                factura.get("proveedor")
            ):
                continue
            resultado.append(factura)

        resultado.sort(
            key=lambda factura: _fecha(factura["fecha"]),
            reverse=True,
        )
        return resultado

    def _datos_informe_exportacion(self):
        facturas = self._facturas_filtradas()
        if not facturas:
            raise ValueError(
                "No hay facturas visibles para exportar "
                "con los filtros actuales."
            )

        filas = []
        for factura in facturas:
            filas.append(
                [
                    factura.get("fecha", ""),
                    (
                        "Crédito"
                        if factura.get("tipo") == "CREDITO"
                        else "Contado"
                    ),
                    factura.get("numero", ""),
                    factura.get("proveedor", ""),
                    int(factura.get("monto", 0)),
                    factura.get("origen", "") or "",
                    factura.get("descripcion", "") or "-",
                    (
                        f"{factura.get('plazo_dias', 0)} días"
                        if factura.get("tipo") == "CREDITO"
                        else "-"
                    ),
                    factura.get("vencimiento", "") or "",
                ]
            )
        return filas

    def _actualizar_informe(self):
        try:
            facturas = self._facturas_filtradas()
        except ValueError as error:
            messagebox.showerror(
                "Filtro inválido",
                str(error),
                parent=self,
            )
            return

        total_contado = sum(
            factura.get("monto", 0)
            for factura in facturas
            if factura.get("tipo") == "CONTADO"
        )
        total_credito = sum(
            factura.get("monto", 0)
            for factura in facturas
            if factura.get("tipo") == "CREDITO"
        )
        self.resumen.configure(
            text=(
                f"  Contado: Gs. {_formatear_monto(total_contado)}"
                f"    |    Crédito: Gs. {_formatear_monto(total_credito)}"
                f"    |    Total: Gs. "
                f"{_formatear_monto(total_contado + total_credito)}"
                f"    |    Cantidad: {len(facturas)}"
            )
        )

        for item in self.tabla.get_children():
            self.tabla.delete(item)

        for factura in facturas:
            plazo = (
                f"{factura.get('plazo_dias', 0)} días"
                if factura.get("tipo") == "CREDITO"
                else "-"
            )
            self.tabla.insert(
                "",
                "end",
                iid=factura.get("id"),
                values=(
                    factura.get("fecha", ""),
                    (
                        "Crédito"
                        if factura.get("tipo") == "CREDITO"
                        else "Contado"
                    ),
                    factura.get("numero", ""),
                    factura.get("proveedor", ""),
                    _formatear_monto(factura.get("monto", 0)),
                    factura.get("origen", "") or "",
                    plazo,
                    factura.get("vencimiento", "") or "",
                ),
            )
        self.factura_seleccionada_id = None

    def _exportar_informe_excel(self):
        try:
            filas = self._datos_informe_exportacion()
        except ValueError as error:
            messagebox.showwarning("Sin resultados", str(error), parent=self)
            return

        periodo = self.filtro_periodo.get().strip()
        ruta = filedialog.asksaveasfilename(
            parent=self,
            title="Guardar informe de facturas",
            defaultextension=".xlsx",
            initialfile=f"Informe_Facturas_{periodo}.xlsx",
            filetypes=[("Archivo Excel", "*.xlsx")],
        )
        if not ruta:
            return

        encabezados = [
            "Fecha",
            "Tipo",
            "Número de factura",
            "Proveedor / local",
            "Monto",
            "Origen",
            "Descripción",
            "Plazo",
            "Vencimiento",
        ]
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Facturas"
        hoja.append(encabezados)

        for celda in hoja[1]:
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center")

        for fila in filas:
            hoja.append(fila)

        for celda in hoja["E"][1:]:
            celda.number_format = "#,##0"

        for indice, ancho in enumerate(
            [13, 12, 22, 32, 16, 18, 40, 14, 16],
            start=1,
        ):
            hoja.column_dimensions[get_column_letter(indice)].width = ancho

        hoja.freeze_panes = "A2"
        hoja.auto_filter.ref = hoja.dimensions

        try:
            libro.save(ruta)
        except OSError as error:
            messagebox.showerror(
                "No se pudo exportar",
                str(error),
                parent=self,
            )
            return

        messagebox.showinfo(
            "Informe exportado",
            f"Se exportaron {len(filas)} facturas a Excel.",
            parent=self,
        )

    def _exportar_informe_pdf(self):
        try:
            filas = self._datos_informe_exportacion()
        except ValueError as error:
            messagebox.showwarning("Sin resultados", str(error), parent=self)
            return

        periodo = self.filtro_periodo.get().strip()
        ruta = filedialog.asksaveasfilename(
            parent=self,
            title="Guardar informe de facturas",
            defaultextension=".pdf",
            initialfile=f"Informe_Facturas_{periodo}.pdf",
            filetypes=[("Documento PDF", "*.pdf")],
        )
        if not ruta:
            return

        estilos = getSampleStyleSheet()
        elementos = [
            Paragraph("Informe documental de facturas", estilos["Title"]),
            Paragraph(
                (
                    f"Período: {periodo} · Tipo: {self.filtro_tipo.get()} · "
                    f"Proveedor: "
                    f"{self.filtro_proveedor.get().strip() or 'Todos'}"
                ),
                estilos["Normal"],
            ),
            Spacer(1, 0.4 * cm),
        ]

        datos_tabla = [[
            "Fecha",
            "Tipo",
            "Factura",
            "Proveedor",
            "Monto",
            "Origen",
            "Vencimiento",
        ]]
        total = 0
        for fila in filas:
            total += int(fila[4])
            datos_tabla.append(
                [
                    fila[0],
                    fila[1],
                    fila[2],
                    fila[3],
                    _formatear_monto(fila[4]),
                    fila[5],
                    fila[8],
                ]
            )

        tabla = Table(
            datos_tabla,
            repeatRows=1,
            colWidths=[
                2.1 * cm,
                1.8 * cm,
                3.2 * cm,
                5.4 * cm,
                2.8 * cm,
                2.6 * cm,
                2.5 * cm,
            ],
        )
        tabla.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#246BFD"),
                    ),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ALIGN", (4, 1), (4, -1), "RIGHT"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#EAF0F7")],
                    ),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        elementos.extend(
            [
                tabla,
                Spacer(1, 0.4 * cm),
                Paragraph(
                    (
                        f"<b>Cantidad:</b> {len(filas)}"
                        f" &nbsp;&nbsp;&nbsp; "
                        f"<b>Total:</b> Gs. {_formatear_monto(total)}"
                    ),
                    estilos["Normal"],
                ),
            ]
        )

        try:
            SimpleDocTemplate(
                ruta,
                pagesize=landscape(A4),
                rightMargin=1 * cm,
                leftMargin=1 * cm,
                topMargin=1 * cm,
                bottomMargin=1 * cm,
                title="Informe de facturas",
            ).build(elementos)
        except (OSError, ValueError) as error:
            messagebox.showerror(
                "No se pudo exportar",
                str(error),
                parent=self,
            )
            return

        messagebox.showinfo(
            "Informe exportado",
            f"Se exportaron {len(filas)} facturas a PDF.",
            parent=self,
        )

    def _eliminar_mes_facturas(self):
        periodo = self.filtro_periodo.get().strip()

        try:
            mes = datetime.strptime(periodo, "%m-%Y")
        except ValueError:
            messagebox.showerror(
                "Período inválido",
                "El período debe tener formato MM-AAAA.",
                parent=self,
            )
            return

        facturas_mes = []
        for _, factura in obtener_facturas():
            try:
                fecha = _fecha(factura.get("fecha", ""))
            except ValueError:
                continue
            if (fecha.month, fecha.year) == (mes.month, mes.year):
                facturas_mes.append(factura)

        if not facturas_mes:
            messagebox.showinfo(
                "Sin facturas",
                f"No hay facturas registradas en {periodo}.",
                parent=self,
            )
            return

        total = sum(
            int(factura.get("monto", 0))
            for factura in facturas_mes
        )

        confirmar = messagebox.askyesno(
            "Eliminar mes completo",
            (
                f"Se eliminarán TODAS las facturas de {periodo}.\n\n"
                f"Cantidad: {len(facturas_mes)}\n"
                f"Total: Gs. {_formatear_monto(total)}\n\n"
                "Esta acción no se puede deshacer. ¿Continuar?"
            ),
            icon="warning",
            parent=self,
        )
        if not confirmar:
            return

        confirmar_final = messagebox.askyesno(
            "Confirmación final",
            (
                f"¿Confirmás definitivamente la eliminación "
                f"del período {periodo}?"
            ),
            icon="warning",
            parent=self,
        )
        if not confirmar_final:
            return

        try:
            eliminadas = eliminar_facturas_por_periodo(periodo)
        except (ValueError, OSError) as error:
            messagebox.showerror(
                "No se pudo eliminar",
                str(error),
                parent=self,
            )
            return

        self._actualizar_informe()
        self.registros_importacion = []
        self.ruta_importacion = None
        self.boton_importar.configure(state="disabled")

        for item in self.tabla_importacion.get_children():
            self.tabla_importacion.delete(item)

        self.etiqueta_archivo.configure(
            text="Seleccioná una planilla de compras."
        )
        self.resumen_importacion.configure(
            text="Sin archivo analizado."
        )

        messagebox.showinfo(
            "Mes eliminado",
            (
                f"Se eliminaron {len(eliminadas)} facturas "
                f"del período {periodo}."
            ),
            parent=self,
        )

    def _ordenar_treeview(self, tabla, columna):
        descendente = getattr(
            tabla,
            "_orden_descendente",
            {},
        ).get(columna, False)

        def clave(item):
            valor = tabla.set(item, columna).strip()

            if columna == "fecha":
                try:
                    return _fecha(valor)
                except ValueError:
                    return datetime.min

            if columna == "monto":
                try:
                    return int(
                        valor.replace(".", "").replace(",", "")
                    )
                except ValueError:
                    return 0

            if columna == "plazo":
                coincidencia = re.search(r"\d+", valor)
                return int(coincidencia.group()) if coincidencia else 0

            return _normalizar(valor)

        items = list(tabla.get_children(""))
        items.sort(key=clave, reverse=descendente)

        for posicion, item in enumerate(items):
            tabla.move(item, "", posicion)

        ordenes = getattr(tabla, "_orden_descendente", {})
        ordenes[columna] = not descendente
        tabla._orden_descendente = ordenes

    def _seleccionar_factura(self, _evento=None):
        seleccion = self.tabla.selection()
        self.factura_seleccionada_id = seleccion[0] if seleccion else None

    def _eliminar_factura(self):
        if not self.factura_seleccionada_id:
            messagebox.showwarning(
                "Sin selección",
                "Seleccioná una factura de la tabla.",
                parent=self,
            )
            return
        if not messagebox.askyesno(
            "Eliminar factura",
            "¿Confirmás la eliminación del registro seleccionado?",
            parent=self,
        ):
            return
        try:
            eliminar_factura_por_id(self.factura_seleccionada_id)
        except OSError as error:
            messagebox.showerror(
                "No se pudo eliminar",
                str(error),
                parent=self,
            )
            return
        self._actualizar_informe()

    def _crear_importador(self):
        tab = self.pestanas.tab("Importar Excel")
        tab.grid_rowconfigure(2, weight=1)
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_columnconfigure(1, weight=0)

        acciones = ctk.CTkFrame(tab, fg_color="transparent")
        acciones.grid(row=0, column=0, sticky="ew", padx=16, pady=16)
        acciones.grid_columnconfigure(0, weight=1)

        self.etiqueta_archivo = ctk.CTkLabel(
            acciones,
            text="Seleccioná una planilla de compras.",
            anchor="w",
            text_color=COLOR_TEXTO_SUAVE,
        )
        self.etiqueta_archivo.grid(
            row=0,
            column=0,
            sticky="ew",
            padx=(0, 12),
        )

        ctk.CTkButton(
            acciones,
            text="Seleccionar Excel",
            command=self._seleccionar_excel,
            fg_color=COLOR_PRIMARIO,
            hover_color=COLOR_PRIMARIO_HOVER,
        ).grid(row=0, column=1, padx=5)

        self.boton_importar = ctk.CTkButton(
            acciones,
            text="Importar válidas",
            command=self._confirmar_importacion,
            fg_color=COLOR_VERDE,
            hover_color="#12835B",
            state="disabled",
        )
        self.boton_importar.grid(row=0, column=2, padx=5)

        ctk.CTkButton(
            acciones,
            text="Eliminar mes completo",
            command=self._eliminar_mes_facturas,
            fg_color=COLOR_ROJO,
            hover_color="#BE3F3F",
        ).grid(row=0, column=3, padx=5)

        self.resumen_importacion = ctk.CTkLabel(
            tab,
            text="Sin archivo analizado.",
            anchor="w",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=COLOR_TEXTO,
            fg_color=COLOR_PANEL_SECUNDARIO,
            corner_radius=10,
        )
        self.resumen_importacion.grid(
            row=1,
            column=0,
            sticky="ew",
            padx=17,
            pady=(0, 12),
        )

        columnas = (
            "estado",
            "hoja",
            "fecha",
            "tipo",
            "numero",
            "proveedor",
            "monto",
            "origen",
        )
        self.tabla_importacion = ttk.Treeview(
            tab,
            columns=columnas,
            show="headings",
            height=18,
        )
        titulos = {
            "estado": "Estado",
            "hoja": "Hoja",
            "fecha": "Fecha",
            "tipo": "Tipo",
            "numero": "Factura",
            "proveedor": "Proveedor",
            "monto": "Monto",
            "origen": "Sucursal",
        }
        anchos = {
            "estado": 90,
            "hoja": 150,
            "fecha": 90,
            "tipo": 80,
            "numero": 130,
            "proveedor": 240,
            "monto": 110,
            "origen": 100,
        }
        for columna in columnas:
            self.tabla_importacion.heading(
                columna,
                text=titulos[columna],
                command=lambda c=columna: self._ordenar_treeview(
                    self.tabla_importacion,
                    c,
                ),
            )
            self.tabla_importacion.column(
                columna,
                width=anchos[columna],
                anchor="center",
            )
        self.tabla_importacion.column("proveedor", anchor="w")
        self.tabla_importacion.grid(
            row=2,
            column=0,
            sticky="nsew",
            padx=(17, 0),
            pady=(0, 15),
        )
        barra_importacion = ttk.Scrollbar(
            tab,
            orient="vertical",
            command=self.tabla_importacion.yview,
        )
        barra_importacion.grid(
            row=2,
            column=1,
            sticky="ns",
            padx=(0, 17),
            pady=(0, 15),
        )
        self.tabla_importacion.configure(
            yscrollcommand=barra_importacion.set
        )

    def _seleccionar_excel(self):
        ruta = filedialog.askopenfilename(
            parent=self,
            title="Seleccionar planilla de facturas",
            filetypes=[("Archivos Excel", "*.xlsx")],
        )
        if not ruta:
            return

        try:
            registros, errores = leer_facturas_excel(ruta)
        except (OSError, ValueError, KeyError, TypeError) as error:
            messagebox.showerror(
                "No se pudo analizar",
                str(error),
                parent=self,
            )
            return

        existentes = {
            _clave_factura(factura)
            for _, factura in obtener_facturas()
        }
        vistos = set()
        preparados = []
        duplicados = 0

        for registro in registros:
            clave = _clave_factura(registro)
            es_duplicado = clave in existentes or clave in vistos
            estado = "DUPLICADO" if es_duplicado else "VÁLIDO"
            if es_duplicado:
                duplicados += 1
            else:
                vistos.add(clave)
            preparados.append((estado, registro))

        self.ruta_importacion = Path(ruta)
        self.registros_importacion = preparados
        self.etiqueta_archivo.configure(
            text=self.ruta_importacion.name
        )

        for item in self.tabla_importacion.get_children():
            self.tabla_importacion.delete(item)

        for indice, (estado, registro) in enumerate(preparados):
            self.tabla_importacion.insert(
                "",
                "end",
                iid=str(indice),
                values=(
                    estado,
                    registro.get("hoja_origen", ""),
                    registro.get("fecha", ""),
                    (
                        "Crédito"
                        if registro.get("tipo") == "CREDITO"
                        else "Contado"
                    ),
                    registro.get("numero", ""),
                    registro.get("proveedor", ""),
                    _formatear_monto(registro.get("monto", 0)),
                    registro.get("origen", "") or "",
                ),
            )

        validos = len(preparados) - duplicados
        self.resumen_importacion.configure(
            text=(
                f"  Válidas: {validos}"
                f"    |    Duplicadas: {duplicados}"
                f"    |    Observaciones: {len(errores)}"
            )
        )
        self.boton_importar.configure(
            state="normal" if validos else "disabled"
        )

        if errores:
            messagebox.showwarning(
                "Planilla analizada",
                (
                    f"Se omitieron {len(errores)} filas incompletas "
                    "o inválidas. Las facturas válidas están en la vista previa."
                ),
                parent=self,
            )

    def _confirmar_importacion(self):
        registros = [
            registro
            for estado, registro in self.registros_importacion
            if estado == "VÁLIDO"
        ]
        if not registros:
            messagebox.showwarning(
                "Sin registros",
                "No hay facturas válidas para importar.",
                parent=self,
            )
            return

        if not messagebox.askyesno(
            "Confirmar importación",
            f"¿Importar {len(registros)} facturas válidas?",
            parent=self,
        ):
            return

        try:
            guardar_facturas(registros)
        except OSError as error:
            messagebox.showerror(
                "No se pudo importar",
                str(error),
                parent=self,
            )
            return

        self.registros_importacion = [
            ("IMPORTADO" if estado == "VÁLIDO" else estado, registro)
            for estado, registro in self.registros_importacion
        ]
        self.boton_importar.configure(state="disabled")
        self._actualizar_informe()
        self._actualizar_proveedores()
        messagebox.showinfo(
            "Importación terminada",
            f"Se importaron {len(registros)} facturas.",
            parent=self,
        )


class VentanaFacturas(ctk.CTkToplevel):
    """Compatibilidad para abrir Facturas como ventana independiente."""

    def __init__(self, master):
        super().__init__(master)
        self.title("Registro de facturas")
        self.geometry("1450x880")
        self.minsize(1100, 700)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        panel = PanelFacturas(self)
        panel.grid(row=0, column=0, sticky="nsew")


def crear_panel_facturas(master):
    return PanelFacturas(master)


def abrir_facturas(master):
    ventana = VentanaFacturas(master)
    ventana.grab_set()
    return ventana
