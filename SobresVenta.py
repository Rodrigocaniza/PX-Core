"""Importación de sobres de venta (PC) desde Excel y comisión de la encargada."""

from __future__ import annotations

import calendar
from pathlib import Path
import shutil

from openpyxl import load_workbook

import Movimientos
import Novedades
from datos import guardar_datos, leer_datos
from ImportadorExcel import (
    buscar_encabezados,
    celda_vacia,
    normalizar,
    parsear_monto,
    texto_seguro,
)


RUTA_SOBRES = Path("Datos") / "sobres_venta.txt"
RUTA_CONFIG_COMISION = Path("Datos") / "config_comision_pc.txt"
UNIDAD_SOBRES = "PC"
PORCENTAJE_COMISION_PREDETERMINADO = 1.0

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}


def crear_linea_sobre(datos):
    return " | ".join([
        str(datos["numero"]),
        datos["periodo"],
        str(datos["monto"]),
        str(datos["comision"]),
        str(datos["monto_final"]),
    ])


def separar_sobre(linea):
    partes = [parte.strip() for parte in linea.split("|")]
    if len(partes) != 5:
        return None
    try:
        return {
            "numero": int(partes[0]),
            "periodo": partes[1],
            "monto": int(partes[2]),
            "comision": int(partes[3]),
            "monto_final": int(partes[4]),
        }
    except ValueError:
        return None


def obtener_sobres_guardados():
    sobres = []
    for linea in leer_datos(RUTA_SOBRES):
        sobre = separar_sobre(linea)
        if sobre is not None:
            sobres.append(sobre)
    return sobres


def leer_porcentaje_comision():
    lineas = leer_datos(RUTA_CONFIG_COMISION)
    if not lineas:
        return PORCENTAJE_COMISION_PREDETERMINADO
    try:
        return float(lineas[0].strip().replace(",", "."))
    except ValueError:
        return PORCENTAJE_COMISION_PREDETERMINADO


def guardar_porcentaje_comision(porcentaje):
    guardar_datos(RUTA_CONFIG_COMISION, [str(porcentaje)])


def crear_resultado():
    return {
        "periodos": {},
        "errores": [],
        "advertencias": [],
        "duplicados": [],
    }


def _periodo_de_hoja(hoja_titulo, anio):
    mes = MESES.get(normalizar(hoja_titulo))
    if mes is None:
        return None
    return f"{mes:02d}-{anio}"


def analizar_excel_sobres(ruta_archivo, anio):
    libro = load_workbook(ruta_archivo, data_only=True)
    resultado = crear_resultado()

    existentes = {}
    for sobre in obtener_sobres_guardados():
        existentes.setdefault(
            (sobre["numero"], sobre["periodo"]), set()
        ).add(sobre["monto"])

    for hoja in libro.worksheets:
        periodo = _periodo_de_hoja(hoja.title, anio)
        if periodo is None:
            continue

        fila_encabezado, columnas = buscar_encabezados(
            hoja, ["N° SOBRE", "MONTO"]
        )
        if fila_encabezado is None:
            continue

        vistos_en_lote = {}
        nuevas = []
        total_monto = 0
        total_comision = 0
        total_monto_final = 0

        for numero_fila in range(fila_encabezado + 1, hoja.max_row + 1):
            fila = hoja[numero_fila]
            celda_numero = fila[0].value

            if celda_vacia(celda_numero):
                continue
            if normalizar(texto_seguro(celda_numero)) == "total":
                continue

            try:
                numero_sobre = int(float(celda_numero))
            except (TypeError, ValueError):
                resultado["errores"].append(
                    f"{hoja.title}, fila {numero_fila}: "
                    "el N° de sobre no es válido"
                )
                continue

            try:
                monto = parsear_monto(fila[1].value)
            except ValueError as exc:
                resultado["errores"].append(
                    f"{hoja.title}, fila {numero_fila}: {exc}"
                )
                continue

            comision_celda = fila[2].value
            comision = (
                0
                if celda_vacia(comision_celda)
                else parsear_monto(comision_celda, permitir_cero=True)
            )

            monto_final_celda = fila[3].value
            monto_final = (
                monto - comision
                if celda_vacia(monto_final_celda)
                else parsear_monto(monto_final_celda, permitir_cero=True)
            )

            clave = (numero_sobre, periodo)
            montos_previos = existentes.get(clave, set())

            if monto in montos_previos or vistos_en_lote.get(clave) == monto:
                resultado["duplicados"].append(
                    f"{hoja.title}, fila {numero_fila}: "
                    f"el sobre N° {numero_sobre} ya estaba importado con "
                    "el mismo monto; se omite"
                )
                continue

            if montos_previos or clave in vistos_en_lote:
                resultado["advertencias"].append(
                    f"{hoja.title}, fila {numero_fila}: "
                    f"el sobre N° {numero_sobre} ya existe con un monto "
                    "distinto; se importó igual, revisalo"
                )

            vistos_en_lote[clave] = monto
            nuevas.append({
                "numero": numero_sobre,
                "periodo": periodo,
                "monto": monto,
                "comision": comision,
                "monto_final": monto_final,
            })
            total_monto += monto
            total_comision += comision
            total_monto_final += monto_final

        if nuevas:
            resultado["periodos"][periodo] = {
                "hoja": hoja.title,
                "nuevas": nuevas,
                "total_monto": total_monto,
                "total_comision": total_comision,
                "total_monto_final": total_monto_final,
            }

    return resultado


def texto_resumen(resultado, porcentaje_comision):
    lineas = []
    for periodo in sorted(resultado["periodos"]):
        datos = resultado["periodos"][periodo]
        comision_encargada = round(
            datos["total_monto"] * porcentaje_comision / 100
        )
        lineas.append(f"PERÍODO {periodo} ({datos['hoja']})")
        lineas.append(f"  Sobres nuevos: {len(datos['nuevas'])}")
        lineas.append(
            f"  Total monto: Gs. "
            f"{Movimientos.formatear_monto(datos['total_monto'])}"
        )
        lineas.append(
            f"  Total comisión asociaciones: Gs. "
            f"{Movimientos.formatear_monto(datos['total_comision'])}"
        )
        lineas.append(
            f"  Total monto final (ingreso PC): Gs. "
            f"{Movimientos.formatear_monto(datos['total_monto_final'])}"
        )
        lineas.append(
            f"  Comisión encargada ({porcentaje_comision}%): Gs. "
            f"{Movimientos.formatear_monto(comision_encargada)}"
        )
        lineas.append("")

    lineas.append(f"ADVERTENCIAS: {len(resultado['advertencias'])}")
    lineas.extend(f"  • {item}" for item in resultado["advertencias"])
    lineas.append(f"DUPLICADOS OMITIDOS: {len(resultado['duplicados'])}")
    lineas.append(f"ERRORES: {len(resultado['errores'])}")
    lineas.extend(f"  • {item}" for item in resultado["errores"])

    return "\n".join(lineas)


def _respaldar_archivos():
    respaldo = {}
    for ruta in (
        RUTA_SOBRES,
        Path(Movimientos.RUTA_MOVIMIENTOS),
        Novedades.RUTA_NOVEDADES,
    ):
        respaldo[ruta] = leer_datos(ruta)
    return respaldo


def _restaurar_archivos(respaldo):
    for ruta, contenido in respaldo.items():
        guardar_datos(ruta, contenido)


def aplicar_resultado(resultado, porcentaje_comision, funcionario=None):
    if resultado["errores"]:
        raise ValueError(
            "hay filas con errores; corregilas antes de importar"
        )

    respaldo = _respaldar_archivos()

    try:
        sobres = leer_datos(RUTA_SOBRES)
        movimientos = leer_datos(str(Movimientos.RUTA_MOVIMIENTOS))
        novedades = leer_datos(Novedades.RUTA_NOVEDADES)

        for periodo, datos in resultado["periodos"].items():
            for sobre in datos["nuevas"]:
                sobres.append(crear_linea_sobre(sobre))

            mes = int(periodo[:2])
            anio = int(periodo[3:])
            ultimo_dia = calendar.monthrange(anio, mes)[1]
            fecha = f"{ultimo_dia:02d}-{mes:02d}-{anio}"

            if datos["total_monto_final"] > 0:
                movimientos.append(
                    f"Ingreso|{fecha}|Externo|{UNIDAD_SOBRES}|"
                    f"{datos['total_monto_final']}|Si"
                )

            comision_encargada = round(
                datos["total_monto"] * porcentaje_comision / 100
            )
            if comision_encargada > 0 and funcionario is not None:
                novedades.append(
                    Novedades.crear_linea_novedad({
                        "cedula": funcionario["cedula"],
                        "nombre": funcionario["nombre"],
                        "tipo": "Comisión",
                        "fecha_inicio": fecha,
                        "fecha_fin": fecha,
                        "monto": comision_encargada,
                        "cubierto_ips": "No",
                        "motivo": (
                            f"Comisión ventas PC {periodo} "
                            f"({porcentaje_comision}%)"
                        ),
                    })
                )

        guardar_datos(RUTA_SOBRES, sobres)
        guardar_datos(str(Movimientos.RUTA_MOVIMIENTOS), movimientos)
        guardar_datos(Novedades.RUTA_NOVEDADES, novedades)
    except Exception:
        _restaurar_archivos(respaldo)
        raise

    guardar_porcentaje_comision(porcentaje_comision)


def seleccionar_encargada_pc():
    funcionarios = [
        funcionario
        for funcionario in Novedades.obtener_funcionarios_activos()
        if funcionario["unidad"] == UNIDAD_SOBRES
    ]
    if not funcionarios:
        funcionarios = Novedades.obtener_funcionarios_activos()

    if not funcionarios:
        print()
        print("No hay funcionarios activos registrados.")
        return None

    print()
    print("¿A quién corresponde la comisión de este mes?")
    for numero, funcionario in enumerate(funcionarios, start=1):
        print(f"{numero}. {funcionario['nombre']} | {funcionario['cargo']}")
    print("0. No asignar comisión")
    print()

    while True:
        opcion = input("Seleccione una opción: ").strip()
        if opcion == "0":
            return None
        if not opcion.isdigit():
            print("Ingresá un número válido.")
            continue
        indice = int(opcion) - 1
        if indice < 0 or indice >= len(funcionarios):
            print("La opción seleccionada no existe.")
            continue
        return funcionarios[indice]


def pedir_porcentaje_comision():
    actual = leer_porcentaje_comision()
    print()
    entrada = input(
        f"Porcentaje de comisión de la encargada [{actual}%]: "
    ).strip()
    if entrada == "":
        return actual
    try:
        return float(entrada.replace(",", "."))
    except ValueError:
        print("Porcentaje inválido, se usará el anterior.")
        return actual


def pedir_anio():
    while True:
        entrada = input("Año de los sobres (AAAA): ").strip()
        if entrada.isdigit() and len(entrada) == 4:
            return int(entrada)
        print("Ingresá un año válido, por ejemplo 2026.")


def importar_sobres_pc():
    print()
    print("====================================")
    print("     IMPORTAR SOBRES DE VENTA PC")
    print("====================================")

    ruta_archivo = input("Ruta del archivo Excel: ").strip()
    if not Path(ruta_archivo).exists():
        print()
        print("No se encontró el archivo.")
        return

    anio = pedir_anio()
    resultado = analizar_excel_sobres(ruta_archivo, anio)

    if not resultado["periodos"]:
        print()
        print("No se encontraron sobres nuevos para importar.")
        if resultado["errores"]:
            print()
            print("\n".join(f"  • {item}" for item in resultado["errores"]))
        return

    porcentaje_comision = pedir_porcentaje_comision()

    print()
    print(texto_resumen(resultado, porcentaje_comision))

    if resultado["errores"]:
        print()
        print("Hay errores. Corregí el Excel y volvé a intentar.")
        return

    confirmacion = input("\n¿Confirmar la importación? (S/N): ").strip()
    if normalizar(confirmacion) != "s":
        print("Importación cancelada.")
        return

    funcionario = seleccionar_encargada_pc()

    try:
        aplicar_resultado(resultado, porcentaje_comision, funcionario)
    except Exception as exc:
        print()
        print(f"No se pudo importar: {exc}")
        return

    print()
    print("Importación completada.")


def abrir_importador_sobres(ventana_padre):
    import customtkinter as ctk
    from tkinter import filedialog, messagebox

    from ImportadorExcel import (
        COLOR_BORDE,
        COLOR_PANEL,
        COLOR_PANEL_SECUNDARIO,
        COLOR_PRIMARIO,
        COLOR_PRIMARIO_HOVER,
        COLOR_TEXTO,
        COLOR_TEXTO_SUAVE,
        COLOR_VERDE,
    )

    ventana = ctk.CTkToplevel(ventana_padre)
    ventana.title("Importar sobres de venta PC")
    ventana.geometry("980x760")
    ventana.minsize(820, 640)
    ventana.transient(ventana_padre)
    ventana.grab_set()
    ventana.grid_columnconfigure(0, weight=1)
    ventana.grid_rowconfigure(3, weight=1)

    funcionarios = [
        funcionario
        for funcionario in Novedades.obtener_funcionarios_activos()
        if funcionario["unidad"] == UNIDAD_SOBRES
    ] or Novedades.obtener_funcionarios_activos()
    nombres_funcionarios = ["Sin asignar comisión"] + [
        funcionario["nombre"] for funcionario in funcionarios
    ]

    estado = {
        "resultado": None,
        "ruta": ctk.StringVar(value=""),
        "anio": ctk.StringVar(value=""),
        "porcentaje": ctk.StringVar(
            value=str(leer_porcentaje_comision())
        ),
        "funcionario": ctk.StringVar(value=nombres_funcionarios[0]),
    }

    encabezado = ctk.CTkFrame(
        ventana, fg_color=COLOR_PANEL, corner_radius=0, border_width=0
    )
    encabezado.grid(row=0, column=0, sticky="ew")
    encabezado.grid_columnconfigure(0, weight=1)

    ctk.CTkLabel(
        encabezado,
        text="Importar sobres de venta PC",
        font=ctk.CTkFont(size=24, weight="bold"),
        text_color=COLOR_TEXTO,
        anchor="w",
    ).grid(row=0, column=0, sticky="ew", padx=28, pady=(22, 4))
    ctk.CTkLabel(
        encabezado,
        text=(
            "Primero se valida todo. Nada se guarda hasta que confirmes "
            "la importación."
        ),
        font=ctk.CTkFont(size=13),
        text_color=COLOR_TEXTO_SUAVE,
        anchor="w",
    ).grid(row=1, column=0, sticky="ew", padx=28, pady=(0, 20))

    seleccion = ctk.CTkFrame(
        ventana,
        fg_color=COLOR_PANEL_SECUNDARIO,
        corner_radius=14,
        border_width=1,
        border_color=COLOR_BORDE,
    )
    seleccion.grid(row=1, column=0, sticky="ew", padx=24, pady=(18, 8))
    seleccion.grid_columnconfigure(0, weight=1)

    ctk.CTkEntry(
        seleccion,
        textvariable=estado["ruta"],
        placeholder_text="Seleccioná el Excel de sobres...",
        height=40,
    ).grid(row=0, column=0, sticky="ew", padx=(16, 8), pady=16)

    def seleccionar():
        ruta = filedialog.askopenfilename(
            parent=ventana,
            title="Seleccionar Excel de sobres",
            filetypes=[("Excel", "*.xlsx")],
        )
        if ruta:
            estado["ruta"].set(ruta)
            estado["resultado"] = None
            boton_importar.configure(state="disabled")

    ctk.CTkButton(
        seleccion,
        text="Elegir archivo",
        command=seleccionar,
        height=40,
        width=140,
        corner_radius=9,
        fg_color=COLOR_PRIMARIO,
        hover_color=COLOR_PRIMARIO_HOVER,
        font=ctk.CTkFont(size=12, weight="bold"),
    ).grid(row=0, column=1, sticky="e", padx=(0, 16), pady=16)

    parametros = ctk.CTkFrame(
        ventana,
        fg_color=COLOR_PANEL_SECUNDARIO,
        corner_radius=14,
        border_width=1,
        border_color=COLOR_BORDE,
    )
    parametros.grid(row=2, column=0, sticky="ew", padx=24, pady=(0, 18))
    for columna in range(3):
        parametros.grid_columnconfigure(columna, weight=1)

    ctk.CTkLabel(
        parametros, text="Año", text_color=COLOR_TEXTO, anchor="w"
    ).grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 0))
    ctk.CTkEntry(
        parametros, textvariable=estado["anio"], height=38
    ).grid(row=1, column=0, sticky="ew", padx=16, pady=(4, 14))

    ctk.CTkLabel(
        parametros,
        text="Comisión de la encargada (%)",
        text_color=COLOR_TEXTO,
        anchor="w",
    ).grid(row=0, column=1, sticky="ew", padx=16, pady=(14, 0))
    ctk.CTkEntry(
        parametros, textvariable=estado["porcentaje"], height=38
    ).grid(row=1, column=1, sticky="ew", padx=16, pady=(4, 14))

    ctk.CTkLabel(
        parametros, text="Asignar a", text_color=COLOR_TEXTO, anchor="w"
    ).grid(row=0, column=2, sticky="ew", padx=16, pady=(14, 0))
    ctk.CTkOptionMenu(
        parametros,
        variable=estado["funcionario"],
        values=nombres_funcionarios,
        height=38,
    ).grid(row=1, column=2, sticky="ew", padx=16, pady=(4, 14))

    cuadro = ctk.CTkTextbox(
        ventana,
        font=ctk.CTkFont(family="Consolas", size=13),
        fg_color=COLOR_PANEL,
        border_width=1,
        border_color=COLOR_BORDE,
        text_color=COLOR_TEXTO,
        wrap="word",
    )
    cuadro.grid(row=3, column=0, sticky="nsew", padx=24, pady=(0, 12))
    cuadro.insert(
        "1.0",
        "Elegí el archivo, completá el año y presioná Analizar Excel.",
    )
    cuadro.configure(state="disabled")

    acciones = ctk.CTkFrame(ventana, fg_color="transparent")
    acciones.grid(row=4, column=0, sticky="ew", padx=24, pady=(0, 22))
    for columna in range(4):
        acciones.grid_columnconfigure(columna, weight=1)

    def mostrar_texto(texto):
        cuadro.configure(state="normal")
        cuadro.delete("1.0", "end")
        cuadro.insert("1.0", texto)
        cuadro.configure(state="disabled")

    def analizar():
        ruta = estado["ruta"].get().strip()
        anio_texto = estado["anio"].get().strip()
        if not ruta:
            messagebox.showwarning(
                "Falta el archivo",
                "Elegí primero el Excel de sobres.",
                parent=ventana,
            )
            return
        if not anio_texto.isdigit() or len(anio_texto) != 4:
            messagebox.showwarning(
                "Año inválido",
                "Ingresá el año con 4 dígitos, por ejemplo 2026.",
                parent=ventana,
            )
            return

        ventana.configure(cursor="watch")
        ventana.update_idletasks()
        resultado = analizar_excel_sobres(ruta, int(anio_texto))
        ventana.configure(cursor="")
        estado["resultado"] = resultado

        try:
            porcentaje = float(
                estado["porcentaje"].get().strip().replace(",", ".")
            )
        except ValueError:
            porcentaje = leer_porcentaje_comision()

        if not resultado["periodos"]:
            mostrar_texto(
                "No se encontraron sobres nuevos para importar.\n\n"
                + "\n".join(f"• {item}" for item in resultado["errores"])
            )
            boton_importar.configure(state="disabled")
            return

        mostrar_texto(texto_resumen(resultado, porcentaje))
        boton_importar.configure(
            state="normal" if not resultado["errores"] else "disabled"
        )

    def importar():
        resultado = estado["resultado"]
        if resultado is None or resultado["errores"]:
            return

        try:
            porcentaje = float(
                estado["porcentaje"].get().strip().replace(",", ".")
            )
        except ValueError:
            messagebox.showwarning(
                "Porcentaje inválido",
                "Ingresá un porcentaje de comisión válido.",
                parent=ventana,
            )
            return

        nombre_funcionario = estado["funcionario"].get()
        funcionario = next(
            (
                f
                for f in funcionarios
                if f["nombre"] == nombre_funcionario
            ),
            None,
        )

        total_nuevas = sum(
            len(datos["nuevas"])
            for datos in resultado["periodos"].values()
        )
        if not messagebox.askyesno(
            "Confirmar importación",
            (
                f"Sobres nuevos: {total_nuevas}\n"
                f"Comisión encargada: {porcentaje}%\n\n"
                "Antes de guardar se creará un respaldo automático.\n\n"
                "¿Continuar?"
            ),
            parent=ventana,
        ):
            return

        try:
            aplicar_resultado(resultado, porcentaje, funcionario)
        except Exception as exc:
            messagebox.showerror(
                "No se pudo importar", str(exc), parent=ventana
            )
            return

        boton_importar.configure(state="disabled")
        messagebox.showinfo(
            "Importación terminada",
            f"Se importaron {total_nuevas} sobres nuevos.",
            parent=ventana,
        )
        mostrar_texto(
            texto_resumen(resultado, porcentaje)
            + "\n\nIMPORTACIÓN COMPLETADA CORRECTAMENTE."
        )

    ctk.CTkButton(
        acciones,
        text="Analizar Excel",
        command=analizar,
        height=42,
        corner_radius=10,
        fg_color=COLOR_PRIMARIO,
        hover_color=COLOR_PRIMARIO_HOVER,
        font=ctk.CTkFont(size=13, weight="bold"),
    ).grid(row=0, column=0, sticky="ew", padx=(0, 8))

    boton_importar = ctk.CTkButton(
        acciones,
        text="Importar sobres",
        command=importar,
        height=42,
        corner_radius=10,
        fg_color=COLOR_VERDE,
        hover_color="#12835B",
        font=ctk.CTkFont(size=13, weight="bold"),
        state="disabled",
    )
    boton_importar.grid(row=0, column=1, sticky="ew", padx=(8, 8))

    ctk.CTkButton(
        acciones,
        text="Ver historial",
        command=lambda: abrir_historial_sobres(ventana),
        height=42,
        corner_radius=10,
        fg_color=COLOR_PRIMARIO,
        hover_color=COLOR_PRIMARIO_HOVER,
        font=ctk.CTkFont(size=13, weight="bold"),
    ).grid(row=0, column=2, sticky="ew", padx=(8, 8))

    ctk.CTkButton(
        acciones,
        text="Cerrar",
        command=ventana.destroy,
        height=42,
        corner_radius=10,
        fg_color=COLOR_PANEL_SECUNDARIO,
        hover_color=COLOR_BORDE,
        text_color=COLOR_TEXTO,
        font=ctk.CTkFont(size=13, weight="bold"),
    ).grid(row=0, column=3, sticky="ew", padx=(8, 0))



def obtener_comisiones_guardadas_periodo(periodo):
    """Obtiene las novedades de comisión generadas por sobres para un período."""
    comisiones = []
    texto_periodo = f"Comisión ventas PC {periodo}"

    for linea in leer_datos(Novedades.RUTA_NOVEDADES):
        datos = Novedades.separar_novedad(linea)
        if datos is None:
            continue
        if (
            datos["tipo"] == "Comisión"
            and texto_periodo.lower() in datos["motivo"].lower()
        ):
            comisiones.append(datos)

    return comisiones


def abrir_historial_sobres(ventana_padre):
    """Muestra los sobres guardados y la comisión generada por período."""
    import customtkinter as ctk
    from tkinter import messagebox, ttk

    from ImportadorExcel import (
        COLOR_BORDE,
        COLOR_PANEL,
        COLOR_PANEL_SECUNDARIO,
        COLOR_PRIMARIO,
        COLOR_PRIMARIO_HOVER,
        COLOR_TEXTO,
        COLOR_TEXTO_SUAVE,
    )

    ventana = ctk.CTkToplevel(ventana_padre)
    ventana.title("Historial de sobres PC")
    ventana.geometry("1080x720")
    ventana.minsize(850, 560)
    ventana.transient(ventana_padre)
    ventana.grid_columnconfigure(0, weight=1)
    ventana.grid_rowconfigure(3, weight=1)

    sobres_guardados = obtener_sobres_guardados()
    periodos = sorted(
        {sobre["periodo"] for sobre in sobres_guardados},
        key=lambda valor: (
            int(valor.split("-")[1]),
            int(valor.split("-")[0]),
        ),
        reverse=True,
    )

    encabezado = ctk.CTkFrame(
        ventana,
        fg_color=COLOR_PANEL,
        corner_radius=0,
    )
    encabezado.grid(row=0, column=0, sticky="ew")
    encabezado.grid_columnconfigure(0, weight=1)

    ctk.CTkLabel(
        encabezado,
        text="Historial de sobres PC",
        text_color=COLOR_TEXTO,
        font=ctk.CTkFont(size=23, weight="bold"),
        anchor="w",
    ).grid(row=0, column=0, sticky="ew", padx=24, pady=(20, 3))

    ctk.CTkLabel(
        encabezado,
        text="Detalle de sobres importados y comisión enviada a RRHH.",
        text_color=COLOR_TEXTO_SUAVE,
        font=ctk.CTkFont(size=13),
        anchor="w",
    ).grid(row=1, column=0, sticky="ew", padx=24, pady=(0, 18))

    filtros = ctk.CTkFrame(
        ventana,
        fg_color=COLOR_PANEL_SECUNDARIO,
        corner_radius=12,
        border_width=1,
        border_color=COLOR_BORDE,
    )
    filtros.grid(row=1, column=0, sticky="ew", padx=20, pady=(16, 8))

    ctk.CTkLabel(
        filtros,
        text="Período",
        text_color=COLOR_TEXTO,
    ).pack(side="left", padx=(16, 8), pady=14)

    filtro_periodo = ctk.CTkComboBox(
        filtros,
        values=["Todos", *periodos],
        width=150,
        state="readonly",
    )
    filtro_periodo.set(periodos[0] if periodos else "Todos")
    filtro_periodo.pack(side="left", pady=14)

    resumen = ctk.CTkLabel(
        ventana,
        text="",
        text_color=COLOR_TEXTO,
        fg_color=COLOR_PANEL_SECUNDARIO,
        corner_radius=10,
        anchor="w",
        justify="left",
        font=ctk.CTkFont(size=13, weight="bold"),
    )
    resumen.grid(row=2, column=0, sticky="ew", padx=20, pady=(0, 8))

    contenedor = ctk.CTkFrame(
        ventana,
        fg_color=COLOR_PANEL,
        corner_radius=12,
        border_width=1,
        border_color=COLOR_BORDE,
    )
    contenedor.grid(row=3, column=0, sticky="nsew", padx=20, pady=(0, 12))
    contenedor.grid_rowconfigure(0, weight=1)
    contenedor.grid_columnconfigure(0, weight=1)

    columnas = (
        "periodo",
        "numero",
        "monto",
        "comision_asociacion",
        "monto_final",
    )
    tabla = ttk.Treeview(
        contenedor,
        columns=columnas,
        show="headings",
        selectmode="browse",
    )
    encabezados = {
        "periodo": "Período",
        "numero": "N° sobre",
        "monto": "Monto",
        "comision_asociacion": "Comisión asociación",
        "monto_final": "Monto final PC",
    }
    anchos = {
        "periodo": 110,
        "numero": 100,
        "monto": 150,
        "comision_asociacion": 170,
        "monto_final": 160,
    }

    for columna in columnas:
        tabla.heading(columna, text=encabezados[columna])
        tabla.column(
            columna,
            width=anchos[columna],
            minwidth=90,
            anchor="center",
        )

    barra_y = ttk.Scrollbar(
        contenedor,
        orient="vertical",
        command=tabla.yview,
    )
    tabla.configure(yscrollcommand=barra_y.set)
    tabla.grid(row=0, column=0, sticky="nsew", padx=(10, 0), pady=10)
    barra_y.grid(row=0, column=1, sticky="ns", pady=10, padx=(0, 10))

    def actualizar():
        for item in tabla.get_children():
            tabla.delete(item)

        periodo_seleccionado = filtro_periodo.get()
        visibles = [
            sobre
            for sobre in sobres_guardados
            if (
                periodo_seleccionado == "Todos"
                or sobre["periodo"] == periodo_seleccionado
            )
        ]

        for indice, sobre in enumerate(visibles):
            tabla.insert(
                "",
                "end",
                iid=str(indice),
                values=(
                    sobre["periodo"],
                    sobre["numero"],
                    Movimientos.formatear_monto(sobre["monto"]),
                    Movimientos.formatear_monto(sobre["comision"]),
                    Movimientos.formatear_monto(sobre["monto_final"]),
                ),
            )

        total_monto = sum(sobre["monto"] for sobre in visibles)
        total_asociacion = sum(sobre["comision"] for sobre in visibles)
        total_final = sum(sobre["monto_final"] for sobre in visibles)

        comisiones_rrhh = []
        if periodo_seleccionado != "Todos":
            comisiones_rrhh = obtener_comisiones_guardadas_periodo(
                periodo_seleccionado
            )

        total_rrhh = sum(item["monto"] for item in comisiones_rrhh)
        nombres = ", ".join(
            sorted({item["nombre"] for item in comisiones_rrhh})
        ) or "Sin comisión asignada"

        resumen.configure(
            text=(
                f"  Sobres: {len(visibles)}    |    "
                f"Monto total: Gs. {Movimientos.formatear_monto(total_monto)}    |    "
                f"Comisión asociaciones: Gs. "
                f"{Movimientos.formatear_monto(total_asociacion)}\n"
                f"  Monto final PC: Gs. {Movimientos.formatear_monto(total_final)}    |    "
                f"Comisión RRHH: Gs. {Movimientos.formatear_monto(total_rrhh)}    |    "
                f"Asignada a: {nombres}"
            )
        )

    filtro_periodo.configure(command=lambda _valor: actualizar())

    acciones = ctk.CTkFrame(ventana, fg_color="transparent")
    acciones.grid(row=4, column=0, sticky="ew", padx=20, pady=(0, 16))

    ctk.CTkButton(
        acciones,
        text="Actualizar",
        width=110,
        command=actualizar,
        fg_color=COLOR_PRIMARIO,
        hover_color=COLOR_PRIMARIO_HOVER,
    ).pack(side="left")

    ctk.CTkButton(
        acciones,
        text="Cerrar",
        width=100,
        command=ventana.destroy,
        fg_color=COLOR_PANEL_SECUNDARIO,
        hover_color=COLOR_BORDE,
        text_color=COLOR_TEXTO,
    ).pack(side="right")

    if not sobres_guardados:
        messagebox.showinfo(
            "Historial de sobres",
            "Todavía no hay sobres guardados.",
            parent=ventana,
        )

    actualizar()

def menu_sobres_pc():
    while True:
        print()
        print("====================================")
        print("         SOBRES DE VENTA PC")
        print("====================================")
        print()
        print("1. Importar sobres desde Excel")
        print("2. Ver sobres importados")
        print("0. Volver")
        print()

        opcion = input("Seleccione una opción: ").strip()

        if opcion == "1":
            importar_sobres_pc()

        elif opcion == "2":
            sobres = obtener_sobres_guardados()
            print()
            print(f"Total de sobres registrados: {len(sobres)}")
            for sobre in sobres[-20:]:
                print(
                    f"  N° {sobre['numero']} | {sobre['periodo']} | "
                    f"Gs. {Movimientos.formatear_monto(sobre['monto'])} | "
                    f"Final Gs. "
                    f"{Movimientos.formatear_monto(sobre['monto_final'])}"
                )

        elif opcion == "0":
            return

        else:
            print()
            print("Opción inválida.")
