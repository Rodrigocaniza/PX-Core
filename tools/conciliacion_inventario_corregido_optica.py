"""Aplica la conciliación contra los inventarios corregidos de la Óptica.

Recalcula el plan desde los archivos y la base, y **antes de escribir** lo compara
contra el plan sellado en el HUMAN_GATE. Si no coincide exactamente, aborta: una
autorización se dio sobre un plan concreto, no sobre lo que el script decida hoy.

No hace `DELETE` de nada. Los artículos que salen del catálogo se compensan a
cero primero y recién después se desactivan, para que nunca quede stock operativo
en algo que ya no aparece en las búsquedas.

    python tools/conciliacion_inventario_corregido_optica.py --entrada <dir> [--base <ruta>] [--confirmar]
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
import sys
import unicodedata
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

RAIZ = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RAIZ))

from modulos.caja_diaria.config import resolve_data_paths  # noqa: E402
from modulos.comercial.application.comercial_controller import (  # noqa: E402
    build_comercial_controller,
)
from modulos.comercial.domain.models import (  # noqa: E402
    ArticleNature, Destination, StockMovement, StockMovementKind,
)

ACTOR = "COMMAND_CENTER/BC-OPTICA-CONCILIACION-INVENTARIO-CORREGIDO-V1-010"
MOTIVO = "ERROR_INVENTARIO"
CORTE = "2026-08-19"
PENDIENTE = "STOCK_INITIAL_PENDING_PHYSICAL_VERIFICATION"
NATURE_CORRECTION = "NATURE_CORRECTION"
RETIRO = "CORRECTED_SOURCE_ABSENCE"

#: Decisión humana: los cuatro son conceptos de Compostura, no inventario físico.
A_SERVICIO = {"2000056": "Par de patillas", "2000070": "Hilo",
              "2000071": "Tornillo", "2000072": "Plaqueta"}
#: Decisión humana: fuera de todo ajuste en esta misión.
INTOCABLES = {"000010", "000037"}

#: El plan aprobado en el HUMAN_GATE. Si el recálculo no da esto, no se escribe.
SELLADO = dict(altas_articulos=41, altas_registros=45, altas_unidades=1064,
               ausentes_reales=773, retirados=766, no_retirables=5,
               comp_asuncion=645, comp_pilar=128, ajustes=37, pos=47, neg=34,
               naturalezas=4, renumerados=1, nunca_existieron=1)

lineas: list[str] = []
fallas: list[str] = []


def registrar(texto: str = "") -> None:
    print(texto, flush=True)
    lineas.append(texto)


def comprobar(condicion: bool, descripcion: str) -> bool:
    registrar(f"  {'OK  ' if condicion else 'FALLA'} {descripcion}")
    if not condicion:
        fallas.append(descripcion)
    return bool(condicion)


def sha256(ruta: Path) -> str:
    return hashlib.sha256(ruta.read_bytes()).hexdigest()


SKU_RE = re.compile(r"^\s*(\d{4,})\s*(.*)$")
PREFIJO = {"ASUNCION": "ASU", "PILAR": "PIL"}
FUENTES = {"PC": "ASUNCION", "P2": "PILAR"}


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", (s or "").lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", " ", s).strip()


def es_global(sku: str) -> bool:
    if 11 <= len(sku) <= 13:
        return True
    if len(sku) == 6 and sku.startswith("00"):
        return True
    if len(sku) == 7 and sku.startswith("2000"):
        return True
    return False


def canonico(sku: str, sucursal: str) -> str:
    return sku if es_global(sku) else f"{PREFIJO[sucursal]}-{sku}"


def variantes(canon: str) -> set[str]:
    """El corregido rellena algunos códigos de barra con un cero adelante."""
    pre, sep, num = canon.rpartition("-")
    base = num if sep else canon
    salida = set()
    for cand in {base.lstrip("0"), "0" + base, base.zfill(13), base.zfill(12), base.zfill(11)}:
        if cand and cand != base:
            salida.add(pre + sep + cand if sep else cand)
    return salida


def leer_corregidos(entrada: Path) -> dict:
    from openpyxl import load_workbook

    filas = {}
    for tag, nombre in (("PC", "Inventario PC.xlsx"), ("P2", "Inventario P2.xlsx")):
        wb = load_workbook(entrada / nombre, data_only=True)
        ws = wb.active
        suc = FUENTES[tag]
        for n, r in enumerate(ws.iter_rows(values_only=True), start=1):
            if n <= 2:
                continue
            art = "" if r[0] is None else str(r[0]).strip()
            if not art:
                continue
            sku = ("" if r[1] is None else str(r[1]).strip()) or (
                SKU_RE.match(art).group(1) if SKU_RE.match(art) else "")
            if not sku:
                continue
            m = SKU_RE.match(art)
            filas[(suc, canonico(sku, suc))] = dict(
                sku=sku, nombre=(m.group(2).strip() if m else art),
                categoria="" if r[2] is None else str(r[2]).strip(),
                marca="" if r[3] is None else str(r[3]).strip(),
                precio="" if r[5] is None else str(r[5]).strip(),
                stock=int(float(r[6])) if r[6] not in (None, "") else None, fila=n)
        wb.close()
    return filas


def leer_anteriores(entrada: Path) -> dict:
    from openpyxl import load_workbook

    COLS = {"PC": dict(a=3, m=4, c=5, s=2), "P2": dict(a=0, c=1, m=2, s=3)}
    filas = {}
    for tag, nombre in (("PC", "PC - Inventario.xlsx"), ("P2", "P2 - Inventario.xlsx")):
        wb = load_workbook(entrada / nombre, data_only=True)
        ws = wb.active
        suc, cv = FUENTES[tag], COLS[tag]
        for n, r in enumerate(ws.iter_rows(values_only=True), start=1):
            if n <= 2:
                continue
            art = "" if r[cv["a"]] is None else str(r[cv["a"]]).strip()
            m = SKU_RE.match(art)
            if not m:
                continue
            stock = "" if r[cv["s"]] is None else str(r[cv["s"]]).strip()
            filas[(suc, canonico(m.group(1), suc))] = dict(
                sku=m.group(1), nombre=m.group(2).strip(),
                categoria="" if r[cv["c"]] is None else str(r[cv["c"]]).strip(),
                stock=int(stock) if stock.isdigit() else None, fila=n)
        wb.close()
    return filas


def radiografia(base: Path) -> dict:
    c = sqlite3.connect(f"file:{base}?mode=ro", uri=True)
    try:
        q = lambda s, *a: c.execute(s, a).fetchone()[0]  # noqa: E731
        return dict(
            articulos=q("SELECT COUNT(*) FROM articles"),
            activos=q("SELECT COUNT(*) FROM articles WHERE active=1"),
            movimientos=q("SELECT COUNT(*) FROM stock_movements"),
            asuncion=q("SELECT COALESCE(SUM(quantity),0) FROM stock_movements WHERE destination='ASUNCION'"),
            pilar=q("SELECT COALESCE(SUM(quantity),0) FROM stock_movements WHERE destination='PILAR'"),
            entradas=q("SELECT COUNT(*) FROM cash_entries"),
            suma_caja=q("SELECT COALESCE(SUM(total),0) FROM cash_entries"),
            sale_items=q("SELECT COUNT(*) FROM sale_items"),
            cash_days=q("SELECT COUNT(*) FROM cash_days"),
            orders=q("SELECT COUNT(*) FROM orders"),
            pendientes=q("SELECT COUNT(*) FROM admin_audit_log WHERE action=?", PENDIENTE),
            integridad=q("PRAGMA integrity_check"),
            fk=len(c.execute("PRAGMA foreign_key_check").fetchall()),
            negativos=q("SELECT COUNT(*) FROM stock_actual WHERE quantity<0"),
            huerfanos=q("SELECT COUNT(*) FROM stock_movements sm LEFT JOIN articles a"
                        " ON a.id=sm.article_id WHERE a.id IS NULL"),
            efectos=q("SELECT COUNT(*) FROM event_effects ee LEFT JOIN domain_events de"
                      " ON de.event_id=ee.event_id WHERE de.event_id IS NULL"),
            duplicados=q("SELECT COUNT(*) FROM (SELECT sku FROM articles GROUP BY sku"
                         " HAVING COUNT(*)>1)"),
            stock_en_retirados=q("SELECT COUNT(*) FROM stock_actual sa JOIN articles a"
                                 " ON a.id=sa.article_id WHERE a.active=0 AND sa.quantity<>0"),
        )
    finally:
        c.close()


def calcular_plan(base: Path, entrada: Path) -> dict:
    COR, ANT = leer_corregidos(entrada), leer_anteriores(entrada)
    c = sqlite3.connect(f"file:{base}?mode=ro", uri=True)
    c.row_factory = sqlite3.Row
    CAT = {f["sku"]: dict(id=f["id"], nature=f["nature"]) for f in
           c.execute("SELECT sku, id, nature FROM articles")}
    LEDGER = {(f["destination"], f["sku"]): f["q"] for f in c.execute(
        "SELECT a.sku, sm.destination, SUM(sm.quantity) q FROM stock_movements sm"
        " JOIN articles a ON a.id=sm.article_id GROUP BY a.sku, sm.destination")}
    c.close()

    aus_brutos = {k for k in ANT if k not in COR}
    nue_brutos = {k for k in COR if k not in ANT}
    renumerados = {}
    for k in sorted(aus_brutos):
        for v in variantes(k[1]):
            if (k[0], v) in nue_brutos:
                renumerados[k] = v
                break
    nunca = {k for k in aus_brutos if k[1] not in CAT}
    ausentes = sorted(aus_brutos - set(renumerados) - nunca)
    nuevos = sorted(k for k in nue_brutos
                    if k[1] not in CAT and k[1] not in set(renumerados.values()))
    ajustes = [(k, LEDGER[k], COR[k]["stock"]) for k in sorted(COR)
               if k in ANT and k in LEDGER and COR[k]["stock"] != LEDGER[k]
               and COR[k]["sku"] not in INTOCABLES]
    return dict(COR=COR, ANT=ANT, CAT=CAT, LEDGER=LEDGER, ausentes=ausentes,
                nuevos=nuevos, ajustes=ajustes, renumerados=renumerados, nunca=nunca)


def main() -> int:  # noqa: C901 - es un procedimiento, se lee de arriba abajo
    parser = argparse.ArgumentParser()
    parser.add_argument("--entrada", required=True, help="dir con los cuatro xlsx")
    parser.add_argument("--base", default=None)
    parser.add_argument("--confirmar", action="store_true")
    args = parser.parse_args()

    entrada = Path(args.entrada)
    base = Path(args.base) if args.base else Path(resolve_data_paths().database)
    antes = radiografia(base)
    sha_antes = sha256(base)

    registrar("CONCILIACION V1-010 -- APLICACION")
    registrar(f"base   : {base}")
    registrar(f"sha256 : {sha_antes}")
    registrar(f"antes  : {antes['articulos']} articulos ({antes['activos']} activos), "
              f"{antes['movimientos']} movimientos, ASU {antes['asuncion']} / PIL {antes['pilar']}")
    registrar()

    plan = calcular_plan(base, entrada)
    COR, CAT, LEDGER = plan["COR"], plan["CAT"], plan["LEDGER"]
    por_canonico: dict[str, list] = {}
    for suc, canon in plan["nuevos"]:
        por_canonico.setdefault(canon, []).append((suc, COR[(suc, canon)]))
    unidades_altas = sum(COR[(s, c)]["stock"] or 0 for s, c in plan["nuevos"]
                         if not COR[(s, c)]["categoria"].lower().startswith("compostura"))

    c = sqlite3.connect(f"file:{base}?mode=ro", uri=True)
    con_stock, sin_stock, sigue_vivo = [], [], []
    ausentes_set = set(plan["ausentes"])
    for suc, canon in plan["ausentes"]:
        art_id = CAT[canon]["id"]
        q = c.execute("SELECT COALESCE(SUM(quantity),0) FROM stock_movements"
                      " WHERE article_id=? AND destination=?", (art_id, suc)).fetchone()[0]
        ventas = c.execute("SELECT COUNT(*) FROM sale_items WHERE article_id=? OR lens_article_id=?",
                           (art_id, art_id)).fetchone()[0]
        otras = dict(c.execute("SELECT destination, SUM(quantity) FROM stock_movements"
                               " WHERE article_id=? AND destination<>? GROUP BY destination",
                               (art_id, suc)))
        if any(v > 0 and (d, canon) not in ausentes_set for d, v in otras.items()) or ventas:
            sigue_vivo.append(canon)
        (con_stock if q > 0 else sin_stock).append((suc, canon, q))
    c.close()
    comp = Counter()
    for suc, canon, q in con_stock:
        comp[suc] += q
    retirables = {canon for _, canon, _ in con_stock + sin_stock if canon not in sigue_vivo}
    pos = sum(cor - act for _, act, cor in plan["ajustes"] if cor > act)
    neg = sum(act - cor for _, act, cor in plan["ajustes"] if cor < act)

    registrar("== el plan recalculado contra el sellado en el HUMAN_GATE ==")
    real = dict(altas_articulos=len(por_canonico), altas_registros=len(plan["nuevos"]),
                altas_unidades=unidades_altas, ausentes_reales=len(plan["ausentes"]),
                retirados=len(retirables), no_retirables=len(set(sigue_vivo)),
                comp_asuncion=comp["ASUNCION"], comp_pilar=comp["PILAR"],
                ajustes=len(plan["ajustes"]), pos=pos, neg=neg,
                naturalezas=len(A_SERVICIO), renumerados=len(plan["renumerados"]),
                nunca_existieron=len(plan["nunca"]))
    for clave, esperado in SELLADO.items():
        comprobar(real[clave] == esperado, f"{clave:18} {real[clave]} (sellado {esperado})")
    if fallas:
        registrar()
        registrar("EL PLAN NO COINCIDE CON EL AUTORIZADO. No se escribe nada.")
        registrar("La autorizacion se dio sobre un plan concreto: adaptarlo solo seria "
                  "ejecutar algo que nadie aprobo.")
        return 1
    registrar()

    if not args.confirmar:
        registrar("NO SE ESCRIBIO NADA: falta --confirmar.")
        return 0

    registrar("== backup verificable ==")
    marca = datetime.now().strftime("%Y%m%d-%H%M%S")
    respaldo = base.parent / "Backups" / f"bc-caja-preconciliacion-{marca}.sqlite3"
    respaldo.parent.mkdir(parents=True, exist_ok=True)
    fuente = sqlite3.connect(f"file:{base}?mode=ro", uri=True)
    try:
        destino = sqlite3.connect(str(respaldo))
        try:
            fuente.backup(destino)
        finally:
            destino.close()
    finally:
        fuente.close()
    registrar(f"  archivo: {respaldo}")
    registrar(f"  sha256 : {sha256(respaldo)}")
    registrar(f"  bytes  : {respaldo.stat().st_size}")
    espejo = radiografia(respaldo)
    comprobar(espejo == antes, "el backup tiene exactamente el mismo contenido que la base")
    comprobar(espejo["integridad"] == "ok" and espejo["fk"] == 0,
              "el backup abre, pasa integrity_check y no tiene violaciones de FK")
    if fallas:
        registrar("El backup no quedo bien. NO se escribe nada.")
        return 1
    registrar()

    cuando = datetime.now(timezone.utc).replace(microsecond=0)
    ctrl = build_comercial_controller(base)
    try:
        registrar("== 1. altas ==")
        creados = {}
        for canon, apar in sorted(por_canonico.items()):
            v = max((x[1] for x in apar), key=lambda y: len(y["nombre"]))
            servicio = v["categoria"].lower().startswith("compostura")
            nat = (ArticleNature.SERVICIO_NO_STOCKEABLE if servicio
                   else ArticleNature.PRODUCTO_STOCKEABLE)
            origen = "; ".join(f"{s}=Inventario {'PC' if s == 'ASUNCION' else 'P2'}.xls "
                               f"fila {x['fila']}" for s, x in apar)
            art = ctrl.guardar_articulo(
                sku=canon, name=v["nombre"], nature=nat, actor=ACTOR,
                notes=f"alta por conciliacion V1-010 @{CORTE}; origen {origen}",
                sale_price=int(v["precio"]) if v["precio"].isdigit() else None)
            for suc, x in apar:
                creados[(suc, canon)] = (art.id, nat, x)
        registrar(f"  {len(por_canonico)} articulos de {len(plan['nuevos'])} registros")
        hecho = ctrl.cargar_stock_inicial(
            [(v[0], k[0], v[2]["stock"]) for k, v in creados.items()
             if v[1] is ArticleNature.PRODUCTO_STOCKEABLE and v[2]["stock"]],
            actor=ACTOR, origen=f"Alta por conciliacion contra el inventario corregido del {CORTE}",
            run_id=f"conciliacion-altas-{CORTE.replace('-', '')}", momento=cuando)
        registrar(f"  stock inicial: {hecho.rows_imported} lineas, {unidades_altas} unidades")
        registrar()

        registrar("== 2. compensacion del stock de los que se retiran ==")
        for suc, canon, q in con_stock:
            ctrl.ledger.registrar(StockMovement(
                article_id=CAT[canon]["id"], destination=Destination(suc),
                kind=StockMovementKind.AJUSTE_NEGATIVO, quantity=q, actor=ACTOR,
                occurred_at=cuando, reason_code=MOTIVO,
                note=(f"RECONCILIACION_INVENTARIO_CORREGIDO / ARTICULO_RETIRADO. No aparece "
                      f"en Inventario {'PC' if suc == 'ASUNCION' else 'P2'}.xls del {CORTE}. "
                      f"Stock que dejaba V1-008: {q}. Se lleva a cero antes de retirarlo. "
                      f"El movimiento original de V1-008 no se toca"),
                document_kind="CONCILIACION_INVENTARIO",
                document_id=f"retiro-{CORTE.replace('-', '')}",
                idempotency_key=f"RETIRO:{CORTE}:{canon}:{suc}"))
        registrar(f"  ASUNCION {comp['ASUNCION']} unidades, PILAR {comp['PILAR']} unidades")
        registrar()

        registrar("== 3. verificar stock cero antes de retirar ==")
        c = sqlite3.connect(f"file:{base}?mode=ro", uri=True)
        pendiente_de_cero = [canon for canon in retirables
                             if c.execute("SELECT COALESCE(SUM(quantity),0) FROM stock_movements"
                                          " WHERE article_id=?",
                                          (CAT[canon]["id"],)).fetchone()[0] != 0]
        c.close()
        comprobar(not pendiente_de_cero,
                  f"los {len(retirables)} a retirar quedaron en cero "
                  f"({len(pendiente_de_cero)} todavia con stock)")
        if fallas:
            registrar("NO se retira nada: hay stock sin compensar.")
            return 1
        registrar()

        registrar("== 4. retiro del catalogo activo ==")
        for canon in sorted(retirables):
            ctrl.desactivar_articulo(
                CAT[canon]["id"], actor=ACTOR,
                motivo=(f"{RETIRO} -> RETIRE_FROM_ACTIVE_CATALOG. No aparece en el inventario "
                        f"corregido del {CORTE}. Su stock se compenso a cero antes de "
                        f"retirarlo. Se conserva toda la historia"))
        registrar(f"  {len(retirables)} articulos retirados; {len(set(sigue_vivo))} no se "
                  f"retiran porque siguen vivos en la otra sucursal")
        registrar()

        registrar("== 5. correcciones de naturaleza ==")
        # El cierre del pendiente se escribe despues de soltar el controlador: dos
        # conexiones de escritura sobre el mismo SQLite se pisan.
        cierres = []
        for sku, nombre in sorted(A_SERVICIO.items()):
            a = ctrl.articulo_por_sku(sku)
            previa = a.nature.value
            ctrl.guardar_articulo(
                article_id=a.id, sku=a.sku, name=a.name,
                nature=ArticleNature.SERVICIO_NO_STOCKEABLE, actor=ACTOR,
                notes=(f"{a.notes} || {NATURE_CORRECTION} V1-010: {previa} -> "
                       f"SERVICIO_NO_STOCKEABLE por definicion operativa. Es un concepto "
                       f"de compostura, no inventario fisico. Las cifras de la fuente "
                       f"quedan solo como evidencia historica: eran centinelas, no "
                       f"conteos. No requiere conteo fisico"))
            cierres.append((a.id, dict(sku=sku, nombre=nombre, de=previa,
                                       a="SERVICIO_NO_STOCKEABLE", cierra=PENDIENTE,
                                       motivo="concepto de Compostura, no inventario fisico",
                                       requiere_conteo=False)))
        for sku in sorted(A_SERVICIO):
            comprobar(ctrl.articulo_por_sku(sku).nature.value == "SERVICIO_NO_STOCKEABLE",
                      f"{sku} {A_SERVICIO[sku]}: SERVICIO_NO_STOCKEABLE")
        registrar()

        registrar("== 6. ajustes del inventario corregido ==")
        for (suc, canon), actual, corregido in plan["ajustes"]:
            delta = corregido - actual
            kind = (StockMovementKind.AJUSTE_POSITIVO if delta > 0
                    else StockMovementKind.AJUSTE_NEGATIVO)
            ctrl.ledger.registrar(StockMovement(
                article_id=CAT[canon]["id"], destination=Destination(suc), kind=kind,
                quantity=abs(delta), actor=ACTOR, occurred_at=cuando, reason_code=MOTIVO,
                note=(f"RECONCILIACION_INVENTARIO_CORREGIDO. Inventario "
                      f"{'PC' if suc == 'ASUNCION' else 'P2'}.xls del {CORTE} "
                      f"fila {COR[(suc, canon)]['fila']}. Stock del sistema: {actual}. "
                      f"Corregido: {corregido}. Delta: {delta:+}"),
                document_kind="CONCILIACION_INVENTARIO",
                document_id=f"conciliacion-{CORTE.replace('-', '')}",
                idempotency_key=f"CONCILIACION:{CORTE}:{canon}:{suc}"))
        registrar(f"  AJUSTE_POSITIVO {pos} unidades, AJUSTE_NEGATIVO {neg} unidades")
    finally:
        ctrl.close()

    from uuid import uuid4
    conexion = sqlite3.connect(str(base))
    try:
        for article_id, detalle in cierres:
            conexion.execute(
                "INSERT INTO admin_audit_log(id, actor, action, target_type, target_id,"
                " result, details_json, recorded_at) VALUES (?,?,?,?,?,?,?,?)",
                (str(uuid4()), ACTOR, NATURE_CORRECTION, "article", article_id, "CONFIRMADO",
                 json.dumps(detalle, ensure_ascii=False, sort_keys=True), cuando.isoformat()))
        conexion.commit()
    finally:
        conexion.close()
    registrar(f"  {len(cierres)} pendientes cerrados por {NATURE_CORRECTION}")

    registrar()
    registrar("== verificacion post-produccion ==")
    despues = radiografia(base)
    esperado = dict(articulos=3595, activos=2829, movimientos=4438, asuncion=6276, pilar=2776)
    for clave, valor in esperado.items():
        comprobar(despues[clave] == valor, f"{clave:12} {despues[clave]} (esperado {valor})")
    comprobar(despues["asuncion"] + despues["pilar"] == 9052,
              f"stock total {despues['asuncion'] + despues['pilar']} (esperado 9052)")
    comprobar(despues["integridad"] == "ok", f"integrity_check: {despues['integridad']}")
    comprobar(despues["fk"] == 0, f"foreign_key_check: {despues['fk']}")
    comprobar(despues["negativos"] == 0, f"stock negativo: {despues['negativos']}")
    comprobar(despues["huerfanos"] == 0, f"huerfanos: {despues['huerfanos']}")
    comprobar(despues["efectos"] == 0, f"efectos sin hecho: {despues['efectos']}")
    comprobar(despues["duplicados"] == 0, f"SKU duplicados: {despues['duplicados']}")
    comprobar(despues["stock_en_retirados"] == 0,
              f"stock operativo en articulos retirados: {despues['stock_en_retirados']}")
    for clave in ("entradas", "suma_caja", "sale_items", "cash_days", "orders"):
        comprobar(despues[clave] == antes[clave], f"Caja, {clave}: {antes[clave]}")
    comprobar(despues["pendientes"] == antes["pendientes"],
              f"los {antes['pendientes']} pendientes de V1-008 siguen registrados")

    c = sqlite3.connect(f"file:{base}?mode=ro", uri=True)
    q = lambda s, *a: c.execute(s, a).fetchone()[0]  # noqa: E731
    ids = ("inventario-inicial-asuncion-20260803", "inventario-inicial-pilar-20260810")
    comprobar(q("SELECT COUNT(*) FROM stock_movements WHERE document_id IN (?,?)", *ids) == 3583
              and q("SELECT COALESCE(SUM(quantity),0) FROM stock_movements"
                    " WHERE document_id IN (?,?)", *ids) == 8748,
              "V1-008 intacta: 3.583 movimientos y 8.748 unidades")
    comprobar(q("SELECT COALESCE(SUM(sm.quantity),0) FROM stock_movements sm JOIN articles a"
                " ON a.id=sm.article_id WHERE a.sku='000010' AND sm.destination='ASUNCION'") == 0,
              "000010 ASUNCION sigue sin unidades")
    comprobar(q("SELECT COALESCE(SUM(sm.quantity),0) FROM stock_movements sm JOIN articles a"
                " ON a.id=sm.article_id WHERE a.sku='000010' AND sm.destination='PILAR'") == 10,
              "000010 PILAR intacto en 10")
    comprobar(dict(c.execute("SELECT sm.destination, SUM(sm.quantity) FROM stock_movements sm"
                             " JOIN articles a ON a.id=sm.article_id WHERE a.sku='000037'"
                             " GROUP BY sm.destination")) == {"ASUNCION": 210, "PILAR": 516}
              and q("SELECT active FROM articles WHERE sku='000037'") == 1,
              "000037 sin stock nuevo y sigue activo: su retiro es del slice 013")
    comprobar(q("SELECT COUNT(*) FROM articles a LEFT JOIN article_categories ac"
                " ON ac.id=a.category_id WHERE a.nature='PRODUCTO_STOCKEABLE'"
                " AND ac.name LIKE 'Compostura%'") == 0,
              "ningun articulo de Compostura queda clasificado como producto")
    c.close()

    registrar()
    registrar(f"sha256 base despues: {sha256(base)}")
    registrar(f"VEREDICTO: {'PASS' if not fallas else 'FALLA'} ({len(fallas)} fallas)")
    for falla in fallas:
        registrar(f"  - {falla}")
    if fallas:
        registrar()
        registrar("LA CONCILIACION NO QUEDO LIMPIA. Para volver atras:")
        registrar("  1. cerrar BC Caja")
        registrar(f"  2. borrar {base.name}, {base.name}-wal y {base.name}-shm")
        registrar(f"  3. copiar {respaldo} sobre {base}")
    (entrada / "APLICACION_PRODUCTIVA.txt").write_text("\n".join(lineas) + "\n", encoding="utf-8")
    return 0 if not fallas else 1


if __name__ == "__main__":
    raise SystemExit(main())
