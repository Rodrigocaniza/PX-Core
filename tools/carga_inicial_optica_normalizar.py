"""Normaliza los dos inventarios reales de la Óptica a la plantilla del sistema.

Lee los XLSX tal como salen del sistema viejo y emite tres archivos que el
mecanismo de carga ya sabe consumir: el catálogo canónico en el formato de
`docs/PLANTILLA_ARTICULOS.csv`, y un recuento por sucursal.

**No escribe en ninguna base.** Lo que decide, y por qué, está documentado en
`artifacts/BC-OPTICA-CARGA-INICIAL-CATALOGO-V1-008/`.

    python tools/carga_inicial_optica_normalizar.py --pc <xlsx> --p2 <xlsx> --salida <dir>
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

#: Qué sucursal es cada archivo. No es una convención de este script: es el
#: vínculo que `cash_register_branches` ya trae sembrado desde las migraciones
#: 018 y 020, y que los pedidos reales corroboran (`orders.branch = 'PC'`).
FUENTES = {
    "PC": dict(hoja="PC", corte="2026-08-03", sucursal="ASUNCION",
               columnas=dict(articulo=3, marca=4, categoria=5, stock=2)),
    "P2": dict(hoja="Sheet1", corte="2026-08-10", sucursal="PILAR",
               columnas=dict(articulo=0, categoria=1, marca=2, stock=3)),
}

REQUIERE_DECISION = "REQUIRES_POLICY_DECISION"

#: La naturaleza sale de la categoría. **Nunca** de la descripción de la fila:
#: el día que alguien escriba «armazón de cristal», deducirla del texto pondría
#: a un cristal a descontar stock y nadie sabría por qué.
NATURE_POR_CATEGORIA = {
    "armazones": "PRODUCTO_STOCKEABLE",
    "lentes de sol": "PRODUCTO_STOCKEABLE",
    "lente de contacto": "PRODUCTO_STOCKEABLE",
    "sujetadores": "PRODUCTO_STOCKEABLE",
    "accesorios": "PRODUCTO_STOCKEABLE",
    "liquidos multiproposit": "PRODUCTO_STOCKEABLE",
    "estuches": "PRODUCTO_STOCKEABLE",
    "estuche lc": "PRODUCTO_STOCKEABLE",
    "limpia cristales": "PRODUCTO_STOCKEABLE",
    "marcadores": "PRODUCTO_STOCKEABLE",
    "medicamentos": "PRODUCTO_STOCKEABLE",
    "organizadores": "PRODUCTO_STOCKEABLE",
    "panos": "PRODUCTO_STOCKEABLE",
    # La «marca» de un cristal es el laboratorio que lo fabrica, y el stock
    # declarado (949, 984, 999, 1971) no es un conteo sino el número que pone un
    # sistema para que algo no se acabe nunca. Un cristal se pide con la receta.
    "cristales": "TRABAJO_BAJO_PEDIDO",
    # Mezcla servicios, tipos de cristal y repuestos físicos en una sola
    # categoría. Cualquier asignación única sería falsa: se resuelve fila por
    # fila en DECISIONES_POR_SKU, con decisión humana o con evidencia.
    "compostura": REQUIERE_DECISION,
    # Filas de PC sin categoría. Parecen armazones por la descripción, que es
    # exactamente el motivo por el que la categoría no decide nada acá.
    "": REQUIERE_DECISION,
}

#: Excepciones resueltas una por una, cada una con de dónde salió. Existen
#: porque la categoría del archivo no alcanzaba: o no había, o era un cajón de
#: sastre. No son un atajo alrededor de la regla — son la regla admitiendo que
#: en estas filas la categoría no informa, y dejando por escrito qué sí informó.
DECISIONES_POR_SKU: dict[tuple[str, str], tuple[str, str, str]] = {
    # -- Compostura: los nueve servicios. Decisión humana, 19/08/2026.
    ("PILAR", "2000101"): ("SERVICIO_NO_STOCKEABLE", "Compostura", "decisión humana"),
    ("PILAR", "2000054"): ("SERVICIO_NO_STOCKEABLE", "Compostura", "decisión humana"),
    ("PILAR", "2000055"): ("SERVICIO_NO_STOCKEABLE", "Compostura", "decisión humana"),
    ("PILAR", "2000057"): ("SERVICIO_NO_STOCKEABLE", "Compostura", "decisión humana"),
    ("PILAR", "2000058"): ("SERVICIO_NO_STOCKEABLE", "Compostura", "decisión humana"),
    ("PILAR", "2000065"): ("SERVICIO_NO_STOCKEABLE", "Compostura", "decisión humana"),
    ("PILAR", "2000149"): ("SERVICIO_NO_STOCKEABLE", "Compostura", "decisión humana"),
    ("PILAR", "2000148"): ("SERVICIO_NO_STOCKEABLE", "Compostura", "decisión humana"),
    ("PILAR", "2000147"): ("SERVICIO_NO_STOCKEABLE", "Compostura", "decisión humana"),
    # -- Compostura: los siete tipos de cristal. Decisión humana, 19/08/2026.
    ("PILAR", "2000060"): ("TRABAJO_BAJO_PEDIDO", "Cristales", "decisión humana"),
    ("PILAR", "2000061"): ("TRABAJO_BAJO_PEDIDO", "Cristales", "decisión humana"),
    ("PILAR", "2000062"): ("TRABAJO_BAJO_PEDIDO", "Cristales", "decisión humana"),
    ("PILAR", "2000063"): ("TRABAJO_BAJO_PEDIDO", "Cristales", "decisión humana"),
    ("PILAR", "2000064"): ("TRABAJO_BAJO_PEDIDO", "Cristales", "decisión humana"),
    ("PILAR", "2000066"): ("TRABAJO_BAJO_PEDIDO", "Cristales", "decisión humana"),
    ("PILAR", "2000067"): ("TRABAJO_BAJO_PEDIDO", "Cristales", "decisión humana"),
    # -- Compostura: los tres repuestos físicos. Decisión humana, 19/08/2026.
    #    La naturaleza es de la decisión; la cantidad NO, y por eso quedan en
    #    CANTIDAD_EN_SUSPENSO: 99.981 hilos y 99.425 tornillos no son un conteo.
    ("PILAR", "2000070"): ("PRODUCTO_STOCKEABLE", "Sujetadores", "decisión humana"),
    ("PILAR", "2000071"): ("PRODUCTO_STOCKEABLE", "Sujetadores", "decisión humana"),
    ("PILAR", "2000072"): ("PRODUCTO_STOCKEABLE", "Sujetadores", "decisión humana"),
    # -- Mostacillas: resuelto por evidencia, no por decisión humana.
    #    Los otros tres artículos llamados «Mostacilla*» del universo están en
    #    Sujetadores o Accesorios, y el más cercano (000012 Mostacilla) comparte
    #    la marca Proray y está en Sujetadores en los dos archivos.
    ("PILAR", "000005"): ("PRODUCTO_STOCKEABLE", "Sujetadores",
                          "evidencia: 000012 Mostacilla, misma marca Proray, en Sujetadores en PC y P2"),
    # -- Las filas sin categoría de PC. Las 3.065 filas del universo cuya
    #    descripción empieza con AC PAT / AC APT / AC PAC son Armazones (2.773)
    #    o Lentes de Sol (289): no hay una tercera posibilidad, y las dos son
    #    PRODUCTO_STOCKEABLE. La naturaleza no depende de cuál sea.
    ("ASUNCION", "100093"): ("PRODUCTO_STOCKEABLE", "Armazones",
                             "evidencia: marca Steffani, en Armazones en las 585 filas donde aparece"),
    ("ASUNCION", "100240"): ("PRODUCTO_STOCKEABLE", "Armazones",
                             "evidencia: marca Betania, en Armazones en las 357 filas donde aparece"),
    #    Sin marca: la naturaleza está determinada, la categoría no. Se deja
    #    vacía en vez de inventarla — un dato ausente se completa después; uno
    #    inventado no se detecta nunca.
    ("ASUNCION", "101181"): ("PRODUCTO_STOCKEABLE", "",
                             "evidencia: prefijo AC PAT, stockeable en las 3.065 filas del universo"),
    ("ASUNCION", "108004"): ("PRODUCTO_STOCKEABLE", "",
                             "evidencia: prefijo AC PAT, stockeable en las 3.065 filas del universo"),
}

#: El artículo entra al catálogo; sus unidades NO. Un número que no es un conteo
#: no puede convertirse en stock: quedaría un depósito que nadie contó y que
#: sólo se corrige compensando.
CANTIDAD_EN_SUSPENSO: dict[tuple[str, str], str] = {
    ("PILAR", "2000070"): "99.981 declarados: valor centinela, no un conteo. Falta recuento real",
    ("PILAR", "2000071"): "99.425 declarados: valor centinela, no un conteo. Falta recuento real",
    ("PILAR", "2000072"): "9.393 declarados: valor centinela, no un conteo. Falta recuento real",
    ("ASUNCION", "000010"): "2.860 declarados: pendiente de confirmación humana",
}

#: Ni la categoría ni la evidencia alcanzan. No entra al catálogo hasta que se
#: decida: puede ser el repuesto físico o el servicio de cambiarlo, y las dos
#: filas vecinas ya cubren la mano de obra.
SIN_RESOLVER: set[tuple[str, str]] = {
    ("PILAR", "2000056"),  # Par de patillas
}

MUEVEN_STOCK = {"PRODUCTO_STOCKEABLE", "PRODUCCION_INTERNA"}
PREFIJO_SUCURSAL = {"ASUNCION": "ASU", "PILAR": "PIL"}

#: El código va pegado o separado de la descripción: «000010 Limpia Cristal» y
#: «107648AC PAT FLEX NEGRO» son la misma forma escrita de dos maneras.
SKU_RE = re.compile(r"^\s*(\d{4,})\s*(.*)$")


def _texto(valor) -> str:
    return "" if valor is None else str(valor).strip()


def _clave_categoria(nombre: str) -> str:
    limpio = unicodedata.normalize("NFKD", (nombre or "").strip().lower())
    limpio = "".join(c for c in limpio if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", limpio)[:22]


def identificador_es_global(sku: str) -> bool:
    """¿El código nombra al mismo artículo en las dos sucursales?

    Se resolvió con los datos, no por preferencia. De los 42 códigos de barras
    compartidos, los 42 son el mismo producto. De los 31 del catálogo interno,
    29 tienen la descripción idéntica. Pero de los 107 códigos de armazón
    compartidos **uno solo** describe el mismo marco: cada sucursal numera sus
    armazones por su cuenta, y tratarlos como uno pegaría dos marcos distintos
    en un artículo para después sumarles stock a algo que no existe.
    """
    if 11 <= len(sku) <= 13:          # código de barras del fabricante
        return True
    if len(sku) == 6 and sku.startswith("00"):    # catálogo interno compartido
        return True
    if len(sku) == 7 and sku.startswith("2000"):  # ídem
        return True
    return False


def leer_hoja(ruta: Path, meta: dict) -> list[dict]:
    from openpyxl import load_workbook

    libro = load_workbook(ruta, read_only=True, data_only=True)
    try:
        hoja = libro[meta["hoja"]] if meta["hoja"] in libro.sheetnames else libro.active
        columnas = meta["columnas"]
        filas = []
        for numero, cruda in enumerate(hoja.iter_rows(values_only=True), start=1):
            if numero <= 2:  # fila 1: la fecha del corte. fila 2: el encabezado
                continue
            articulo = _texto(cruda[columnas["articulo"]])
            if not articulo:
                continue  # fila vacía de la exportación, no un artículo
            filas.append(dict(
                fila_fisica=numero, articulo=articulo,
                marca=_texto(cruda[columnas["marca"]]),
                categoria=_texto(cruda[columnas["categoria"]]),
                stock=_texto(cruda[columnas["stock"]])))
        return filas
    finally:
        libro.close()


def normalizar(archivos: dict[str, Path]) -> tuple[list[dict], list[dict]]:
    registros, rechazos = [], []
    for tag, ruta in archivos.items():
        meta = FUENTES[tag]
        for fila in leer_hoja(ruta, meta):
            encontrado = SKU_RE.match(fila["articulo"])
            if not encontrado:
                rechazos.append(dict(fuente=tag, fila=fila["fila_fisica"],
                                     articulo=fila["articulo"],
                                     motivo="sin código: no se puede derivar un SKU"))
                continue
            sku, nombre = encontrado.group(1), encontrado.group(2).strip()
            if not nombre:
                rechazos.append(dict(fuente=tag, fila=fila["fila_fisica"],
                                     articulo=fila["articulo"],
                                     motivo="código sin descripción: no hay nombre de artículo"))
                continue
            try:
                cantidad = int(fila["stock"])
            except ValueError:
                cantidad = None
            clave = (meta["sucursal"], sku)
            if clave in SIN_RESOLVER:
                rechazos.append(dict(fuente=tag, fila=fila["fila_fisica"],
                                     articulo=fila["articulo"],
                                     motivo="sin resolver: espera decisión humana"))
                continue
            categoria, motivo = fila["categoria"], ""
            if clave in DECISIONES_POR_SKU:
                nature, categoria, motivo = DECISIONES_POR_SKU[clave]
            else:
                nature = NATURE_POR_CATEGORIA.get(_clave_categoria(fila["categoria"]),
                                                  REQUIERE_DECISION)
            registros.append(dict(
                sku=sku, nombre=nombre, categoria=categoria, marca=fila["marca"],
                sucursal=meta["sucursal"], corte=meta["corte"], nature=nature,
                decision=motivo, categoria_original=fila["categoria"],
                stock_en_suspenso=CANTIDAD_EN_SUSPENSO.get(clave, ""),
                stock_inicial=cantidad, fuente_archivo=ruta.name,
                fuente_fila=fila["fila_fisica"], texto_original=fila["articulo"]))
    return registros, rechazos


def consolidar(registros: list[dict]) -> tuple[list[dict], dict, list[dict]]:
    """Un artículo por identidad canónica, y el recuento aparte.

    Catálogo y stock son dos cosas distintas: acá se separan y el sistema las
    escribe en dos pasos, porque que un artículo exista no significa que haya
    uno en el depósito.
    """
    por_canonico = defaultdict(list)
    for registro in registros:
        registro["sku_canonico"] = (
            registro["sku"] if identificador_es_global(registro["sku"])
            else f"{PREFIJO_SUCURSAL[registro['sucursal']]}-{registro['sku']}")
        por_canonico[registro["sku_canonico"]].append(registro)

    catalogo, pendientes = [], []
    for sku, iguales in sorted(por_canonico.items()):
        if any(r["nature"] == REQUIERE_DECISION for r in iguales):
            pendientes.extend(iguales)
            continue
        # Gana la descripción más informativa; la otra queda anotada con su
        # archivo y su fila, para que se pueda volver a la planilla.
        principal = max(iguales, key=lambda r: (len(r["nombre"]), r["corte"]))
        variantes = [r for r in iguales
                     if r["nombre"].strip().lower() != principal["nombre"].strip().lower()]
        notas = "; ".join(
            f"variante en {r['sucursal']} ({r['fuente_archivo']} fila {r['fuente_fila']}): "
            f"{r['nombre']}" for r in variantes)
        origen = " | ".join(
            f"{r['sucursal']}={r['fuente_archivo']}#{r['fuente_fila']}@{r['corte']}"
            for r in iguales)
        decidido = "; ".join(
            f"naturaleza por {r['decision']}"
            + (f" (el archivo la tenía en «{r['categoria_original']}»)"
               if r["categoria_original"] and r["categoria_original"] != r["categoria"] else "")
            for r in iguales if r.get("decision"))
        suspendido = "; ".join(
            f"cantidad en suspenso en {r['sucursal']}: {r['stock_en_suspenso']}"
            for r in iguales if r.get("stock_en_suspenso"))
        anotaciones = " || ".join(x for x in (notas, decidido, suspendido) if x)
        catalogo.append(dict(
            sku=sku, name=principal["nombre"], nature=principal["nature"],
            category=principal["categoria"], brand=principal["marca"],
            sale_price="", location="", min_stock="", barcode="", unit="UNIDAD",
            notes=(anotaciones + " || " if anotaciones else "") + "origen: " + origen))

    recuentos = {s: [] for s in PREFIJO_SUCURSAL}
    for articulo in catalogo:
        for registro in por_canonico[articulo["sku"]]:
            if registro["nature"] not in MUEVEN_STOCK:
                continue  # un servicio o un trabajo bajo pedido no lleva stock
            if registro.get("stock_en_suspenso"):
                continue  # el artículo existe; sus unidades esperan un conteo real
            if not registro["stock_inicial"] or registro["stock_inicial"] <= 0:
                continue
            recuentos[registro["sucursal"]].append(dict(
                sku=articulo["sku"], cantidad=registro["stock_inicial"],
                fuente=f"{registro['fuente_archivo']}#{registro['fuente_fila']}",
                corte=registro["corte"]))
    return catalogo, recuentos, pendientes


CAMPOS = ["sku", "name", "nature", "category", "brand", "sale_price", "location",
          "min_stock", "barcode", "unit", "notes"]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--pc", required=True, help="PC - Inventario.xlsx (Asunción)")
    parser.add_argument("--p2", required=True, help="P2 - Inventario.xlsx (Pilar)")
    parser.add_argument("--salida", required=True)
    args = parser.parse_args()

    salida = Path(args.salida)
    salida.mkdir(parents=True, exist_ok=True)
    registros, rechazos = normalizar({"PC": Path(args.pc), "P2": Path(args.p2)})
    catalogo, recuentos, pendientes = consolidar(registros)

    with (salida / "catalogo_canonico.csv").open("w", encoding="utf-8-sig", newline="") as archivo:
        escritor = csv.DictWriter(archivo, fieldnames=CAMPOS)
        escritor.writeheader()
        escritor.writerows(catalogo)
    for sucursal, lineas in recuentos.items():
        (salida / f"recuento_{sucursal}.json").write_text(
            json.dumps(lineas, ensure_ascii=False, indent=1), encoding="utf-8")

    globales = sum(1 for a in catalogo if identificador_es_global(a["sku"]))
    print(f"registros normalizados : {len(registros)}")
    print(f"filas rechazadas       : {len(rechazos)}")
    print(f"articulos canonicos    : {len(catalogo)} "
          f"({globales} globales, {len(catalogo) - globales} con prefijo de sucursal)")
    for sucursal, lineas in recuentos.items():
        print(f"recuento {sucursal:9}: {len(lineas):>5} lineas, "
              f"{sum(x['cantidad'] for x in lineas):>7} unidades")
    suspendidos = [r for r in registros if r.get("stock_en_suspenso")]
    if suspendidos:
        print(f"unidades en suspenso   : {len(suspendidos)} articulos, "
              f"{sum(r['stock_inicial'] or 0 for r in suspendidos)} unidades declaradas que NO entran")
        for r in suspendidos:
            print(f"    {r['sucursal']:9} {r['sku']:>8} {r['nombre'][:28]:30} "
                  f"declara {r['stock_inicial']:>7} -- {r['stock_en_suspenso']}")
    print(f"en espera de decision  : {len(pendientes)}")
    for clave, cuantos in Counter(
            (r["sucursal"], r["categoria"] or "(sin categoria)") for r in pendientes).most_common():
        print(f"    {clave[0]:9} {clave[1]:22} {cuantos:>4}")
    for rechazo in rechazos:
        print(f"    RECHAZO {rechazo['fuente']} fila {rechazo['fila']:>5}: "
              f"{rechazo['motivo']} -- {rechazo['articulo'][:44]!r}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
