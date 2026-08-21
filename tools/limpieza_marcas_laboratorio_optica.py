# -*- coding: utf-8 -*-
"""Saca el laboratorio de la columna «Marca», ahora que tiene su propio campo.

En las planillas de la Óptica la «Marca» de un cristal trae el laboratorio, y
así entró al catálogo: treinta y un artículos tienen hoy «Laboratorio Optilab»
o «Laboratorio Servi Optical» donde debería ir el fabricante. La migración 028
le dio al laboratorio su lugar —`articles.default_laboratory_id`— y desde
entonces la marca dejó de ser el único sitio donde ese dato podía estar.

Esto limpia la marca. No toca el laboratorio, ni el stock, ni una línea de
venta, ni la naturaleza de nada. Y no inventa marcas: donde la fuente corregida
del 19/08 trae una marca real, se usa esa; donde no trae ninguna, la marca queda
en blanco, que es lo que era.

    python tools/limpieza_marcas_laboratorio_optica.py [--base <ruta>]
        [--fuente-corregida <Inventario P2.xlsx>] [--confirmar]

Sin `--confirmar` no escribe: imprime el plan y sale. Ese es el dry-run.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
import sys
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4

RAIZ = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(RAIZ))

from modulos.caja_diaria.config import resolve_data_paths  # noqa: E402
from modulos.comercial.application.comercial_controller import (  # noqa: E402
    build_comercial_controller,
)

ACCION = "BRAND_LABORATORY_CLEANUP"
ACTOR = "COMMAND_CENTER/BC-OPTICA-LIMPIEZA-MARCAS-LABORATORIO-V1-015"
EVIDENCIA = (RAIZ / "artifacts" / "BC-OPTICA-LIMPIEZA-MARCAS-LABORATORIO-V1-015"
             / "EVIDENCIA_FUENTE_CORREGIDA.json")

#: Las dos grafías que en producción nombran un laboratorio en vez de una marca,
#: y el laboratorio que nombra cada una. Es una lista cerrada a propósito: no se
#: limpia «toda marca que se parezca a un laboratorio», porque ese criterio
#: barrería «Optica San Cayetano» —que es una óptica, no un laboratorio— y
#: cualquier fabricante que algún día se llame parecido. Lo que autoriza a
#: limpiar es la evidencia de V1-008 y V1-012 sobre estas dos, no un parecido.
MARCAS_QUE_SON_LABORATORIO = {
    "laboratorio optilab": "Laboratorio Optilab",
    "laboratorio servi optical": "ServiOptica",
}

#: Sólo en esta categoría se da por confirmado que la marca era el laboratorio.
#: En Cristales la planilla usaba la columna «Marca» para eso, y la fuente
#: corregida del 19/08 sigue haciéndolo: no hay marca real en ningún lado.
CATEGORIA_CONFIRMADA = "Cristales"

CONFIRMADO = "LABORATORY_IN_BRAND_CONFIRMED"
FUENTE_REAL = "CORRECTED_SOURCE_HAS_REAL_BRAND"
MARCA_VALIDA = "VALID_BRAND"
AMBIGUO = "AMBIGUOUS"
SIN_ACCION = "NO_ACTION"

SKU_RE = re.compile(r"^\s*(\d{4,})\s*(.*)$")

lineas: list[str] = []
fallas: list[str] = []


def registrar(texto: str = "") -> None:
    print(texto, flush=True)
    lineas.append(texto)


def comprobar(condicion: bool, descripcion: str) -> bool:
    registrar(f"  {'OK  ' if condicion else 'FALLA'} {descripcion}")
    if not condicion:
        fallas.append(descripcion)
    return bool(condicion)


def sha256(ruta: Path) -> str:
    return hashlib.sha256(ruta.read_bytes()).hexdigest()


def norm(s: str | None) -> str:
    s = unicodedata.normalize("NFKD", (s or "").lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", " ", s).strip()


# --------------------------------------------------------------------------
# Lectura
# --------------------------------------------------------------------------

def radiografia(base: Path) -> dict:
    """Todo lo que esta misión promete no mover."""
    c = sqlite3.connect(f"file:{base}?mode=ro", uri=True)
    try:
        q = lambda s, *a: c.execute(s, a).fetchone()[0]  # noqa: E731
        return dict(
            articulos=q("SELECT COUNT(*) FROM articles"),
            activos=q("SELECT COUNT(*) FROM articles WHERE active=1"),
            movimientos=q("SELECT COUNT(*) FROM stock_movements"),
            asuncion=q("SELECT COALESCE(SUM(quantity),0) FROM stock_movements"
                       " WHERE destination='ASUNCION'"),
            pilar=q("SELECT COALESCE(SUM(quantity),0) FROM stock_movements"
                    " WHERE destination='PILAR'"),
            entradas=q("SELECT COUNT(*) FROM cash_entries"),
            suma_caja=q("SELECT COALESCE(SUM(total),0) FROM cash_entries"),
            sale_items=q("SELECT COUNT(*) FROM sale_items"),
            laboratorios_en_lineas=q(
                "SELECT COUNT(*) FROM sale_items WHERE COALESCE(laboratory,'') <> ''"),
            categorias=q("SELECT COUNT(*) FROM article_categories"),
            marcas=q("SELECT COUNT(*) FROM brands"),
            laboratorios=q("SELECT COUNT(*) FROM laboratories"),
            con_default=q("SELECT COUNT(*) FROM articles"
                          " WHERE default_laboratory_id IS NOT NULL"),
            integridad=q("PRAGMA integrity_check"),
            fk=len(c.execute("PRAGMA foreign_key_check").fetchall()),
            negativos=q("SELECT COUNT(*) FROM stock_actual WHERE quantity<0"),
            huerfanos=q("SELECT COUNT(*) FROM stock_movements sm LEFT JOIN articles a"
                        " ON a.id=sm.article_id WHERE a.id IS NULL"),
            efectos=q("SELECT COUNT(*) FROM event_effects ee LEFT JOIN domain_events de"
                      " ON de.event_id=ee.event_id WHERE de.event_id IS NULL"),
        )
    finally:
        c.close()


def foto_articulos(base: Path) -> dict[str, dict]:
    c = sqlite3.connect(f"file:{base}?mode=ro", uri=True)
    c.row_factory = sqlite3.Row
    try:
        cats = {r[0]: r[1] for r in c.execute("SELECT id, name FROM article_categories")}
        marcas = {r[0]: r[1] for r in c.execute("SELECT id, name FROM brands")}
        salida = {}
        for fila in c.execute("SELECT * FROM articles"):
            d = dict(fila)
            d.pop("updated_at", None)  # cambia por definición al escribir
            d["_categoria"] = cats.get(d["category_id"])
            d["_marca"] = marcas.get(d["brand_id"])
            salida[d["sku"]] = d
        return salida
    finally:
        c.close()


def diferencias(antes: dict[str, dict], despues: dict[str, dict]) -> list[tuple]:
    salida = []
    for sku in sorted(set(antes) | set(despues)):
        a, d = antes.get(sku), despues.get(sku)
        if a is None or d is None:
            salida.append((sku, "existencia", a is not None, d is not None))
            continue
        for campo in a:
            if campo.startswith("_"):
                continue
            if a[campo] != d[campo]:
                salida.append((sku, campo, a[campo], d[campo]))
    return salida


def leer_fuente_corregida(ruta: Path) -> dict[str, str]:
    """SKU -> marca, tal como quedó en la planilla corregida del 19/08.

    Encabezado en la fila 2, datos desde la 3, y la columna «Articulo» trae el
    código pegado al nombre. Es el mismo formato que perfiló V1-010.
    """
    from openpyxl import load_workbook

    libro = load_workbook(ruta, data_only=True)
    hoja = libro.active
    try:
        encabezado = [str(c or "").strip() for c in
                      next(hoja.iter_rows(min_row=2, max_row=2, values_only=True))]
        if "Articulo" not in encabezado or "Marca" not in encabezado:
            raise SystemExit(f"la fuente {ruta.name} no tiene «Articulo» y «Marca»"
                             f" en la fila 2: {encabezado}")
        col_art = encabezado.index("Articulo")
        col_marca = encabezado.index("Marca")
        salida: dict[str, str] = {}
        for n, fila in enumerate(hoja.iter_rows(values_only=True), start=1):
            if n <= 2:
                continue
            art = str(fila[col_art] or "").strip()
            if not art:
                continue
            m = SKU_RE.match(art)
            if not m:
                continue
            salida[m.group(1)] = str(fila[col_marca] or "").strip()
        return salida
    finally:
        libro.close()


def evidencia_registrada() -> dict:
    return json.loads(EVIDENCIA.read_text(encoding="utf-8"))


# --------------------------------------------------------------------------
# Clasificación
# --------------------------------------------------------------------------

def clasificar(articulo: dict, marca_corregida: str | None,
               marca_destino_compostura: str | None) -> tuple[str, str | None, str]:
    """Devuelve (clase, marca objetivo, evidencia).

    La marca objetivo sólo la mira quien va a escribir, y sólo para las clases
    que cambian algo. Vaciar la marca es `CONFIRMADO`, nunca un objetivo nulo:
    la diferencia entre «dejala como está» y «vaciala» la decide la clase, para
    que no dependa de si un valor vino en `None`.
    """
    marca = articulo["_marca"]
    categoria = articulo["_categoria"]

    if not marca:
        return SIN_ACCION, None, "ya no tiene marca: nada que limpiar"

    if norm(marca) not in MARCAS_QUE_SON_LABORATORIO:
        return MARCA_VALIDA, marca, f"«{marca}» no nombra a ninguno de los laboratorios"

    laboratorio = MARCAS_QUE_SON_LABORATORIO[norm(marca)]

    # 1. La fuente corregida manda, si trae una marca que no sea otro laboratorio.
    if marca_corregida and norm(marca_corregida) not in MARCAS_QUE_SON_LABORATORIO:
        return (FUENTE_REAL, marca_corregida,
                f"la fuente corregida del 19/08 dice «{marca_corregida}»")
    if (marca_corregida is not None and not marca_corregida
            and categoria == CATEGORIA_CONFIRMADA):
        # La fuente corregida está a mano y deja la marca vacía. Para un cristal
        # eso confirma lo que ya sabíamos. Fuera de Cristales no confirma nada:
        # una celda en blanco es «no sé», no «no tiene marca», y limpiar por una
        # celda vacía sería inventar una decisión que nadie tomó.
        return (CONFIRMADO, None,
                "la fuente corregida del 19/08 deja la marca vacía")

    # 2. Sin marca real en la fuente, sólo Cristales está confirmado.
    if categoria == CATEGORIA_CONFIRMADA:
        return (CONFIRMADO, None,
                f"cristal: la planilla usaba «Marca» para el laboratorio"
                f" («{laboratorio}»), y hoy eso vive en default_laboratory_id")

    # 3. Compostura, con la marca de la fuente corregida ya registrada en V1-012.
    if marca_destino_compostura and categoria in ("Compostura", "Composturas"):
        return (FUENTE_REAL, marca_destino_compostura,
                f"Compostura: la fuente corregida del 19/08 le da"
                f" «{marca_destino_compostura}» a las 21 filas de la categoría"
                " (registrado en V1-012; desde Casa no se releyó el archivo)")

    # 4. Todo lo demás espera a una persona.
    return (AMBIGUO, marca,
            f"«{marca}» nombra un laboratorio, pero el artículo está en"
            f" «{categoria}» y ninguna fuente corregida disponible le da"
            " una marca real")


def construir_plan(base: Path,
                   corregidas: dict[str, str] | None) -> tuple[list[dict], dict]:
    ev = evidencia_registrada()
    destino_compostura = (ev.get("marca_destino_compostura") or {}).get("nombre")
    articulos = foto_articulos(base)
    marcas_por_nombre: dict[str, dict] = {}
    c = sqlite3.connect(f"file:{base}?mode=ro", uri=True)
    try:
        for mid, nombre, activa in c.execute("SELECT id, name, active FROM brands"):
            marcas_por_nombre[norm(nombre)] = dict(id=mid, name=nombre, active=activa)
        laboratorios = [r[0] for r in c.execute("SELECT name FROM laboratories")]
    finally:
        c.close()

    plan = []
    for sku, art in sorted(articulos.items()):
        if not art["_marca"] or norm(art["_marca"]) not in MARCAS_QUE_SON_LABORATORIO:
            continue
        corregida = corregidas.get(sku) if corregidas is not None else None
        clase, objetivo, evidencia = clasificar(art, corregida, destino_compostura)
        destino = marcas_por_nombre.get(norm(objetivo)) if objetivo else None
        cambia = clase in (CONFIRMADO, FUENTE_REAL)
        plan.append(dict(
            sku=sku, id=art["id"], nombre=art["name"], categoria=art["_categoria"],
            naturaleza=art["nature"], activo=art["active"],
            marca_actual=art["_marca"], clase=clase,
            marca_propuesta=(None if clase == CONFIRMADO else objetivo),
            marca_destino_id=(None if clase == CONFIRMADO else (destino or {}).get("id")),
            marca_destino_activa=(None if clase == CONFIRMADO
                                  else (destino or {}).get("active")),
            default_laboratory_id=art["default_laboratory_id"],
            evidencia=evidencia, cambia=cambia))
    return plan, dict(laboratorios=laboratorios, marcas=marcas_por_nombre)


def marcas_sospechosas_no_listadas(catalogo: dict) -> list[str]:
    """Marcas que nombran un laboratorio y que esta misión no tiene autorizadas.

    Existe para que una marca-laboratorio nueva se vea, no para limpiarla. Si
    aparece algo acá, es una decisión de catálogo, no un caso más del lote.
    """
    labs = {norm(n) for n in catalogo["laboratorios"]}
    return sorted(m["name"] for clave, m in catalogo["marcas"].items()
                  if clave in labs and clave not in MARCAS_QUE_SON_LABORATORIO)


# --------------------------------------------------------------------------

def imprimir_plan(plan: list[dict], origen_fuente: str) -> None:
    registrar("== los casos, uno por uno ==")
    registrar(f"fuente de marcas corregidas: {origen_fuente}")
    registrar()
    cab = (f"{'SKU':9s} | {'Nombre':26s} | {'Categoria':11s} | {'Marca actual':26s}"
           f" | {'Marca propuesta':21s} | {'Lab. default':12s} | Accion")
    registrar(cab)
    registrar("-" * len(cab))
    for p in plan:
        if not p["cambia"]:
            propuesta = "(sin cambio)"
        elif p["clase"] == CONFIRMADO:
            propuesta = "(en blanco)"
        else:
            propuesta = p["marca_propuesta"] or "(en blanco)"
        lab = "si" if p["default_laboratory_id"] else "(ninguno)"
        registrar(f"{p['sku']:9s} | {p['nombre'][:26]:26s} | {(p['categoria'] or '')[:11]:11s}"
                  f" | {p['marca_actual'][:26]:26s} | {propuesta[:21]:21s}"
                  f" | {lab:12s} | {p['clase']}")
    registrar()
    for p in plan:
        registrar(f"  {p['sku']}: {p['evidencia']}")
    registrar()


def resumen(plan: list[dict]) -> dict:
    conteo: dict[str, int] = {}
    for p in plan:
        conteo[p["clase"]] = conteo.get(p["clase"], 0) + 1
    return conteo


def _volcar(salida: str | None) -> None:
    if salida:
        Path(salida).write_text("\n".join(lineas) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base", default=None)
    parser.add_argument("--fuente-corregida", default=None,
                        help="Inventario P2 corregido del 19/08. Si no se pasa se usa"
                             " la evidencia registrada en artifacts.")
    parser.add_argument("--confirmar", action="store_true")
    parser.add_argument("--salida", default=None, help="dónde dejar este informe")
    args = parser.parse_args()

    base = Path(args.base) if args.base else Path(resolve_data_paths().database)
    registrar("LIMPIEZA DE MARCAS QUE EN REALIDAD SON LABORATORIOS")
    registrar(f"base   : {base}")
    if not base.exists():
        registrar("La base no existe. No se escribe nada.")
        return 1
    registrar(f"sha256 : {sha256(base)}")

    corregidas = None
    origen = ("evidencia registrada en artifacts (V1-010 y V1-012): la fuente"
              " corregida vive en la PC de la Óptica")
    if args.fuente_corregida:
        ruta = Path(args.fuente_corregida)
        if not ruta.exists():
            registrar(f"La fuente corregida {ruta} no existe. No se escribe nada.")
            return 1
        corregidas = leer_fuente_corregida(ruta)
        origen = f"{ruta.name} (sha256 {sha256(ruta)[:16]}, {len(corregidas)} filas)"
    registrar()

    antes_totales = radiografia(base)
    antes_articulos = foto_articulos(base)
    plan, catalogo = construir_plan(base, corregidas)

    imprimir_plan(plan, origen)

    conteo = resumen(plan)
    registrar("== resumen ==")
    for clase in (CONFIRMADO, FUENTE_REAL, MARCA_VALIDA, AMBIGUO, SIN_ACCION):
        registrar(f"  {clase:34s} {conteo.get(clase, 0)}")
    a_cambiar = [p for p in plan if p["cambia"]]
    registrar(f"  {'artículos que cambian':34s} {len(a_cambiar)}")
    registrar()

    registrar("== guardas previas ==")
    comprobar(all(p["categoria"] for p in plan),
              "todos los casos conservan su categoría")
    for p in a_cambiar:
        if p["clase"] != FUENTE_REAL:
            continue
        comprobar(p["marca_destino_id"] is not None,
                  f"{p['sku']}: la marca «{p['marca_propuesta']}» ya existe"
                  " (no se crea ninguna)")
        comprobar(p["marca_destino_activa"] == 1,
                  f"{p['sku']}: «{p['marca_propuesta']}» está activa")
    nuevas = marcas_sospechosas_no_listadas(catalogo)
    comprobar(not nuevas,
              f"ninguna marca-laboratorio fuera de la lista autorizada ({nuevas})")

    ambiguos = [p for p in plan if p["clase"] == AMBIGUO]
    if ambiguos:
        registrar()
        registrar("== HUMAN_GATE: no se tocan ==")
        for p in ambiguos:
            registrar(f"  {p['sku']} «{p['nombre']}» ({p['categoria']},"
                      f" {p['naturaleza']}): marca «{p['marca_actual']}»")
            registrar(f"      {p['evidencia']}")
    if fallas:
        registrar()
        registrar("Alguna guarda falló. No se escribe nada.")
        _volcar(args.salida)
        return 1
    registrar()

    if not args.confirmar:
        registrar("DRY-RUN: no se escribió nada. Falta --confirmar.")
        registrar(f"Esto vaciaría la marca de {conteo.get(CONFIRMADO, 0)} artículos y la"
                  f" reemplazaría por una marca real en {conteo.get(FUENTE_REAL, 0)}.")
        registrar("Ningún otro campo se nombra, así que ningún otro campo cambia.")
        _volcar(args.salida)
        return 0

    if not a_cambiar:
        registrar("NO SE ESCRIBE NADA: ya está todo limpio.")
        registrar("Idempotencia: una segunda corrida no cambia nada.")
        _volcar(args.salida)
        return 0

    registrar("== backup verificable ==")
    sello = datetime.now().strftime("%Y%m%d-%H%M%S")
    respaldo = base.parent / "Backups" / f"bc-caja-premarcas-{sello}.sqlite3"
    respaldo.parent.mkdir(parents=True, exist_ok=True)
    origen_c = sqlite3.connect(f"file:{base}?mode=ro", uri=True)
    try:
        destino_c = sqlite3.connect(str(respaldo))
        try:
            origen_c.backup(destino_c)
        finally:
            destino_c.close()
    finally:
        origen_c.close()
    registrar(f"  archivo: {respaldo}")
    registrar(f"  sha256 : {sha256(respaldo)}")
    comprobar(radiografia(respaldo) == antes_totales,
              "el backup tiene el mismo contenido que la base")
    if fallas:
        registrar("El backup no quedó bien. No se escribe nada.")
        _volcar(args.salida)
        return 1
    registrar()

    cuando = datetime.now(timezone.utc).replace(microsecond=0)
    registrar("== limpieza ==")
    ctrl = build_comercial_controller(base)
    try:
        for p in a_cambiar:
            # Un solo campo nombrado. `actualizar_articulo` lee el resto y lo
            # vuelve a dejar como estaba: es la operación que faltaba en V1-010.
            quedo = ctrl.actualizar_articulo(p["id"], actor=ACTOR,
                                             brand_id=p["marca_destino_id"])
            registrar(f"  {p['sku']} «{p['nombre'][:24]}»: «{p['marca_actual']}» -> "
                      f"{p['marca_propuesta'] or '(en blanco)'}"
                      f"   [{quedo.nature.value}, laboratorio"
                      f" {'intacto' if quedo.default_laboratory_id == p['default_laboratory_id'] else 'CAMBIADO'}]")
    finally:
        ctrl.close()

    # La bitácora, en su propia conexión y después de cerrar el controlador:
    # dos conexiones escribiendo a la vez es como se traba SQLite.
    conexion = sqlite3.connect(str(base))
    try:
        for p in a_cambiar:
            conexion.execute(
                "INSERT INTO admin_audit_log(id, actor, action, target_type, target_id,"
                " result, details_json, recorded_at) VALUES (?,?,?,?,?,?,?,?)",
                (str(uuid4()), ACTOR, ACCION, "article", p["id"], p["clase"],
                 json.dumps(dict(
                     sku=p["sku"], nombre=p["nombre"], categoria=p["categoria"],
                     marca_anterior=p["marca_actual"], marca_nueva=p["marca_propuesta"],
                     clase=p["clase"], evidencia=p["evidencia"], fuente=origen,
                     alcance=("solo brand_id: ni laboratorio, ni stock, ni naturaleza,"
                              " ni una linea de venta"),
                     laboratorio_por_defecto_intacto=p["default_laboratory_id"]),
                     ensure_ascii=False, sort_keys=True),
                 cuando.isoformat()))
        conexion.commit()
    finally:
        conexion.close()
    registrar()

    registrar("== verificación ==")
    despues_totales = radiografia(base)
    despues_articulos = foto_articulos(base)

    cambios = diferencias(antes_articulos, despues_articulos)
    esperados = {(p["sku"], "brand_id") for p in a_cambiar}
    inesperados = [c for c in cambios if (c[0], c[1]) not in esperados]
    comprobar(not inesperados,
              f"sólo cambió brand_id, y sólo donde correspondía ({inesperados})")
    comprobar(len(cambios) == len(esperados),
              f"cambios: {len(cambios)}, esperados {len(esperados)}")

    for clave in ("articulos", "activos", "movimientos", "asuncion", "pilar",
                  "entradas", "suma_caja", "sale_items", "laboratorios_en_lineas",
                  "categorias", "marcas", "laboratorios", "con_default"):
        comprobar(antes_totales[clave] == despues_totales[clave],
                  f"{clave}: {antes_totales[clave]} sin cambio")
    comprobar(despues_totales["integridad"] == "ok", "integrity_check ok")
    comprobar(despues_totales["fk"] == 0, "FK 0")
    comprobar(despues_totales["negativos"] == 0, "negativos 0")
    comprobar(despues_totales["huerfanos"] == 0, "huérfanos 0")
    comprobar(despues_totales["efectos"] == 0, "efectos sin hecho 0")

    plan_despues, _ = construir_plan(base, corregidas)
    comprobar(not [p for p in plan_despues if p["cambia"]],
              "idempotencia: una segunda corrida no cambiaría nada")
    registrar()
    registrar(f"rollback si hiciera falta: copiar {respaldo.name} sobre {base.name}")
    _volcar(args.salida)
    return 1 if fallas else 0


if __name__ == "__main__":
    raise SystemExit(main())
