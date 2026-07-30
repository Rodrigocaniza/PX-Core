"""Persistencia y exportación de planillas de asociaciones.

Este módulo NO registra movimientos económicos ni modifica cierres.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from uuid import uuid4
import json

from datos import leer_datos, guardar_datos

BASE_DIR = Path(__file__).resolve().parent
DATOS_DIR = BASE_DIR / "Datos"
RUTA_CONFIG = DATOS_DIR / "asociaciones_config.txt"
RUTA_PLANILLAS = DATOS_DIR / "asociaciones_planillas.txt"
RUTA_DETALLES = DATOS_DIR / "asociaciones_detalles.txt"
CARPETA_LOGOS = DATOS_DIR / "logos_asociaciones"

TIPOS_DESCUENTO = (
    "SIN_DESCUENTO",
    "SOLO_PRIMERA_CUOTA",
    "TODAS_LAS_CUOTAS",
    "MANUAL",
)


def _asegurar_carpetas():
    DATOS_DIR.mkdir(parents=True, exist_ok=True)
    CARPETA_LOGOS.mkdir(parents=True, exist_ok=True)


def _leer_json(ruta):
    _asegurar_carpetas()
    registros = []
    for linea in leer_datos(ruta):
        linea = linea.strip()
        if not linea:
            continue
        try:
            registros.append(json.loads(linea))
        except json.JSONDecodeError:
            continue
    return registros


def _guardar_json(ruta, registros):
    _asegurar_carpetas()
    guardar_datos(
        ruta,
        [json.dumps(r, ensure_ascii=False, separators=(",", ":")) for r in registros],
    )


def limpiar_texto(valor, campo, obligatorio=True):
    texto = str(valor or "").strip()
    if obligatorio and not texto:
        raise ValueError(f"Completá el campo {campo}.")
    if "\n" in texto or "\r" in texto:
        texto = " ".join(texto.splitlines())
    return texto


def validar_periodo(periodo):
    valor = limpiar_texto(periodo, "período")
    try:
        datetime.strptime(valor, "%m-%Y")
    except ValueError as error:
        raise ValueError("El período debe tener formato MM-AAAA.") from error
    return valor


def convertir_entero(valor, campo, permitir_cero=True):
    texto = str(valor or "").strip().replace(".", "").replace(",", "").replace(" ", "")
    if not texto:
        return 0 if permitir_cero else _error(f"Completá el campo {campo}.")
    if not texto.isdigit():
        raise ValueError(f"{campo} debe contener solo números.")
    numero = int(texto)
    if numero < 0 or (numero == 0 and not permitir_cero):
        raise ValueError(f"{campo} debe ser mayor que cero.")
    return numero


def convertir_porcentaje(valor):
    texto = str(valor or "0").strip().replace(",", ".")
    try:
        numero = float(texto)
    except ValueError as error:
        raise ValueError("El porcentaje no es válido.") from error
    if numero < 0 or numero > 100:
        raise ValueError("El porcentaje debe estar entre 0 y 100.")
    return numero


def _error(mensaje):
    raise ValueError(mensaje)


def listar_configuraciones():
    return _leer_json(RUTA_CONFIG)


def obtener_configuracion(config_id):
    return next((x for x in listar_configuraciones() if x["id"] == config_id), None)


def guardar_configuracion(datos):
    registros = listar_configuraciones()
    config_id = datos.get("id") or f"ASC-{uuid4().hex[:10].upper()}"
    tipo = limpiar_texto(datos.get("tipo_descuento"), "tipo de descuento")
    if tipo not in TIPOS_DESCUENTO:
        raise ValueError("Tipo de descuento inválido.")

    registro = {
        "id": config_id,
        "asociacion": limpiar_texto(datos.get("asociacion"), "asociación"),
        "local": limpiar_texto(datos.get("local"), "local"),
        "titulo": limpiar_texto(datos.get("titulo"), "título"),
        "texto_liquidacion": limpiar_texto(
            datos.get("texto_liquidacion"), "texto de liquidación"
        ),
        "logo": str(datos.get("logo") or "").strip(),
        "tipo_descuento": tipo,
        "porcentaje": convertir_porcentaje(datos.get("porcentaje", 0)),
        "permite_ajuste_manual": bool(datos.get("permite_ajuste_manual", True)),
        "categorias": [
            limpiar_texto(x, "categoría")
            for x in datos.get("categorias", [])
            if str(x).strip()
        ] or ["Funcionarios"],
    }

    for i, actual in enumerate(registros):
        if actual["id"] == config_id:
            registros[i] = registro
            break
    else:
        registros.append(registro)

    _guardar_json(RUTA_CONFIG, registros)
    return registro


def eliminar_configuracion(config_id):
    planillas = [x for x in listar_planillas() if x["config_id"] == config_id]
    if planillas:
        raise ValueError("No se puede eliminar: la configuración tiene planillas.")
    registros = [x for x in listar_configuraciones() if x["id"] != config_id]
    _guardar_json(RUTA_CONFIG, registros)


def listar_planillas(periodo=None, config_id=None, categoria=None):
    registros = _leer_json(RUTA_PLANILLAS)
    if periodo:
        registros = [x for x in registros if x["periodo"] == periodo]
    if config_id:
        registros = [x for x in registros if x["config_id"] == config_id]
    if categoria:
        registros = [x for x in registros if x["categoria"] == categoria]
    return registros


def obtener_planilla(planilla_id):
    return next((x for x in listar_planillas() if x["id"] == planilla_id), None)


def guardar_planilla(datos):
    registros = listar_planillas()
    planilla_id = datos.get("id") or f"PLA-{uuid4().hex[:10].upper()}"
    config = obtener_configuracion(datos.get("config_id"))
    if not config:
        raise ValueError("Seleccioná una asociación/local válido.")

    categoria = limpiar_texto(datos.get("categoria"), "categoría")
    if categoria not in config["categorias"]:
        raise ValueError("La categoría no pertenece a la configuración elegida.")

    registro = {
        "id": planilla_id,
        "config_id": config["id"],
        "periodo": validar_periodo(datos.get("periodo")),
        "categoria": categoria,
        "descuento_manual": convertir_entero(
            datos.get("descuento_manual", 0), "descuento manual"
        ),
        "observacion": limpiar_texto(
            datos.get("observacion", ""), "observación", obligatorio=False
        ),
        "creado": datos.get("creado") or datetime.now().isoformat(timespec="seconds"),
    }

    duplicada = next(
        (
            x for x in registros
            if x["id"] != planilla_id
            and x["config_id"] == registro["config_id"]
            and x["periodo"] == registro["periodo"]
            and x["categoria"] == registro["categoria"]
        ),
        None,
    )
    if duplicada:
        raise ValueError("Ya existe una planilla para ese local, período y categoría.")

    for i, actual in enumerate(registros):
        if actual["id"] == planilla_id:
            registros[i] = registro
            break
    else:
        registros.append(registro)

    _guardar_json(RUTA_PLANILLAS, registros)
    return registro


def eliminar_planilla(planilla_id):
    _guardar_json(
        RUTA_PLANILLAS,
        [x for x in listar_planillas() if x["id"] != planilla_id],
    )
    _guardar_json(
        RUTA_DETALLES,
        [x for x in listar_detalles() if x["planilla_id"] != planilla_id],
    )


def listar_detalles(planilla_id=None):
    registros = _leer_json(RUTA_DETALLES)
    if planilla_id:
        registros = [x for x in registros if x["planilla_id"] == planilla_id]
    return registros



def normalizar_numero_cuota(valor):
    """Acepta 1, 1/10 o fechas de Excel usadas para mostrar cuota/total."""
    if hasattr(valor, "day") and hasattr(valor, "month"):
        return f"{valor.day}/{valor.month}"

    texto = str(valor or "").strip().replace(" ", "")
    if not texto:
        raise ValueError("Completá el número de cuota.")

    texto = texto.replace("-", "/")
    if "/" in texto:
        partes = texto.split("/")
        if len(partes) != 2 or not all(p.isdigit() for p in partes):
            raise ValueError("La cuota debe tener formato 1 o 1/10.")
        actual, total = map(int, partes)
        if actual <= 0 or total <= 0 or actual > total:
            raise ValueError("La relación de cuota no es válida.")
        return f"{actual}/{total}"

    if texto.isdigit():
        numero = int(texto)
        if numero <= 0:
            raise ValueError("El número de cuota debe ser mayor que cero.")
        return str(numero)

    raise ValueError("La cuota debe tener formato 1 o 1/10.")


def es_primera_cuota(valor):
    texto = normalizar_numero_cuota(valor)
    return int(texto.split("/")[0]) == 1

def guardar_detalle(datos):
    planilla = obtener_planilla(datos.get("planilla_id"))
    if not planilla:
        raise ValueError("La planilla seleccionada no existe.")

    registros = listar_detalles()
    detalle_id = datos.get("id") or f"DET-{uuid4().hex[:10].upper()}"
    total_compra = convertir_entero(datos.get("total_compra"), "total de compra")
    monto_cuota = convertir_entero(
        datos.get("monto_cuota"), "monto de cuota", permitir_cero=False
    )
    saldo = datos.get("saldo_pendiente")
    saldo_pendiente = (
        convertir_entero(saldo, "saldo pendiente")
        if str(saldo or "").strip()
        else max(total_compra - monto_cuota, 0)
    )

    registro = {
        "id": detalle_id,
        "planilla_id": planilla["id"],
        "legajo": limpiar_texto(datos.get("legajo"), "legajo"),
        "nombre": limpiar_texto(datos.get("nombre"), "nombre y apellido"),
        "numero_cuota": normalizar_numero_cuota(datos.get("numero_cuota")),
        "total_compra": total_compra,
        "monto_cuota": monto_cuota,
        "saldo_pendiente": saldo_pendiente,
    }

    for i, actual in enumerate(registros):
        if actual["id"] == detalle_id:
            registros[i] = registro
            break
    else:
        registros.append(registro)

    _guardar_json(RUTA_DETALLES, registros)
    return registro


def eliminar_detalle(detalle_id):
    _guardar_json(
        RUTA_DETALLES,
        [x for x in listar_detalles() if x["id"] != detalle_id],
    )



def eliminar_periodo(periodo, config_id=None):
    """Elimina todas las planillas y detalles del período indicado."""
    periodo = validar_periodo(periodo)
    planillas = listar_planillas(periodo=periodo, config_id=config_id)
    ids = {x["id"] for x in planillas}
    if not ids:
        return 0

    _guardar_json(
        RUTA_PLANILLAS,
        [x for x in listar_planillas() if x["id"] not in ids],
    )
    _guardar_json(
        RUTA_DETALLES,
        [x for x in listar_detalles() if x["planilla_id"] not in ids],
    )
    return len(ids)


def _texto_celda(valor):
    return str(valor or "").strip()


def _categoria_desde_titulo(titulo):
    texto = _texto_celda(titulo)
    if " - " in texto:
        return texto.rsplit(" - ", 1)[-1].strip()
    return texto


def _periodo_desde_texto(texto, nombre_hoja=""):
    import re
    meses = {
        "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
        "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
        "septiembre": 9, "setiembre": 9, "octubre": 10,
        "noviembre": 11, "diciembre": 12,
    }
    combinado = f"{_texto_celda(texto)} {_texto_celda(nombre_hoja)}".lower()
    patron = r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|setiembre|octubre|noviembre|diciembre)\s+(?:de\s+)?(\d{4})"
    coincidencia = re.search(patron, combinado)
    if coincidencia:
        return f"{meses[coincidencia.group(1)]:02d}-{coincidencia.group(2)}"

    coincidencia = re.search(r"\b(0[1-9]|1[0-2])[-/](\d{4})\b", combinado)
    if coincidencia:
        return f"{coincidencia.group(1)}-{coincidencia.group(2)}"
    raise ValueError("No se pudo identificar el período del Excel.")


def importar_excel(ruta_excel, config_id, reemplazar_periodo=False):
    """Importa uno o varios bloques de planillas desde un libro Excel."""
    from openpyxl import load_workbook

    config = obtener_configuracion(config_id)
    if not config:
        raise ValueError("Seleccioná una asociación/local válido.")

    libro = load_workbook(ruta_excel, data_only=True)
    bloques = []

    for hoja in libro.worksheets:
        fila = 1
        while fila <= hoja.max_row:
            titulo = _texto_celda(hoja.cell(fila, 1).value)
            if not titulo.lower().startswith("sres."):
                fila += 1
                continue

            categoria = _categoria_desde_titulo(titulo)
            fila_liquidacion = fila + 1
            periodo = _periodo_desde_texto(
                hoja.cell(fila_liquidacion, 1).value,
                hoja.title,
            )

            fila_encabezado = None
            for candidata in range(fila + 1, min(fila + 7, hoja.max_row) + 1):
                encabezado = _texto_celda(hoja.cell(candidata, 1).value).lower()
                if "legajo" in encabezado:
                    fila_encabezado = candidata
                    break
            if fila_encabezado is None:
                raise ValueError(
                    f"No se encontró el encabezado de columnas para {categoria}."
                )

            detalles = []
            actual = fila_encabezado + 1
            while actual <= hoja.max_row:
                legajo = hoja.cell(actual, 1).value
                nombre = _texto_celda(hoja.cell(actual, 2).value)
                etiqueta = _texto_celda(hoja.cell(actual, 4).value).lower()

                if etiqueta in {"sub total", "subtotal", "total"} or etiqueta.startswith("descuento"):
                    break
                if not _texto_celda(legajo) and not nombre:
                    break

                cuota = normalizar_numero_cuota(hoja.cell(actual, 3).value)
                total_compra = convertir_entero(
                    hoja.cell(actual, 4).value, "total de compra"
                )
                monto_cuota = convertir_entero(
                    hoja.cell(actual, 5).value,
                    "monto de cuota",
                    permitir_cero=False,
                )
                saldo_valor = hoja.cell(actual, 6).value
                saldo_texto = _texto_celda(saldo_valor)
                saldo = (
                    0 if saldo_texto in {"", "-"}
                    else convertir_entero(saldo_valor, "saldo pendiente")
                )
                detalles.append({
                    "legajo": _texto_celda(legajo),
                    "nombre": nombre,
                    "numero_cuota": cuota,
                    "total_compra": total_compra,
                    "monto_cuota": monto_cuota,
                    "saldo_pendiente": saldo,
                })
                actual += 1

            if detalles:
                bloques.append({
                    "categoria": categoria,
                    "periodo": periodo,
                    "detalles": detalles,
                    "hoja": hoja.title,
                })
            fila = max(actual + 1, fila + 1)

    if not bloques:
        raise ValueError("El Excel no contiene planillas reconocibles.")

    periodos = {x["periodo"] for x in bloques}
    if reemplazar_periodo:
        for periodo in periodos:
            eliminar_periodo(periodo, config_id=config_id)
    else:
        existentes = [
            x for x in bloques
            if listar_planillas(
                periodo=x["periodo"],
                config_id=config_id,
                categoria=x["categoria"],
            )
        ]
        if existentes:
            raise ValueError(
                "Ya existen planillas de ese período. "
                "Usá 'Eliminar mes completo' o activá reemplazo."
            )

    # Agrega categorías nuevas detectadas sin perder las existentes.
    categorias = list(config["categorias"])
    for bloque in bloques:
        if bloque["categoria"] not in categorias:
            categorias.append(bloque["categoria"])
    if categorias != config["categorias"]:
        config = dict(config)
        config["categorias"] = categorias
        guardar_configuracion(config)

    resumen = {"planillas": 0, "registros": 0, "periodos": sorted(periodos)}
    for bloque in bloques:
        planilla = guardar_planilla({
            "config_id": config_id,
            "periodo": bloque["periodo"],
            "categoria": bloque["categoria"],
            "descuento_manual": 0,
            "observacion": f"Importado desde {Path(ruta_excel).name}",
        })
        for detalle in bloque["detalles"]:
            guardar_detalle({"planilla_id": planilla["id"], **detalle})
            resumen["registros"] += 1
        resumen["planillas"] += 1
    return resumen


def siguiente_periodo(periodo):
    """Devuelve el mes siguiente manteniendo el formato MM-AAAA."""
    fecha = datetime.strptime(validar_periodo(periodo), "%m-%Y")
    if fecha.month == 12:
        return f"01-{fecha.year + 1}"
    return f"{fecha.month + 1:02d}-{fecha.year}"


def siguiente_numero_cuota(numero_cuota):
    """Incrementa 1 a 2 y 1/10 a 2/10."""
    texto = normalizar_numero_cuota(numero_cuota)
    if "/" not in texto:
        return str(int(texto) + 1)

    actual, total = map(int, texto.split("/"))
    if actual >= total:
        return None
    return f"{actual + 1}/{total}"


def generar_mes_siguiente(planilla_id):
    """Genera la planilla siguiente usando únicamente saldos pendientes.

    Conserva el total original de la compra, incrementa el número de cuota,
    ajusta la última cuota al saldo disponible y descuenta esa cuota del saldo.
    """
    origen = obtener_planilla(planilla_id)
    if not origen:
        raise ValueError("La planilla seleccionada no existe.")

    periodo_destino = siguiente_periodo(origen["periodo"])
    existente = next(
        (
            p for p in listar_planillas(
                periodo=periodo_destino,
                config_id=origen["config_id"],
                categoria=origen["categoria"],
            )
        ),
        None,
    )
    if existente:
        raise ValueError(
            "Ya existe una planilla para el mes siguiente y la misma categoría."
        )

    detalles_origen = listar_detalles(planilla_id)
    continuados = []
    finalizados = 0

    for detalle in detalles_origen:
        saldo_anterior = int(detalle["saldo_pendiente"])
        cuota_siguiente = siguiente_numero_cuota(detalle["numero_cuota"])

        if saldo_anterior <= 0 or cuota_siguiente is None:
            finalizados += 1
            continue

        monto_siguiente = min(int(detalle["monto_cuota"]), saldo_anterior)
        saldo_nuevo = max(saldo_anterior - monto_siguiente, 0)

        continuados.append({
            "legajo": detalle["legajo"],
            "nombre": detalle["nombre"],
            "numero_cuota": cuota_siguiente,
            "total_compra": detalle["total_compra"],
            "monto_cuota": monto_siguiente,
            "saldo_pendiente": saldo_nuevo,
        })

    if not continuados:
        raise ValueError(
            "No hay registros con saldo pendiente para generar el mes siguiente."
        )

    nueva = guardar_planilla({
        "config_id": origen["config_id"],
        "periodo": periodo_destino,
        "categoria": origen["categoria"],
        "descuento_manual": 0,
        "observacion": f"Generada desde {origen['periodo']}",
    })

    try:
        for detalle in continuados:
            guardar_detalle({
                **detalle,
                "planilla_id": nueva["id"],
            })
    except Exception:
        eliminar_planilla(nueva["id"])
        raise

    return {
        "planilla": nueva,
        "periodo_origen": origen["periodo"],
        "periodo_destino": periodo_destino,
        "continuados": len(continuados),
        "finalizados": finalizados,
    }




def generar_meses_siguientes(planilla_ids):
    """Genera el mes siguiente de varias planillas como una sola operación."""
    ids = list(dict.fromkeys(planilla_ids or []))
    if not ids:
        raise ValueError("Seleccioná al menos una planilla.")

    errores = []
    destinos = set()

    for planilla_id in ids:
        origen = obtener_planilla(planilla_id)
        if not origen:
            errores.append(f"Planilla inexistente: {planilla_id}")
            continue

        periodo_destino = siguiente_periodo(origen["periodo"])
        clave_destino = (
            origen["config_id"],
            periodo_destino,
            origen["categoria"].casefold(),
        )
        if clave_destino in destinos:
            errores.append(
                f"{origen['categoria']} {periodo_destino}: destino repetido "
                "dentro de la selección."
            )
        destinos.add(clave_destino)

        existente = listar_planillas(
            periodo=periodo_destino,
            config_id=origen["config_id"],
            categoria=origen["categoria"],
        )
        if existente:
            errores.append(
                f"{origen['categoria']} {periodo_destino}: ya existe."
            )

        continuables = 0
        for detalle in listar_detalles(planilla_id):
            if (
                int(detalle["saldo_pendiente"]) > 0
                and siguiente_numero_cuota(detalle["numero_cuota"]) is not None
            ):
                continuables += 1

        if continuables == 0:
            errores.append(
                f"{origen['categoria']} {origen['periodo']}: "
                "no tiene saldos pendientes."
            )

    if errores:
        raise ValueError(
            "No se generó ninguna planilla:\n\n- " + "\n- ".join(errores)
        )

    resultados = []
    creadas = []
    try:
        for planilla_id in ids:
            resultado = generar_mes_siguiente(planilla_id)
            resultados.append(resultado)
            creadas.append(resultado["planilla"]["id"])
    except Exception:
        for creada_id in creadas:
            eliminar_planilla(creada_id)
        raise

    return {
        "planillas": len(resultados),
        "continuados": sum(x["continuados"] for x in resultados),
        "finalizados": sum(x["finalizados"] for x in resultados),
        "periodos": sorted({x["periodo_destino"] for x in resultados}),
        "resultados": resultados,
    }





def calcular_saldo_pendiente(planilla_id):
    return sum(
        int(d["saldo_pendiente"])
        for d in listar_detalles(planilla_id)
    )

def calcular_totales(planilla_id, descuento_manual_override=None):
    planilla = obtener_planilla(planilla_id)
    if not planilla:
        raise ValueError("La planilla no existe.")
    config = obtener_configuracion(planilla["config_id"])
    detalles = listar_detalles(planilla_id)

    subtotal = sum(x["monto_cuota"] for x in detalles)
    tipo = config["tipo_descuento"]
    porcentaje = float(config["porcentaje"])

    if tipo == "SOLO_PRIMERA_CUOTA":
        base = sum(
            x["monto_cuota"] for x in detalles
            if es_primera_cuota(x["numero_cuota"])
        )
        descuento = round(base * porcentaje / 100)
    elif tipo == "TODAS_LAS_CUOTAS":
        base = subtotal
        descuento = round(base * porcentaje / 100)
    elif tipo == "MANUAL":
        base = 0
        descuento = (
            convertir_entero(descuento_manual_override, "descuento manual")
            if descuento_manual_override is not None
            else planilla["descuento_manual"]
        )
    else:
        base = 0
        descuento = 0

    return {
        "subtotal": subtotal,
        "base_descuento": base,
        "descuento": min(descuento, subtotal),
        "total": max(subtotal - descuento, 0),
        "cantidad": len(detalles),
    }


UNIDADES = (
    "", "UN", "DOS", "TRES", "CUATRO", "CINCO", "SEIS", "SIETE", "OCHO", "NUEVE",
    "DIEZ", "ONCE", "DOCE", "TRECE", "CATORCE", "QUINCE", "DIECISEIS",
    "DIECISIETE", "DIECIOCHO", "DIECINUEVE", "VEINTE", "VEINTIUNO",
    "VEINTIDOS", "VEINTITRES", "VEINTICUATRO", "VEINTICINCO", "VEINTISEIS",
    "VEINTISIETE", "VEINTIOCHO", "VEINTINUEVE",
)
DECENAS = ("", "", "VEINTE", "TREINTA", "CUARENTA", "CINCUENTA", "SESENTA", "SETENTA", "OCHENTA", "NOVENTA")
CENTENAS = ("", "CIENTO", "DOSCIENTOS", "TRESCIENTOS", "CUATROCIENTOS", "QUINIENTOS", "SEISCIENTOS", "SETECIENTOS", "OCHOCIENTOS", "NOVECIENTOS")


def _grupo_letras(n):
    if n == 0:
        return ""
    if n == 100:
        return "CIEN"
    partes = []
    if n >= 100:
        partes.append(CENTENAS[n // 100])
        n %= 100
    if n < 30:
        if n:
            partes.append(UNIDADES[n])
    else:
        decena, unidad = divmod(n, 10)
        partes.append(DECENAS[decena] + (f" Y {UNIDADES[unidad]}" if unidad else ""))
    return " ".join(partes)


def numero_en_letras(numero):
    n = int(numero)
    if n == 0:
        return "CERO GUARANÍES"
    partes = []
    millones, resto = divmod(n, 1_000_000)
    miles, unidades = divmod(resto, 1_000)
    if millones:
        partes.append("UN MILLON" if millones == 1 else f"{_grupo_letras(millones)} MILLONES")
    if miles:
        partes.append("MIL" if miles == 1 else f"{_grupo_letras(miles)} MIL")
    if unidades:
        partes.append(_grupo_letras(unidades))
    return " ".join(partes) + " GUARANÍES"


def exportar_excel(planilla_id, ruta_destino):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    planilla = obtener_planilla(planilla_id)
    config = obtener_configuracion(planilla["config_id"])
    detalles = listar_detalles(planilla_id)
    totales = calcular_totales(planilla_id)

    wb = Workbook()
    ws = wb.active
    ws.title = planilla["categoria"][:31]
    ws.sheet_view.showGridLines = False

    if config.get("logo") and Path(config["logo"]).exists():
        img = Image(config["logo"])
        img.width, img.height = 120, 60
        ws.add_image(img, "A1")

    ws.merge_cells("A1:F1")
    ws["A1"] = config["titulo"]
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"{config['texto_liquidacion']} {planilla['periodo']}"
    ws["A2"].font = Font(size=12, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:F3")
    ws["A3"] = planilla["categoria"].upper()
    ws["A3"].font = Font(size=12, bold=True)
    ws["A3"].alignment = Alignment(horizontal="center")

    encabezados = ["Legajo", "Nombre y Apellido", "Cuota", "Total compra", "Monto cuota", "Saldo"]
    for col, texto in enumerate(encabezados, 1):
        celda = ws.cell(5, col, texto)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1F4E78")
        celda.alignment = Alignment(horizontal="center")

    thin = Side(style="thin", color="808080")
    for fila, detalle in enumerate(detalles, 6):
        valores = [
            detalle["legajo"], detalle["nombre"], detalle["numero_cuota"],
            detalle["total_compra"], detalle["monto_cuota"], detalle["saldo_pendiente"],
        ]
        for col, valor in enumerate(valores, 1):
            celda = ws.cell(fila, col, valor)
            celda.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            celda.alignment = Alignment(horizontal="left" if col == 2 else "center")
            if col >= 4:
                celda.number_format = '#,##0'

    fila = 6 + len(detalles)
    ws.cell(fila, 4, "SUBTOTAL").font = Font(bold=True)
    ws.cell(fila, 5, totales["subtotal"]).number_format = '#,##0'
    fila += 1
    if totales["descuento"]:
        ws.cell(fila, 4, "DESCUENTO").font = Font(bold=True)
        ws.cell(fila, 5, totales["descuento"]).number_format = '#,##0'
        fila += 1
    ws.cell(fila, 4, "TOTAL").font = Font(bold=True)
    ws.cell(fila, 5, totales["total"]).font = Font(bold=True)
    ws.cell(fila, 5).number_format = '#,##0'
    fila += 2
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
    ws.cell(fila, 1, numero_en_letras(totales["total"]))
    ws.cell(fila, 1).font = Font(bold=True)
    ws.cell(fila, 1).alignment = Alignment(horizontal="center")

    widths = [14, 36, 10, 16, 16, 16]
    for i, ancho in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = ancho
    ws.freeze_panes = "A6"
    wb.save(ruta_destino)


def exportar_pdf(planilla_id, ruta_destino):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    planilla = obtener_planilla(planilla_id)
    config = obtener_configuracion(planilla["config_id"])
    detalles = listar_detalles(planilla_id)
    totales = calcular_totales(planilla_id)

    doc = SimpleDocTemplate(
        str(ruta_destino), pagesize=landscape(A4),
        leftMargin=12*mm, rightMargin=12*mm, topMargin=10*mm, bottomMargin=10*mm,
    )
    estilos = getSampleStyleSheet()
    centro = ParagraphStyle("Centro", parent=estilos["Normal"], alignment=1, fontSize=10)
    elementos = []

    if config.get("logo") and Path(config["logo"]).exists():
        elementos.append(Image(config["logo"], width=35*mm, height=18*mm))
    elementos.append(Paragraph(f"<b>{config['titulo']}</b>", estilos["Title"]))
    elementos.append(Paragraph(
        f"<b>{config['texto_liquidacion']} {planilla['periodo']}</b>", centro
    ))
    elementos.append(Paragraph(f"<b>{planilla['categoria'].upper()}</b>", centro))
    elementos.append(Spacer(1, 5*mm))

    datos = [["Legajo", "Nombre y Apellido", "Cuota", "Total compra", "Monto cuota", "Saldo"]]
    for d in detalles:
        datos.append([
            d["legajo"], d["nombre"], d["numero_cuota"],
            f"{d['total_compra']:,}".replace(",", "."),
            f"{d['monto_cuota']:,}".replace(",", "."),
            f"{d['saldo_pendiente']:,}".replace(",", "."),
        ])
    tabla = Table(datos, colWidths=[28*mm, 80*mm, 20*mm, 34*mm, 34*mm, 34*mm], repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), .5, colors.grey),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("ALIGN", (1,1), (1,-1), "LEFT"),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 4*mm))

    resumen = [["SUBTOTAL", f"{totales['subtotal']:,}".replace(",", ".")]]
    if totales["descuento"]:
        resumen.append(["DESCUENTO", f"{totales['descuento']:,}".replace(",", ".")])
    resumen.append(["TOTAL", f"{totales['total']:,}".replace(",", ".")])
    t_resumen = Table(resumen, colWidths=[55*mm, 35*mm], hAlign="RIGHT")
    t_resumen.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), .5, colors.grey),
        ("FONTNAME", (0,0), (-1,-1), "Helvetica-Bold"),
        ("ALIGN", (1,0), (1,-1), "RIGHT"),
    ]))
    elementos.append(t_resumen)
    elementos.append(Spacer(1, 4*mm))
    elementos.append(Paragraph(f"<b>{numero_en_letras(totales['total'])}</b>", centro))
    doc.build(elementos)
