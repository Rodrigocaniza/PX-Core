"""Paso 1: perfil de los dos inventarios corregidos, y diferencias con los anteriores."""
import pickle, re, unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from openpyxl import load_workbook

W = Path(r"C:\Users\Striker\AppData\Local\Temp\claude\c--Users-Striker-Desktop-Proyecto-X-PX-Core\55fef905-3b6d-44f0-bf35-8b120350484f\scratchpad\m010")
VIEJO = Path(r"C:\Users\Striker\AppData\Local\Temp\claude\c--Users-Striker-Desktop-Proyecto-X-PX-Core\55fef905-3b6d-44f0-bf35-8b120350484f\scratchpad\m008")

out = []
def reg(t=""):
    out.append(str(t))

def txt(v):
    return "" if v is None else str(v).strip()

def norm(s):
    s = unicodedata.normalize("NFKD", (s or "").lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", " ", s).strip()

SKU_RE = re.compile(r"^\s*(\d{4,})\s*(.*)$")
NUEVO = {"PC": "ASUNCION", "P2": "PILAR"}

# --- leer corregidos ---
nuevo = {}
for tag, nombre in (("PC", "Inventario PC.xlsx"), ("P2", "Inventario P2.xlsx")):
    wbv = load_workbook(W / nombre, data_only=True)
    wbf = load_workbook(W / nombre, data_only=False)
    wsv, wsf = wbv.active, wbf.active
    filas, formulas = [], 0
    for n, r in enumerate(wsv.iter_rows(values_only=True), start=1):
        if n <= 2:
            continue
        art = txt(r[0])
        if not art:
            continue
        filas.append(dict(n=n, articulo=art, cod_barra=txt(r[1]), categoria=txt(r[2]),
                          marca=txt(r[3]), costo=txt(r[4]), precio=txt(r[5]), stock=txt(r[6])))
    for fila in wsf.iter_rows():
        for celda in fila:
            if isinstance(celda.value, str) and celda.value.startswith("="):
                formulas += 1
    nuevo[tag] = dict(filas=filas, fisicas=wsv.max_row, cols=wsv.max_column,
                      merges=len(wsv.merged_cells.ranges), formulas=formulas,
                      encabezado=[txt(c) for c in next(wsv.iter_rows(min_row=2, max_row=2, values_only=True))],
                      fecha=txt(wsv.cell(row=1, column=1).value))
    wbv.close(); wbf.close()

# --- leer viejos ---
raw_viejo = pickle.load(open(VIEJO / "raw.pkl", "rb"))
COLS_V = {"PC": dict(a=3, m=4, c=5, s=2), "P2": dict(a=0, c=1, m=2, s=3)}
viejo = {}
for tag in ("PC", "P2"):
    cv = COLS_V[tag]
    filas = []
    for n, r in enumerate(raw_viejo[tag], start=1):
        if n <= 2 or not txt(r[cv["a"]]):
            continue
        filas.append(dict(n=n, articulo=txt(r[cv["a"]]), categoria=txt(r[cv["c"]]),
                          marca=txt(r[cv["m"]]), stock=txt(r[cv["s"]])))
    viejo[tag] = filas


def sku_de(articulo, cod_barra=""):
    if cod_barra:
        return cod_barra
    m = SKU_RE.match(articulo)
    return m.group(1) if m else ""


reg("=" * 78)
reg("PASO 1 -- PERFIL DE LOS INVENTARIOS CORREGIDOS")
reg("=" * 78)
for tag in ("PC", "P2"):
    d = nuevo[tag]
    fs = d["filas"]
    reg()
    reg(f"### Inventario {tag}.xls -> {NUEVO[tag]}")
    reg(f"  cabecera fila 1        : {d['fecha']!r}")
    reg(f"  encabezado             : {d['encabezado']}")
    reg(f"  filas fisicas          : {d['fisicas']}")
    reg(f"  filas con articulo     : {len(fs)}")
    reg(f"  filas vacias           : {d['fisicas'] - 2 - len(fs)}")
    reg(f"  celdas combinadas      : {d['merges']}")
    reg(f"  formulas               : {d['formulas']}")
    codigos = [sku_de(f["articulo"], f["cod_barra"]) for f in fs]
    sin_codigo = [f for f, c in zip(fs, codigos) if not c]
    reg(f"  filas sin codigo       : {len(sin_codigo)}")
    sin_desc = [f for f in fs if not SKU_RE.match(f['articulo']) or not SKU_RE.match(f['articulo']).group(2).strip()]
    reg(f"  filas sin descripcion  : {len(sin_desc)}")
    dup = {k: v for k, v in Counter(codigos).items() if v > 1 and k}
    reg(f"  codigos duplicados     : {len(dup)} {list(dup.items())[:6]}")
    nombres = Counter(f["articulo"] for f in fs)
    reg(f"  descripciones repetidas: {len({k for k, v in nombres.items() if v > 1})}")
    coincide = sum(1 for f in fs if f["cod_barra"] and f["articulo"].startswith(f["cod_barra"]))
    reg(f"  'Cod. Barra' coincide con el prefijo de 'Articulo': {coincide} de {len(fs)}")
    stocks = []
    for f in fs:
        try:
            stocks.append(int(float(f["stock"])))
        except ValueError:
            pass
    reg(f"  stock numerico         : {len(stocks)} de {len(fs)}")
    reg(f"  stock <= 0             : {sum(1 for s in stocks if s <= 0)}   (=0: {sum(1 for s in stocks if s == 0)}, <0: {sum(1 for s in stocks if s < 0)})")
    reg(f"  unidades totales       : {sum(stocks)}")
    reg(f"  min / mediana / max    : {min(stocks)} / {sorted(stocks)[len(stocks)//2]} / {max(stocks)}")
    reg(f"  top 10 stock           : {sorted(stocks, reverse=True)[:10]}")
    con_precio = sum(1 for f in fs if f["precio"] and f["precio"] != "0")
    con_costo = sum(1 for f in fs if f["costo"] and f["costo"] != "0")
    reg(f"  con PrecioA            : {con_precio}   con CostoA: {con_costo}")
    cats = Counter(f["categoria"] or "(sin categoria)" for f in fs)
    reg(f"  categorias ({len(cats)}):")
    for k, v in cats.most_common():
        u = sum(int(float(f["stock"])) for f in fs if (f["categoria"] or "(sin categoria)") == k and f["stock"])
        reg(f"      {k[:30]:32} {v:>5} filas  {u:>8} unidades")
    marcas = Counter(f["marca"] or "(sin marca)" for f in fs)
    reg(f"  marcas distintas       : {len([m for m in marcas if m != '(sin marca)'])} (sin marca: {marcas.get('(sin marca)', 0)})")
    tot = [f for f in fs if re.search(r"\b(total|subtotal|suma|resumen)\b", norm(f["articulo"]))]
    reg(f"  filas total/subtotal   : {len(tot)}")

reg()
reg("=" * 78)
reg("DIFERENCIAS ESTRUCTURALES CON LAS FUENTES DE LA MISION 008")
reg("=" * 78)
reg(f"  {'':22} {'ANTES':>28} {'AHORA':>28}")
reg(f"  {'archivo PC':22} {'PC - Inventario.xlsx':>28} {'Inventario PC.xls':>28}")
reg(f"  {'corte declarado':22} {'2026-08-03':>28} {nuevo['PC']['fecha'][:10]:>28}")
reg(f"  {'PC filas con articulo':22} {len(viejo['PC']):>28} {len(nuevo['PC']['filas']):>28}")
reg(f"  {'P2 filas con articulo':22} {len(viejo['P2']):>28} {len(nuevo['P2']['filas']):>28}")
reg(f"  {'columnas':22} {'Casilla/Zona/Stock/Art/Marca/Cat/Obs':>28}")
reg(f"  {'':22} {'':>28} {'Art/CodBarra/Cat/Marca/Costo/Precio/Stock':>28}")
reg()
reg("  columnas NUEVAS que antes no existian: Cod. Barra, CostoA, PrecioA")
reg("  columnas que DESAPARECEN: Casilla, Zona, Observacion (venian vacias en las 2.586 filas)")

pickle.dump(dict(nuevo=nuevo, viejo=viejo), open(W / "perfil.pkl", "wb"))
(W / "perfil.txt").write_text("\n".join(out), encoding="utf-8")
print("\n".join(out))
