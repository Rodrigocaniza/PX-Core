"""Cargar el catálogo real de la Óptica sin tipearlo y sin inventarlo.

Son unos 2.000 artículos en Asunción y 1.000 y pico en Pilar. Cargarlos a mano
no es una opción, y cargarlos mal es peor que no cargarlos: un catálogo con
naturalezas adivinadas pondría a un cristal a descontar stock.

Dos reglas gobiernan este módulo.

**Nada se infiere.** Si el archivo no dice la naturaleza, la fila se rechaza. No
se deduce de que la descripción diga «armazón», porque el día que alguien
escriba «armazón de cristal» el sistema va a estar equivocado y nadie va a saber
por qué.

**Catálogo no es stock.** Aplicar el archivo crea artículos y ni una sola
unidad. Las unidades entran por un recuento, que es otro hecho, con su fecha,
su responsable y su motivo.
"""

from __future__ import annotations

import csv
import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping, Sequence


class CargaInicialError(ValueError):
    """El archivo o el momento no permiten cargar."""


#: Lo que toda fila tiene que traer. Sin esto no hay artículo que crear.
COLUMNAS_OBLIGATORIAS = ("sku", "name", "nature")

#: Lo que puede venir y puede faltar. Faltar no es un error: es un dato que
#: todavía no existe, y se guarda como ausente en vez de como cero.
COLUMNAS_OPCIONALES = (
    "category", "brand", "sale_price", "location", "min_stock", "barcode",
    "unit", "notes",
)

_ENTEROS = ("sale_price", "min_stock")


def _texto(valor: Any) -> str:
    return str(valor if valor is not None else "").strip()


def _entero_opcional(valor: Any, columna: str, fila: int) -> int | None:
    """Un número, o nada. Nunca un cero inventado.

    Cero y «no sé» no son lo mismo: un precio de venta en cero significa que se
    regala, y eso casi nunca es lo que el archivo quería decir.
    """
    crudo = _texto(valor).replace(".", "").replace(" ", "")
    if not crudo:
        return None
    try:
        numero = int(crudo)
    except ValueError as exc:
        raise CargaInicialError(
            f"fila {fila}: «{columna}» tiene que ser un número entero, "
            f"y dice {valor!r}") from exc
    if numero < 0:
        raise CargaInicialError(f"fila {fila}: «{columna}» no puede ser negativo")
    return numero


def leer_archivo_de_articulos(ruta: str | Path) -> list[dict[str, Any]]:
    """Lee un CSV o un XLSX de artículos y devuelve filas normalizadas.

    No decide nada sobre ellas: eso lo hace el plan. Acá sólo se convierte el
    archivo en algo que el planificador pueda mirar, y se rechaza de entrada un
    archivo que no tenga las columnas mínimas — porque un archivo sin `nature`
    no es un archivo incompleto, es otro archivo.
    """
    ruta = Path(ruta)
    if not ruta.exists():
        raise CargaInicialError(f"no existe el archivo {ruta}")

    if ruta.suffix.lower() in (".xlsx", ".xlsm"):
        crudas = _leer_excel(ruta)
    elif ruta.suffix.lower() == ".csv":
        crudas = _leer_csv(ruta)
    else:
        raise CargaInicialError(
            f"formato no soportado: {ruta.suffix}. Se aceptan .csv y .xlsx")

    if not crudas:
        raise CargaInicialError("el archivo no tiene ni una fila de datos")

    faltantes = [c for c in COLUMNAS_OBLIGATORIAS if c not in crudas[0]]
    if faltantes:
        raise CargaInicialError(
            f"al archivo le faltan columnas obligatorias: {', '.join(faltantes)}. "
            f"Las mínimas son {', '.join(COLUMNAS_OBLIGATORIAS)}")

    filas = []
    for numero, cruda in enumerate(crudas, start=1):
        fila: dict[str, Any] = {}
        for columna in COLUMNAS_OBLIGATORIAS + COLUMNAS_OPCIONALES:
            valor = cruda.get(columna)
            if columna in _ENTEROS:
                fila[columna] = _entero_opcional(valor, columna, numero)
            else:
                fila[columna] = _texto(valor)
        if not any(fila[c] for c in COLUMNAS_OBLIGATORIAS):
            continue  # fila totalmente vacía al final de la planilla
        filas.append(fila)
    return filas


def _leer_csv(ruta: Path) -> list[dict[str, Any]]:
    with ruta.open(encoding="utf-8-sig", newline="") as archivo:
        lector = csv.DictReader(archivo)
        return [{(k or "").strip().lower(): v for k, v in fila.items()}
                for fila in lector]


def _leer_excel(ruta: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - openpyxl ya es dependencia
        raise CargaInicialError("falta openpyxl para leer archivos Excel") from exc

    libro = load_workbook(ruta, read_only=True, data_only=True)
    hoja = libro.active
    filas = hoja.iter_rows(values_only=True)
    try:
        encabezado = [str(c or "").strip().lower() for c in next(filas)]
    except StopIteration:
        return []
    return [dict(zip(encabezado, fila)) for fila in filas]


def sha256_de(ruta: str | Path) -> str:
    """Huella del archivo. Es lo que impide cargar el mismo dos veces."""
    digest = hashlib.sha256()
    with Path(ruta).open("rb") as archivo:
        for bloque in iter(lambda: archivo.read(1 << 20), b""):
            digest.update(bloque)
    return digest.hexdigest()


@dataclass(frozen=True)
class CorridaDeCarga:
    """Una carga que ocurrió, con lo necesario para auditarla después."""

    id: str
    file_name: str
    file_sha256: str
    unit: str
    rows_processed: int
    rows_imported: int
    rows_skipped: int
    error_count: int
    result: str
    administrator: str
    recorded_at: str


@dataclass(frozen=True)
class ResumenDeCompletitud:
    """Qué trae el archivo y qué le falta, antes de decidir nada.

    Distinguir «no vino» de «vino vacío» es la diferencia entre un catálogo que
    se puede completar después y uno en el que nadie sabe qué falta.
    """

    filas: int
    con_precio: int
    sin_precio: int
    con_ubicacion: int
    sin_ubicacion: int
    con_categoria: int
    con_marca: int
    con_codigo_de_barras: int
    columnas_ausentes: tuple[str, ...]

    @property
    def pendientes(self) -> tuple[str, ...]:
        """Lo que va a haber que completar a mano después de cargar."""
        faltantes = []
        if self.sin_precio:
            faltantes.append(f"{self.sin_precio} sin precio de venta")
        if self.sin_ubicacion:
            faltantes.append(f"{self.sin_ubicacion} sin ubicación")
        return tuple(faltantes)


def resumir_completitud(filas: Sequence[Mapping[str, Any]],
                        columnas_ausentes: Sequence[str] = ()) -> ResumenDeCompletitud:
    def contar(columna: str) -> int:
        return sum(1 for fila in filas if fila.get(columna))

    return ResumenDeCompletitud(
        filas=len(filas),
        con_precio=contar("sale_price"),
        sin_precio=len(filas) - contar("sale_price"),
        con_ubicacion=contar("location"),
        sin_ubicacion=len(filas) - contar("location"),
        con_categoria=contar("category"),
        con_marca=contar("brand"),
        con_codigo_de_barras=contar("barcode"),
        columnas_ausentes=tuple(columnas_ausentes),
    )
