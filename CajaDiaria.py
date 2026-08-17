"""Caja diaria de la óptica: UI legacy adaptada a Core + SQLite.

La ventana conserva el flujo CustomTkinter conocido. Toda operación nueva de
UI usa ``modulos.caja_diaria``; las funciones TXT permanecen temporalmente
solo como compatibilidad y caracterización, sin doble escritura.
"""

from __future__ import annotations

import os
import sys
import calendar
import uuid
from dataclasses import replace
from datetime import date, datetime, timedelta
from pathlib import Path

import customtkinter as ctk
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

from ImportadorExcel import (
    COLOR_BORDE,
    COLOR_PANEL,
    COLOR_PANEL_SECUNDARIO,
    COLOR_PRIMARIO,
    COLOR_PRIMARIO_HOVER,
    COLOR_ROJO,
    COLOR_TEXTO,
    COLOR_TEXTO_SUAVE,
    COLOR_VERDE,
    celda_vacia,
    normalizar,
    parsear_fecha,
    parsear_monto,
    texto_seguro,
)
from datos import guardar_datos, leer_datos
from Movimientos import UNIDADES, formatear_monto
from modulos.caja_diaria.bootstrap import build_cash_day_controller
from modulos.caja_diaria.config import resolve_data_paths
from modulos.caja_diaria.domain.models import (
    BUSINESS_TIMEZONE, SaleItem, client_balance_from_classification,
)
from modulos.caja_diaria.domain.errors import InvalidCashDayError
from modulos.caja_diaria.domain.tracking import NextAction, TrackingStatus
from modulos.caja_diaria.application.tracking_service import (
    ETIQUETAS_ESTADO as ETIQUETAS_ESTADO_UI,
    GRUPOS_SEGUIMIENTO,
)
from modulos.caja_diaria.application.services import FILTRO_REQUIEREN_ATENCION
from modulos.caja_diaria.ui.controller import friendly_error
from modulos.caja_diaria.ui.privacy import FinancialPrivacy


RUTA_CAJA_DIARIA = "Datos/caja_diaria.txt"
RUTA_ARQUEO = "Datos/arqueo_caja.txt"
UNIDAD_POR_DEFECTO = "PC"
DESCRIPCION_CAJA_INICIAL = "CAJA INICIAL"
PLANTILLA_RECETA_OBSERVACIONES = (
    "Nombre:\n"
    "Armazón:\n"
    "Cristal:\n"
    "OD:\n"
    "OI:\n"
    "ADD:\n"
    "Altura:\n"
    "DI:\n"
    "N.º FACTURA:\n"
    "RAZÓN SOCIAL:\n"
    "RUC:"
)


def observaciones_son_plantilla_neutra(valor) -> bool:
    """La plantilla inicial no representa trabajo manual pendiente."""
    return str(valor or "") == PLANTILLA_RECETA_OBSERVACIONES

# Billetes y monedas en guaraníes. Ajustar si cambia el circulante.
DENOMINACIONES = [
    100000, 50000, 20000, 10000, 5000, 2000, 1000, 500, 100, 50,
]


def perfil_visual(ancho: int, alto: int, escala_dpi: float = 1.0) -> dict:
    """Métricas deterministas; Full HD mejora legibilidad sin duplicar la UI."""
    full_hd = ancho >= 1700 and alto >= 900
    if not full_hd:
        return {
            "nombre": "compacto", "ventana": "1366x768", "fuente": 9,
            "fuente_label": 8, "fuente_seccion": 11, "fuente_kpi": 14,
            "campo_alto": 24, "fila": 27, "izquierda": 570, "separacion": 16,
            "cabecera_alto": 50, "kpi_alto": 74, "form_alto": 400,
            "toolbar_alto": 44, "grilla_alto": 354, "acciones_alto": 38,
            "contenido_ancho": 1330,
        }
    # Tk/CustomTkinter ya aplican el escalado DPI del sistema. No multiplicamos
    # nuevamente por 125%, evitando doble escala y solapamientos.
    return {
        "nombre": "full-hd", "ventana": "1920x1080", "fuente": 12,
        "fuente_label": 10, "fuente_seccion": 14, "fuente_kpi": 20,
        "campo_alto": 34, "fila": 32, "izquierda": 750, "separacion": 18,
        "cabecera_alto": 66, "kpi_alto": 104, "form_alto": 550,
        "toolbar_alto": 58, "grilla_alto": 590, "acciones_alto": 50,
            "contenido_ancho": 1890,
    }


# RC18: la cabecera muestra seis importes. Los tres primeros gobiernan la
# decision operativa del dia; los tres restantes son respaldo y se agrupan
# aparte para bajar el ruido visual. Solo afecta presentacion.
KPI_PRINCIPALES = (
    ("ventas", "Venta", "#1672E8"),
    ("efectivo", "Efectivo", "#18A66A"),
    ("esperado", "Esperado", "#0F5FB9"),
)
KPI_SECUNDARIOS = (
    ("tarjeta", "Tarj./Transf.", "#52657D"),
    ("gastos", "Gastos", "#B45309"),
    ("entregado", "Entregado", "#6B5B95"),
)


def metricas_resumen_kpi(perfil: dict) -> dict:
    """Tamanos del resumen de caja derivados del perfil visual vigente."""
    if perfil["nombre"] == "full-hd":
        # ViewSonic 24" a 1920x1080: el importe principal se lee de pie.
        return {
            "fuente_principal": perfil["fuente_kpi"],
            "fuente_secundaria": max(10, perfil["fuente_kpi"] - 7),
            "fuente_titulo": perfil["fuente_label"],
            # Piso; el alto real lo fija el contenido medido del resumen.
            "cabecera_alto": 52,
        }
    # 1366x768 conserva la altura de cabecera ya validada en RC15/RC17.
    return {
        "fuente_principal": max(10, perfil["fuente_kpi"] - 3),
        "fuente_secundaria": max(9, perfil["fuente_kpi"] - 5),
        "fuente_titulo": perfil["fuente_label"],
        "cabecera_alto": 42,
    }


#: Ultimo recurso si no se encuentra VERSION.txt. Debe coincidir con
#: pilot/package_docs/VERSION.txt; hay una prueba que lo verifica.
VERSION_APLICACION = "1.0.0-rc.27"


def version_aplicacion() -> str:
    """Version mostrada en el pie, leida del VERSION.txt que acompaña al programa.

    Estaba cableada y el paquete quedo autoidentificandose con una version
    vieja tras instalar. Leerla del archivo que el build ya empaqueta hace que
    programa y paquete no puedan divergir.
    """
    candidatos = [
        Path(getattr(sys, "_MEIPASS", "")).parent / "VERSION.txt",   # junto al .exe
        Path(sys.executable).parent / "VERSION.txt",
        Path(__file__).resolve().parent / "pilot" / "package_docs" / "VERSION.txt",
    ]
    for ruta in candidatos:
        try:
            primera = ruta.read_text(encoding="utf-8").splitlines()[0].strip()
        except (OSError, IndexError, ValueError):
            continue
        if primera.startswith("BC Caja "):
            return primera[len("BC Caja "):].strip()
    return VERSION_APLICACION


def area_trabajo_windows() -> tuple[int, int, int, int] | None:
    """RectÃ¡ngulo Ãºtil primario, excluida la barra de tareas de Windows."""
    if sys.platform != "win32":
        return None
    try:
        import ctypes
        from ctypes import wintypes

        rect = wintypes.RECT()
        if ctypes.windll.user32.SystemParametersInfoW(0x0030, 0, ctypes.byref(rect), 0):
            return rect.left, rect.top, rect.right, rect.bottom
    except (AttributeError, OSError):
        pass
    return None

CAMPOS = [
    "fecha", "unidad", "descripcion", "sobre", "arm_org", "cod",
    "armazon", "cristal", "receta_dr", "total", "efectivo",
    "tarjeta_cheque", "ordenes", "cuotas", "saldo", "gastos", "origen",
]
MOVEMENT_COLUMN_SPECS = (
    ("hora", "Hora", 52, "center"),
    ("descripcion", "Cliente", 150, "w"),
    ("cliente_telefono", "Teléfono", 95, "center"),
    ("tipo_resumen", "Tipo/Resumen", 115, "w"),
    ("sobre", "Comprobante", 85, "center"),
    ("total", "Total", 80, "center"),
    ("efectivo", "Efectivo", 80, "center"),
    ("tarjeta_transferencia", "Tarj./Transf.", 110, "center"),
    ("monto_convenio", "A cobrar conv.", 105, "center"),
    ("cuotas", "Cuotas", 55, "center"),
    ("saldo", "Saldo", 80, "center"),
    ("vendedora", "Vendedora", 85, "center"),
    ("estado", "Estado", 85, "center"),
)
COLUMNAS_OPERATIVAS = [
    (key, title, width) for key, title, width, _anchor in MOVEMENT_COLUMN_SPECS
]
PRODUCTO_TRABAJO = (
    ("arm_org", "Tipo / Producto", 150), ("cod", "Código", 90),
    ("laboratorio", "Laboratorio", 140),
    ("armazon", "P. Armazón / Desc. %", 125),
    ("cristal", "P. Cristal / Desc. %", 125),
    ("receta_dr", "Receta / Doctor", 170),
)
COBRO_PAGO = (
    ("efectivo", "Efectivo", 125), ("transferencia", "Transferencia", 125),
    ("tarjeta_cheque", "Tarjeta / Cheque", 135),
    ("ordenes", "Orden / Convenio", 150), ("monto_convenio", "Monto convenio", 125),
    ("cuotas", "Cuotas", 75), ("total", "Total de la venta", 150),
    ("saldo", "Saldo cliente", 125),
)
CAMPOS_MONETARIOS_UI = (
    "caja_inicial", "armazon", "cristal", "total", "efectivo",
    "tarjeta_cheque", "transferencia", "monto_convenio", "saldo", "salida_monto",
)


def formatear_importe_ui(valor):
    """Formatea enteros monetarios para edición visual sin alterar su valor."""
    texto = "" if valor is None else str(valor).strip()
    if not texto:
        return ""
    limpio = texto.replace(".", "").replace(" ", "")
    if not limpio.isdigit():
        return texto
    return f"{int(limpio):,}".replace(",", ".")


def sumar_importes_formulario(armazon, cristal):
    """Total visible canónico: únicamente precio de armazón más cristal."""
    return parsear_monto(armazon or "0", permitir_cero=True) + parsear_monto(
        cristal or "0", permitir_cero=True
    )


def sumar_medios_no_efectivo(tarjeta_cheque, transferencia):
    """Combina los medios no efectivos en el campo canónico persistido."""
    return parsear_monto(tarjeta_cheque or "0", permitir_cero=True) + parsear_monto(
        transferencia or "0", permitir_cero=True
    )


def calcular_saldo_pendiente(total, efectivo, tarjeta_cheque, transferencia, monto_convenio=0):
    """Saldo visible; un sobrepago nunca se representa como deuda negativa."""
    montos = [
        parsear_monto(valor or "0", permitir_cero=True)
        for valor in (total, efectivo, tarjeta_cheque, transferencia, monto_convenio)
    ]
    tarjeta_transferencia = montos[2] + montos[3]
    pendiente_antes_convenio = max(0, montos[0] - montos[1] - tarjeta_transferencia)
    if montos[4] > pendiente_antes_convenio:
        raise ValueError("El monto convenio excede el total pendiente.")
    return client_balance_from_classification(
        montos[0], montos[1], tarjeta_transferencia, montos[4]
    )


def construir_item_producto_visible(valores):
    """Convierte el producto actualmente visible en el primer item de la venta."""
    item = SaleItem(
        description=valores.get("arm_org") or valores.get("cod") or "Producto",
        code=valores.get("cod", ""), item_type=valores.get("arm_org", ""),
        frame_price=valores.get("armazon", ""), lens_price=valores.get("cristal", ""),
        frame_discount_percent=valores.get("descuento_armazon", ""),
        lens_discount_percent=valores.get("descuento_cristal", ""),
        no_cost=str(valores.get("sin_costo", "0")) in {"1", "true", "True"},
        laboratory=valores.get("laboratorio", ""),
        prescription_doctor=valores.get("receta_dr", ""),
    )
    if not item.no_cost and item.reference_subtotal <= 0:
        raise ValueError("El producto debe tener un precio de armazón o cristal.")
    return item


def completar_items_para_guardar(valores, items):
    """Conserva N items o promueve el producto visible para la venta habitual."""
    return tuple(items) if items else (construir_item_producto_visible(valores),)


def escribir_importe_formateado(campo, valor):
    """Escribe un importe legible en UI sin alterar su valor numérico."""
    texto = formatear_importe_ui(valor)
    campo.delete(0, "end")
    campo.insert(0, texto)
    return texto


def leer_valores_formulario(campos):
    """Lee únicamente controles de entrada; los botones no forman parte del payload."""
    valores = {}
    for clave, campo in campos.items():
        getter = getattr(campo, "get", None)
        if not callable(getter):
            continue
        if isinstance(campo, ctk.CTkTextbox):
            valores[clave] = getter("1.0", "end-1c").strip()
        else:
            valores[clave] = getter().strip()
    return valores


def resumen_venta_en_curso(cliente, items, privacidad):
    """Modelo visual del draft; no persiste ni altera sus artículos."""
    total = sum(item.subtotal for item in items)
    return {
        "cliente": str(cliente or "").strip() or "Cliente sin nombre",
        "cantidad": len(items),
        "total": formatear_monto(total),
        "estado": "EN CURSO",
        "items": tuple(
            (item.description, formatear_monto(item.subtotal))
            for item in items
        ),
    }


def mostrar_error_guardado(error, parent=None):
    """Hace visible cualquier rechazo del guardado usando mensajes seguros."""
    titulo, detalle = friendly_error(error)
    messagebox.showerror(titulo, detalle, parent=parent)


def formatear_diferencia_ui(diferencia):
    prefijo = "+" if diferencia > 0 else ""
    return prefijo + formatear_monto(diferencia)


def describir_diferencia_arqueo(diferencia):
    if diferencia == 0:
        return "ARQUEO CONFORME", "Caja conforme"
    if diferencia < 0:
        return "ARQUEO CON DIFERENCIA", f"Faltan {formatear_monto(abs(diferencia))}"
    return "ARQUEO CON DIFERENCIA", f"Sobran {formatear_monto(diferencia)}"

# ------------------------------------------------------------------
# Compatibilidad legacy TXT (no usada por el flujo normal de UI)
# ------------------------------------------------------------------

def _texto_archivo(valor):
    """Evita romper el formato TXT delimitado por barras."""
    return texto_seguro(valor).replace("|", "/")


def registro_a_linea(registro):
    return "|".join(_texto_archivo(registro.get(campo, "")) for campo in CAMPOS)


def linea_a_registro(linea):
    partes = linea.split("|")
    partes += [""] * (len(CAMPOS) - len(partes))
    return dict(zip(CAMPOS, partes))


def registros_del_dia(fecha, unidad):
    registros = []
    for linea in leer_datos(RUTA_CAJA_DIARIA):
        registro = linea_a_registro(linea)
        if registro["fecha"] == fecha and registro["unidad"] == unidad:
            registros.append(registro)
    return registros


def agregar_registros(nuevos):
    lineas = leer_datos(RUTA_CAJA_DIARIA)
    lineas.extend(registro_a_linea(registro) for registro in nuevos)
    guardar_datos(RUTA_CAJA_DIARIA, lineas)


def eliminar_dia(fecha, unidad):
    lineas = leer_datos(RUTA_CAJA_DIARIA)
    restantes = [
        linea for linea in lineas
        if not (
            linea_a_registro(linea)["fecha"] == fecha
            and linea_a_registro(linea)["unidad"] == unidad
        )
    ]
    eliminados = len(lineas) - len(restantes)
    guardar_datos(RUTA_CAJA_DIARIA, restantes)
    return eliminados


# ------------------------------------------------------------------
# Importación desde Excel (una hoja = un día)
# ------------------------------------------------------------------

def _num_o_cero(valor, errores, hoja_titulo, fila_num, nombre_campo):
    if celda_vacia(valor):
        return 0
    try:
        return parsear_monto(valor, permitir_cero=True)
    except ValueError as exc:
        errores.append(
            f"Hoja '{hoja_titulo}' fila {fila_num}: "
            f"{nombre_campo} inválido ('{valor}') - {exc}."
        )
        return 0


def _valor_celda(fila, indice):
    return fila[indice].value if indice < len(fila) else None


def analizar_hoja(hoja, unidad, fecha_por_defecto=None):
    """Devuelve (registros, caja_inicial, totales_hoja, errores) de una hoja."""

    registros = []
    errores = []
    caja_inicial = None
    declarados = None
    efectivo_final_declarado = None

    fila_encabezado = None
    for numero in range(1, min(hoja.max_row, 6) + 1):
        valores = {normalizar(celda.value) for celda in hoja[numero]}
        if "total" in valores and "efectivo" in valores:
            fila_encabezado = numero
            break

    if fila_encabezado is None:
        errores.append(
            f"Hoja '{hoja.title}': no se encontró la fila de encabezados "
            "(TOTAL / Efectivo). Se omite la hoja."
        )
        return [], None, None, errores

    # Las hojas futuras del libro real son plantillas con formulas en cero.
    # No constituyen dias operativos y no deben crear cajas vacias ni errores.
    filas_operativas = []
    for numero in range(fila_encabezado + 1, hoja.max_row + 1):
        fila = hoja[numero]
        descripcion_normalizada = normalizar(_valor_celda(fila, 0))
        if descripcion_normalizada in (
            "", "totales", "efectivo final", normalizar(DESCRIPCION_CAJA_INICIAL)
        ):
            continue
        if any(not celda_vacia(_valor_celda(fila, indice)) for indice in range(4, 14)):
            filas_operativas.append(numero)
    if not filas_operativas:
        return [], None, None, []

    fecha_celda = hoja.cell(row=1, column=1).value
    try:
        fecha, _ = parsear_fecha(fecha_celda)
    except ValueError:
        if fecha_por_defecto is None:
            errores.append(
                f"Hoja '{hoja.title}': la fecha en A1 ('{fecha_celda}') es "
                "inválida y no hay fecha de respaldo. Se omite la hoja."
            )
            return [], None, None, errores
        fecha = fecha_por_defecto
        errores.append(
            f"Hoja '{hoja.title}': la fecha en A1 ('{fecha_celda}') es "
            f"inválida. Se usó {fecha} por posición de la hoja. "
            "Revisala manualmente."
        )

    suma_total = suma_efectivo_ventas = suma_tarjeta = suma_gastos = 0

    for numero in range(fila_encabezado + 1, hoja.max_row + 1):
        fila = hoja[numero]

        if all(celda_vacia(celda.value) for celda in fila):
            continue

        descripcion = texto_seguro(_valor_celda(fila, 0))
        sobre = texto_seguro(_valor_celda(fila, 1))
        arm_org = texto_seguro(_valor_celda(fila, 2))
        cod = texto_seguro(_valor_celda(fila, 3))
        armazon = texto_seguro(_valor_celda(fila, 4))
        cristal = texto_seguro(_valor_celda(fila, 5))
        receta_dr = texto_seguro(_valor_celda(fila, 6))
        total_val = _valor_celda(fila, 7)
        efectivo_val = _valor_celda(fila, 8)
        tarjeta_val = _valor_celda(fila, 9)
        ordenes = texto_seguro(_valor_celda(fila, 10))
        cuotas = texto_seguro(_valor_celda(fila, 11))
        saldo = texto_seguro(_valor_celda(fila, 12))
        gastos_val = _valor_celda(fila, 13)

        if normalizar(descripcion) == normalizar(DESCRIPCION_CAJA_INICIAL):
            monto = _num_o_cero(
                efectivo_val, errores, hoja.title, numero, "caja inicial"
            )
            if caja_inicial is not None:
                errores.append(
                    f"Hoja '{hoja.title}' fila {numero}: hay más de una "
                    "fila de CAJA INICIAL, se usó la última encontrada."
                )
            caja_inicial = monto
            continue

        if normalizar(descripcion) == "totales":
            valores_declarados = (total_val, efectivo_val, tarjeta_val, gastos_val)
            if not any(isinstance(valor, str) and valor.startswith("=") for valor in valores_declarados):
                declarados = {
                    indice: _num_o_cero(valor, errores, hoja.title, numero, campo)
                    for indice, (valor, campo) in enumerate(zip(
                        valores_declarados,
                        ("TOTAL declarado", "Efectivo declarado", "Tarj./Cheq. declarado", "Gastos declarados"),
                    ))
                    if not celda_vacia(valor)
                }
            continue

        if normalizar(descripcion) in ("efectivo final", ""):
            # Fila de totales o vacía: no se importa como movimiento,
            # más abajo se recalcula y se compara contra lo declarado.
            if not celda_vacia(efectivo_val) and not (
                isinstance(efectivo_val, str) and efectivo_val.startswith("=")
            ):
                efectivo_final_declarado = _num_o_cero(
                    efectivo_val, errores, hoja.title, numero, "Efectivo final declarado"
                )
            continue

        total = _num_o_cero(total_val, errores, hoja.title, numero, "TOTAL")
        efectivo = _num_o_cero(
            efectivo_val, errores, hoja.title, numero, "Efectivo"
        )
        tarjeta = _num_o_cero(
            tarjeta_val, errores, hoja.title, numero, "Tarj./Cheq."
        )
        gastos = _num_o_cero(gastos_val, errores, hoja.title, numero, "Gastos")

        campos_con_datos = (
            armazon, cristal, ordenes, cuotas, saldo,
            total, efectivo, tarjeta, gastos,
        )
        if all(celda_vacia(v) for v in campos_con_datos):
            # Fila puramente informativa (ej. "Salida 19:30hs").
            continue

        registros.append({
            "fecha": fecha,
            "unidad": unidad,
            "descripcion": descripcion,
            "sobre": sobre,
            "arm_org": arm_org,
            "cod": cod,
            "armazon": armazon,
            "cristal": cristal,
            "receta_dr": receta_dr,
            "total": total,
            "efectivo": efectivo,
            "tarjeta_cheque": tarjeta,
            "ordenes": ordenes,
            "cuotas": cuotas,
            "saldo": saldo,
            "gastos": gastos,
            "origen": "excel",
        })

        suma_total += total
        suma_efectivo_ventas += efectivo
        suma_tarjeta += tarjeta
        suma_gastos += gastos

    if caja_inicial is None:
        errores.append(
            f"Hoja '{hoja.title}': no se encontró la fila CAJA INICIAL. "
            "Se asumió 0."
        )
        caja_inicial = 0

    totales_hoja = {
        "total": suma_total,
        "efectivo": suma_efectivo_ventas,
        "tarjeta_cheque": suma_tarjeta,
        "gastos": suma_gastos,
        "efectivo_final": caja_inicial + suma_efectivo_ventas - suma_gastos,
    }

    calculados_declarables = (
        suma_total,
        caja_inicial + suma_efectivo_ventas,
        suma_tarjeta,
        suma_gastos,
    )
    diferencias_declaradas = {
        indice: (valor, calculados_declarables[indice])
        for indice, valor in (declarados or {}).items()
        if valor != calculados_declarables[indice]
    }
    if diferencias_declaradas:
        errores.append(
            f"Hoja '{hoja.title}': los TOTALES declarados no coinciden con los "
            f"recalculados: {diferencias_declaradas}."
        )
    if efectivo_final_declarado is not None and efectivo_final_declarado != totales_hoja["efectivo_final"]:
        errores.append(
            f"Hoja '{hoja.title}': el efectivo final declarado {efectivo_final_declarado} "
            f"no coincide con el recalculado {totales_hoja['efectivo_final']}."
        )

    return registros, caja_inicial, totales_hoja, errores


def _fecha_siguiente(fecha_texto):
    try:
        fecha_dt = datetime.strptime(fecha_texto, "%d-%m-%Y")
    except (TypeError, ValueError):
        return None
    return (fecha_dt + timedelta(days=1)).strftime("%d-%m-%Y")


def analizar_archivo(ruta, unidad=UNIDAD_POR_DEFECTO):
    libro = load_workbook(ruta, data_only=True)
    resultado = {
        "por_dia": {},
        "errores": [],
        "duplicados": [],
    }

    existentes = {
        (registro["fecha"], registro["unidad"])
        for linea in leer_datos(RUTA_CAJA_DIARIA)
        for registro in [linea_a_registro(linea)]
    }

    fechas_archivo = set()
    fecha_previa = None

    for hoja in libro.worksheets:
        fecha_respaldo = _fecha_siguiente(fecha_previa) if fecha_previa else None
        registros, caja_inicial, totales, errores = analizar_hoja(
            hoja,
            unidad,
            fecha_por_defecto=fecha_respaldo,
        )
        resultado["errores"].extend(errores)

        if not registros:
            continue

        fecha = registros[0]["fecha"] if registros else fecha_respaldo
        if fecha is None:
            resultado["errores"].append(
                f"Hoja '{hoja.title}': no fue posible determinar la fecha."
            )
            continue

        fecha_previa = fecha

        if fecha in fechas_archivo:
            resultado["duplicados"].append(fecha)
            resultado["errores"].append(
                f"Hoja '{hoja.title}': la fecha {fecha} está repetida "
                "dentro del mismo Excel y fue omitida."
            )
            continue

        fechas_archivo.add(fecha)

        if (fecha, unidad) in existentes:
            resultado["duplicados"].append(fecha)
            continue

        resultado["por_dia"][fecha] = {
            "registros": registros,
            "caja_inicial": caja_inicial,
            "totales": totales,
            "hoja": hoja.title,
            "unidad": unidad,
        }

    return resultado

def aplicar_importacion(resultado):
    nuevos = []
    for fecha, datos in resultado["por_dia"].items():
        nuevos.append({
            "fecha": fecha,
            "unidad": datos.get("unidad", UNIDAD_POR_DEFECTO),
            "descripcion": DESCRIPCION_CAJA_INICIAL,
            "sobre": "", "arm_org": "", "cod": "", "armazon": "",
            "cristal": "", "receta_dr": "",
            "total": "", "efectivo": datos["caja_inicial"],
            "tarjeta_cheque": "", "ordenes": "", "cuotas": "",
            "saldo": "", "gastos": "", "origen": "excel",
        })
        nuevos.extend(datos["registros"])
    agregar_registros(nuevos)
    return len(nuevos)


# ------------------------------------------------------------------
# Cierre del día y arqueo
# ------------------------------------------------------------------

def calcular_cierre(fecha, unidad):
    registros = registros_del_dia(fecha, unidad)
    caja_inicial = 0
    suma_total = suma_efectivo = suma_tarjeta = suma_gastos = 0

    for registro in registros:
        if normalizar(registro["descripcion"]) == normalizar(
            DESCRIPCION_CAJA_INICIAL
        ):
            caja_inicial = int(registro["efectivo"] or 0)
            continue
        suma_total += int(registro["total"] or 0)
        suma_efectivo += int(registro["efectivo"] or 0)
        suma_tarjeta += int(registro["tarjeta_cheque"] or 0)
        suma_gastos += int(registro["gastos"] or 0)

    return {
        "caja_inicial": caja_inicial,
        "total": suma_total,
        "efectivo_ventas": suma_efectivo,
        "tarjeta_cheque": suma_tarjeta,
        "gastos": suma_gastos,
        "efectivo_esperado": caja_inicial + suma_efectivo - suma_gastos,
        "cantidad_registros": len(
            [r for r in registros if normalizar(r["descripcion"])
             != normalizar(DESCRIPCION_CAJA_INICIAL)]
        ),
    }


def registrar_arqueo(fecha, unidad, conteo):
    """Guarda un solo arqueo por fecha y unidad; el nuevo reemplaza al anterior."""

    total_contado = sum(
        denominacion * cantidad for denominacion, cantidad in conteo.items()
    )
    cierre = calcular_cierre(fecha, unidad)
    diferencia = total_contado - cierre["efectivo_esperado"]
    estado = (
        "OK" if diferencia == 0
        else "SOBRA" if diferencia > 0
        else "FALTA"
    )
    detalle = ";".join(
        f"{denominacion}x{cantidad}"
        for denominacion, cantidad in sorted(conteo.items(), reverse=True)
        if cantidad
    )

    linea = "|".join([
        fecha,
        unidad,
        str(total_contado),
        str(cierre["efectivo_esperado"]),
        str(diferencia),
        estado,
        detalle,
    ])

    lineas = []
    for existente in leer_datos(RUTA_ARQUEO):
        partes = existente.split("|")
        if len(partes) >= 2 and partes[0] == fecha and partes[1] == unidad:
            continue
        lineas.append(existente)

    lineas.append(linea)
    guardar_datos(RUTA_ARQUEO, lineas)

    return {
        "total_contado": total_contado,
        "efectivo_esperado": cierre["efectivo_esperado"],
        "diferencia": diferencia,
        "estado": estado,
    }


# TODO (fase 2): una vez validada la importación, generar automáticamente
# en Movimientos.RUTA_MOVIMIENTOS un Ingreso (efectivo + tarjeta/cheque) y
# un Egreso (gastos) por día/unidad, evitando duplicar si ya se generó.


# ------------------------------------------------------------------
# Interfaz gráfica
# ------------------------------------------------------------------

def abrir_caja_diaria(ventana_padre, controller=None, usar_ventana_raiz=False):
    """Abre la UI conocida usando Core + SQLite para toda operación nueva."""
    controller = controller or build_cash_day_controller()
    ctk.set_appearance_mode("light")
    ventana = ventana_padre if usar_ventana_raiz else ctk.CTkToplevel(ventana_padre)
    ventana.title("Caja diaria - Óptica")
    tamano_forzado = os.environ.get("BC_CAJA_WINDOW_SIZE", "").lower().split("x")
    area_trabajo = area_trabajo_windows()
    if area_trabajo:
        area_x, area_y, area_derecha, area_inferior = area_trabajo
        ancho_trabajo = area_derecha - area_x
        alto_trabajo = area_inferior - area_y
        barra_tareas = max(0, ventana.winfo_screenheight() - alto_trabajo)
    else:
        area_x = area_y = barra_tareas = 0
        ancho_trabajo, alto_trabajo = ventana.winfo_screenwidth(), ventana.winfo_screenheight()
    tamano_de_prueba = len(tamano_forzado) == 2 and all(
        valor.isdigit() for valor in tamano_forzado
    )
    if tamano_de_prueba:
        ancho_disponible, alto_disponible = map(int, tamano_forzado)
    else:
        ancho_disponible, alto_disponible = ancho_trabajo, alto_trabajo
    perfil = perfil_visual(ancho_disponible, alto_disponible)
    # El perfil describe la UI interna. La ventana conserva siempre el marco
    # nativo: maximizar no equivale a forzar una geometría fullscreen.
    ancho_logico = min(ancho_disponible, ancho_trabajo)
    alto_objetivo = alto_disponible - barra_tareas if tamano_de_prueba else alto_disponible
    alto_logico = min(alto_objetivo, alto_trabajo)
    if tamano_de_prueba:
        ventana.geometry(f"{ancho_logico}x{alto_logico}+{area_x}+{area_y}")
    else:
        ancho_inicial = min(1440, max(1100, ancho_logico - 120))
        alto_inicial = min(900, max(680, alto_logico - 100))
        ventana.geometry(f"{ancho_inicial}x{alto_inicial}+40+30")
    ventana.minsize(1100, 680)
    ventana.resizable(True, True)
    ventana.overrideredirect(False)
    ventana.attributes("-fullscreen", False)
    if sys.platform == "win32" and not tamano_de_prueba:
        ventana.after_idle(lambda: ventana.state("zoomed"))

    color_fondo = "#EAF2FB"
    color_panel = "#F4F8FD"
    color_panel_alto = "#F8FBFF"
    color_borde_suave = "#B9CDE5"
    color_texto = "#162238"
    color_suave = "#607089"
    color_azul = "#1672E8"
    color_verde = "#18A66A"
    color_naranja = "#F59E0B"
    ventana.configure(fg_color=color_fondo)

    barra_superior = ctk.CTkFrame(
        ventana, height=(40 if perfil["nombre"] == "full-hd" else 30), fg_color="#0F5FB9", corner_radius=0
    )
    barra_superior.pack(fill="x", padx=0, pady=0)
    barra_superior.pack_propagate(False)
    ctk.CTkLabel(
        barra_superior, text="BC", width=28, height=22, corner_radius=5,
        fg_color="#FFFFFF", text_color="#0F5FB9",
        font=ctk.CTkFont(size=(18 if perfil["nombre"] == "full-hd" else 14), weight="bold"), anchor="center",
    ).pack(side="left", padx=(16, 8))
    ctk.CTkLabel(
        barra_superior, text="BC Caja Diaria   │   Óptica Central",
        text_color="#FFFFFF", font=ctk.CTkFont(size=(18 if perfil["nombre"] == "full-hd" else 14), weight="bold"),
    ).pack(side="left", padx=(0, 12))

    # RC25: la alerta del circuito, en la franja que la operadora tiene siempre
    # delante. Antes habia que entrar a Seguimiento para enterarse de que habia
    # 15 trabajos esperando, es decir, habia que sospecharlo primero; ahora lo
    # dice la propia pantalla y el clic lleva exactamente a esos trabajos.
    # Va aca y no en la cabecera de Caja porque ahi competia por el ancho con
    # los seis importes y terminaba recortandolos.
    aviso_seguimiento = ctk.CTkButton(
        barra_superior, text="", height=(26 if perfil["nombre"] == "full-hd" else 22),
        fg_color="#FDECEC", text_color="#A32626", hover_color="#FBDADA",
        corner_radius=6, anchor="w",
        font=ctk.CTkFont(size=(13 if perfil["nombre"] == "full-hd" else 11),
                         weight="bold"),
        command=lambda: ir_a_pendientes_sucursal(),
    )

    privacidad = FinancialPrivacy(timeout_seconds=300)
    estado_admin = {"session": None, "window": None}
    navegacion = ctk.CTkFrame(
        ventana, height=(44 if perfil["nombre"] == "full-hd" else 32), fg_color="#F5F9FE", corner_radius=0,
        border_width=1, border_color=color_borde_suave,
    )
    navegacion.pack(fill="x")
    navegacion.pack_propagate(False)
    ctk.CTkLabel(
        barra_superior, text="⚙  Configuración        📅  " + date.today().strftime("Hoy, %d/%m/%Y"),
        text_color="#DCEBFF", font=ctk.CTkFont(size=11, weight="bold"),
    ).pack(side="right", padx=18)

    pestañas = ctk.CTkTabview(
        ventana, fg_color=color_fondo, border_width=0, corner_radius=0
    )
    pestañas.pack(fill="both", expand=True, padx=8, pady=(2, 4))
    tab_importar = pestañas.add("Importar Excel")
    tab_manual = pestañas.add("Cargar manual")
    tab_pedidos = pestañas.add("Pedidos")
    tab_seguimiento = pestañas.add("Seguimiento")
    tab_arqueo = pestañas.add("Arqueo")
    tab_historial = pestañas.add("Historial")
    pestañas._segmented_button.grid_forget()
    for fila_oculta in (0, 1, 2):
        pestañas.grid_rowconfigure(fila_oculta, minsize=0, weight=0)

    botones_navegacion = {}

    def seleccionar_pestaña(nombre):
        pestañas.set(nombre)
        for etiqueta, boton in botones_navegacion.items():
            boton.configure(
                fg_color="#EAF3FF" if etiqueta == nombre else "transparent",
                text_color=color_azul if etiqueta == nombre else color_suave,
            )

    for nombre, etiqueta_nav in (
        ("Cargar manual", "▣  Caja diaria"),
        ("Pedidos", "📦  Pedidos"),
        ("Seguimiento", "🚚  Seguimiento"),
        ("Arqueo", "▤  Arqueo"),
        ("Importar Excel", "▣  Importar Excel"),
        ("Historial", "Historial"),
    ):
        if nombre in ("Arqueo", "Importar Excel"):
            continue
        boton = ctk.CTkButton(
            navegacion, text=etiqueta_nav, width=(190 if perfil["nombre"] == "full-hd" else 150), height=(42 if perfil["nombre"] == "full-hd" else 30), corner_radius=0,
            fg_color="transparent", hover_color=color_panel_alto,
            text_color=color_suave, font=ctk.CTkFont(size=(14 if perfil["nombre"] == "full-hd" else 11), weight="bold"),
            command=lambda destino=nombre: seleccionar_pestaña(destino),
        )
        boton.pack(side="left", padx=(16 if nombre == "Cargar manual" else 0, 0), pady=0)
        botones_navegacion[nombre] = boton
    seleccionar_pestaña("Cargar manual")

    boton_administrador = ctk.CTkButton(
        barra_superior, text="Administrador", width=125, height=26,
        fg_color="#FFFFFF", text_color="#0F5FB9", hover_color="#DCEBFF",
        command=lambda: abrir_acceso_administrador(),
    )
    boton_administrador.pack(side="right", padx=(4, 8))

    def mostrar_error(error):
        titulo, detalle = friendly_error(error)
        messagebox.showerror(titulo, detalle, parent=ventana)

    # ---- Importar Excel ----
    estado = {"ruta": ctk.StringVar(), "resultado": None}

    fila_sup = ctk.CTkFrame(tab_importar, fg_color="transparent")
    fila_sup.pack(fill="x", padx=8, pady=8)
    ctk.CTkEntry(
        fila_sup, textvariable=estado["ruta"], placeholder_text="Excel..."
    ).pack(side="left", fill="x", expand=True, padx=(0, 8))

    combo_unidad = ctk.CTkComboBox(fila_sup, values=UNIDADES)
    combo_unidad.set(UNIDAD_POR_DEFECTO)
    combo_unidad.pack(side="left", padx=(0, 8))

    cuadro = ctk.CTkTextbox(tab_importar, fg_color=COLOR_PANEL_SECUNDARIO[1])
    cuadro.pack(fill="both", expand=True, padx=8, pady=(0, 8))

    def mostrar(texto):
        cuadro.configure(state="normal")
        cuadro.delete("1.0", "end")
        cuadro.insert("1.0", texto)
        cuadro.configure(state="disabled")

    def resumen_texto(resultado):
        lineas = [f"Días detectados: {len(resultado['por_dia'])}"]
        for fecha, datos in sorted(resultado["por_dia"].items()):
            t = datos["totales"]
            lineas.append(
                f"- {fecha} ({datos['hoja']}): "
                f"{len(datos['registros'])} registros | "
                f"Total {formatear_monto(t['total'])} | "
                f"Efectivo final {formatear_monto(t['efectivo_final'])}"
            )
        if resultado["duplicados"]:
            lineas.append("")
            lineas.append(
                "Ya cargados (se omiten): "
                + ", ".join(sorted(set(resultado["duplicados"])))
            )
        if resultado["errores"]:
            lineas.append("")
            lineas.append("AVISOS / ERRORES:")
            lineas.extend(f"- {e}" for e in resultado["errores"])
        return "\n".join(lineas)

    def elegir():
        ruta = filedialog.askopenfilename(
            parent=ventana, title="Seleccionar caja diaria",
            filetypes=[("Excel", "*.xlsx")],
        )
        if ruta:
            estado["ruta"].set(ruta)
            analizar()

    def analizar():
        ruta = estado["ruta"].get().strip()
        if not ruta:
            messagebox.showwarning(
                "Falta el archivo", "Seleccioná el Excel.", parent=ventana
            )
            return
        try:
            resultado = analizar_archivo(ruta, combo_unidad.get())
        except Exception as exc:
            messagebox.showerror("No se pudo leer", str(exc), parent=ventana)
            return
        estado["resultado"] = resultado
        mostrar(resumen_texto(resultado))
        boton_importar.configure(
            state="normal"
            if resultado["por_dia"] and not resultado["errores"]
            else "disabled"
        )

    def importar():
        session = estado_admin.get("session")
        if session is None:
            messagebox.showerror(
                "Acceso restringido", "La importación requiere una sesión administrativa.", parent=ventana
            )
            return
        try:
            controller.admin.require(session.token)
        except Exception as exc:
            estado_admin["session"] = None
            mostrar_error(exc)
            return
        resultado = estado["resultado"]
        if not resultado or not resultado["por_dia"]:
            return
        if resultado["errores"]:
            messagebox.showerror(
                "Revisar importación",
                "Corregí los avisos antes de importar. No se guardó ningún dato.",
                parent=ventana,
            )
            return
        if not messagebox.askyesno(
            "Confirmar", f"Se cargarán {len(resultado['por_dia'])} días. "
            "Los datos se guardarán en SQLite. ¿Continuar?",
            parent=ventana,
        ):
            return
        try:
            resumen = controller.import_legacy_analysis(resultado)
            controller.admin.register_import(
                session.token, Path(estado["ruta"].get()), resumen, combo_unidad.get()
            )
        except Exception as exc:
            mostrar_error(exc)
            return
        messagebox.showinfo(
            "Listo",
            f"Se importaron {resumen.entries} registros en {resumen.days} días.",
            parent=ventana,
        )
        boton_importar.configure(state="disabled")

    ctk.CTkButton(
        fila_sup, text="Elegir Excel", command=elegir,
        fg_color=COLOR_PRIMARIO, hover_color=COLOR_PRIMARIO_HOVER,
    ).pack(side="left", padx=(0, 8))
    boton_importar = ctk.CTkButton(
        fila_sup, text="Importar", command=importar, state="disabled",
        fg_color=COLOR_VERDE, hover_color="#12835B",
    )
    boton_importar.pack(side="left")

    def mostrar_panel_administrador(session):
        panel = ctk.CTkToplevel(ventana)
        estado_admin["window"] = panel
        panel.title("Administrador V1")
        panel.geometry("900x640")
        panel.transient(ventana); panel.grab_set()
        ctk.CTkLabel(panel, text=f"ADMINISTRADOR V1   ·   {session.username}",
                     font=ctk.CTkFont(size=18, weight="bold"), text_color="#0F5FB9").pack(anchor="w", padx=18, pady=12)
        tabs = ctk.CTkTabview(panel); tabs.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        sections = {name: tabs.add(name) for name in (
            "Importación de datos", "Usuarios y permisos", "Sucursal y caja", "Responsables",
            "Arqueos", "Notificaciones de cierre", "Auditoría", "Envíos pendientes",
        )}
        ctk.CTkLabel(sections["Importación de datos"], text="Importación protegida por sesión administrativa.",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(36, 12))
        ctk.CTkButton(sections["Importación de datos"], text="Abrir importación de datos",
                      command=lambda: (panel.grab_release(), panel.withdraw(), seleccionar_pestaña("Importar Excel"))).pack()
        ctk.CTkLabel(sections["Usuarios y permisos"], text="Rol ADMIN\nPermisos sensibles validados por el servicio.", justify="left").pack(anchor="w", padx=20, pady=20)
        branch = controller.admin.setting("branch")
        branch_name = ctk.CTkEntry(sections["Sucursal y caja"], placeholder_text="Sucursal", width=280)
        branch_name.insert(0, branch.get("branch", "")); branch_name.pack(pady=(30, 8))
        cashbox_name = ctk.CTkEntry(sections["Sucursal y caja"], placeholder_text="Caja", width=280)
        cashbox_name.insert(0, branch.get("cashbox", "")); cashbox_name.pack(pady=8)
        ctk.CTkButton(sections["Sucursal y caja"], text="Guardar configuración", command=lambda: (
            controller.admin.update_setting(session.token, "branch", {"branch": branch_name.get(), "cashbox": cashbox_name.get()}),
            messagebox.showinfo("Administrador", "Configuración guardada.", parent=panel))).pack(pady=10)
        ctk.CTkLabel(sections["Responsables"], text="Responsables autorizados\nLa identidad canónica queda en la caja y cada arqueo.",
                     justify="center").pack(pady=35)
        counting = controller.admin.setting("counting")
        tolerance_entry = ctk.CTkEntry(sections["Arqueos"], placeholder_text="Tolerancia", width=180)
        tolerance_entry.insert(0, str(counting.get("tolerance", 0))); tolerance_entry.pack(pady=(30, 8))
        limit_entry = ctk.CTkEntry(sections["Arqueos"], placeholder_text="Límite autorización", width=180)
        limit_entry.insert(0, str(counting.get("admin_limit", 0))); limit_entry.pack(pady=8)
        blind_var = ctk.BooleanVar(value=bool(counting.get("blind_close", True)))
        ctk.CTkCheckBox(sections["Arqueos"], text="Conteo ciego de cierre", variable=blind_var).pack(pady=8)
        ctk.CTkButton(sections["Arqueos"], text="Guardar política", command=lambda: (
            controller.admin.update_setting(session.token, "counting", {"blind_close": bool(blind_var.get()),
                "tolerance": int(tolerance_entry.get() or 0), "reason_mode": "ANY_DIFFERENCE",
                "admin_limit": int(limit_entry.get() or 0)}),
            messagebox.showinfo("Administrador", "Política guardada.", parent=panel))).pack(pady=10)
        mail = controller.admin.setting("mail"); mail_section = sections["Notificaciones de cierre"]
        mail_enabled = ctk.BooleanVar(value=bool(mail.get("enabled", False)))
        ctk.CTkCheckBox(mail_section, text="Activar envío al cierre", variable=mail_enabled).pack(pady=(18, 5))
        recipient = ctk.CTkEntry(mail_section, placeholder_text="Destinatario confirmado", width=330); recipient.insert(0, mail.get("recipient", "")); recipient.pack(pady=5)
        host = ctk.CTkEntry(mail_section, placeholder_text="SMTP host", width=330); host.insert(0, mail.get("host", "")); host.pack(pady=5)
        mail_user = ctk.CTkEntry(mail_section, placeholder_text="Usuario SMTP", width=330); mail_user.insert(0, mail.get("username", "")); mail_user.pack(pady=5)
        mail_secret = ctk.CTkEntry(mail_section, placeholder_text="Contraseña de aplicación", show="•", width=330); mail_secret.pack(pady=5)
        def guardar_mail():
            controller.admin.update_setting(session.token, "mail", {"enabled": bool(mail_enabled.get()),
                "recipient": recipient.get().strip(), "cc": [], "subject": "Cierre {fecha} - {sucursal}",
                "host": host.get().strip(), "port": 587, "username": mail_user.get().strip(), "secret_ref": "smtp"})
            if mail_secret.get(): controller.admin.set_mail_secret(session.token, mail_secret.get()); mail_secret.delete(0, "end")
            messagebox.showinfo("Administrador", "Notificaciones guardadas.", parent=panel)
        ctk.CTkButton(mail_section, text="Guardar notificación", command=guardar_mail).pack(pady=6)
        ctk.CTkButton(mail_section, text="Probar envío / reintentar", command=lambda: messagebox.showinfo(
            "Prueba", f"Enviados: {controller.admin.retry_outbox()}", parent=panel)).pack(pady=3)
        audit_box = ctk.CTkTextbox(sections["Auditoría"]); audit_box.pack(fill="both", expand=True, padx=10, pady=10)
        for row in controller.admin.audit_rows(session.token): audit_box.insert("end", f"{row['recorded_at']}  {row['actor']}  {row['action']}  {row['result']}\n")
        audit_box.configure(state="disabled")
        ctk.CTkLabel(sections["Envíos pendientes"], text="Cola persistente: PENDIENTE / ENVIADO / ERROR / NO CONFIGURADO").pack(pady=(35, 12))
        ctk.CTkButton(sections["Envíos pendientes"], text="Reintentar", command=lambda: controller.admin.process_outbox()).pack()

    def abrir_acceso_administrador(configuracion_inicial=False):
        dialog = ctk.CTkToplevel(ventana); first = configuracion_inicial or not controller.admin.has_admin()
        dialog.title("Configuración inicial administrativa" if first else "Acceso Administrador")
        dialog.geometry("430x340" if first else "430x270"); dialog.resizable(False, False); dialog.transient(ventana); dialog.grab_set()
        ctk.CTkLabel(dialog, text="CONFIGURACIÓN INICIAL SEGURA" if first else "ACCESO ADMINISTRADOR",
                     font=ctk.CTkFont(size=17, weight="bold"), text_color="#0F5FB9").pack(pady=(22, 10))
        if first: ctk.CTkLabel(dialog, text="Credencial local con hash seguro. No compartir por chat.").pack(pady=4)
        username_entry = ctk.CTkEntry(dialog, placeholder_text="Administrador", width=300); username_entry.pack(pady=7)
        password_entry = ctk.CTkEntry(dialog, placeholder_text="Contraseña", show="•", width=300); password_entry.pack(pady=7)
        confirm_entry = ctk.CTkEntry(dialog, placeholder_text="Confirmar contraseña", show="•", width=300) if first else None
        if confirm_entry: confirm_entry.pack(pady=7)
        def submit():
            try:
                if first and password_entry.get() != confirm_entry.get(): raise ValueError("Las contraseñas no coinciden.")
                session = (controller.admin.create_initial_admin(username_entry.get(), password_entry.get()) if first
                           else controller.admin.authenticate(username_entry.get(), password_entry.get()))
            except Exception as exc:
                password_entry.delete(0, "end"); messagebox.showerror("Acceso denegado", str(exc), parent=dialog); return
            estado_admin["session"] = session; password_entry.delete(0, "end"); dialog.destroy(); mostrar_panel_administrador(session)
        ctk.CTkButton(dialog, text="Configurar" if first else "Ingresar", command=submit).pack(pady=12)
        password_entry.bind("<Return>", lambda _event: submit()); username_entry.focus_set()

    # ---- Caja operativa (disposición tipo planilla) ----
    campos_manual = {}
    estado_edicion = {"entry_id": None, "guardando": False, "caja_abierta": True}
    items_venta = []
    item_editando = {"index": None}

    cabecera = ctk.CTkFrame(tab_manual, fg_color=color_panel, corner_radius=7)
    cabecera.pack(fill="x", padx=4, pady=(2, 2))
    cabecera.grid_propagate(False)
    cabecera.grid_columnconfigure(8, weight=1)
    # RC18: el rotulo de seccion y las etiquetas de contexto quedan en un
    # segundo plano tipografico para que los importes dominen la lectura.
    fuente_chrome_cabecera = perfil["fuente_label"]
    ctk.CTkLabel(
        cabecera, text="RESUMEN DE CAJA", text_color=COLOR_TEXTO_SUAVE,
        font=ctk.CTkFont(size=fuente_chrome_cabecera, weight="bold")
    ).grid(row=0, column=0, sticky="w", padx=(10, 8), pady=4)

    controles_cabecera = [
        ("fecha", "Fecha", 130),
        ("unidad", "Sucursal", 130),
        ("caja_inicial", "Caja inicial", 150),
    ]
    for indice, (clave, etiqueta, ancho) in enumerate(controles_cabecera):
        ctk.CTkLabel(
            cabecera, text=etiqueta, text_color=COLOR_TEXTO_SUAVE,
            font=ctk.CTkFont(size=fuente_chrome_cabecera, weight="bold"),
        ).grid(
            row=0, column=indice * 2 + 1, sticky="w", padx=(4, 3), pady=4
        )
        if clave == "unidad":
            campo = ctk.CTkComboBox(cabecera, values=UNIDADES, width=ancho, height=max(27, perfil["campo_alto"]))
            campo.set(UNIDAD_POR_DEFECTO)
        else:
            campo = ctk.CTkEntry(cabecera, width=ancho, height=max(27, perfil["campo_alto"]))
        campo.grid(row=0, column=indice * 2 + 2, padx=(0, 8), pady=4)
        campos_manual[clave] = campo
    campos_manual["fecha"].insert(0, date.today().strftime("%d-%m-%Y"))
    campos_manual["caja_inicial"].bind("<Key>", lambda _event: "break")

    aviso_entregas = ctk.CTkButton(
        cabecera, text="Trabajos a entregar: 0", width=150,
        height=max(27, perfil["campo_alto"]),
        font=ctk.CTkFont(size=fuente_chrome_cabecera, weight="bold"),
        fg_color="#FFF3CD", text_color="#7A4B00", border_width=1,
        border_color="#E6B85C", hover_color="#FFE5A3",
        command=lambda: abrir_pedidos_desde_alerta(),
    )
    aviso_entregas.grid(row=0, column=7, sticky="w", padx=(2, 6), pady=4)

    columnas_operativas = COLUMNAS_OPERATIVAS

    formulario = ctk.CTkFrame(tab_manual, fg_color="transparent", corner_radius=0)
    columnas_bloque = 3
    for columna in range(columnas_bloque):
        formulario.grid_columnconfigure(columna, weight=1, uniform="bloques_operador")

    secciones_formulario = (
        ("1", "CLIENTE Y COMPROBANTE", (
            ("descripcion", "Cliente", 185), ("cliente_documento", "CI / RUC", 100),
            ("cliente_telefono", "Teléfono", 140), ("sobre", "Sobre / Trabajo", 100),
            ("fecha_entrega", "Fecha de entrega", 150), ("vendedora", "Vendedora *", 140),
        )),
        ("2", "DETALLE DE VENTA", PRODUCTO_TRABAJO),
        ("3", "PAGO", COBRO_PAGO),
    )
    secciones_widgets = {}
    for indice_seccion, (numero, titulo, columnas) in enumerate(secciones_formulario):
        seccion = ctk.CTkFrame(
            formulario, fg_color="#F7FAFF", corner_radius=10,
            border_width=2, border_color="#8FB3D9",
        )
        secciones_widgets[titulo] = seccion
        seccion.grid_propagate(False)
        fila_bloque = 0
        columna_bloque = indice_seccion
        seccion.grid(row=fila_bloque, column=columna_bloque, sticky="nsew", padx=6, pady=2)
        formulario.grid_rowconfigure(fila_bloque, weight=1)
        seccion.grid_columnconfigure(0, weight=0, minsize=(116 if perfil["nombre"] == "full-hd" else 72))
        seccion.grid_columnconfigure(1, weight=1)
        es_pago = titulo == "PAGO"
        es_detalle = titulo == "DETALLE DE VENTA"
        if es_pago or es_detalle:
            seccion.grid_columnconfigure(2, weight=0, minsize=(116 if perfil["nombre"] == "full-hd" else 72))
            seccion.grid_columnconfigure(3, weight=1)
        ctk.CTkLabel(
            seccion, text=numero, width=18, height=18, corner_radius=9,
            fg_color=color_azul, text_color="#FFFFFF",
            font=ctk.CTkFont(size=10, weight="bold"),
        ).grid(row=0, column=0, sticky="w", padx=(10, 0), pady=(8, 4))
        ctk.CTkLabel(
            seccion, text=titulo, height=22, text_color=color_texto, anchor="w",
            font=ctk.CTkFont(size=perfil["fuente_seccion"], weight="bold"),
        ).grid(row=0, column=0, columnspan=2, sticky="w", padx=(38, 10), pady=(8, 4))
        for indice_campo, (clave, etiqueta, ancho) in enumerate(columnas):
            if es_pago:
                posiciones = {
                    "efectivo": (1, 0), "transferencia": (1, 2),
                    "tarjeta_cheque": (2, 0),
                    "ordenes": (3, 0), "monto_convenio": (3, 2),
                    "cuotas": (4, 0), "total": (5, 0), "saldo": (5, 2),
                }
                fila_campo, columna_etiqueta = posiciones[clave]
                columna_campo = columna_etiqueta + 1
                expansion = 3 if clave in ("tarjeta_cheque", "cuotas") else 1
            elif es_detalle:
                posiciones = {
                    "arm_org": (1, 0), "cod": (2, 0), "laboratorio": (3, 0),
                    "armazon": (4, 0), "cristal": (4, 3), "receta_dr": (5, 0),
                }
                fila_campo, columna_etiqueta = posiciones[clave]
                columna_campo = columna_etiqueta + 1
                expansion = 5 if clave in ("arm_org", "cod", "laboratorio", "receta_dr") else 1
            else:
                fila_campo = indice_campo + 1
                columna_etiqueta = 0
                columna_campo = 1
                expansion = 1
            ctk.CTkLabel(
                seccion, text=etiqueta, height=perfil["campo_alto"], text_color=color_suave, anchor="w",
                font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
            ).grid(row=fila_campo, column=columna_etiqueta, sticky="ew", padx=(10, 5), pady=2)
            if clave == "vendedora":
                campo = ctk.CTkComboBox(
                    seccion, values=["Seleccionar...", "Ana", "Belén", "Carla", "Diana"],
                    width=(ancho if perfil["nombre"] == "full-hd" else min(ancho, 88)), height=perfil["campo_alto"], fg_color="#FFFFFF",
                    border_color=color_borde_suave,
                )
                campo.set("Seleccionar...")
            else:
                campo = ctk.CTkEntry(
                    seccion, width=(ancho if perfil["nombre"] == "full-hd" else min(ancho, 88)), height=perfil["campo_alto"], fg_color="#FFFFFF",
                    border_width=1, border_color=color_borde_suave, text_color=color_texto,
                    font=ctk.CTkFont(size=perfil["fuente"]),
                )
            campo.grid(
                row=fila_campo, column=columna_campo, columnspan=expansion,
                sticky="ew", padx=(0, 10), pady=2,
            )
            campos_manual[clave] = campo

    detalle_venta = secciones_widgets["DETALLE DE VENTA"]
    detalle_venta.grid_columnconfigure(2, weight=0, minsize=44)
    detalle_venta.grid_columnconfigure(3, weight=0, minsize=72)
    detalle_venta.grid_columnconfigure(4, weight=1)
    detalle_venta.grid_columnconfigure(5, weight=0, minsize=44)
    for clave, columna in (("descuento_armazon", 2), ("descuento_cristal", 5)):
        campos_manual[clave] = ctk.CTkEntry(
            detalle_venta, width=42, height=perfil["campo_alto"],
            placeholder_text="0", justify="center", fg_color="#FFF8E7",
            border_width=1, border_color="#D6A84B", text_color=color_texto,
            font=ctk.CTkFont(size=perfil["fuente"]),
        )
        campos_manual[clave].grid(row=4, column=columna, sticky="e", padx=(2, 8), pady=2)
    campos_manual["sin_costo"] = ctk.CTkCheckBox(
        detalle_venta, text="Artículo sin costo", onvalue="1", offvalue="0",
        height=perfil["campo_alto"], text_color=color_texto,
        font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
    )
    campos_manual["sin_costo"].grid(
        row=7, column=2, columnspan=4, sticky="e", padx=10, pady=(3, 7)
    )
    for clave in ("total", "saldo"):
        campos_manual[clave].configure(
            fg_color="#E7F1FC", border_color="#7DA9D7", text_color="#0F5FB9",
            font=ctk.CTkFont(size=perfil["fuente"], weight="bold"),
        )
        campos_manual[clave].bind("<Key>", lambda _event: "break")
    campos_manual["fecha_entrega"].configure(placeholder_text="dd-mm-aaaa")

    def abrir_selector_fecha_entrega():
        selector = ctk.CTkToplevel(ventana)
        selector.title("Seleccionar fecha de entrega")
        selector.resizable(False, False)
        hoy = date.today()
        estado_mes = {"año": hoy.year, "mes": hoy.month}
        cuerpo_calendario = ctk.CTkFrame(selector, fg_color="#FFFFFF")
        cuerpo_calendario.pack(fill="both", expand=True, padx=10, pady=10)

        def elegir(dia):
            valor = date(estado_mes["año"], estado_mes["mes"], dia)
            campos_manual["fecha_entrega"].delete(0, "end")
            campos_manual["fecha_entrega"].insert(0, valor.strftime("%d-%m-%Y"))
            selector.destroy()

        def renderizar(delta=0):
            mes = estado_mes["mes"] + delta
            año = estado_mes["año"]
            if mes < 1: año, mes = año - 1, 12
            if mes > 12: año, mes = año + 1, 1
            estado_mes.update(año=año, mes=mes)
            for widget in cuerpo_calendario.winfo_children(): widget.destroy()
            ctk.CTkButton(cuerpo_calendario, text="‹", width=32, command=lambda: renderizar(-1)).grid(row=0, column=0)
            ctk.CTkLabel(cuerpo_calendario, text=f"{calendar.month_name[mes].capitalize()} {año}", width=190,
                         font=ctk.CTkFont(weight="bold")).grid(row=0, column=1, columnspan=5)
            ctk.CTkButton(cuerpo_calendario, text="›", width=32, command=lambda: renderizar(1)).grid(row=0, column=6)
            for columna, nombre in enumerate(("Lu", "Ma", "Mi", "Ju", "Vi", "Sá", "Do")):
                ctk.CTkLabel(cuerpo_calendario, text=nombre, width=34).grid(row=1, column=columna, pady=3)
            for fila, semana in enumerate(calendar.monthcalendar(año, mes), start=2):
                for columna, dia in enumerate(semana):
                    if dia:
                        ctk.CTkButton(cuerpo_calendario, text=str(dia), width=34, height=28,
                                      command=lambda valor=dia: elegir(valor)).grid(row=fila, column=columna, padx=1, pady=1)
        renderizar()
        selector.transient(ventana)
        selector.grab_set()

    campo_fecha_entrega = campos_manual["fecha_entrega"]
    info_fecha_entrega = campo_fecha_entrega.grid_info()
    campo_fecha_entrega.grid_configure(padx=(0, 42))
    ctk.CTkButton(
        campo_fecha_entrega.master, text="📅", width=32, height=perfil["campo_alto"],
        command=abrir_selector_fecha_entrega, fg_color="#E7F1FC", text_color="#0F5FB9",
    ).grid(row=info_fecha_entrega["row"], column=info_fecha_entrega["column"], sticky="e", padx=(0, 10), pady=2)

    zona_secundaria = ctk.CTkFrame(tab_manual, fg_color="#F8FAFD", corner_radius=7,
                                   border_width=1, border_color=color_borde_suave)
    ctk.CTkLabel(zona_secundaria, text="SALIDA DE CAJA", text_color=color_suave,
                 font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold")).pack(side="left", padx=(8, 5))
    # La autoría no se solicita dos veces: el servicio registra al responsable
    # canónico de la caja/sesión en la auditoría de la salida.
    campos_manual["salida_tipo"] = ctk.CTkComboBox(
        zona_secundaria, values=["Gasto", "Entrega administración"], width=145,
        height=perfil["campo_alto"],
    )
    campos_manual["salida_tipo"].set("Gasto")
    campos_manual["salida_tipo"].pack(side="left", padx=2, pady=4)
    for clave, ancho, placeholder in (
        ("salida_concepto", 190, "Concepto"), ("salida_monto", 105, "Monto"),
        ("salida_observacion", 190, "Observación"),
    ):
        campos_manual[clave] = ctk.CTkEntry(
            zona_secundaria, width=ancho, height=perfil["campo_alto"],
            placeholder_text=placeholder,
        )
        campos_manual[clave].pack(side="left", padx=2, pady=4)
    campos_manual["accion_salida"] = ctk.CTkButton(
        zona_secundaria, text="Guardar salida", width=105, height=perfil["campo_alto"],
        fg_color="#FFF7ED", text_color="#B45309", border_width=1,
        border_color="#F2C69F", hover_color="#FFEBD7",
    )
    campos_manual["accion_salida"].pack(side="left", padx=2, pady=4)

    # El draft es una zona propia: nunca cuenta como movimiento persistido.
    lista_productos = ctk.CTkFrame(tab_manual, fg_color="#F7FAFF", corner_radius=9,
                                   border_width=2, border_color="#8FB3D9")
    ctk.CTkLabel(lista_productos, text="VENTA EN CURSO", text_color=color_texto,
                 font=ctk.CTkFont(size=perfil["fuente_seccion"], weight="bold")).place(x=12, y=6)
    cuerpo_draft = ctk.CTkFrame(lista_productos, fg_color="transparent")
    cuerpo_draft.pack(fill="both", expand=True, padx=8, pady=(28, 6))
    cuerpo_draft.grid_columnconfigure(0, weight=1)
    cuerpo_draft.grid_rowconfigure(0, weight=1)
    panel_items = ctk.CTkFrame(cuerpo_draft, fg_color="transparent")
    panel_items.grid(row=0, column=0, sticky="nsew")
    panel_items.grid_rowconfigure(0, weight=1)
    panel_items.grid_columnconfigure(0, weight=1)
    columnas_items = ("producto", "codigo", "tipo", "armazon", "cristal", "subtotal")
    grilla_items = ttk.Treeview(
        panel_items, columns=columnas_items, show="headings", height=2,
        style="Draft.Treeview",
    )
    for clave, titulo, ancho in (
        ("producto", "Artículo", 160), ("codigo", "Código", 72),
        ("tipo", "Tipo", 82), ("armazon", "P. Armazón", 84),
        ("cristal", "P. Cristal", 84), ("subtotal", "Subtotal", 90),
    ):
        anchor = "w" if clave in ("producto", "tipo") else "center"
        grilla_items.heading(clave, text=titulo, anchor=anchor)
        grilla_items.column(
            clave, width=ancho, minwidth=70, anchor=anchor,
            stretch=clave == "producto",
        )
    scroll_items = ttk.Scrollbar(panel_items, orient="vertical", command=grilla_items.yview)
    grilla_items.configure(yscrollcommand=scroll_items.set)
    grilla_items.grid(row=0, column=0, sticky="nsew")
    scroll_items.grid(row=0, column=1, sticky="ns")
    acciones_item = ctk.CTkFrame(lista_productos, fg_color="transparent")
    acciones_item.place(relx=0.98, y=3, anchor="ne")
    panel_total_draft = ctk.CTkFrame(
        tab_manual, fg_color="#E7F1FC", corner_radius=8,
        border_width=1, border_color="#7DA9D7",
    )
    panel_total_draft.pack_propagate(False)
    panel_total_draft.place_forget()
    ctk.CTkLabel(
        panel_total_draft, text="OBSERVACIONES", text_color="#42627F",
        font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
    ).pack(anchor="w", padx=10, pady=(7, 2))
    campos_manual["notas"] = ctk.CTkTextbox(
        panel_total_draft, fg_color="#FFFFFF", border_width=1,
        border_color="#8FB3D9", wrap="word", font=ctk.CTkFont(size=perfil["fuente"]),
    )
    campos_manual["notas"].pack(fill="both", expand=True, padx=8, pady=(0, 8))
    campos_manual["notas"].insert("1.0", PLANTILLA_RECETA_OBSERVACIONES)
    total_draft_var = ctk.StringVar(value="0")

    bloque_producto = formulario
    columna_guardar = None
    claves_operacion = tuple(
        clave for clave in (
            "descripcion", "cliente_telefono", "cliente_documento", "sobre",
            "fecha_entrega", "vendedora", "arm_org", "cod", "laboratorio",
            "armazon", "cristal", "descuento_armazon", "descuento_cristal",
            "sin_costo", "receta_dr", "total", "efectivo",
            "tarjeta_cheque", "transferencia", "ordenes", "monto_convenio",
            "cuotas", "saldo", "notas",
        ) if clave in campos_manual
    )
    orden_teclado = list(claves_operacion)
    for indice, clave in enumerate(orden_teclado[:-1]):
        siguiente = orden_teclado[indice + 1]
        campos_manual[clave].bind(
            "<Return>", lambda _event, destino=siguiente: campos_manual[destino].focus_set()
        )

    estado_operativo = ctk.CTkFrame(cabecera, fg_color="transparent")
    estado_operativo.grid(row=0, column=9, sticky="e", padx=4, pady=4)

    # RC18: los controles de estado acompanan la escala del perfil en vez de
    # quedar fijos en 20/24 px, que a 24" se leian como texto residual.
    fuente_estado = perfil["fuente_label"]
    alto_estado = max(20, perfil["campo_alto"] - 12)
    alto_boton_estado = max(24, perfil["campo_alto"] - 6)
    estado_caja = ctk.CTkLabel(
        estado_operativo, text="SIN CONSULTAR", width=120, height=alto_estado,
        corner_radius=5,
        fg_color=color_panel_alto, text_color=COLOR_TEXTO_SUAVE,
        font=ctk.CTkFont(size=fuente_estado, weight="bold"),
    )
    estado_caja.pack(side="left", padx=(0, 5))
    boton_cerrar_caja = ctk.CTkButton(
        estado_operativo, text="Cerrar caja", width=92, height=alto_boton_estado,
        font=ctk.CTkFont(size=fuente_estado, weight="bold"),
        fg_color="#B42318", hover_color="#8F1C13", state="disabled",
        command=lambda: cerrar_caja(),
    )
    boton_cerrar_caja.pack(side="left")
    boton_arqueo = ctk.CTkButton(
        estado_operativo, text="Arqueo", width=78, height=alto_boton_estado,
        font=ctk.CTkFont(size=fuente_estado, weight="bold"),
        fg_color="#0F5FB9", hover_color="#0B4D98",
        command=lambda: abrir_modal_arqueo(),
    )
    boton_arqueo.pack(side="left", padx=(5, 0))
    # RC18: jerarquia visual de KPI. Los importes que gobiernan la operacion
    # (Venta, Efectivo, Esperado) se leen a distancia en ViewSonic 24"; el resto
    # se agrupa como informacion secundaria. No cambia ningun calculo.
    resumen_compacto = ctk.CTkFrame(cabecera, fg_color="transparent")
    resumen_compacto.grid(row=0, column=8, sticky="e", padx=2, pady=2)
    etiquetas_kpi = {}
    metricas_kpi = metricas_resumen_kpi(perfil)
    fuente_kpi_principal = metricas_kpi["fuente_principal"]
    fuente_kpi_secundaria = metricas_kpi["fuente_secundaria"]
    fuente_kpi_titulo = metricas_kpi["fuente_titulo"]

    def agregar_kpi(padre, clave, titulo, color, principal):
        indicador = ctk.CTkFrame(
            padre, fg_color="#FFFFFF" if principal else "#F1F5FA", corner_radius=5,
        )
        indicador.pack(side="left", padx=(0, 3 if principal else 2))
        # CTkLabel pide 28 px de alto por defecto, sin relacion con la fuente:
        # sin altura explicita la tarjeta ocupaba 59 px en todos los perfiles.
        tam_titulo = fuente_kpi_titulo if principal else max(8, fuente_kpi_titulo - 1)
        tam_valor = fuente_kpi_principal if principal else fuente_kpi_secundaria
        ctk.CTkLabel(
            indicador, text=titulo, text_color=color_suave,
            height=tam_titulo + 6,
            font=ctk.CTkFont(size=tam_titulo, weight="bold"),
        ).pack(padx=6 if principal else 4, pady=(1, 0))
        valor = ctk.CTkLabel(
            indicador, text="—", text_color=color,
            height=tam_valor + 8,
            font=ctk.CTkFont(size=tam_valor, weight="bold"),
        )
        valor.pack(padx=6 if principal else 4, pady=(0, 2 if principal else 1))
        etiquetas_kpi[clave] = valor

    bloque_kpi_principal = ctk.CTkFrame(resumen_compacto, fg_color="transparent")
    bloque_kpi_principal.pack(side="left")
    for clave, titulo, color in KPI_PRINCIPALES:
        agregar_kpi(bloque_kpi_principal, clave, titulo, color, True)

    # height=2 explicito: CTkFrame pide 200 px por defecto y, con fill="y",
    # ese pedido estiraba la fila de la cabecera y desplazaba los importes.
    ctk.CTkFrame(resumen_compacto, width=1, height=2, fg_color="#D8E3EF").pack(
        side="left", fill="y", padx=(3, 4), pady=3,
    )

    bloque_kpi_secundario = ctk.CTkFrame(resumen_compacto, fg_color="transparent")
    bloque_kpi_secundario.pack(side="left")
    for clave, titulo, color in KPI_SECUNDARIOS:
        agregar_kpi(bloque_kpi_secundario, clave, titulo, color, False)
    def formatear_campo_monetario(clave):
        campo = campos_manual[clave]
        texto = formatear_importe_ui(campo.get())
        if texto != campo.get():
            escribir_importe_formateado(campo, texto)

    def recalcular_total_visible(_event=None):
        if items_venta:
            total = sum(item.subtotal for item in items_venta)
            total_draft_var.set(formatear_monto(total))
            campo_total = campos_manual["total"]
            campo_total.delete(0, "end")
            campo_total.insert(0, formatear_importe_ui(total))
            recalcular_saldo_visible()
            return
        try:
            visible = SaleItem(
                frame_price=campos_manual["armazon"].get(),
                lens_price=campos_manual["cristal"].get(),
                frame_discount_percent=campos_manual["descuento_armazon"].get(),
                lens_discount_percent=campos_manual["descuento_cristal"].get(),
                no_cost=campos_manual["sin_costo"].get() == "1",
            )
            total = visible.subtotal
        except (TypeError, ValueError):
            return
        total_draft_var.set(formatear_monto(total))
        campo_total = campos_manual["total"]
        campo_total.delete(0, "end")
        campo_total.insert(0, formatear_importe_ui(total))
        recalcular_saldo_visible()

    def refrescar_items():
        for row in grilla_items.get_children():
            grilla_items.delete(row)
        for index, item in enumerate(items_venta):
            grilla_items.insert("", "end", iid=str(index), values=(
                item.description, item.code,
                ("SIN COSTO" if item.no_cost else item.item_type),
                formatear_monto(item.frame_final_price),
                formatear_monto(item.lens_final_price),
                formatear_monto(item.subtotal),
            ))
        recalcular_total_visible()
        try:
            refrescar_grilla(controller.load_day(
                campos_manual["fecha"].get().strip(), campos_manual["unidad"].get().strip()
            ))
        except (NameError, Exception):
            pass

    def agregar_producto():
        try:
            item = construir_item_producto_visible(
                {clave: campos_manual[clave].get() for clave in (
                    "arm_org", "cod", "armazon", "cristal", "descuento_armazon",
                    "descuento_cristal", "sin_costo", "laboratorio", "receta_dr"
                )}
            )
            if not item.no_cost and item.reference_subtotal <= 0:
                raise ValueError("El producto debe tener un precio de armazón o cristal.")
        except Exception as exc:
            messagebox.showwarning("Producto inválido", str(exc), parent=ventana)
            return
        index = item_editando["index"]
        if index is None:
            items_venta.append(item)
        else:
            items_venta[index] = replace(item, id=items_venta[index].id)
            item_editando["index"] = None
        for clave in ("arm_org", "cod", "armazon", "cristal",
                      "descuento_armazon", "descuento_cristal",
                      "laboratorio", "receta_dr"):
            campos_manual[clave].delete(0, "end")
        campos_manual["sin_costo"].deselect()
        refrescar_items()

    def editar_item():
        selected = grilla_items.selection()
        if not selected:
            return
        index = int(selected[0]); item = items_venta[index]; item_editando["index"] = index
        values = {"arm_org": item.item_type, "cod": item.code, "armazon": item.frame_price,
                  "cristal": item.lens_price,
                  "descuento_armazon": item.frame_discount_percent,
                  "descuento_cristal": item.lens_discount_percent,
                  "laboratorio": item.laboratory,
                  "receta_dr": item.prescription_doctor}
        for clave, value in values.items():
            campos_manual[clave].delete(0, "end"); campos_manual[clave].insert(0, "" if value is None else str(value))
        campos_manual["sin_costo"].select() if item.no_cost else campos_manual["sin_costo"].deselect()

    def quitar_item():
        selected = grilla_items.selection()
        if selected:
            items_venta.pop(int(selected[0])); item_editando["index"] = None; refrescar_items()

    detalle_venta = secciones_widgets["DETALLE DE VENTA"]
    boton_agregar_articulo = ctk.CTkButton(
        detalle_venta, text="+ Agregar artículo", height=perfil["campo_alto"],
        command=agregar_producto, fg_color="#1672E8", hover_color="#0F5FC7",
        font=ctk.CTkFont(size=perfil["fuente"], weight="bold"),
    )
    boton_agregar_articulo.grid(
        row=7, column=0, columnspan=2, sticky="ew", padx=10, pady=(3, 7)
    )
    ctk.CTkButton(
        acciones_item, text="Editar artículo seleccionado", width=165, height=22,
        command=editar_item,
    ).pack(side="left", padx=(0, 3))
    ctk.CTkButton(
        acciones_item, text="Quitar", width=70, height=22,
        command=quitar_item, fg_color="#D9534F",
    ).pack(side="left", padx=1)

    def recalcular_saldo_visible(_event=None):
        try:
            saldo = calcular_saldo_pendiente(
                campos_manual["total"].get(), campos_manual["efectivo"].get(),
                campos_manual["tarjeta_cheque"].get(), campos_manual["transferencia"].get(),
                campos_manual["monto_convenio"].get(),
            )
        except (TypeError, ValueError):
            return
        campo = campos_manual["saldo"]
        campo.delete(0, "end")
        campo.insert(0, formatear_importe_ui(saldo))

    for clave in CAMPOS_MONETARIOS_UI:
        campos_manual[clave].bind(
            "<FocusOut>", lambda _event, campo=clave: formatear_campo_monetario(campo), add="+"
        )
    for clave in ("armazon", "cristal", "descuento_armazon", "descuento_cristal"):
        campos_manual[clave].bind("<KeyRelease>", recalcular_total_visible, add="+")
        campos_manual[clave].bind("<FocusOut>", recalcular_total_visible, add="+")
    campos_manual["sin_costo"].configure(command=recalcular_total_visible)
    for clave in ("total", "efectivo", "tarjeta_cheque", "transferencia", "monto_convenio"):
        campos_manual[clave].bind("<KeyRelease>", recalcular_saldo_visible, add="+")
        campos_manual[clave].bind("<FocusOut>", recalcular_saldo_visible, add="+")
    filtro_movimientos = ctk.StringVar(value="Todos")
    busqueda_movimientos = ctk.StringVar()
    toolbar_movimientos = ctk.CTkFrame(
        tab_manual, fg_color=color_panel, corner_radius=8,
        border_width=1, border_color=color_borde_suave,
    )
    ctk.CTkLabel(
        toolbar_movimientos, text="Movimientos del día", text_color=color_texto,
        font=ctk.CTkFont(size=(18 if perfil["nombre"] == "full-hd" else 14), weight="bold"),
    ).pack(side="left", padx=12)
    entrada_busqueda = ctk.CTkEntry(
        toolbar_movimientos, textvariable=busqueda_movimientos,
        placeholder_text="Buscar movimiento…", width=(290 if perfil["nombre"] == "full-hd" else 205), height=(38 if perfil["nombre"] == "full-hd" else 28),
        fg_color="#F7F9FC", border_color=color_borde_suave,
    )
    entrada_busqueda.pack(side="left", padx=(8, 6), pady=6)
    botones_filtro = {}
    for nombre_filtro in ("Todos", "Ventas", "Salidas", "Pendientes"):
        boton_filtro = ctk.CTkButton(
            toolbar_movimientos, text=nombre_filtro, width=(82 if perfil["nombre"] == "full-hd" else 68), height=(38 if perfil["nombre"] == "full-hd" else 28),
            corner_radius=4, fg_color="#EAF3FF" if nombre_filtro == "Todos" else "transparent",
            text_color=color_azul if nombre_filtro == "Todos" else color_suave,
            border_width=1, border_color=color_borde_suave,
            command=lambda valor=nombre_filtro: aplicar_filtro_movimientos(valor),
        )
        boton_filtro.pack(side="left", padx=1, pady=6)
        botones_filtro[nombre_filtro] = boton_filtro
    marco_grilla = ctk.CTkFrame(tab_manual, fg_color=color_panel, corner_radius=5)
    marco_grilla.pack(fill="both", expand=True, padx=6, pady=3)
    estilo = ttk.Style(ventana)
    estilo.theme_use("clam")
    estilo.configure(
        "Caja.Treeview", rowheight=perfil["fila"], font=("Segoe UI", perfil["fuente"]),
        background="#FFFFFF", fieldbackground="#FFFFFF", foreground="#24324A",
        borderwidth=0, relief="flat",
    )
    estilo.map(
        "Caja.Treeview",
        background=[("selected", "#245DA8")],
        foreground=[("selected", "#FFFFFF")],
    )
    estilo.configure(
        "Caja.Treeview.Heading", font=("Segoe UI", perfil["fuente"], "bold"),
        background="#EDF3FA", foreground="#33425B", relief="flat", padding=(4, 5),
    )
    estilo.map("Caja.Treeview.Heading", background=[("active", "#245DA8")])
    estilo.configure(
        "Draft.Treeview", rowheight=(27 if perfil["nombre"] == "full-hd" else 22),
        font=("Segoe UI", perfil["fuente"]), background="#FFFFFF",
        fieldbackground="#FFFFFF", foreground="#24324A", borderwidth=0,
    )
    estilo.configure(
        "Draft.Treeview.Heading", font=("Segoe UI", perfil["fuente"], "bold"),
        background="#EDF3FA", foreground="#33425B", relief="flat", padding=(3, 2),
    )
    claves_grilla = [clave for clave, _, _ in columnas_operativas] + ["acciones"]
    grilla_caja = ttk.Treeview(
        marco_grilla, columns=claves_grilla, show="headings", style="Caja.Treeview"
    )
    for clave, etiqueta, ancho, anchor in MOVEMENT_COLUMN_SPECS:
        grilla_caja.heading(clave, text=etiqueta, anchor=anchor)
        ancho_perfil = int(ancho * (1.24 if perfil["nombre"] == "full-hd" else 1.0))
        grilla_caja.column(
            clave, width=ancho_perfil, minwidth=65, stretch=False, anchor=anchor
        )
    grilla_caja.heading("acciones", text="Acciones", anchor="center")
    grilla_caja.column("acciones", width=105, minwidth=105, stretch=False, anchor="center")
    grilla_caja.tag_configure("voided", foreground="#E0717C", background="#241824")
    grilla_caja.tag_configure("expense", foreground="#F5A3AA")
    grilla_caja.tag_configure("pending", foreground="#3B2A05", background="#FFF3CD")
    grilla_caja.tag_configure("draft", foreground="#174A7E", background="#EAF3FF")
    grilla_caja.tag_configure("draft_item", foreground="#42627F", background="#F5F9FE")
    scroll_horizontal = ttk.Scrollbar(
        marco_grilla, orient="horizontal", command=grilla_caja.xview
    )
    scroll_vertical = ttk.Scrollbar(
        marco_grilla, orient="vertical", command=grilla_caja.yview
    )
    grilla_caja.configure(
        xscrollcommand=scroll_horizontal.set, yscrollcommand=scroll_vertical.set
    )

    def ajustar_columnas_movimientos(ancho_disponible):
        """Distribuye el excedente sin regalarlo a Cliente ni forzar scroll."""
        bases = {clave: ancho for clave, _titulo, ancho, _anchor in MOVEMENT_COLUMN_SPECS}
        bases["acciones"] = 105
        maximos = {clave: int(ancho * 1.18) for clave, ancho in bases.items()}
        disponible = max(sum(bases.values()), int(ancho_disponible) - 22)
        extra = max(0, disponible - sum(bases.values()))
        capacidad = sum(maximos[k] - bases[k] for k in bases)
        for clave, base in bases.items():
            adicional = min(maximos[clave] - base, round(extra * (maximos[clave] - base) / capacidad)) if capacidad else 0
            grilla_caja.column(clave, width=base + adicional, minwidth=base, stretch=False)
        total = sum(int(grilla_caja.column(clave, "width")) for clave in bases)
        if total <= disponible:
            scroll_horizontal.grid_remove()
        else:
            scroll_horizontal.grid()
    grilla_caja.grid(row=0, column=0, sticky="nsew")
    scroll_vertical.grid(row=0, column=1, sticky="ns")
    scroll_horizontal.grid(row=1, column=0, sticky="ew")
    marco_grilla.grid_rowconfigure(0, weight=1)
    marco_grilla.grid_columnconfigure(0, weight=1)
    # RC29: las tres partes de la jornada se calculan por separado para que
    # Historial pueda darles jerarquia visual distinta. `texto_estado` las
    # compone exactamente como antes, asi que el aviso de cierre no cambia una
    # coma y ninguna cifra se recalcula en dos lugares.
    def estado_dia(cash_day):
        return "ABIERTO" if cash_day.status.value == "OPEN" else "CERRADO"

    def resumen_economico_dia(cash_day):
        totales = cash_day.totals()
        return (
            f"Efectivo actual  {formatear_monto(totales.expected_cash)}    "
            f"Total ventas  {formatear_monto(totales.total)}\n"
            f"Gastos  {formatear_monto(totales.expenses)}    "
            f"Entregado a administración  {formatear_monto(totales.withdrawals)}    "
            f"Efectivo final  {formatear_monto(totales.expected_cash)}"
        )

    def detalle_sesion_dia(cash_day):
        """Apertura, cierre, duracion y hora extra. Vacio si sigue abierta."""
        if cash_day.closed_at is None or cash_day.session_duration_seconds is None:
            return ""
        apertura = cash_day.opened_at.astimezone(BUSINESS_TIMEZONE).strftime("%H:%M:%S")
        cierre = cash_day.closed_at.astimezone(BUSINESS_TIMEZONE).strftime("%H:%M:%S")
        horas, resto = divmod(cash_day.session_duration_seconds, 3600)
        minutos = resto // 60
        if cash_day.overtime_triggered is None:
            extra = "Hora extra: política pendiente para este día"
        elif cash_day.overtime_triggered:
            extra = f"Hora extra: SÍ · {cash_day.overtime_minutes} min"
        else:
            extra = "Hora extra: NO"
        return (
            f"Apertura real  {apertura}    Cierre real  {cierre}\n"
            f"Duración  {horas:02d}:{minutos:02d}    {extra}"
        )

    def texto_estado(cash_day):
        texto = f"{estado_dia(cash_day)}   ·   {resumen_economico_dia(cash_day)}"
        detalle = detalle_sesion_dia(cash_day)
        return f"{texto}\n\n{detalle}" if detalle else texto

    def tiene_saldo_cliente(entry):
        return entry.client_balance_amount > 0

    def valores_fila(entry):
        importe = lambda value: formatear_monto(value or 0)
        item_count = len(entry.effective_items)
        resumen = f"{item_count} producto{'s' if item_count != 1 else ''}"
        if entry.outflow_type == "ENTREGA_ADMINISTRACION":
            resumen = "Entrega administración"
        elif entry.outflow_type == "GASTO" or entry.expenses:
            resumen = f"Gasto · {entry.description}"
        estado_fila = "ANULADO" if entry.status.value == "VOIDED" else "SALIDA" if entry.outflow_type else "PENDIENTE" if tiene_saldo_cliente(entry) else "COBRADO"
        total_visible = (entry.expenses or entry.withdrawal or 0) if entry.outflow_type else (entry.total or 0)
        return (
            entry.created_at.astimezone(BUSINESS_TIMEZONE).strftime("%H:%M"),
            entry.description, entry.customer_phone, resumen, entry.envelope,
            importe(total_visible), importe(entry.cash), importe(entry.card_check),
            importe(entry.agreement_amount), entry.installments,
            importe(entry.client_balance_amount), entry.saleswoman, estado_fila,
            "Editar · Anular",
        )
    estado_render_grilla = {"generacion": 0}
    TAMANO_LOTE_GRILLA = 250

    def refrescar_grilla(cash_day):
        estado_render_grilla["generacion"] += 1
        generacion = estado_render_grilla["generacion"]
        for item in grilla_caja.get_children():
            grilla_caja.delete(item)
        consulta = busqueda_movimientos.get().strip().casefold()
        filtro = filtro_movimientos.get()
        movimientos = []
        for entry in sorted(cash_day.entries, key=lambda item: (item.created_at, item.id)):
            if consulta and consulta not in " ".join(str(value) for value in valores_fila(entry)).casefold():
                continue
            if filtro == "Ventas" and ((entry.expenses or 0) > 0 or (entry.withdrawal or 0) > 0):
                continue
            if filtro == "Salidas" and not entry.outflow_type:
                continue
            if filtro == "Pendientes" and not tiene_saldo_cliente(entry):
                continue
            if entry.status.value == "VOIDED":
                tags = ("voided",)
            elif entry.expenses or entry.withdrawal:
                tags = ("expense",)
            elif tiene_saldo_cliente(entry):
                tags = ("pending",)
            else:
                tags = ()
            movimientos.append((entry.id, list(valores_fila(entry)), tags))

        def insertar_lote(inicio=0):
            if generacion != estado_render_grilla["generacion"]:
                return
            fin = min(inicio + TAMANO_LOTE_GRILLA, len(movimientos))
            for entry_id, values, tags in movimientos[inicio:fin]:
                grilla_caja.insert("", "end", iid=entry_id, values=values, tags=tags)
            if fin < len(movimientos):
                ventana.after_idle(lambda: insertar_lote(fin))

        insertar_lote()

    def aplicar_filtro_movimientos(nombre):
        filtro_movimientos.set(nombre)
        for etiqueta, boton in botones_filtro.items():
            activo = etiqueta == nombre
            boton.configure(
                fg_color="#EAF3FF" if activo else "transparent",
                text_color=color_azul if activo else color_suave,
            )
        try:
            refrescar_grilla(controller.load_day(
                campos_manual["fecha"].get().strip(), campos_manual["unidad"].get().strip()
            ))
        except Exception:
            pass

    entrada_busqueda.bind("<KeyRelease>", lambda _event: aplicar_filtro_movimientos(filtro_movimientos.get()))
    def actualizar_estado(cash_day):
        totales = cash_day.totals()
        abierta = cash_day.status.value == "OPEN"
        estado_caja.configure(
            text="Estado: ABIERTO" if abierta else "Estado: CERRADO",
            fg_color="#123B2C" if abierta else "#3A2630",
            text_color=color_verde if abierta else "#E0717C",
        )
        saldo_pendiente = 0
        cobrar_convenio = 0
        for entry in cash_day.entries:
            if entry.status.value != "ACTIVE":
                continue
            try:
                saldo_pendiente += entry.client_balance_amount
                cobrar_convenio += entry.agreement_amount or 0
            except (TypeError, ValueError):
                pass
        mostrar_importe = lambda value: privacidad.display(formatear_monto(value))
        escribir_importe_formateado(campos_manual["caja_inicial"], cash_day.opening_cash)
        etiquetas_kpi["ventas"].configure(text=mostrar_importe(totales.total))
        etiquetas_kpi["efectivo"].configure(text=mostrar_importe(totales.cash))
        etiquetas_kpi["tarjeta"].configure(text=mostrar_importe(totales.card_check))
        etiquetas_kpi["gastos"].configure(text=mostrar_importe(totales.expenses))
        etiquetas_kpi["entregado"].configure(text=mostrar_importe(totales.withdrawals))
        etiquetas_kpi["esperado"].configure(text=mostrar_importe(totales.expected_cash))
        etiqueta_resumen_secundario.configure(
            text=f"Saldo cliente {mostrar_importe(saldo_pendiente)}  ·  "
                 f"Convenios {mostrar_importe(cobrar_convenio)}"
        )
        refrescar_grilla(cash_day)
        try:
            refrescar_avisos()
        except NameError:
            pass
        estado_control = "normal" if cash_day.status.value == "OPEN" else "disabled"
        estado_edicion["caja_abierta"] = cash_day.status.value == "OPEN"
        boton_cerrar_caja.configure(state="normal" if abierta else "disabled")
        for clave in claves_operacion:
            campos_manual[clave].configure(state=estado_control)
        campos_manual["transferencia"].configure(state=estado_control)
        campos_manual["notas"].configure(state=estado_control)
        campos_manual["cliente_documento"].configure(state=estado_control)
        campos_manual["cliente_telefono"].configure(state=estado_control)
        campos_manual["fecha_entrega"].configure(state=estado_control)
        campos_manual["vendedora"].configure(state=estado_control)
        for clave_salida in (
            "salida_tipo", "salida_concepto", "salida_monto",
            "salida_observacion",
        ):
            campos_manual[clave_salida].configure(state=estado_control)
        boton_guardar.configure(state=estado_control)
        boton_salida.configure(state=estado_control)

    def solicitar_conteo_obligatorio(titulo, *, esperado=None, ciego=False):
        result = {"quantities": None}
        dialog = ctk.CTkToplevel(ventana); dialog.title(titulo); dialog.geometry("610x500")
        dialog.resizable(False, False); dialog.transient(ventana); dialog.grab_set()
        ctk.CTkLabel(dialog, text=titulo.upper(), font=ctk.CTkFont(size=17, weight="bold"),
                     text_color="#0F5FB9").pack(pady=(15, 4))
        subtitle = "Contá primero el efectivo real." if ciego else "Ingresá las cantidades por denominación."
        ctk.CTkLabel(dialog, text=subtitle).pack(pady=(0, 8))
        grid = ctk.CTkFrame(dialog, fg_color="#FFFFFF"); grid.pack(fill="x", padx=14, pady=4)
        fields = {}
        for index, denomination in enumerate(DENOMINACIONES):
            column = (index // 5) * 2; row = index % 5
            ctk.CTkLabel(grid, text=formatear_monto(denomination), width=100, anchor="e").grid(row=row, column=column, padx=8, pady=6)
            entry = ctk.CTkEntry(grid, width=80, placeholder_text="0"); entry.grid(row=row, column=column + 1, padx=8, pady=6)
            fields[denomination] = entry
        total_label = ctk.CTkLabel(dialog, text="Efectivo contado: 0", font=ctk.CTkFont(size=14, weight="bold"))
        total_label.pack(pady=8)
        def values(): return {denomination: int(field.get().strip() or "0") for denomination, field in fields.items()}
        def preview(_event=None):
            try: total = sum(denomination * quantity for denomination, quantity in values().items())
            except ValueError: total_label.configure(text="Revisá las cantidades"); return
            total_label.configure(text=f"Efectivo contado: {formatear_monto(total)}")
        for field in fields.values(): field.bind("<KeyRelease>", preview, add="+")
        def accept():
            try: quantities = values()
            except ValueError: messagebox.showerror("Conteo inválido", "Use cantidades enteras.", parent=dialog); return
            counted = sum(denomination * quantity for denomination, quantity in quantities.items())
            if esperado is not None:
                difference = counted - esperado
                if not messagebox.askyesno(
                    "Confirmar arqueo", f"Efectivo esperado: {formatear_monto(esperado)}\n"
                    f"Efectivo contado: {formatear_monto(counted)}\n"
                    f"Diferencia: {formatear_diferencia_ui(difference)}\n\n¿Confirmar este conteo?\nElegí No para contar nuevamente.", parent=dialog
                ): return
            result["quantities"] = quantities; dialog.destroy()
        actions = ctk.CTkFrame(dialog, fg_color="transparent"); actions.pack(pady=10)
        ctk.CTkButton(actions, text="Cancelar", fg_color="#6B7280", command=dialog.destroy).pack(side="left", padx=5)
        ctk.CTkButton(actions, text="Confirmar conteo", command=accept).pack(side="left", padx=5)
        dialog.protocol("WM_DELETE_WINDOW", dialog.destroy); ventana.wait_window(dialog)
        return result["quantities"]

    def abrir_o_consultar():
        try:
            cash_day = controller.load_day(campos_manual["fecha"].get().strip(), campos_manual["unidad"].get().strip())
            aviso = None
        except Exception as exc:
            quantities = solicitar_conteo_obligatorio("Arqueo de apertura")
            if quantities is None:
                return
            try:
                unit = campos_manual["unidad"].get().strip()
                responsible = os.environ.get("BC_CAJA_RESPONSABLE", "").strip() or f"Caja {unit.upper()}"
                cash_day = controller.admin.open_from_count(
                    campos_manual["fecha"].get().strip(), unit, quantities, responsible, str(uuid.uuid4())
                )
                aviso = None
            except Exception as open_error:
                mostrar_error(open_error); return
        actualizar_estado(cash_day)
        diferencia = controller.opening_difference_message(cash_day)
        if diferencia:
            messagebox.showwarning("Diferencia de apertura", diferencia, parent=ventana)
        if aviso:
            messagebox.showinfo("Caja existente", aviso, parent=ventana)
        boton_abrir.pack_forget()

    def cerrar_caja():
        try:
            cash_day_abierta = controller.load_day(
                campos_manual["fecha"].get().strip(),
                campos_manual["unidad"].get().strip(),
            )
        except Exception as exc:
            mostrar_error(exc)
            return
        totales = cash_day_abierta.totals()
        quantities = solicitar_conteo_obligatorio(
            "Arqueo de cierre", esperado=totales.expected_cash,
            ciego=bool(controller.admin.setting("counting").get("blind_close", True)),
        )
        if quantities is None: return
        counted = sum(denomination * quantity for denomination, quantity in quantities.items())
        difference = counted - totales.expected_cash
        reason = ""
        if difference:
            reason = simpledialog.askstring("Diferencia de cierre", "Motivo obligatorio de la diferencia:", parent=ventana) or ""
            if not reason.strip(): return
        try:
            responsible = controller.canonical_responsible(cash_day_abierta)
            cash_day, closing_count, mail_status = controller.admin.close_with_count(
                cash_day_abierta.id, quantities, responsible, str(uuid.uuid4()), reason=reason,
                admin_token=(estado_admin["session"].token if estado_admin.get("session") else ""),
            )
        except Exception as exc:
            mostrar_error(exc)
            return
        actualizar_estado(cash_day)
        messagebox.showinfo("Caja cerrada", texto_estado(cash_day) + f"\n\nCorreo: {mail_status}", parent=ventana)
        if controller.last_warning:
            messagebox.showwarning("Backup pendiente", controller.last_warning, parent=ventana)

    def hay_cambios_sin_guardar():
        if estado_edicion["entry_id"] or estado_salida["entry_id"] or items_venta:
            return True
        for clave in claves_operacion:
            valor = (campos_manual[clave].get("1.0", "end-1c")
                     if clave == "notas" else campos_manual[clave].get())
            if str(valor).strip() and not (
                (clave == "vendedora" and valor == "Seleccionar...")
                or (clave == "fecha" and valor == date.today().strftime("%d-%m-%Y"))
                or (clave == "unidad" and valor == UNIDAD_POR_DEFECTO)
                or (clave == "sin_costo" and str(valor) == "0")
                or (clave == "notas" and observaciones_son_plantilla_neutra(valor))
            ):
                return True
        for clave in ("salida_concepto", "salida_monto", "salida_observacion"):
            if campos_manual[clave].get().strip():
                return True
        return False

    def limpiar_operacion(confirmar=True):
        if confirmar and hay_cambios_sin_guardar() and not messagebox.askyesno(
            "Descartar cambios",
            "Hay cambios sin guardar. ¿Querés limpiar toda la carga?",
            parent=ventana,
        ):
            return
        for clave in claves_operacion:
            if clave == "sin_costo":
                campos_manual[clave].deselect()
            elif clave == "notas":
                campos_manual[clave].delete("1.0", "end")
                campos_manual[clave].insert("1.0", PLANTILLA_RECETA_OBSERVACIONES)
            elif clave != "vendedora":
                campos_manual[clave].delete(0, "end")
        for clave in ("cliente_documento", "cliente_telefono", "fecha_entrega"):
            campos_manual[clave].delete(0, "end")
        campos_manual["vendedora"].set("Seleccionar...")
        items_venta.clear()
        item_editando["index"] = None
        estado_edicion["entry_id"] = None
        boton_guardar.configure(text="Guardar venta  —  F9")
        boton_cancelar.pack_forget()
        grilla_caja.selection_remove(grilla_caja.selection())
        limpiar_salida()
        refrescar_items()
        campos_manual["descripcion"].focus_set()

    def guardar_manual():
        if estado_edicion["guardando"]:
            return
        estado_edicion["guardando"] = True
        boton_guardar.configure(state="disabled")
        try:
            recalcular_total_visible()
            recalcular_saldo_visible()
            valores = leer_valores_formulario(campos_manual)
            if not items_venta:
                if estado_edicion["entry_id"]:
                    raise ValueError(
                        "La venta editada no puede quedar sin productos. "
                        "Cancelá la edición o anulá la venta completa."
                    )
                items_venta[:] = completar_items_para_guardar(valores, items_venta)
                refrescar_items()
                valores = leer_valores_formulario(campos_manual)
            valores["items"] = tuple(items_venta)
            valores["tarjeta_cheque"] = str(sumar_medios_no_efectivo(
                valores["tarjeta_cheque"], valores["transferencia"]
            ))
            if estado_edicion["entry_id"]:
                motivo_edicion = simpledialog.askstring(
                    "Edición auditada", "Motivo obligatorio de la edición:", parent=ventana
                )
                if not str(motivo_edicion or "").strip():
                    return
                responsable = os.environ.get("USERNAME") or os.environ.get("USER") or ""
                if not responsable:
                    responsable = simpledialog.askstring(
                        "Edición auditada", "Usuario responsable:", parent=ventana
                    )
                if not str(responsable or "").strip():
                    return
                cash_day, _ = controller.update_manual_entry(
                    estado_edicion["entry_id"], valores,
                    reason=motivo_edicion, user=responsable,
                )
            else:
                cash_day, _ = controller.add_manual_entry(valores)
        except Exception as exc:
            mostrar_error_guardado(exc, ventana)
            return
        finally:
            estado_edicion["guardando"] = False
            boton_guardar.configure(
                state="normal" if estado_edicion["caja_abierta"] else "disabled"
            )
        actualizar_estado(cash_day)
        messagebox.showinfo(
            "Guardado",
            "Movimiento actualizado." if estado_edicion["entry_id"] else "Movimiento agregado.",
            parent=ventana,
        )
        estado_edicion["entry_id"] = None
        boton_guardar.configure(text="Guardar movimiento")
        boton_cancelar.pack_forget()
        limpiar_operacion(confirmar=False)

    atributos_ui = {
        "descripcion": "description", "sobre": "envelope", "arm_org": "frame_origin",
        "cod": "code", "armazon": "frame", "cristal": "lens",
        "laboratorio": "laboratory", "receta_dr": "prescription_doctor",
        "total": "total", "efectivo": "cash", "tarjeta_cheque": "card_check",
        "ordenes": "orders", "monto_convenio": "agreement_amount",
        "cuotas": "installments", "saldo": "balance",
        "gastos": "expenses", "notas": "source_reference",
        "cliente_documento": "customer_document", "cliente_telefono": "customer_phone", "vendedora": "saleswoman",
        "fecha_entrega": "delivery_date",
    }

    def movimiento_seleccionado():
        seleccion = grilla_caja.selection()
        if not seleccion:
            messagebox.showwarning("Seleccioná una fila", "Elegí un movimiento de la grilla.", parent=ventana)
            return None, None
        try:
            cash_day = controller.load_day(
                campos_manual["fecha"].get().strip(), campos_manual["unidad"].get().strip()
            )
        except Exception as exc:
            mostrar_error(exc)
            return None, None
        entry = next((item for item in cash_day.entries if item.id == seleccion[0]), None)
        return cash_day, entry

    def editar_seleccionado(_event=None):
        cash_day, entry = movimiento_seleccionado()
        if entry is None:
            return
        if cash_day.status.value != "OPEN" or entry.status.value != "ACTIVE":
            messagebox.showwarning("No editable", "La fila está cerrada o anulada.", parent=ventana)
            return
        if entry.outflow_type:
            estado_salida["entry_id"] = entry.id
            campos_manual["salida_tipo"].set(
                "Entrega administración"
                if entry.outflow_type == "ENTREGA_ADMINISTRACION" else "Gasto"
            )
            for clave, valor in (
                ("salida_concepto", entry.description),
                ("salida_monto", entry.expenses or entry.withdrawal or 0),
                ("salida_observacion", entry.observations or entry.source_reference),
            ):
                campo = campos_manual[clave]
                campo.delete(0, "end")
                campo.insert(0, formatear_importe_ui(valor) if clave == "salida_monto" else str(valor or ""))
            boton_salida.configure(text="Actualizar salida")
            campos_manual["salida_concepto"].focus_set()
            return
        cargar_para_editar(cash_day, entry)


    def cancelar_edicion():
        estado_edicion["entry_id"] = None
        boton_guardar.configure(text="Guardar movimiento")
        boton_cancelar.pack_forget()
        limpiar_operacion()

    def anular_seleccionado():
        cash_day, entry = movimiento_seleccionado()
        if entry is None:
            return
        if cash_day.status.value != "OPEN" or entry.status.value != "ACTIVE":
            messagebox.showwarning("No anulable", "La fila está cerrada o ya fue anulada.", parent=ventana)
            return
        motivo = simpledialog.askstring("Anular movimiento", "Motivo de anulación:", parent=ventana)
        if not motivo:
            return
        if not messagebox.askyesno("Confirmar anulación", f"¿Anular '{entry.description}'?", parent=ventana):
            return
        try:
            controller.void_entry(
                cash_day.business_date.strftime("%d-%m-%Y"), cash_day.unit, entry.id, motivo,
                user=controller.canonical_responsible(cash_day),
            )
            actualizar_estado(controller.load_day(
                cash_day.business_date.strftime("%d-%m-%Y"), cash_day.unit
            ))
        except Exception as exc:
            mostrar_error(exc)

    acciones = ctk.CTkFrame(tab_manual, fg_color=color_panel, corner_radius=6)
    acciones.pack(fill="x", padx=6, pady=(1, 2))
    acciones_primarias = ctk.CTkFrame(acciones, fg_color="transparent")
    acciones_primarias.pack(fill="x", padx=4, pady=3)
    boton_abrir = ctk.CTkButton(
        estado_operativo, text="ABRIR / CONSULTAR", command=abrir_o_consultar,
        width=118, height=alto_boton_estado, corner_radius=4, fg_color=color_azul,
        hover_color="#1D65C5", font=ctk.CTkFont(size=fuente_estado, weight="bold"),
    )
    boton_abrir.pack(side="left", padx=(0, 5), before=estado_caja)

    def mostrar_boton_abrir(_event=None):
        if not boton_abrir.winfo_manager():
            boton_abrir.pack(side="left", padx=(0, 5), before=estado_caja)

    for clave in ("fecha", "unidad", "caja_inicial"):
        campos_manual[clave].bind("<FocusIn>", mostrar_boton_abrir, add="+")

    def refrescar_estado_consultado(_event=None):
        try:
            actualizar_estado(controller.load_day(
                campos_manual["fecha"].get().strip(), campos_manual["unidad"].get().strip()
            ))
        except Exception:
            estado_caja.configure(text="SIN CONSULTAR")

    for clave in ("fecha", "unidad"):
        campos_manual[clave].bind("<FocusOut>", refrescar_estado_consultado, add="+")
    def alternar_privacidad():
        privacidad.show() if privacidad.hidden else privacidad.hide()
        boton_privacidad.configure(text="👁 Mostrar totales" if privacidad.hidden else "👁 Ocultar totales")
        refrescar_estado_consultado()
        refrescar_items()

    def aplicar_privacidad_campos():
        # El modo privacidad sólo afecta KPIs globales, nunca la operación en curso.
        return None

    boton_privacidad = ctk.CTkButton(
        barra_superior, text="👁 Ocultar totales", width=145, height=28,
        fg_color="transparent", text_color="#FFFFFF", border_width=1,
        border_color="#8EB9EA", hover_color="#0B4D98", command=alternar_privacidad,
    )
    boton_privacidad.pack(side="right", padx=6)

    def registrar_actividad(_event=None):
        privacidad.activity()

    def revisar_auto_privacidad():
        estaba_oculta = privacidad.hidden
        if privacidad.check_timeout() and not estaba_oculta:
            boton_privacidad.configure(text="👁 Mostrar totales")
            refrescar_estado_consultado()
            refrescar_items()
        ventana.after(1000, revisar_auto_privacidad)

    for evento in ("<KeyPress>", "<Button>", "<Motion>"):
        ventana.bind_all(evento, registrar_actividad, add="+")
    ventana.after(1000, revisar_auto_privacidad)
    pago = secciones_widgets["PAGO"]
    boton_guardar = ctk.CTkButton(
        pago, text="Guardar venta  —  F9", command=guardar_manual,
        height=perfil["campo_alto"], fg_color=color_azul, hover_color="#0F5FC7",
        font=ctk.CTkFont(size=perfil["fuente"], weight="bold"),
    )
    boton_guardar.grid(
        row=6, column=0, columnspan=4, sticky="ew", padx=10, pady=(5, 9)
    )
    estado_salida = {"entry_id": None}

    def limpiar_salida():
        estado_salida["entry_id"] = None
        campos_manual["salida_tipo"].set("Gasto")
        for clave in (
            "salida_concepto", "salida_monto", "salida_observacion",
        ):
            campos_manual[clave].delete(0, "end")
        boton_salida.configure(text="Guardar salida")

    def guardar_salida_integrada():
        tipo = (
            "ENTREGA_ADMINISTRACION"
            if campos_manual["salida_tipo"].get() == "Entrega administración"
            else "GASTO"
        )
        try:
            if estado_salida["entry_id"]:
                motivo = simpledialog.askstring(
                    "Edición auditada", "Motivo obligatorio:", parent=ventana
                )
                if not str(motivo or "").strip():
                    return
                cash_day, _ = controller.update_outflow(
                    campos_manual["fecha"].get().strip(), campos_manual["unidad"].get().strip(),
                    estado_salida["entry_id"], tipo,
                    campos_manual["salida_concepto"].get(), campos_manual["salida_monto"].get(),
                    observations=campos_manual["salida_observacion"].get(),
                    performed_by="", reason=motivo,
                )
            else:
                cash_day, _ = controller.add_outflow(
                    campos_manual["fecha"].get().strip(), campos_manual["unidad"].get().strip(),
                    tipo, campos_manual["salida_concepto"].get(),
                    campos_manual["salida_monto"].get(),
                    observations=campos_manual["salida_observacion"].get(),
                    performed_by="",
                )
        except Exception as exc:
            mostrar_error(exc)
            return
        actualizar_estado(cash_day)
        messagebox.showinfo("Salida guardada", "La salida se actualizó correctamente.", parent=ventana)
        limpiar_salida()

    boton_salida = campos_manual["accion_salida"]
    boton_salida.configure(command=guardar_salida_integrada)
    boton_limpiar = ctk.CTkButton(
        zona_secundaria, text="Limpiar todo", command=limpiar_operacion,
        width=95, height=perfil["campo_alto"], fg_color="#6D3AC1",
        text_color="#FFFFFF", border_width=1, border_color="#53299B",
        hover_color="#57309A",
    )
    boton_limpiar.pack(side="left", padx=(5, 2), pady=4)
    campos_manual["salida_monto"].bind(
        "<Return>", lambda _event: guardar_salida_integrada()
    )
    boton_cancelar = ctk.CTkButton(
        zona_secundaria, text="Cancelar edición", command=cancelar_edicion,
        fg_color=COLOR_PANEL_SECUNDARIO[1], hover_color=COLOR_PRIMARIO_HOVER,
    )
    def accion_en_fila(event):
        if grilla_caja.identify_region(event.x, event.y) != "cell":
            return
        if grilla_caja.identify_column(event.x) != f"#{len(claves_grilla)}":
            return
        item = grilla_caja.identify_row(event.y)
        if not item:
            return
        grilla_caja.selection_set(item)
        izquierda, _, ancho, _ = grilla_caja.bbox(item, "acciones")
        if event.x < izquierda + ancho / 2:
            editar_seleccionado()
        else:
            anular_seleccionado()

    def desplazar_movimientos(event):
        pasos = -1 if event.delta > 0 else 1
        grilla_caja.yview_scroll(pasos, "units")
        return "break"

    grilla_caja.bind("<MouseWheel>", desplazar_movimientos, add="+")
    grilla_items.bind("<MouseWheel>", lambda event: (
        grilla_items.yview_scroll(-1 if event.delta > 0 else 1, "units"), "break"
    )[1], add="+")
    grilla_caja.bind("<Double-1>", editar_seleccionado)
    grilla_caja.bind("<ButtonRelease-1>", accion_en_fila, add="+")
    ventana.bind("<F2>", editar_seleccionado)
    ventana.bind("<F3>", lambda _event: anular_seleccionado())
    ventana.bind("<F9>", lambda _event: guardar_manual())
    ventana.bind("<F12>", lambda _event: cerrar_caja())
    campos_manual["fecha"].bind(
        "<Return>", lambda _event: campos_manual["unidad"].focus_set()
    )
    campos_manual["unidad"].bind(
        "<Return>", lambda _event: campos_manual["caja_inicial"].focus_set()
    )
    campos_manual["caja_inicial"].bind(
        "<Return>", lambda _event: abrir_o_consultar()
    )

    pie = ctk.CTkFrame(tab_manual, fg_color="transparent", height=18)
    pie.pack(fill="x", padx=8, pady=(0, 1))
    ruta_datos = resolve_data_paths().root
    etiqueta_pie = ctk.CTkLabel(
        pie,
        text=f"BC Caja {version_aplicacion()}   ·   Datos: {ruta_datos}",
        anchor="w", text_color=COLOR_TEXTO_SUAVE, font=ctk.CTkFont(size=9),
    )
    etiqueta_pie.pack(side="left", fill="x", expand=True)
    etiqueta_resumen_secundario = ctk.CTkLabel(
        pie, text="Saldo cliente —  ·  Convenios —", anchor="center",
        text_color=COLOR_TEXTO_SUAVE, font=ctk.CTkFont(size=9, weight="bold"),
    )
    etiqueta_resumen_secundario.pack(side="left", padx=12)
    reloj = ctk.CTkLabel(
        pie, text="", anchor="e", text_color=COLOR_TEXTO_SUAVE,
        font=ctk.CTkFont(size=9),
    )
    reloj.pack(side="right")
    # Macro-layout UX-006: header context, KPI cards, five-section form,
    # movements table with native scrollbars and no visible pagination.
    for bloque in (
        cabecera, formulario, lista_productos, panel_total_draft, zona_secundaria,
        toolbar_movimientos, marco_grilla, acciones, pie,
    ):
        bloque.pack_forget()
    ancho_total = min(perfil["contenido_ancho"], ancho_logico - 22)
    y_cabecera = 4
    es_full_hd = perfil["nombre"] == "full-hd"
    alto_cabecera = max(
        metricas_kpi["cabecera_alto"] if es_full_hd else 36,
        resumen_compacto.winfo_reqheight() + 8,
    )
    alto_form = 310 if es_full_hd else 210
    alto_draft = 220 if es_full_hd else 160
    alto_secundario = 40 if es_full_hd else 36
    faltante_vertical = 0 if es_full_hd else max(0, 768 - alto_logico)
    separacion_vertical = 5 if es_full_hd else (0 if faltante_vertical else 3)
    reduccion_grilla = max(0, faltante_vertical - (15 if faltante_vertical else 0))
    alto_grilla = 220 if es_full_hd else max(140, 167 - reduccion_grilla)
    y_form = y_cabecera + alto_cabecera + separacion_vertical
    y_draft = y_form + alto_form + separacion_vertical
    y_secundario = y_draft + alto_draft + separacion_vertical
    y_toolbar = y_secundario + alto_secundario + separacion_vertical
    y_grilla = y_toolbar + perfil["toolbar_alto"]
    cabecera.configure(width=ancho_total, height=alto_cabecera)
    cabecera.place(x=4, y=4)
    formulario.configure(width=ancho_total, height=alto_form)
    formulario.grid_propagate(False)
    formulario.place(x=4, y=y_form)
    ancho_columna_izquierda = int(ancho_total * 3 / 5)
    lista_productos.configure(width=ancho_total, height=alto_draft + alto_secundario + separacion_vertical)
    lista_productos.pack_propagate(False)
    lista_productos.place(x=4, y=y_draft)
    zona_secundaria.configure(width=ancho_columna_izquierda, height=alto_secundario)
    zona_secundaria.pack_propagate(False)
    zona_secundaria.place(x=4, y=y_secundario)
    acciones.place_forget()
    toolbar_movimientos.configure(
        width=ancho_total, height=perfil["toolbar_alto"],
    )
    toolbar_movimientos.place(x=4, y=y_toolbar)
    if not es_full_hd:
        campos_manual["salida_concepto"].configure(width=140)
        campos_manual["salida_monto"].configure(width=85)
        campos_manual["salida_observacion"].configure(width=110)
        boton_salida.configure(width=95)
        entrada_busqueda.configure(width=130)
        for boton in botones_filtro.values():
            boton.configure(width=55)
    alto_grilla += perfil["acciones_alto"] if es_full_hd else 31
    marco_grilla.configure(width=ancho_total, height=alto_grilla)
    marco_grilla.grid_propagate(False)
    marco_grilla.place(x=4, y=y_grilla)
    pie.configure(width=ancho_total - 8, height=24 if perfil["nombre"] == "full-hd" else 18)
    pie.place(
        x=8,
        y=y_grilla + alto_grilla + 3,
    )
    # RC18: la cabecera full-hd crece para alojar los KPI jerarquizados; el
    # presupuesto vertical restante lo absorbe la grilla, que a 1080p sobra.
    alto_cabecera_full_hd = metricas_resumen_kpi(perfil_visual(1920, 1080))["cabecera_alto"]
    estado_layout = {"after": None, "metricas": {}}

    def aplicar_macro_layout():
        """Recalcula posiciones con el area cliente Tk ya renderizada."""
        estado_layout["after"] = None
        ventana.update_idletasks()
        ancho_cliente = max(1, tab_manual.winfo_width())
        alto_cliente = max(1, tab_manual.winfo_height())
        ancho_actual = min(perfil["contenido_ancho"], ancho_cliente - 8)
        x_actual = max(4, (ancho_cliente - ancho_actual) // 2)
        full_hd_actual = ancho_cliente >= 1700 and alto_cliente >= 850
        if full_hd_actual:
            alto_cab, alto_tot = alto_cabecera_full_hd, 0
            form_preferido, form_minimo = 310, 280
            draft_preferido, draft_minimo = 220, 110
            alto_sec, sep = 40, 5
        else:
            alto_cab, alto_tot = 42, 0
            form_preferido, form_minimo = 212, 172
            # Baseline RC5 conservado para trazabilidad: draft_preferido, draft_minimo = 182, 145
            draft_preferido, draft_minimo = 195, 175
            alto_sec, sep = 32, 1
        # RC18: el perfil tipografico proviene de la pantalla y el alto de
        # cabecera de la ventana. Una ventana reducida en un monitor grande
        # conserva las tarjetas grandes, asi que la cabecera se ajusta al
        # contenido real del resumen en vez de a una constante por rama.
        alto_cab = max(alto_cab, resumen_compacto.winfo_reqheight() + 8)
        margen_inferior = 4
        alto_pie_actual = max(18, pie.winfo_reqheight())
        fila_renderizada = int(estilo.lookup("Caja.Treeview", "rowheight"))
        filas_minimas = 5
        alto_grilla_minimo = 46 + (filas_minimas * fila_renderizada) + scroll_horizontal.winfo_reqheight() + 4
        extra_toolbar = alto_sec + sep + perfil["toolbar_alto"]
        fijos_sin_form_draft = 4 + alto_cab + sep + alto_tot + sep + sep + extra_toolbar
        presupuesto_form_draft = max(
            form_minimo + draft_minimo,
            alto_cliente - alto_pie_actual - margen_inferior - 2
            - alto_grilla_minimo - fijos_sin_form_draft,
        )
        alto_form_actual = min(
            form_preferido,
            max(form_minimo, presupuesto_form_draft - draft_minimo),
        )
        draft_actual = min(
            draft_preferido,
            max(draft_minimo, presupuesto_form_draft - alto_form_actual),
        )
        y_cab = 4
        y_tot = y_cab + alto_cab + sep
        y_form_actual = y_tot + alto_tot + sep
        y_draft_actual = y_form_actual + alto_form_actual + sep
        y_sec = y_draft_actual + draft_actual + sep
        y_toolbar_actual = y_sec + alto_sec + sep
        y_grid = y_toolbar_actual + perfil["toolbar_alto"]
        y_footer = alto_cliente - alto_pie_actual - margen_inferior
        alto_grid = max(1, y_footer - 2 - y_grid)
        cabecera.configure(width=ancho_actual, height=alto_cab); cabecera.place(x=x_actual, y=y_cab)
        formulario.configure(width=ancho_actual, height=alto_form_actual); formulario.place(x=x_actual, y=y_form_actual)
        ancho_izquierdo_actual = int(ancho_actual * 3 / 5)
        ancho_derecho_actual = ancho_actual - ancho_izquierdo_actual - 8
        lista_productos.configure(width=ancho_izquierdo_actual, height=draft_actual); lista_productos.place(x=x_actual, y=y_draft_actual)
        alto_observaciones = y_toolbar_actual - y_draft_actual - sep
        panel_total_draft.configure(
            width=ancho_derecho_actual, height=alto_observaciones,
        )
        panel_total_draft.place(
            x=x_actual + ancho_izquierdo_actual + 8, y=y_draft_actual,
        )
        ancho_tabla = max(1, ancho_izquierdo_actual - 34)
        for clave, proporcion in (
            ("producto", 0.35), ("codigo", 0.12), ("tipo", 0.18),
            ("armazon", 0.12), ("cristal", 0.12), ("subtotal", 0.11),
        ):
            grilla_items.column(
                clave, width=max(72, int(ancho_tabla * proporcion)), stretch=False,
            )
        zona_secundaria.configure(width=ancho_izquierdo_actual, height=alto_sec); zona_secundaria.place(x=x_actual, y=y_sec)
        acciones.place_forget()
        toolbar_movimientos.configure(width=ancho_actual, height=perfil["toolbar_alto"]); toolbar_movimientos.place(x=x_actual, y=y_toolbar_actual)
        ajustar_columnas_movimientos(ancho_actual)
        marco_grilla.configure(width=ancho_actual, height=alto_grid); marco_grilla.place(x=x_actual, y=y_grid)
        pie.configure(width=ancho_actual, height=alto_pie_actual); pie.place(x=x_actual, y=y_footer)
        estado_layout["metricas"] = {"cliente": (ancho_cliente, alto_cliente), "grilla_y": y_grid, "grilla_alto": alto_grid, "pie_y": y_footer, "pie_alto": alto_pie_actual, "footer_bottom": y_footer + alto_pie_actual, "required": y_footer + alto_pie_actual + margen_inferior, "draft_alto": draft_actual, "overflow": max(0, y_footer + alto_pie_actual + margen_inferior - alto_cliente)}

    def programar_macro_layout(_event=None):
        if estado_layout["after"] is not None:
            ventana.after_cancel(estado_layout["after"])
        estado_layout["after"] = ventana.after_idle(aplicar_macro_layout)

    tab_manual.bind("<Configure>", programar_macro_layout, add="+")
    ventana.after_idle(aplicar_macro_layout)
    def actualizar_reloj():
        if reloj.winfo_exists():
            reloj.configure(text=datetime.now().strftime("%H:%M:%S"))
            ventana.after(1000, actualizar_reloj)

    actualizar_reloj()
    # ---- Arqueo ----
    superior = ctk.CTkFrame(
        tab_arqueo, fg_color="#FFFFFF", corner_radius=8,
        border_width=1, border_color=color_borde_suave,
    )
    superior.pack(fill="x", padx=12, pady=(10, 6))
    ctk.CTkLabel(superior, text="Resumen esperado", font=ctk.CTkFont(size=14, weight="bold")).pack(
        side="left", padx=12, pady=10
    )
    entrada_fecha = ctk.CTkEntry(superior, width=120)
    entrada_fecha.insert(0, date.today().strftime("%d-%m-%Y"))
    entrada_fecha.pack(side="left", padx=6)
    combo_unidad_arqueo = ctk.CTkComboBox(superior, values=UNIDADES, width=140)
    combo_unidad_arqueo.set(UNIDAD_POR_DEFECTO)
    combo_unidad_arqueo.pack(side="left", padx=6)
    etiqueta_esperado = ctk.CTkLabel(
        superior, text="Efectivo esperado por sistema: —",
        font=ctk.CTkFont(size=13, weight="bold"), text_color=color_azul,
    )
    etiqueta_esperado.pack(side="right", padx=16)

    cuerpo_arqueo = ctk.CTkFrame(tab_arqueo, fg_color="transparent")
    cuerpo_arqueo.pack(fill="both", expand=True, padx=12, pady=4)
    cuerpo_arqueo.grid_columnconfigure(0, weight=3)
    cuerpo_arqueo.grid_columnconfigure(1, weight=2)

    conteo_frame = ctk.CTkFrame(
        cuerpo_arqueo, fg_color="#FFFFFF", corner_radius=8,
        border_width=1, border_color=color_borde_suave,
    )
    conteo_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
    ctk.CTkLabel(
        conteo_frame, text="Conteo físico", font=ctk.CTkFont(size=14, weight="bold")
    ).grid(row=0, column=0, columnspan=3, sticky="w", padx=12, pady=(8, 4))
    for columna, titulo in enumerate(("Denominación", "Cantidad", "Subtotal")):
        ctk.CTkLabel(
            conteo_frame, text=titulo, text_color=color_suave,
            font=ctk.CTkFont(size=9, weight="bold"),
        ).grid(row=1, column=columna, padx=12, pady=2, sticky="w")

    campos_conteo = {}
    etiquetas_subtotal = {}
    estado_arqueo = {"esperado": None}
    for indice, denominacion in enumerate(DENOMINACIONES, start=2):
        ctk.CTkLabel(conteo_frame, text=formatear_monto(denominacion)).grid(
            row=indice, column=0, sticky="w", padx=12, pady=2
        )
        campo = ctk.CTkEntry(conteo_frame, width=100, height=25, placeholder_text="0")
        campo.grid(row=indice, column=1, padx=12, pady=2)
        subtotal = ctk.CTkLabel(conteo_frame, text="0", width=120, anchor="e")
        subtotal.grid(row=indice, column=2, padx=12, pady=2, sticky="e")
        campos_conteo[denominacion] = campo
        etiquetas_subtotal[denominacion] = subtotal

    resultado_frame = ctk.CTkFrame(
        cuerpo_arqueo, fg_color="#FFFFFF", corner_radius=8,
        border_width=1, border_color=color_borde_suave,
    )
    resultado_frame.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
    ctk.CTkLabel(
        resultado_frame, text="Resultado del arqueo",
        font=ctk.CTkFont(size=14, weight="bold"),
    ).pack(anchor="w", padx=16, pady=(14, 10))
    etiqueta_contado = ctk.CTkLabel(resultado_frame, text="Efectivo contado: 0", anchor="w")
    etiqueta_contado.pack(fill="x", padx=16, pady=5)
    etiqueta_diferencia = ctk.CTkLabel(resultado_frame, text="Diferencia: —", anchor="w")
    etiqueta_diferencia.pack(fill="x", padx=16, pady=5)
    etiqueta_estado_arqueo = ctk.CTkLabel(
        resultado_frame, text="Consultá una caja", anchor="w", justify="left",
        corner_radius=6, fg_color="#EEF3F8",
    )
    etiqueta_estado_arqueo.pack(fill="x", padx=16, pady=(10, 16), ipady=10)

    def cantidades_arqueo():
        cantidades = {}
        for denominacion, campo in campos_conteo.items():
            texto = campo.get().strip() or "0"
            cantidades[denominacion] = int(texto)
        return cantidades

    def actualizar_previsualizacion(_event=None):
        try:
            cantidades = cantidades_arqueo()
        except ValueError:
            etiqueta_estado_arqueo.configure(
                text="Revisá las cantidades", fg_color="#FFF1F2", text_color=COLOR_ROJO
            )
            return
        contado = sum(denominacion * cantidad for denominacion, cantidad in cantidades.items())
        for denominacion, cantidad in cantidades.items():
            etiquetas_subtotal[denominacion].configure(
                text=formatear_monto(denominacion * cantidad)
            )
        etiqueta_contado.configure(text=f"Efectivo contado: {formatear_monto(contado)}")
        esperado = estado_arqueo["esperado"]
        if esperado is None:
            return
        diferencia = contado - esperado
        estado_texto, detalle = describir_diferencia_arqueo(diferencia)
        etiqueta_diferencia.configure(
            text=f"Diferencia: {formatear_diferencia_ui(diferencia)}"
        )
        etiqueta_estado_arqueo.configure(
            text=f"{estado_texto}\n{detalle}",
            fg_color="#EAF8F1" if diferencia == 0 else "#FFF7E6",
            text_color=color_verde if diferencia == 0 else "#B45309",
        )

    def consultar_arqueo():
        try:
            cash_day = controller.load_day(
                entrada_fecha.get().strip(), combo_unidad_arqueo.get().strip()
            )
        except Exception as exc:
            mostrar_error(exc)
            return
        estado_arqueo["esperado"] = cash_day.totals().expected_cash
        etiqueta_esperado.configure(
            text="Efectivo esperado por sistema: "
            + formatear_monto(estado_arqueo["esperado"])
        )
        actualizar_previsualizacion()

    def guardar_arqueo():
        try:
            resultado = controller.record_cash_count(
                entrada_fecha.get().strip(), combo_unidad_arqueo.get().strip(),
                cantidades_arqueo(),
            )
        except Exception as exc:
            mostrar_error(exc)
            return
        estado_arqueo["esperado"] = resultado.expected_total
        actualizar_previsualizacion()
        if resultado.difference:
            messagebox.showwarning(
                "Arqueo con diferencia",
                describir_diferencia_arqueo(resultado.difference)[1]
                + ". El arqueo fue guardado correctamente.",
                parent=ventana,
            )
        else:
            messagebox.showinfo(
                "Arqueo conforme", "El arqueo fue guardado correctamente.", parent=ventana
            )

    for campo in campos_conteo.values():
        campo.bind("<KeyRelease>", actualizar_previsualizacion, add="+")

    botones_arqueo = ctk.CTkFrame(resultado_frame, fg_color="transparent")
    botones_arqueo.pack(fill="x", padx=12, pady=8)
    ctk.CTkButton(
        botones_arqueo, text="Consultar caja", command=consultar_arqueo,
        fg_color="#FFFFFF", text_color=color_texto, border_width=1,
        border_color=color_borde_suave,
    ).pack(side="left", padx=4)
    ctk.CTkButton(
        botones_arqueo, text="Guardar arqueo", command=guardar_arqueo,
        fg_color=COLOR_PRIMARIO, hover_color=COLOR_PRIMARIO_HOVER,
    ).pack(side="left", padx=4)

    def abrir_modal_arqueo():
        """Abre el mismo flujo de arqueo para la caja activa sin salir de Caja diaria."""
        try:
            cash_day = controller.load_day(
                campos_manual["fecha"].get().strip(), campos_manual["unidad"].get().strip()
            )
        except Exception as exc:
            mostrar_error(exc)
            return

        modal = ctk.CTkToplevel(ventana)
        modal.title("Arqueo de caja")
        modal.geometry("760x620")
        modal.resizable(False, False)
        modal.transient(ventana)
        modal.grab_set()
        modal.bind("<Escape>", lambda _event: modal.destroy())

        totals = cash_day.totals()
        latest = None
        quantities = {}
        responsible = controller.canonical_responsible(cash_day)

        header = ctk.CTkFrame(modal, fg_color="#0F5FB9", corner_radius=0)
        header.pack(fill="x")
        ctk.CTkLabel(
            header, text="ARQUEO DE CAJA", text_color="#FFFFFF",
            font=ctk.CTkFont(size=17, weight="bold"),
        ).pack(side="left", padx=18, pady=12)
        ctk.CTkLabel(
            header,
            text=f"{cash_day.business_date.strftime('%d-%m-%Y')}  ·  {cash_day.unit}",
            text_color="#DCEBFF", font=ctk.CTkFont(size=12, weight="bold"),
        ).pack(side="right", padx=18)

        summary = ctk.CTkFrame(modal, fg_color="#F4F8FD", corner_radius=8)
        summary.pack(fill="x", padx=12, pady=(10, 6))
        summary.grid_columnconfigure((0, 1, 2), weight=1)
        summary_values = (
            ("Caja inicial", cash_day.opening_cash),
            ("Cobros en efectivo", totals.cash),
            ("Gastos", totals.expenses),
            ("Entregas administración", totals.withdrawals),
            ("Efectivo esperado", totals.expected_cash),
        )
        for index, (label, value) in enumerate(summary_values):
            cell = ctk.CTkFrame(summary, fg_color="#FFFFFF", corner_radius=5)
            cell.grid(row=index // 3, column=index % 3, sticky="ew", padx=4, pady=4)
            ctk.CTkLabel(cell, text=label, text_color=color_suave, font=ctk.CTkFont(size=9)).pack(pady=(4, 0))
            ctk.CTkLabel(cell, text=formatear_monto(value), text_color=color_azul,
                         font=ctk.CTkFont(size=12, weight="bold")).pack(pady=(0, 4))
        identity = ctk.CTkFrame(summary, fg_color="#FFFFFF", corner_radius=5)
        identity.grid(row=1, column=2, sticky="ew", padx=4, pady=4)
        ctk.CTkLabel(identity, text="Responsable canónico", text_color=color_suave,
                     font=ctk.CTkFont(size=9)).pack(pady=(4, 0))
        ctk.CTkLabel(identity, text=responsible, text_color=color_texto,
                     font=ctk.CTkFont(size=11, weight="bold")).pack(pady=(0, 4))

        body = ctk.CTkFrame(modal, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=12, pady=4)
        count_panel = ctk.CTkFrame(body, fg_color="#FFFFFF", corner_radius=8)
        count_panel.pack(side="left", fill="both", expand=True, padx=(0, 5))
        result_panel = ctk.CTkFrame(body, fg_color="#FFFFFF", corner_radius=8, width=270)
        result_panel.pack(side="right", fill="y", padx=(5, 0))
        result_panel.pack_propagate(False)
        ctk.CTkLabel(count_panel, text="Conteo físico", font=ctk.CTkFont(size=14, weight="bold")).grid(
            row=0, column=0, columnspan=4, sticky="w", padx=12, pady=8
        )
        modal_fields = {}
        for index, denomination in enumerate(DENOMINACIONES):
            block = index // 5
            row = index % 5 + 1
            column = block * 2
            ctk.CTkLabel(count_panel, text=formatear_monto(denomination), width=85, anchor="e").grid(
                row=row, column=column, padx=(10, 4), pady=5
            )
            field = ctk.CTkEntry(count_panel, width=65, height=28, placeholder_text="0")
            field.grid(row=row, column=column + 1, padx=(0, 10), pady=5)
            if quantities.get(denomination):
                field.insert(0, str(quantities[denomination]))
            if latest:
                field.configure(state="disabled")
            modal_fields[denomination] = field

        counted_label = ctk.CTkLabel(result_panel, text="Efectivo contado: 0", anchor="w")
        counted_label.pack(fill="x", padx=14, pady=(18, 7))
        difference_label = ctk.CTkLabel(result_panel, text="Diferencia: —", anchor="w")
        difference_label.pack(fill="x", padx=14, pady=7)
        status_label = ctk.CTkLabel(result_panel, text="", justify="left", wraplength=235,
                                    corner_radius=6, fg_color="#EEF3F8")
        status_label.pack(fill="x", padx=14, pady=10, ipady=10)

        def modal_quantities():
            if latest:
                return dict(latest.quantities)
            return {denomination: int(field.get().strip() or "0")
                    for denomination, field in modal_fields.items()}

        def preview(_event=None):
            try:
                values = modal_quantities()
            except ValueError:
                status_label.configure(text="Revisá las cantidades", fg_color="#FFF1F2")
                return
            counted = sum(denomination * quantity for denomination, quantity in values.items())
            difference = counted - totals.expected_cash
            state_text, detail = describir_diferencia_arqueo(difference)
            counted_label.configure(text=f"Efectivo contado: {formatear_monto(counted)}")
            difference_label.configure(text=f"Diferencia: {formatear_diferencia_ui(difference)}")
            status_label.configure(
                text=(f"{state_text}\n{detail}\nEstado: "
                      f"{'GUARDADO' if latest else 'BORRADOR'}  ·  Caja {cash_day.status.value}"),
                fg_color="#EAF8F1" if difference == 0 else "#FFF7E6",
            )

        def save():
            try:
                result = controller.admin.record_count(
                    cash_day.id, "INTERMEDIATE", modal_quantities(), responsible,
                    str(uuid.uuid4()),
                )
            except Exception as exc:
                mostrar_error(exc)
                return
            messagebox.showinfo(
                "Arqueo guardado", "El arqueo quedó guardado sin crear duplicados.", parent=modal
            )
            modal.destroy()

        for field in modal_fields.values():
            field.bind("<KeyRelease>", preview, add="+")
        preview()
        actions = ctk.CTkFrame(result_panel, fg_color="transparent")
        actions.pack(side="bottom", fill="x", padx=10, pady=12)
        ctk.CTkButton(actions, text="Cerrar", command=modal.destroy, fg_color="#6B7280").pack(
            side="left", padx=4
        )
        ctk.CTkButton(actions, text="Guardar arqueo", command=save,
                      state="disabled" if latest else "normal").pack(side="left", padx=4)
        modal.after_idle(lambda: next(iter(modal_fields.values())).focus_set() if not latest else None)
    # ---- Historial / edición / anulación ----
    filtros_historial = ctk.CTkFrame(tab_historial, fg_color="transparent")
    filtros_historial.pack(fill="x", padx=8, pady=8)
    ctk.CTkLabel(filtros_historial, text="Desde").pack(side="left")
    entrada_historial = ctk.CTkEntry(filtros_historial, width=105)
    entrada_historial.insert(0, date.today().strftime("%d-%m-%Y"))
    entrada_historial.pack(side="left", padx=(4, 8))
    ctk.CTkLabel(filtros_historial, text="Hasta").pack(side="left")
    entrada_historial_hasta = ctk.CTkEntry(filtros_historial, width=105)
    entrada_historial_hasta.insert(0, date.today().strftime("%d-%m-%Y"))
    entrada_historial_hasta.pack(side="left", padx=(4, 8))
    unidad_historial = ctk.CTkComboBox(filtros_historial, values=UNIDADES, width=140)
    unidad_historial.set(UNIDAD_POR_DEFECTO)
    unidad_historial.pack(side="left", padx=8)

    resumen_historial = ctk.CTkLabel(
        tab_historial, text="", justify="left", text_color=COLOR_TEXTO_SUAVE
    )
    resumen_historial.pack(fill="x", padx=8, pady=(0, 4), anchor="w")
    lista_historial = ctk.CTkScrollableFrame(tab_historial, fg_color="#F3F6FA")
    lista_historial.pack(fill="both", expand=True, padx=8, pady=8)

    def cargar_para_editar(cash_day, entry):
        if entry.outflow_type:
            grilla_caja.selection_set(entry.id)
            editar_seleccionado()
            return
        if cash_day.status.value != "OPEN" or entry.status.value != "ACTIVE":
            messagebox.showwarning(
                "No editable", "La fila está cerrada o anulada.", parent=ventana
            )
            return
        valores = {
            "fecha": cash_day.business_date.strftime("%d-%m-%Y"),
            "unidad": cash_day.unit,
            "caja_inicial": cash_day.opening_cash,
            "descripcion": entry.description,
            "sobre": entry.envelope,
            "arm_org": entry.frame_origin,
            "cod": entry.code,
            "armazon": entry.frame,
            "cristal": entry.lens,
            "laboratorio": entry.laboratory,
            "receta_dr": entry.prescription_doctor,
            "total": entry.total,
            "efectivo": entry.cash,
            "tarjeta_cheque": entry.card_check,
            "ordenes": entry.orders,
            "monto_convenio": entry.agreement_amount,
            "cuotas": entry.installments,
            "saldo": entry.balance,
            "gastos": entry.expenses,
            "notas": entry.observations,
            "cliente_documento": entry.customer_document,
            "cliente_telefono": entry.customer_phone,
            "vendedora": entry.saleswoman,
            "fecha_entrega": (
                entry.delivery_date.strftime("%d-%m-%Y") if entry.delivery_date else ""
            ),
        }
        for clave, valor in valores.items():
            if clave not in campos_manual:
                continue
            campo = campos_manual[clave]
            if clave in ("unidad", "vendedora"):
                campo.set("" if valor is None else str(valor))
            elif clave == "notas":
                campo.delete("1.0", "end")
                campo.insert("1.0", "" if valor is None else str(valor))
            else:
                campo.delete(0, "end")
                campo.insert(0, formatear_importe_ui(valor)
                             if clave in CAMPOS_MONETARIOS_UI else "" if valor is None else str(valor))
        estado_edicion["entry_id"] = entry.id
        items_venta[:] = list(entry.effective_items)
        item_editando["index"] = None
        refrescar_items()
        boton_guardar.configure(text="Guardar cambios")
        boton_cancelar.pack(side="left", padx=3)
        pestañas.set("Cargar manual")
        campos_manual["descripcion"].focus_set()
    def anular_desde_historial(cash_day, entry):
        motivo = simpledialog.askstring(
            "Anular movimiento",
            "Motivo de anulación:",
            parent=ventana,
        )
        if not motivo:
            return
        if not messagebox.askyesno(
            "Confirmar anulación",
            f"¿Anular '{entry.description}'? El registro quedará en el historial.",
            parent=ventana,
        ):
            return
        try:
            controller.void_entry(
                cash_day.business_date.strftime("%d-%m-%Y"), cash_day.unit, entry.id, motivo,
                user=os.environ.get("USERNAME") or os.environ.get("USER") or "",
            )
        except Exception as exc:
            mostrar_error(exc)
            return
        consultar_historial()

    def editar_caja(cash_day):
        nuevo = simpledialog.askinteger(
            "Editar caja", "Nueva caja inicial:", initialvalue=cash_day.opening_cash,
            minvalue=0, parent=ventana,
        )
        if nuevo is None or nuevo == cash_day.opening_cash:
            return
        motivo = simpledialog.askstring(
            "Corrección auditada", "Motivo obligatorio:", parent=ventana
        )
        if not motivo:
            messagebox.showwarning("Motivo requerido", "La corrección no fue guardada.", parent=ventana)
            return
        usuario = simpledialog.askstring(
            "Corrección auditada", "Usuario responsable:", parent=ventana
        )
        if not usuario:
            return
        try:
            controller.correct_opening_cash(
                cash_day.business_date.strftime("%d-%m-%Y"), cash_day.unit,
                nuevo, motivo, usuario,
            )
        except Exception as exc:
            mostrar_error(exc)
            return
        consultar_historial()

    def consultar_historial():
        try:
            cash_days = controller.list_history_range(
                entrada_historial.get().strip(), entrada_historial_hasta.get().strip(),
                unidad_historial.get().strip(),
            )
        except Exception as exc:
            mostrar_error(exc)
            return
        for widget in lista_historial.winfo_children():
            widget.destroy()
        resumen_historial.configure(text=f"{len(cash_days)} jornadas en el período")
        # RC29: cada jornada es una tarjeta y todo lo del dia cuelga de ella.
        #
        # Antes la cabecera y las filas eran hermanas dentro del scroll, y la
        # unica pista de donde terminaba un dia era el padding. Ahora el bloque
        # tiene borde propio y separacion entre jornadas, asi que se distingue
        # sin leer una sola cifra. Nada de esto cambia contenido: mismas
        # jornadas, mismo orden, mismos numeros.
        for cash_day in cash_days:
            tarjeta = ctk.CTkFrame(
                lista_historial, fg_color="#FFFFFF", corner_radius=8,
                border_width=1, border_color=color_borde_suave)
            # Marca de pertenencia para la sonda visual.
            tarjeta._bc_jornada_historial = True
            tarjeta.pack(fill="x", padx=4, pady=(0, 10))

            cabecera = ctk.CTkFrame(tarjeta, fg_color="transparent")
            cabecera.pack(fill="x", padx=10, pady=(8, 2))
            # La fecha manda; el estado la acompaña como chip compacto.
            ctk.CTkLabel(
                cabecera, text=cash_day.business_date.strftime("%d-%m-%Y"),
                anchor="w", text_color=color_texto,
                font=ctk.CTkFont(size=perfil["fuente"] + 3, weight="bold"),
            ).pack(side="left")
            abierto = estado_dia(cash_day) == "ABIERTO"
            ctk.CTkLabel(
                cabecera, text=f"  {estado_dia(cash_day)}  ", corner_radius=7,
                fg_color="#FFF3CD" if abierto else "#DDF5E8",
                text_color="#7A4B00" if abierto else "#17633A", height=20,
                font=ctk.CTkFont(size=max(8, perfil["fuente_label"] - 1), weight="bold"),
            ).pack(side="left", padx=(10, 0))
            ctk.CTkButton(
                cabecera, text="Editar caja", width=90, height=28,
                command=lambda d=cash_day: editar_caja(d),
            ).pack(side="right")

            ctk.CTkLabel(
                tarjeta, text=resumen_economico_dia(cash_day), anchor="w",
                justify="left", text_color=color_texto,
                font=ctk.CTkFont(size=perfil["fuente"]),
            ).pack(fill="x", padx=10, pady=(0, 2))

            detalle_sesion = detalle_sesion_dia(cash_day)
            if detalle_sesion:
                # Secundario y compacto: acompaña, no abre otro bloque.
                ctk.CTkLabel(
                    tarjeta, text=detalle_sesion.replace("\n", "    "), anchor="w",
                    justify="left", text_color=color_suave,
                    font=ctk.CTkFont(size=max(8, perfil["fuente_label"] - 1)),
                ).pack(fill="x", padx=10, pady=(0, 2))

            for indice_entry, entry in enumerate(cash_day.entries):
                es_anulado = entry.status.value == "VOIDED"
                color_fila = "#FDECEC" if es_anulado else ("#FFFFFF" if indice_entry % 2 == 0 else "#EEF4FB")
                fila = ctk.CTkFrame(tarjeta, fg_color=color_fila)
                fila.pack(fill="x", padx=10, pady=1)
                estado_texto = "ANULADO" if entry.status.value == "VOIDED" else "ACTIVO"
                detalle = (
                f"{entry.description} | Total {formatear_monto(entry.total or 0)} | "
                f"Efectivo {formatear_monto(entry.cash or 0)} | "
                f"Tarj./Cheq. {formatear_monto(entry.card_check or 0)} | "
                f"Gastos {formatear_monto(entry.expenses or 0)} | {estado_texto}"
                )
                if entry.void_reason:
                    detalle += f" ({entry.void_reason})"
                ctk.CTkLabel(
                    fila, text=detalle, anchor="w",
                    text_color="#A32626" if es_anulado else "#132238",
                ).pack(
                side="left", fill="x", expand=True
                )
                habilitado = (
                "normal"
                if cash_day.status.value == "OPEN" and entry.status.value == "ACTIVE"
                else "disabled"
                )
                ctk.CTkButton(
                fila, text="Editar", width=65, state=habilitado,
                command=lambda d=cash_day, e=entry: cargar_para_editar(d, e),
                ).pack(side="left", padx=2)
                ctk.CTkButton(
                fila, text="Anular", width=65, state=habilitado,
                fg_color=COLOR_ROJO,
                command=lambda d=cash_day, e=entry: anular_desde_historial(d, e),
                ).pack(side="left", padx=2)
            # Aire al pie de la tarjeta: la ultima fila no toca el borde.
            ctk.CTkFrame(tarjeta, fg_color="transparent", height=6).pack(fill="x")

    def rango_rapido(dias=None, mes=False):
        hoy = date.today()
        desde = hoy.replace(day=1) if mes else hoy - timedelta(days=(dias or 1) - 1)
        for campo, valor in ((entrada_historial, desde), (entrada_historial_hasta, hoy)):
            campo.delete(0, "end")
            campo.insert(0, valor.strftime("%d-%m-%Y"))
        consultar_historial()

    for texto, comando in (
        ("Hoy", lambda: rango_rapido(1)),
        ("7 días", lambda: rango_rapido(7)),
        ("Este mes", lambda: rango_rapido(mes=True)),
    ):
        ctk.CTkButton(
            filtros_historial, text=texto, width=68, command=comando,
            fg_color="#FFFFFF", text_color=color_texto, border_width=1,
            border_color=color_borde_suave,
        ).pack(side="right", padx=2)

    ctk.CTkButton(
        filtros_historial,
        text="Consultar",
        command=consultar_historial,
        fg_color=COLOR_PRIMARIO,
        hover_color=COLOR_PRIMARIO_HOVER,
    ).pack(side="left", padx=8)

    # ---- Pedidos: alineación, chips accesibles y reversión auditada ----
    barra_pedidos = ctk.CTkFrame(tab_pedidos, fg_color="transparent")
    barra_pedidos.pack(fill="x", padx=10, pady=10)
    # RC28: Pedidos abre en lo que requiere atencion, no en una hoja en blanco.
    # Es el mismo grupo que cuenta la alerta de la cabecera.
    filtro_pedidos = ctk.StringVar(value=FILTRO_REQUIEREN_ATENCION)

    # Contexto compacto: que se esta mostrando y como salir del filtro. Solo
    # aparece cuando hay filtro; con "Todos" no tiene nada que decir.
    contexto_pedidos = ctk.CTkFrame(tab_pedidos, fg_color="#EEF4FB", corner_radius=6)
    etiqueta_contexto_pedidos = ctk.CTkLabel(
        contexto_pedidos, text="", anchor="w", text_color=color_texto,
        font=ctk.CTkFont(size=perfil["fuente_label"] + 1, weight="bold"))
    etiqueta_contexto_pedidos.pack(side="left", padx=(10, 14), pady=6)
    ctk.CTkButton(
        contexto_pedidos, text="Ver todos", width=110, height=28,
        fg_color="#FFFFFF", text_color=color_azul, border_width=1,
        border_color=color_borde_suave, hover_color="#EAF3FF",
        font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
        command=lambda: refrescar_pedidos("Todos"),
    ).pack(side="left", pady=6)

    marco_pedidos = ctk.CTkFrame(tab_pedidos, fg_color="#FFFFFF")
    marco_pedidos.pack(fill="both", expand=True, padx=10, pady=(0, 8))
    marco_pedidos.grid_rowconfigure(0, weight=1)
    marco_pedidos.grid_columnconfigure(0, weight=1)
    columnas_pedido = ("entrega", "cliente", "telefono", "documento", "sobre", "sucursal", "vendedora", "origen", "estado")
    grilla_pedidos = ttk.Treeview(marco_pedidos, columns=columnas_pedido, show="headings", style="Caja.Treeview")
    alineacion_pedidos = {
        "entrega": "center", "cliente": "w", "telefono": "center",
        "documento": "center", "sobre": "center", "sucursal": "center",
        "vendedora": "center", "origen": "center", "estado": "center",
    }
    # RC22: el estado del pedido pasa a color de fila del propio Treeview.
    # Antes se dibujaba como chip flotante sobre el frame contenedor y quedaba
    # desacoplado al desplazar o repintar.
    for estado_pedido, (fondo_pedido, _borde, texto_pedido) in {
        "PENDIENTE": ("#FFF1CC", "#E6A23C", "#8A4B08"),
        "LISTO": ("#DCEEFF", "#82B7E8", "#174A7E"),
        "ENTREGADO": ("#DDF5E8", "#79C99E", "#17633A"),
        "ANULADO": ("#FDECEC", "#E5A3A3", "#A32626"),
    }.items():
        grilla_pedidos.tag_configure(
            f"estado_{estado_pedido}", background=fondo_pedido, foreground=texto_pedido)
    for clave, titulo, ancho in (
        ("entrega", "Entrega", 100), ("cliente", "Cliente", 220),
        ("telefono", "Teléfono", 125), ("documento", "CI/RUC", 120),
        ("sobre", "Sobre", 75), ("sucursal", "Sucursal", 100),
        ("vendedora", "Vendedora", 120), ("origen", "Origen", 90),
        ("estado", "Estado", 110),
    ):
        anchor = alineacion_pedidos[clave]
        grilla_pedidos.heading(clave, text=titulo, anchor=anchor)
        grilla_pedidos.column(clave, width=ancho, minwidth=ancho, anchor=anchor, stretch=False)
    vacio_pedidos = ctk.CTkLabel(
        marco_pedidos, text="", justify="center", text_color=color_suave,
        font=ctk.CTkFont(size=perfil["fuente_label"] + 1))

    scroll_pedidos = ttk.Scrollbar(marco_pedidos, orient="vertical")
    grilla_pedidos.grid(row=0, column=0, sticky="nsew", padx=(5, 0), pady=5)
    scroll_pedidos.grid(row=0, column=1, sticky="ns", padx=(0, 5), pady=5)

    scroll_pedidos.configure(command=grilla_pedidos.yview)
    grilla_pedidos.configure(yscrollcommand=scroll_pedidos.set)

    def refrescar_pedidos(nombre=None):
        if nombre:
            filtro_pedidos.set(nombre)
        activo = filtro_pedidos.get()
        # La caja se atiende a si misma: por defecto se muestran los pedidos de
        # su propia sucursal. `Todos` es la salida explicita a la vista global.
        caja = contexto_sucursal["caja"] or None
        sucursal = caja if activo != "Todos" else None
        for item in grilla_pedidos.get_children():
            grilla_pedidos.delete(item)
        pedidos = controller.list_orders(activo, branch=sucursal)
        for pedido in pedidos:
            grilla_pedidos.insert("", "end", iid=pedido.id,
                                  tags=(f"estado_{pedido.status.value}",), values=(
                pedido.delivery_date.strftime("%d-%m-%Y"), pedido.customer_name,
                pedido.customer_phone, pedido.customer_document, pedido.envelope, pedido.branch,
                pedido.saleswoman, pedido.origin.value, pedido.status.value,
            ))
        # Que se esta mostrando, y de donde. Sin esto, una lista corta se lee
        # como "no hay nada" cuando en realidad hay un filtro puesto.
        if activo == "Todos":
            contexto_pedidos.pack_forget()
        else:
            etiqueta_contexto_pedidos.configure(
                text=f"Mostrando: {activo} ({len(pedidos)})"
                     + (f"  ·  Caja {sucursal}" if sucursal else ""))
            contexto_pedidos.pack(fill="x", padx=10, pady=(0, 6), before=marco_pedidos)
        if not pedidos:
            vacio_pedidos.configure(text=(
                "No hay pedidos pendientes."
                if activo == FILTRO_REQUIEREN_ATENCION else
                f"No hay pedidos en «{activo}»."))
            vacio_pedidos.place(relx=0.5, rely=0.42, anchor="center")
        else:
            vacio_pedidos.place_forget()
        actualizar_botones_pedido()

    def abrir_pedidos_desde_alerta():
        """La alerta lleva su propio filtro: abre exactamente sus pedidos."""
        seleccionar_pestaña("Pedidos")
        refrescar_pedidos(aviso_pedidos.get("filtro") or FILTRO_REQUIEREN_ATENCION)

    for nombre in (FILTRO_REQUIEREN_ATENCION, "Hoy", "Atrasados", "Próximos", "Todos"):
        ctk.CTkButton(
            barra_pedidos, text=nombre, width=95,
            command=lambda valor=nombre: refrescar_pedidos(valor),
        ).pack(side="left", padx=3)

    botones_estado_pedido = {}

    def estado_pedido_seleccionado():
        seleccion = grilla_pedidos.selection()
        return str(grilla_pedidos.set(seleccion[0], "estado")) if seleccion else ""

    def actualizar_botones_pedido(_event=None):
        estado = estado_pedido_seleccionado()
        botones_estado_pedido["PENDIENTE"].configure(state="normal" if estado in {"LISTO", "ENTREGADO"} else "disabled")
        botones_estado_pedido["LISTO"].configure(state="normal" if estado == "PENDIENTE" else "disabled")
        botones_estado_pedido["ENTREGADO"].configure(state="normal" if estado == "LISTO" else "disabled")

    def cambiar_estado_pedido(estado):
        seleccion = grilla_pedidos.selection()
        if not seleccion:
            messagebox.showwarning("Seleccioná un pedido", "Elegí una fila.", parent=ventana)
            return
        actual = estado_pedido_seleccionado()
        if actual == estado:
            return
        motivo = "Cambio operativo"
        responsable = os.environ.get("USERNAME") or os.environ.get("USER") or "Sistema"
        if actual == "ENTREGADO" and estado == "PENDIENTE":
            if not messagebox.askyesno(
                "Revertir entrega", "¿Corregir este pedido entregado y devolverlo a PENDIENTE?",
                parent=ventana,
            ):
                return
            motivo = simpledialog.askstring(
                "Corrección auditada", "Motivo obligatorio de la corrección:", parent=ventana,
            )
            if not str(motivo or "").strip():
                return
            if not responsable or responsable == "Sistema":
                responsable = simpledialog.askstring(
                    "Corrección auditada", "Usuario responsable:", parent=ventana,
                )
            if not str(responsable or "").strip():
                return
        try:
            controller.update_order_status(
                seleccion[0], estado, reason=motivo, responsible=responsable,
            )
            refrescar_pedidos()
            refrescar_avisos()
        except Exception as exc:
            mostrar_error(exc)

    for estado, texto_boton, color in (
        ("PENDIENTE", "Marcar pendiente", "#D97706"),
        ("LISTO", "Marcar listo", color_verde),
        ("ENTREGADO", "Marcar entregado", color_azul),
    ):
        boton = ctk.CTkButton(
            barra_pedidos, text=texto_boton, width=118,
            command=lambda destino=estado: cambiar_estado_pedido(destino), fg_color=color,
        )
        boton.pack(side="right", padx=3)
        botones_estado_pedido[estado] = boton
    grilla_pedidos.bind("<<TreeviewSelect>>", actualizar_botones_pedido, add="+")
    actualizar_botones_pedido()
    # La sucursal sale del vinculo persistente de esta caja, no de la cajera ni
    # del combo de unidad del formulario. Si nadie la asigno, no se adivina.
    # Se resuelve aca, antes que los avisos, porque tanto la cabecera de Caja
    # como la pestaña de Seguimiento preguntan en que local estan.
    def caja_instalada():
        try:
            return str(controller.admin.setting("branch").get("cashbox", "")).strip()
        except Exception:
            return ""

    contexto_sucursal = {
        "caja": caja_instalada(),
        "sucursal": None,
        "todas": False,
    }
    contexto_sucursal["sucursal"] = (
        controller.tracking.branch_of_register(contexto_sucursal["caja"])
        if contexto_sucursal["caja"] else None
    )
    #: Alerta que la cabecera esta mostrando ahora, para que el clic sepa a que
    #: grupo llevar. Se rellena en cada refresco.
    aviso_principal = {}
    #: Lo mismo para la alerta de Pedidos: cantidad y filtro que la origino.
    aviso_pedidos = {}

    def refrescar_avisos():
        # RC28: la alerta y la vista que abre el clic salen de la misma
        # consulta. Antes la alerta sumaba `Hoy` mas `Atrasados` y el clic
        # filtraba solo `Hoy`, asi que con los vencidos abria en blanco y la
        # operadora tenia que volver a buscar lo que el sistema ya sabia.
        alerta = controller.orders_alert(branch=contexto_sucursal["caja"] or None)
        aviso_pedidos.update(alerta)
        pendientes = alerta["cantidad"]
        aviso_entregas.configure(
            text=f"⚠ Trabajos {pendientes}",
            fg_color="#FFE5A3" if pendientes else "#F7FAFF",
            text_color="#7A4B00" if pendientes else COLOR_TEXTO_SUAVE,
        )
        refrescar_aviso_seguimiento()

    def refrescar_aviso_seguimiento():
        """`⚠ 15 por recibir desde Pilar — clic para ver`, o nada.

        Se muestra la alerta mas urgente de esta sucursal, una sola: dos o tres
        avisos compitiendo en la cabecera vuelven a obligar a decidir cual
        mirar, que es justo lo que la alerta viene a evitar.
        """
        sucursal = contexto_sucursal["sucursal"]
        if not sucursal:
            aviso_seguimiento.pack_forget()
            return
        try:
            pendientes = controller.tracking.pending_actions_for_branch(sucursal)
        except Exception:
            # La cabecera de Caja no se cae porque el circuito falle.
            aviso_seguimiento.pack_forget()
            return
        principal = pendientes.get("principal")
        if not principal:
            aviso_seguimiento.pack_forget()
            return
        aviso_principal.update(principal)
        critico = principal["clave"] == "atrasados"
        aviso_seguimiento.configure(
            text=f"⚠ {principal['texto']} — clic para ver",
            fg_color="#FDECEC" if critico else "#FFF3CD",
            text_color="#A32626" if critico else "#7A4B00",
            border_color="#E5A3A3" if critico else "#E6B85C",
            hover_color="#FBDADA" if critico else "#FFE5A3")
        aviso_seguimiento.pack(side="left", padx=(4, 12))

    def ir_a_pendientes_sucursal():
        """Abre Seguimiento ya filtrado en los trabajos que originaron el aviso."""
        contexto_alerta["filtro"] = aviso_principal.get("filtro") or "Atrasados"
        contexto_alerta["grupo"] = aviso_principal.get("grupo")
        seleccionar_pestaña("Seguimiento")
        ir_a_atrasados()

    refrescar_avisos()

    # ---- Seguimiento RC19: Pilar -> Asuncion -> laboratorio -> Pilar ----
    # Vista operativa, no panel de metricas: arriba las excepciones, en la
    # grilla el laboratorio con su linea y su WhatsApp ya resueltos.
    seguimiento = ctk.CTkFrame(tab_seguimiento, fg_color="transparent")
    seguimiento.pack(fill="both", expand=True, padx=10, pady=8)

    alerta_seguimiento = ctk.CTkLabel(
        seguimiento, text="", height=30, corner_radius=6, anchor="w",
        fg_color="#FDECEC", text_color="#A32626",
        font=ctk.CTkFont(size=perfil["fuente_label"] + 2, weight="bold"),
    )

    resumen_seguimiento = ctk.CTkFrame(seguimiento, fg_color="transparent")
    resumen_seguimiento.pack(fill="x", pady=(0, 6))
    INDICADORES_SEGUIMIENTO = (
        ("atrasados", "Atrasados", "#B42318"),
        ("por_recibir_en_asuncion", "Por recibir", "#B45309"),
        ("en_laboratorio", "En laboratorio", "#0F5FB9"),
        ("confirmados_para_manana", "Confirmados", "#6B5B95"),
        ("listos_para_enviar_a_pilar", "Listos p/ Pilar", color_verde),
        ("en_transito_a_pilar", "En tránsito", "#52657D"),
    )
    etiquetas_seguimiento = {}
    for clave, titulo, color in INDICADORES_SEGUIMIENTO:
        critico = clave == "atrasados"
        tarjeta = ctk.CTkFrame(
            resumen_seguimiento, fg_color="#FFFFFF" if critico else "#F1F5FA",
            corner_radius=6,
        )
        tarjeta.pack(side="left", padx=(0, 6))
        ctk.CTkLabel(
            tarjeta, text=titulo, text_color=color_suave,
            height=perfil["fuente_label"] + 6,
            font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
        ).pack(padx=8, pady=(2, 0))
        valor = ctk.CTkLabel(
            tarjeta, text="0", text_color=color,
            height=(perfil["fuente_kpi"] if critico else perfil["fuente"]) + 8,
            font=ctk.CTkFont(
                size=perfil["fuente_kpi"] if critico else perfil["fuente"], weight="bold",
            ),
        )
        valor.pack(padx=8, pady=(0, 3))
        etiquetas_seguimiento[clave] = (tarjeta, valor)

    recepcion_seguimiento = ctk.CTkLabel(
        resumen_seguimiento, text="", text_color=color_texto, anchor="e",
        font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
    )
    recepcion_seguimiento.pack(side="right", padx=8)

    barra_seguimiento = ctk.CTkFrame(seguimiento, fg_color="transparent")
    barra_seguimiento.pack(fill="x", pady=(0, 6))
    # La vista normal muestra lo que queda por hacer. Los terminados siguen
    # consultables, pero no compiten con el trabajo pendiente.
    filtro_seguimiento = ctk.StringVar(value="Activos")
    FILTROS_SEGUIMIENTO = (
        ("Activos", None), ("Atrasados", "OVERDUE"),
        ("Por recibir", "ENVIADO_DESDE_PILAR"), ("En laboratorio", "EN_LABORATORIO"),
        ("Listos p/ Pilar", "RECIBIDO_DEL_LABORATORIO"), ("En tránsito", "ENVIADO_A_PILAR"),
        ("Recibidos en Asunción", "RECIBIDO_EN_ASUNCION"),
        ("Completados", "COMPLETADOS"), ("Todos", "TODOS"),
    )
    # RC25: de nueve filtros a tres. Los que faltan no desaparecieron: eran
    # etapas, y las etapas ahora se leen como secciones dentro de la propia
    # lista. Lo que queda aca es el alcance, que es otro eje: que tan atras
    # mirar. La operadora no elige etapa, ve todas las suyas de una vez.
    FILTROS_VISIBLES = ("Activos", "Completados", "Todos")
    botones_seguimiento = {}
    for nombre_filtro, _valor in FILTROS_SEGUIMIENTO:
        if nombre_filtro not in FILTROS_VISIBLES:
            continue
        boton_seg = ctk.CTkButton(
            barra_seguimiento, text=nombre_filtro, width=100, height=30,
            fg_color="transparent", hover_color="#EAF3FF", text_color=color_suave,
            border_width=1, border_color=color_borde_suave,
            font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
            command=lambda valor=nombre_filtro: refrescar_seguimiento(valor),
        )
        boton_seg.pack(side="left", padx=2)
        botones_seguimiento[nombre_filtro] = boton_seg

    etiqueta_sucursal = ctk.CTkLabel(
        barra_seguimiento, text="", text_color=color_texto,
        font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
    )
    etiqueta_sucursal.pack(side="right", padx=(10, 6))

    # RC27: el alcance de sucursal deja la barra principal.
    #
    # Seguimiento abre en la sucursal de esta caja, que es lo que la operadora
    # necesita casi siempre. Ver todas las sucursales es una consulta
    # administrativa y ocasional, asi que vive en `Más` y no ocupa un boton
    # permanente al lado de las acciones del dia.
    def alternar_alcance_sucursal():
        contexto_sucursal["todas"] = not contexto_sucursal["todas"]
        refrescar_seguimiento()

    combo_lab_filtro = ctk.CTkComboBox(
        barra_seguimiento, values=["Todos los laboratorios"], width=190, height=30,
        command=lambda _valor: refrescar_seguimiento(),
    )
    combo_lab_filtro.set("Todos los laboratorios")
    combo_lab_filtro.pack(side="right", padx=(6, 0))

    # RC20: los dos accesos que hacen la vista autosuficiente. Son botones, no
    # formularios permanentes: la pantalla sigue siendo la grilla operativa.
    ctk.CTkButton(
        barra_seguimiento, text="Laboratorios", width=130, height=30,
        fg_color="#FFFFFF", text_color=color_azul, border_width=1,
        border_color=color_borde_suave, hover_color="#EAF3FF",
        font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
        command=lambda: abrir_abm_laboratorios(),
    ).pack(side="right", padx=(6, 0))
    ctk.CTkButton(
        barra_seguimiento, text="+ Nuevo envío desde Pilar", width=205, height=30,
        fg_color=color_azul, hover_color="#0F5FC7",
        font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
        command=lambda: abrir_nuevo_envio_pilar(),
    ).pack(side="right", padx=(6, 0))

    # RC25: la recepcion del lote, en una sola linea y con sus dos
    # discrepancias a mano. Aparece solo mientras hay un lote sin terminar de
    # recibir: cuando la recepcion cerro, deja de ocupar pantalla. Recibir no
    # tiene boton propio, porque ya es la "Acción siguiente" de un trabajo
    # ENVIADO DESDE PILAR; aca van las dos salidas que no son el caso normal.
    barra_recepcion = ctk.CTkFrame(seguimiento, fg_color="#FFF9EC", corner_radius=6)
    etiqueta_conciliacion = ctk.CTkLabel(
        barra_recepcion, text="", anchor="w", text_color="#7A4B00",
        font=ctk.CTkFont(size=perfil["fuente_label"] + 1, weight="bold"),
    )
    etiqueta_conciliacion.pack(side="left", padx=(10, 14), pady=6)

    boton_no_llego = ctk.CTkButton(
        barra_recepcion, text="No llegó", width=120, height=30,
        fg_color="#FFFFFF", text_color="#A32626", border_width=1,
        border_color="#E5A3A3", hover_color="#FDECEC",
        font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
        command=lambda: marcar_no_llego(),
    )
    boton_no_llego.pack(side="left", padx=(0, 6), pady=6)

    ctk.CTkButton(
        barra_recepcion, text="+ No estaba en lista", width=180, height=30,
        fg_color="#FFFFFF", text_color=color_azul, border_width=1,
        border_color=color_borde_suave, hover_color="#EAF3FF",
        font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
        command=lambda: abrir_no_estaba_en_lista(),
    ).pack(side="left", pady=6)

    marco_seguimiento = ctk.CTkFrame(seguimiento, fg_color="#FFFFFF", corner_radius=6)
    marco_seguimiento.pack(fill="both", expand=True)
    marco_seguimiento.grid_rowconfigure(0, weight=0)
    marco_seguimiento.grid_rowconfigure(1, weight=1)

    vacio_seguimiento = ctk.CTkLabel(
        marco_seguimiento, text="", justify="center", text_color=color_suave,
        font=ctk.CTkFont(size=perfil["fuente_label"] + 1),
    )
    marco_seguimiento.grid_columnconfigure(0, weight=1)
    # RC21: la tabla queda orientada al circuito logistico. Vendedora sale de
    # la vista (el dato sigue en el dominio y en ventas) y el ancho de Estado
    # se dimensiona al caso mas largo, "CONFIRMADO PARA MAÑANA · RECIBIDO DEL
    # LABORATORIO", para no truncar la etapa fisica.
    # RC26: el telefono entra entre Cliente y Tipo de trabajo. Estaba solo en
    # el detalle, asi que llamar a un cliente obligaba a abrir su ficha; ahora
    # se lee en la misma linea. El ancho sale de las demas, no del total: la
    # tabla tiene que seguir entrando en 1366 sin scroll horizontal.
    COLUMNAS_SEGUIMIENTO = (
        ("sel", "", 34, "center", False),
        ("sobre", "Sobre", 90, "center", False),
        ("cliente", "Cliente", 170, "w", False),
        ("telefono", "Teléfono", 110, "w", False),
        ("tipo", "Tipo de trabajo", 150, "w", False),
        ("laboratorio", "Laboratorio", 150, "w", False),
        ("estado", "Estado", 290, "w", False),
        ("observacion", "Observación", 240, "w", True),
    )
    # Paleta de estados, alineada con RC18: fondo, borde y texto por etapa.
    COLORES_ESTADO_SEGUIMIENTO = {
        "ENVIADO DESDE PILAR": ("#FFF3CD", "#E6B85C", "#7A4B00"),
        "RECIBIDO EN ASUNCIÓN": ("#E7F1FC", "#7DA9D7", "#0F5FB9"),
        "EN LABORATORIO": ("#EDE9F7", "#A99BD1", "#4B3B87"),
        "RECIBIDO DEL LABORATORIO": ("#DDF5E8", "#79C99E", "#17633A"),
        "ENVIADO A PILAR": ("#E4F0FF", "#8FB3D9", "#1E4E8C"),
        "RECIBIDO EN PILAR": ("#D8F3E4", "#4FB884", "#0F5132"),
    }
    COLOR_CHIP_ATRASADO = ("#FDECEC", "#E5A3A3", "#A32626")
    COLOR_CHIP_CONFIRMADO = ("#F3EFFA", "#B9A7E0", "#5B3FA8")

    # Una tabla vacia debe explicar por que lo esta, no dejar a la operadora
    # dudando de si el envio se guardo.
    MENSAJES_VACIO = {
        "Activos": "No hay trabajos en circuito.\n"
                   "Usá  + Nuevo envío desde Pilar  para cargar el lote de una consulta.",
        "Atrasados": "Ningún trabajo está atrasado.",
        "Por recibir": "No hay trabajos pendientes de recibir en Asunción.",
        "En laboratorio": "No hay trabajos en laboratorio.",
        "Listos p/ Pilar": "No hay trabajos listos para enviar a Pilar.",
        "En tránsito": "No hay trabajos en tránsito a Pilar.",
        "Recibidos en Asunción": "No hay trabajos esperando envío a laboratorio.",
        "Completados": "Todavía no hay trabajos recibidos en Pilar.",
        "Todos": "No hay trabajos registrados en el seguimiento.",
    }

    # RC22: la tabla deja de ser un Treeview con chips flotantes.
    #
    # Antes el estado se dibujaba como widget `.place()` sobre el frame que
    # contiene la grilla: vivia en otra capa que el contenido scrollable, no se
    # recortaba al viewport y solo se reubicaba en los eventos interceptados a
    # mano, de modo que cualquier repintado no previsto lo dejaba flotando.
    #
    # Ahora cada fila es un widget real dentro de un frame scrollable: el chip
    # es hijo de su propia fila, asi que se desplaza, se recorta, se repinta y
    # se destruye con ella. No queda ninguna capa flotante que sincronizar.
    # Encabezado y lista van en filas distintas del grid: compartiendo celda,
    # el encabezado tapaba la primera fila de datos.
    encabezado_seguimiento = ctk.CTkFrame(
        marco_seguimiento, fg_color="#EEF4FB", corner_radius=0,
        height=perfil["fuente"] + 20,
    )
    encabezado_seguimiento.grid(row=0, column=0, sticky="ew", padx=5, pady=(5, 0))
    encabezado_seguimiento.grid_propagate(False)

    lista_seguimiento = ctk.CTkScrollableFrame(
        marco_seguimiento, fg_color="#FFFFFF", corner_radius=0,
    )
    lista_seguimiento.grid(row=1, column=0, sticky="nsew", padx=5, pady=(0, 5))
    lista_seguimiento.grid_columnconfigure(0, weight=1)
    for indice, (_clave, titulo, ancho, anclaje, expandible) in enumerate(COLUMNAS_SEGUIMIENTO):
        encabezado_seguimiento.grid_columnconfigure(
            indice, weight=1 if expandible else 0, minsize=ancho)
        ctk.CTkLabel(
            encabezado_seguimiento, text=titulo, anchor="w" if anclaje == "w" else "center",
            text_color=color_texto,
            font=ctk.CTkFont(size=perfil["fuente"], weight="bold"),
        ).grid(row=0, column=indice, sticky="ew", padx=(10, 6), pady=6)
    # El encabezado no lleva peso: con peso, la fila crecia y el frame quedaba
    # centrado dejando bandas vacias arriba y abajo. Todo el sobrante va a la
    # lista, que es la que debe estirarse.
    marco_seguimiento.grid_rowconfigure(0, weight=0)
    marco_seguimiento.grid_rowconfigure(1, weight=1)

    panel_atrasados = ctk.CTkFrame(seguimiento, fg_color="#FFF7F7", corner_radius=6)
    etiqueta_atrasados = ctk.CTkLabel(
        panel_atrasados, text="", anchor="w", justify="left", text_color="#A32626",
        font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
    )
    etiqueta_atrasados.pack(fill="x", padx=10, pady=6)

    acciones_seguimiento = ctk.CTkFrame(seguimiento, fg_color="transparent")
    acciones_seguimiento.pack(fill="x", pady=(6, 0))

    estado_seguimiento = {"filas": {}, "widgets": {}, "secciones": {},
                          "seleccion": None, "marcados": set()}
    # `foco` es el grupo al que la operadora entro desde una alerta o desde un
    # encabezado. None significa la vista normal: todos sus grupos a la vez.
    contexto_alerta = {"filtro": "Atrasados", "grupo": None}
    contexto_grupo = {"foco": None}
    # Lote cuya recepcion esta abierta: es al que se cuelga un fisico que
    # aparecio sin figurar en la lista.
    estado_recepcion = {"shipment": None}

    def enfocar_grupo(clave):
        """Abre exactamente un grupo, o vuelve a la vista completa."""
        contexto_grupo["foco"] = None if contexto_grupo["foco"] == clave else clave
        refrescar_seguimiento()

    def alternar_marca(work_id):
        marcados = estado_seguimiento["marcados"]
        if work_id in marcados:
            marcados.discard(work_id)
        else:
            marcados.add(work_id)
        estado_seguimiento["seleccion"] = work_id
        actualizar_acciones_seguimiento()

    def aplicar_marcas_visibles():
        """Sincroniza los tildes con la seleccion, sin reconstruir la tabla.

        Marcar o desmarcar no cambia que filas hay: solo su tilde. Reconstruir
        la tabla entera para eso costaba 440 widgets y casi un segundo de
        pantalla en blanco, que es el parpadeo que se veia al usar
        `Seleccionar todo` o `Limpiar selección`.
        """
        marcados = estado_seguimiento["marcados"]
        for identificador, widgets in estado_seguimiento["widgets"].items():
            marca = widgets.get("marca")
            if marca is None:
                continue
            if identificador in marcados:
                marca.select()
            else:
                marca.deselect()
        actualizar_acciones_seguimiento()

    def seleccionar_todo():
        estado_seguimiento["marcados"] = set(estado_seguimiento["filas"])
        aplicar_marcas_visibles()

    def limpiar_seleccion():
        estado_seguimiento["marcados"].clear()
        aplicar_marcas_visibles()

    def trabajo_seleccionado():
        return estado_seguimiento["filas"].get(estado_seguimiento["seleccion"])

    def seleccionar_fila(work_id):
        estado_seguimiento["seleccion"] = work_id
        for identificador, widgets in estado_seguimiento["widgets"].items():
            elegido = identificador == work_id
            widgets["fila"].configure(
                fg_color=widgets["fondo_activo"] if elegido else widgets["fondo"])
        actualizar_acciones_seguimiento()

    def responsable_actual():
        return os.environ.get("BC_CAJA_RESPONSABLE") or os.environ.get("USERNAME") or "Operadora"

    def refrescar_seguimiento(nombre=None):
        if nombre:
            filtro_seguimiento.set(nombre)
        activo = filtro_seguimiento.get()
        for etiqueta, boton_filtro in botones_seguimiento.items():
            seleccionado = etiqueta == activo
            boton_filtro.configure(
                fg_color="#EAF3FF" if seleccionado else "transparent",
                text_color=color_azul if seleccionado else color_suave,
            )
        catalogo = controller.tracking.list_laboratories()
        combo_lab_filtro.configure(
            values=["Todos los laboratorios"] + [lab.name for lab in catalogo]
        )
        elegido = combo_lab_filtro.get()
        laboratory_id = next(
            (lab.id for lab in catalogo if lab.name == elegido), None,
        )
        valor_filtro = dict(FILTROS_SEGUIMIENTO).get(activo)
        if valor_filtro == "COMPLETADOS":
            alcance, estado, solo_atrasados = "COMPLETADOS", None, False
        elif valor_filtro == "TODOS":
            alcance, estado, solo_atrasados = "TODOS", None, False
        elif valor_filtro == "OVERDUE":
            alcance, estado, solo_atrasados = "ACTIVOS", None, True
        else:
            alcance, estado, solo_atrasados = "ACTIVOS", valor_filtro, False
        sucursal = contexto_sucursal["sucursal"]
        local = None if (contexto_sucursal["todas"] or not sucursal) else sucursal
        tablero = controller.tracking.board(
            status=estado, only_overdue=solo_atrasados,
            laboratory_id=laboratory_id, scope=alcance,
            responsible_branch=local, group=contexto_grupo["foco"],
        )
        if not sucursal:
            etiqueta_sucursal.configure(
                text=f"Caja {contexto_sucursal['caja'] or '—'} sin sucursal asignada",
                text_color="#B45309")
        else:
            etiqueta_sucursal.configure(
                text=("Todas las sucursales" if contexto_sucursal["todas"]
                      else f"Sucursal: {sucursal}"),
                text_color=color_texto)
        # RC26: las filas se reutilizan en vez de reconstruirse.
        #
        # Cada fila son ~29 widgets Tk y la tabla entera ronda los 440.
        # Destruirlos y volver a crearlos en cada refresco costaba ~0,9 s con
        # la tabla vacia a la vista: eso es lo que se veia como parpadeo, y
        # ocurria incluso cuando la lista era exactamente la misma de antes.
        #
        # Ahora el widget de cada trabajo se crea una sola vez y despues solo
        # se repinta: se le cambian textos, colores y posicion. Solo se crean
        # los trabajos que aparecen y solo se destruyen los que dejan de
        # figurar, que es lo unico que de verdad cambio.
        lista_seguimiento.grid_remove()
        try:
            def crear_fila(identificador):
                """Esqueleto de una fila. Se construye una vez por trabajo."""
                marco = ctk.CTkFrame(
                    lista_seguimiento, corner_radius=0, height=perfil["fuente"] + 22)
                # Marca de pertenencia: la sonda verifica que cada chip cuelgue
                # de una fila y no de una capa flotante.
                marco._bc_fila_seguimiento = True
                marco.grid_propagate(False)
                for columna, (_c, _t, ancho, _a, expandible) in enumerate(
                        COLUMNAS_SEGUIMIENTO):
                    marco.grid_columnconfigure(
                        columna, weight=1 if expandible else 0, minsize=ancho)
                marca = ctk.CTkCheckBox(
                    marco, text="", width=22, checkbox_width=18, checkbox_height=18,
                    command=lambda w=identificador: alternar_marca(w))
                marca.grid(row=0, column=0, padx=(8, 0), pady=4)
                celdas = []
                for columna in range(1, 6):
                    etiqueta = ctk.CTkLabel(
                        marco, text="", anchor="w",
                        font=ctk.CTkFont(size=perfil["fuente"]))
                    etiqueta.grid(row=0, column=columna, sticky="ew",
                                  padx=(10, 6), pady=5)
                    celdas.append(etiqueta)
                # La operadora lee la novedad sin abrir el detalle.
                observacion = ctk.CTkLabel(
                    marco, text="", anchor="w", text_color=color_suave,
                    font=ctk.CTkFont(size=perfil["fuente"]))
                observacion.grid(row=0, column=7, sticky="ew", padx=(10, 6), pady=5)
                # El chip vive dentro de la fila: se desplaza y se destruye con ella.
                celda_estado = ctk.CTkFrame(marco, fg_color="transparent")
                celda_estado.grid(row=0, column=6, sticky="w", padx=(10, 6), pady=4)
                fuente_chip = ctk.CTkFont(
                    size=max(8, perfil["fuente_label"] - 1), weight="bold")
                chip_alerta = ctk.CTkLabel(
                    celda_estado, text="", corner_radius=7, height=20, font=fuente_chip)
                chip_estado = ctk.CTkLabel(
                    celda_estado, text="", corner_radius=7, height=20, font=fuente_chip)
                chip_estado.pack(side="left")
                for widget in (marco, marca, *celdas, observacion, celda_estado,
                               chip_alerta, chip_estado):
                    widget.bind("<Button-1>",
                                lambda _e, w=identificador: seleccionar_fila(w), add="+")
                    widget.bind("<Double-Button-1>",
                                lambda _e, w=identificador: (seleccionar_fila(w),
                                                             abrir_detalle_trabajo()),
                                add="+")
                return {"fila": marco, "marca": marca, "celdas": celdas,
                        "observacion": observacion, "chip_alerta": chip_alerta,
                        "chip_estado": chip_estado,
                        "fondo": "#FFFFFF", "fondo_activo": "#DCEBFA"}

            def pintar_fila(fila, indice, posicion):
                """Deja la fila diciendo lo que corresponde, sin recrearla."""
                identificador = fila.work.id
                widgets = estado_seguimiento["widgets"].get(identificador)
                if widgets is None:
                    widgets = crear_fila(identificador)
                    estado_seguimiento["widgets"][identificador] = widgets
                fondo = "#FFF5F5" if fila.overdue else (
                    "#FFFFFF" if posicion % 2 else "#FAFCFE")
                widgets["fondo"] = fondo
                elegida = estado_seguimiento["seleccion"] == identificador
                widgets["fila"].configure(
                    fg_color=widgets["fondo_activo"] if elegida else fondo)
                widgets["fila"].grid(row=indice, column=0, sticky="ew", pady=(0, 1))
                if identificador in estado_seguimiento["marcados"]:
                    widgets["marca"].select()
                else:
                    widgets["marca"].deselect()
                color_fila = "#A32626" if fila.overdue else color_texto
                for etiqueta, texto in zip(widgets["celdas"], (
                    fila.envelope, fila.customer_name, fila.customer_phone or "—",
                    fila.work_type, fila.laboratory_name,
                )):
                    etiqueta.configure(text=texto, text_color=color_fila)
                widgets["observacion"].configure(text=fila.observation)
                if fila.alert:
                    fondo_chip, _b, texto_chip = (
                        COLOR_CHIP_ATRASADO if fila.overdue else COLOR_CHIP_CONFIRMADO)
                    widgets["chip_alerta"].configure(
                        text=f"  {fila.alert}  ", fg_color=fondo_chip,
                        text_color=texto_chip)
                    widgets["chip_alerta"].pack(
                        side="left", padx=(0, 4), before=widgets["chip_estado"])
                else:
                    widgets["chip_alerta"].pack_forget()
                fondo_chip, _b, texto_chip = COLORES_ESTADO_SEGUIMIENTO.get(
                    fila.physical_status, ("#EEF4FB", "#B9CDE5", "#132238"))
                widgets["chip_estado"].configure(
                    text=f"  {fila.physical_status}  ", fg_color=fondo_chip,
                    text_color=texto_chip)
                estado_seguimiento["filas"][identificador] = fila

            def pintar_grupo(clave_grupo, titulo_grupo, cantidad, indice):
                """Encabezado de seccion, tambien reutilizado."""
                cabecera_grupo = estado_seguimiento["secciones"].get(clave_grupo)
                if cabecera_grupo is None:
                    cabecera_grupo = ctk.CTkFrame(
                        lista_seguimiento, corner_radius=4,
                        height=perfil["fuente"] + 14)
                    cabecera_grupo.grid_propagate(False)
                    etiqueta = ctk.CTkLabel(
                        cabecera_grupo, text="", anchor="w",
                        font=ctk.CTkFont(size=perfil["fuente"], weight="bold"))
                    etiqueta.pack(side="left", padx=6, pady=2)
                    for widget in (cabecera_grupo, etiqueta):
                        widget.configure(cursor="hand2")
                        widget.bind("<Button-1>",
                                    lambda _e, g=clave_grupo: enfocar_grupo(g), add="+")
                    cabecera_grupo._bc_etiqueta = etiqueta
                    estado_seguimiento["secciones"][clave_grupo] = cabecera_grupo
                enfocado = contexto_grupo["foco"] == clave_grupo
                cabecera_grupo.configure(
                    fg_color="#EEF4FB" if enfocado else "#F7FAFF")
                cabecera_grupo._bc_etiqueta.configure(
                    text=f"  {titulo_grupo}  ·  {cantidad}"
                         + ("     (viendo solo este grupo — clic para ver todo)"
                            if enfocado else ""),
                    text_color=color_azul if enfocado else color_texto)
                cabecera_grupo.grid(
                    row=indice, column=0, sticky="ew", pady=(6 if indice else 0, 2))

            # RC25: la lista se recorre como el circuito. Cada grupo abre con su
            # encabezado y su cantidad, de modo que la operadora ve de una vez que
            # etapas tienen trabajo esperandola sin cambiar de pantalla ni elegir
            # un filtro. Dentro de cada grupo mandan las excepciones, que es el
            # orden en que el tablero ya devolvio las filas.
            estado_seguimiento["filas"].clear()
            por_grupo = {clave: [] for clave, _t, _e in GRUPOS_SEGUIMIENTO}
            for fila in tablero["rows"]:
                por_grupo[fila.group].append(fila)
            indice = 0
            for clave_grupo, titulo_grupo, _etapa in GRUPOS_SEGUIMIENTO:
                filas_grupo = por_grupo[clave_grupo]
                if not filas_grupo:
                    cabecera = estado_seguimiento["secciones"].get(clave_grupo)
                    if cabecera is not None:
                        cabecera.grid_remove()
                    continue
                pintar_grupo(clave_grupo, titulo_grupo, len(filas_grupo), indice)
                indice += 1
                for posicion, fila in enumerate(filas_grupo):
                    pintar_fila(fila, indice, posicion)
                    indice += 1
            # Solo desaparece de la pantalla lo que de verdad dejo de estar.
            for identificador in set(estado_seguimiento["widgets"]) - set(
                    estado_seguimiento["filas"]):
                estado_seguimiento["widgets"].pop(identificador)["fila"].destroy()
        finally:
            lista_seguimiento.grid()
        estado_seguimiento["marcados"] &= set(estado_seguimiento["filas"])
        if estado_seguimiento["seleccion"] not in estado_seguimiento["filas"]:
            estado_seguimiento["seleccion"] = None
        for clave, _titulo, _color in INDICADORES_SEGUIMIENTO:
            tarjeta, valor = etiquetas_seguimiento[clave]
            cantidad = tablero["summary"][clave]
            valor.configure(text=str(cantidad))
            if clave == "atrasados":
                tarjeta.configure(fg_color="#FDECEC" if cantidad else "#FFFFFF")
        # La conciliacion se muestra mientras la recepcion sigue abierta. Es
        # una linea, no un tablero: `Declarados 15 · Recibidos 14 · No llegó 1
        # · Extra 1` alcanza para saber si el lote cerro.
        recepcion_actual = controller.tracking.current_reception(sucursal)
        estado_recepcion["shipment"] = (
            recepcion_actual["shipment"].id if recepcion_actual["shipment"] else None)
        if recepcion_actual["shipment"] is not None:
            etiqueta_conciliacion.configure(text=recepcion_actual["line"])
            barra_recepcion.pack(fill="x", pady=(0, 6), before=marco_seguimiento)
            # Una sola cuenta de recepcion a la vez. El resumen de arriba mide
            # lo que hay en pantalla y la conciliacion mide el lote declarado:
            # juntos daban dos totales distintos para la misma palabra.
            recepcion_seguimiento.pack_forget()
        else:
            barra_recepcion.pack_forget()
            recepcion = tablero["reception"]
            recepcion_seguimiento.configure(
                text="Enviados: {enviados}    Recibidos: {recibidos}    "
                     "Falta recibir: {falta_recibir}".format(**recepcion)
            )
            recepcion_seguimiento.pack(side="right", padx=8)
        # Las alertas son de este local: se calculan por proxima accion
        # pendiente de su sucursal, no por existencia global del trabajo.
        pendientes = (
            controller.tracking.pending_actions_for_branch(sucursal)
            if sucursal and not contexto_sucursal["todas"] else None
        )
        if pendientes and pendientes["alertas"]:
            principal = pendientes["alertas"][0]
            contexto_alerta["filtro"] = principal["filtro"]
            contexto_alerta["grupo"] = principal["grupo"]
            resto = len(pendientes["alertas"]) - 1
            texto = f"  ⚠  {principal['texto']}"
            if resto:
                texto += f"   ·   y {resto} pendiente(s) más en {sucursal}"
            alerta_seguimiento.configure(
                text=texto + "   ·   clic para ver estos",
                fg_color="#FDECEC" if principal["clave"] == "atrasados" else "#FFF7E8",
                text_color="#A32626" if principal["clave"] == "atrasados" else "#7A4B00")
            alerta_seguimiento.pack(fill="x", pady=(0, 6), before=resumen_seguimiento)
        elif tablero["alert"] and contexto_sucursal["todas"]:
            contexto_alerta["filtro"] = "Atrasados"
            contexto_alerta["grupo"] = None
            alerta_seguimiento.configure(
                text="  ⚠  " + tablero["alert"] + "   ·   clic para ver solo estos",
                fg_color="#FDECEC", text_color="#A32626")
            alerta_seguimiento.pack(fill="x", pady=(0, 6), before=resumen_seguimiento)
        else:
            alerta_seguimiento.pack_forget()
        if not tablero["rows"]:
            # RC27: una tabla vacia tiene que decir de que sucursal esta
            # hablando. Antes se leia como "no hay nada" y en realidad podia
            # haber trabajo en el otro local: la operadora no tiene por que
            # deducir que la vista esta acotada a su caja.
            if activo == "Activos" and sucursal and not contexto_sucursal["todas"]:
                texto_vacio = (f"No hay trabajos pendientes en {sucursal.title()}."
                               "\nUsá  Más ▸ Ver todas las sucursales  si "
                               "necesitás mirar el otro local.")
            else:
                texto_vacio = MENSAJES_VACIO.get(
                    activo, "No hay trabajos que mostrar con este filtro.")
            vacio_seguimiento.configure(text=texto_vacio)
            vacio_seguimiento.place(relx=0.5, rely=0.42, anchor="center")
        else:
            vacio_seguimiento.place_forget()
        grupos = tablero["overdue_groups"]
        if grupos:
            etiqueta_atrasados.configure(text="     ".join(
                "{name} — {count} atrasado{plural}   ☎ {phone_line}   ✆ {whatsapp}".format(
                    plural="" if grupo["count"] == 1 else "s", **grupo,
                )
                for grupo in grupos
            ))
            panel_atrasados.pack(fill="x", pady=(6, 0), before=acciones_seguimiento)
        else:
            panel_atrasados.pack_forget()
        # La cabecera de Caja y esta pestaña leen la misma fuente y se refrescan
        # juntas: no pueden quedar diciendo cosas distintas.
        refrescar_aviso_seguimiento()
        actualizar_acciones_seguimiento()

    def marcar_no_llego():
        """Lo que figuraba en el envio y no aparecio, marcado en bloque.

        No avanza de etapa: queda ligado al lote como NO LLEGÓ y bloquea su
        propio avance hasta que aparezca o se corrija.
        """
        ids = seleccion_actual()
        if not ids:
            messagebox.showwarning(
                "No llegó", "Marcá los trabajos que no aparecieron en el envío.",
                parent=ventana)
            return
        if not messagebox.askyesno(
                "No llegó",
                f"¿Marcar {len(ids)} trabajo(s) como NO LLEGÓ?\n\n"
                "Quedan ligados al envío y no avanzan hasta resolverse.",
                parent=ventana):
            return
        try:
            controller.tracking.mark_batch_not_arrived(
                ids, responsible=responsable_actual())
        except Exception as exc:
            mostrar_error(exc)
            return
        refrescar_seguimiento()

    def abrir_no_estaba_en_lista():
        """El fisico que aparecio sin figurar: se busca el pedido y se reutiliza.

        Primero se busca por Sobre o cliente entre los pedidos ya cargados. Si
        el pedido existe —y casi siempre existe— no se vuelve a escribir
        cliente ni receta: se cuelga ese mismo pedido al lote.
        """
        dialogo = ctk.CTkToplevel(ventana)
        dialogo.title("Recibir un trabajo que no estaba en la lista")
        dialogo.geometry("760x520")
        dialogo.transient(ventana)
        dialogo.grab_set()

        ctk.CTkLabel(
            dialogo, text="Buscá el trabajo por Sobre o por cliente",
            text_color=color_texto,
            font=ctk.CTkFont(size=perfil["fuente_seccion"], weight="bold"),
        ).pack(anchor="w", padx=14, pady=(12, 2))
        ctk.CTkLabel(
            dialogo, text="El pedido ya está cargado: no hace falta volver a "
                          "escribir cliente ni receta.",
            text_color=color_suave, font=ctk.CTkFont(size=perfil["fuente_label"]),
        ).pack(anchor="w", padx=14, pady=(0, 8))

        fila_busqueda = ctk.CTkFrame(dialogo, fg_color="transparent")
        fila_busqueda.pack(fill="x", padx=14)
        campo_busqueda = ctk.CTkEntry(
            fila_busqueda, height=32, placeholder_text="Sobre o cliente")
        campo_busqueda.pack(side="left", fill="x", expand=True)

        lista = ctk.CTkScrollableFrame(dialogo, fg_color="#FFFFFF")
        lista.pack(fill="both", expand=True, padx=14, pady=(8, 6))
        elegido = {"order_id": None}
        estado_busqueda = ctk.CTkLabel(
            dialogo, text="", anchor="w", text_color=color_suave,
            font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"))
        estado_busqueda.pack(anchor="w", padx=14)

        def elegir(order_id, etiqueta):
            elegido["order_id"] = order_id
            estado_busqueda.configure(
                text=f"Seleccionado: {etiqueta}", text_color=color_verde)

        def buscar(_event=None):
            for hijo in lista.winfo_children():
                hijo.destroy()
            elegido["order_id"] = None
            termino = campo_busqueda.get().strip()
            if not termino:
                estado_busqueda.configure(
                    text="Escribí un sobre o un cliente para buscar.",
                    text_color=color_suave)
                return
            try:
                candidatos = controller.tracking.search_receivable_orders(termino)
            except Exception as exc:
                mostrar_error(exc)
                return
            if not candidatos:
                estado_busqueda.configure(
                    text="Ningún pedido pendiente coincide. "
                         "Puede que ya esté en el circuito.",
                    text_color="#B45309")
                return
            estado_busqueda.configure(
                text=f"{len(candidatos)} pedido(s). Elegí el que llegó.",
                text_color=color_suave)
            for pedido in candidatos:
                etiqueta = (f"{pedido.envelope or '—'}   ·   {pedido.customer_name}"
                            f"   ·   {pedido.branch}")
                ctk.CTkButton(
                    lista, text=etiqueta, anchor="w", height=30,
                    fg_color="transparent", text_color=color_texto,
                    hover_color="#EAF3FF",
                    font=ctk.CTkFont(size=perfil["fuente"]),
                    command=lambda o=pedido.id, e=etiqueta: elegir(o, e),
                ).pack(fill="x", pady=1)

        ctk.CTkButton(
            fila_busqueda, text="Buscar", width=110, height=32,
            fg_color=color_azul, hover_color="#0F5FC7",
            font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
            command=buscar,
        ).pack(side="left", padx=(6, 0))
        campo_busqueda.bind("<Return>", buscar, add="+")

        def confirmar():
            if not elegido["order_id"]:
                messagebox.showwarning(
                    "Elegí el trabajo", "Buscá y seleccioná el pedido que llegó.",
                    parent=dialogo)
                return
            try:
                controller.tracking.add_unlisted_reception(
                    elegido["order_id"], responsible=responsable_actual(),
                    shipment_id=estado_recepcion["shipment"])
            except Exception as exc:
                mostrar_error(exc)
                return
            dialogo.destroy()
            refrescar_seguimiento()

        acciones = ctk.CTkFrame(dialogo, fg_color="transparent")
        acciones.pack(fill="x", padx=14, pady=12, side="bottom")
        ctk.CTkButton(
            acciones, text="Recibir este trabajo", width=200, height=34,
            fg_color=color_verde, hover_color="#128A57",
            font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
            command=confirmar,
        ).pack(side="left")
        ctk.CTkButton(
            acciones, text="Cancelar", width=120, height=34, fg_color="#52657D",
            command=dialogo.destroy,
        ).pack(side="right")
        campo_busqueda.focus_set()
        return dialogo

    def accion_seguimiento(operacion):
        fila = trabajo_seleccionado()
        if fila is None:
            messagebox.showwarning(
                "Seleccioná un trabajo", "Elegí una fila.", parent=ventana,
            )
            return
        try:
            operacion(fila)
        except Exception as exc:
            mostrar_error(exc)
            return
        refrescar_seguimiento()

    def enviar_a_laboratorio(fila):
        catalogo = controller.tracking.list_laboratories(only_active=True)
        if not catalogo:
            raise InvalidCashDayError(
                "Cargá al menos un laboratorio antes de enviar trabajos."
            )
        nombres = ", ".join(lab.name for lab in catalogo)
        elegido = simpledialog.askstring(
            "Enviar a laboratorio", f"Laboratorio ({nombres}):", parent=ventana,
        )
        laboratorio = next(
            (lab for lab in catalogo
             if lab.name.casefold() == str(elegido or "").strip().casefold()), None,
        )
        if laboratorio is None:
            return
        fecha = simpledialog.askstring(
            "Fecha esperada", "Fecha esperada (dd-mm-aaaa):", parent=ventana,
            initialvalue=date.today().strftime("%d-%m-%Y"),
        )
        if not str(fecha or "").strip():
            return
        hora = simpledialog.askstring(
            "Hora esperada", "Hora esperada (HH:MM):", parent=ventana,
            initialvalue=controller.tracking.default_expected_time().strftime("%H:%M"),
        )
        controller.tracking.send_to_laboratory(
            fila.work.id, laboratorio.id, expected_date=fecha, expected_time=hora or None,
            responsible=responsable_actual(),
        )

    def abrir_novedad(work_ids):
        """Una misma respuesta del laboratorio puede aplicar a varios trabajos."""
        if not work_ids:
            messagebox.showwarning(
                "Novedad", "Marcá uno o varios trabajos.", parent=ventana)
            return
        dialogo = ctk.CTkToplevel(ventana)
        dialogo.title(f"Novedad · {len(work_ids)} trabajo(s)")
        dialogo.geometry("560x420")
        dialogo.transient(ventana)
        dialogo.grab_set()

        ctk.CTkLabel(
            dialogo, text=f"Registrar novedad para {len(work_ids)} trabajo(s)",
            text_color=color_texto,
            font=ctk.CTkFont(size=perfil["fuente_seccion"], weight="bold"),
        ).pack(anchor="w", padx=14, pady=(12, 6))

        medio = ctk.StringVar(value="LLAMADA")
        fila_medio = ctk.CTkFrame(dialogo, fg_color="transparent")
        fila_medio.pack(fill="x", padx=14)
        ctk.CTkLabel(
            fila_medio, text="Medio", text_color=color_suave,
            font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
        ).pack(side="left", padx=(0, 8))
        for etiqueta, valor in (("Llamada", "LLAMADA"), ("WhatsApp", "WHATSAPP"),
                                ("Otro", "OTRO")):
            ctk.CTkRadioButton(
                fila_medio, text=etiqueta, variable=medio, value=valor,
                font=ctk.CTkFont(size=perfil["fuente"]),
            ).pack(side="left", padx=6)

        ctk.CTkLabel(
            dialogo, text="Respuesta del laboratorio", text_color=color_suave,
            font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
        ).pack(anchor="w", padx=14, pady=(12, 2))
        campo_respuesta = ctk.CTkEntry(dialogo, height=32)
        campo_respuesta.pack(fill="x", padx=14)

        plazo = {"fecha": None, "hora": None}
        etiqueta_plazo = ctk.CTkLabel(
            dialogo, text="Sin nuevo plazo", text_color=color_texto,
            font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"))

        def fijar(fecha, hora, texto):
            plazo["fecha"], plazo["hora"] = fecha, hora
            etiqueta_plazo.configure(text=texto)

        def elegir_manual():
            fecha = simpledialog.askstring(
                "Nuevo plazo", "Fecha (dd-mm-aaaa):", parent=dialogo,
                initialvalue=date.today().strftime("%d-%m-%Y"))
            if not str(fecha or "").strip():
                return
            hora = simpledialog.askstring(
                "Nuevo plazo", "Hora (HH:MM):", parent=dialogo,
                initialvalue=controller.tracking.default_expected_time().strftime("%H:%M"))
            fijar(fecha, hora or None, f"Nuevo plazo: {fecha} {hora or ''}".strip())

        atajos = ctk.CTkFrame(dialogo, fg_color="transparent")
        atajos.pack(fill="x", padx=14, pady=(12, 4))
        hoy = date.today()
        manana = hoy + timedelta(days=1)
        for texto, accion in (
            ("Más tarde hoy", lambda: fijar(
                hoy.strftime("%d-%m-%Y"), "18:00", "Nuevo plazo: hoy 18:00")),
            ("Mañana", lambda: fijar(
                manana.strftime("%d-%m-%Y"),
                controller.tracking.default_expected_time().strftime("%H:%M"),
                "Nuevo plazo: mañana "
                + controller.tracking.default_expected_time().strftime("%H:%M"))),
            ("Elegir fecha/hora", elegir_manual),
            ("Solo observación", lambda: fijar(None, None, "Sin nuevo plazo")),
        ):
            ctk.CTkButton(
                atajos, text=texto, width=132, height=30,
                fg_color="#FFFFFF", text_color=color_azul, border_width=1,
                border_color=color_borde_suave, hover_color="#EAF3FF",
                font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
                command=accion,
            ).pack(side="left", padx=3)
        etiqueta_plazo.pack(anchor="w", padx=14, pady=(8, 0))

        def guardar():
            respuesta = campo_respuesta.get().strip()
            try:
                for work_id in work_ids:
                    controller.tracking.register_contact(
                        work_id, operator=responsable_actual(), channel=medio.get(),
                        result=respuesta, next_expected_date=plazo["fecha"],
                        next_expected_time=plazo["hora"])
            except Exception as exc:
                mostrar_error(exc)
                return
            dialogo.destroy()
            refrescar_seguimiento()

        acciones = ctk.CTkFrame(dialogo, fg_color="transparent")
        acciones.pack(fill="x", padx=14, pady=14, side="bottom")
        ctk.CTkButton(
            acciones, text="Guardar novedad", width=170, height=34,
            fg_color=color_verde, hover_color="#128A57",
            font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
            command=guardar,
        ).pack(side="left")
        ctk.CTkButton(
            acciones, text="Cancelar", width=120, height=34, fg_color="#52657D",
            command=dialogo.destroy,
        ).pack(side="right")
        return dialogo

    def pedir_motivo(titulo, detalle):
        """Motivo obligatorio. Sin texto no hay excepcion ni correccion."""
        motivo = simpledialog.askstring(titulo, detalle, parent=ventana)
        if motivo is None:
            return None
        if not motivo.strip():
            messagebox.showwarning(
                titulo, "El motivo es obligatorio: queda en el historial.",
                parent=ventana)
            return None
        return motivo.strip()

    def marcar_queda_a_confirmar():
        """El trabajo esta en la óptica pero el cliente todavia no confirmó."""
        ids = seleccion_actual()
        if not ids:
            messagebox.showwarning(
                "Queda a confirmar", "Marcá uno o varios trabajos.", parent=ventana)
            return
        nota = simpledialog.askstring(
            "Queda a confirmar",
            f"{len(ids)} trabajo(s).\n\n¿Qué se está esperando?\n"
            "Por ejemplo: Cliente confirma mañana · Esperando llamada ·\n"
            "Falta confirmar cristal · Esperando autorización",
            parent=ventana)
        if nota is None:
            return
        try:
            controller.tracking.mark_awaiting_confirmation(
                ids, responsible=responsable_actual(), note=nota)
        except Exception as exc:
            mostrar_error(exc)
            return
        refrescar_seguimiento()

    def resolver_confirmacion(ids):
        """Dos caminos válidos: confirmó o canceló. Ninguno más."""
        dialogo = ctk.CTkToplevel(ventana)
        dialogo.title(f"Resolver confirmación · {len(ids)} trabajo(s)")
        dialogo.geometry("560x260")
        dialogo.transient(ventana)
        dialogo.grab_set()
        ctk.CTkLabel(
            dialogo, text=f"¿Qué pasó con {len(ids)} trabajo(s)?",
            text_color=color_texto,
            font=ctk.CTkFont(size=perfil["fuente_seccion"], weight="bold"),
        ).pack(anchor="w", padx=16, pady=(16, 4))
        ctk.CTkLabel(
            dialogo, text="Estaban esperando la confirmación del cliente.",
            text_color=color_suave, font=ctk.CTkFont(size=perfil["fuente_label"]),
        ).pack(anchor="w", padx=16, pady=(0, 14))

        def confirmo():
            datos = pedir_destino_laboratorio(len(ids))
            if datos is None:
                return
            try:
                controller.tracking.resolve_confirmation_confirmed(
                    ids, responsible=responsable_actual(), **datos)
            except Exception as exc:
                mostrar_error(exc)
                return
            dialogo.destroy()
            refrescar_seguimiento()

        def cancelo():
            motivo = pedir_motivo(
                "Canceló",
                "Motivo (obligatorio).\nPor ejemplo: devuelto a exhibición, "
                "cliente desistió, trabajo sin efecto:")
            if motivo is None:
                return
            try:
                controller.tracking.resolve_confirmation_cancelled(
                    ids, responsible=responsable_actual(), reason=motivo)
            except Exception as exc:
                mostrar_error(exc)
                return
            dialogo.destroy()
            refrescar_seguimiento()

        ctk.CTkButton(
            dialogo, text="Confirmó  —  enviar a laboratorio", height=42,
            fg_color=color_verde, hover_color="#128A57",
            font=ctk.CTkFont(size=perfil["fuente_label"] + 1, weight="bold"),
            command=confirmo,
        ).pack(fill="x", padx=16, pady=(0, 8))
        ctk.CTkButton(
            dialogo, text="Canceló  —  cerrar por excepción", height=42,
            fg_color="#FFFFFF", text_color="#A32626", border_width=1,
            border_color="#E5A3A3", hover_color="#FDECEC",
            font=ctk.CTkFont(size=perfil["fuente_label"] + 1, weight="bold"),
            command=cancelo,
        ).pack(fill="x", padx=16)
        ctk.CTkButton(
            dialogo, text="Todavía no se sabe", height=32, fg_color="#52657D",
            font=ctk.CTkFont(size=perfil["fuente_label"]),
            command=dialogo.destroy,
        ).pack(fill="x", padx=16, pady=12, side="bottom")
        return dialogo

    def cerrar_por_excepcion():
        """Salida para lo que no llegó a completarse. Nunca el cierre normal."""
        ids = seleccion_actual()
        if not ids:
            messagebox.showwarning(
                "Cerrar por excepción", "Marcá uno o varios trabajos.", parent=ventana)
            return
        motivo = pedir_motivo(
            "Cerrar por excepción",
            f"{len(ids)} trabajo(s).\n\nMotivo (obligatorio).\n"
            "Cancelación, devolución a exhibición, trabajo sin efecto\n"
            "o corrección administrativa:")
        if motivo is None:
            return
        if not messagebox.askyesno(
                "Cerrar por excepción",
                f"¿Cerrar {len(ids)} trabajo(s) por excepción?\n\n"
                f"Motivo: {motivo}\nResponsable: {responsable_actual()}\n\n"
                "Queda registrado en el historial del trabajo.",
                parent=ventana):
            return
        try:
            for work_id in ids:
                controller.tracking.close_by_exception(
                    work_id, responsible=responsable_actual(), reason=motivo)
        except Exception as exc:
            mostrar_error(exc)
            return
        refrescar_seguimiento()

    def corregir_estado():
        """Solo retrocesos declarados, con motivo y responsable."""
        fila = trabajo_seleccionado()
        if fila is None:
            messagebox.showwarning(
                "Corregir estado", "Elegí una fila.", parent=ventana)
            return
        try:
            destinos = controller.tracking.correctable_targets(fila.work.id)
        except Exception as exc:
            mostrar_error(exc)
            return
        if not destinos:
            messagebox.showinfo(
                "Corregir estado",
                f"{fila.physical_status} no admite retroceso.\n\n"
                "Si hace falta un ajuste, usá Cerrar por excepción.",
                parent=ventana)
            return
        opciones = {ETIQUETAS_ESTADO_UI[e]: e for e in destinos}
        elegido = simpledialog.askstring(
            "Corregir estado",
            f"{fila.envelope} está en {fila.physical_status}.\n\n"
            f"Retroceder a ({', '.join(opciones)}):",
            parent=ventana, initialvalue=next(iter(opciones)))
        destino = opciones.get(str(elegido or "").strip().upper())
        if destino is None:
            return
        motivo = pedir_motivo(
            "Corregir estado",
            f"{fila.envelope}: {fila.physical_status} → {ETIQUETAS_ESTADO_UI[destino]}\n\n"
            "Motivo (obligatorio):")
        if motivo is None:
            return
        try:
            controller.tracking.correct_status(
                fila.work.id, destino,
                responsible=responsable_actual(), reason=motivo)
        except Exception as exc:
            mostrar_error(exc)
            return
        refrescar_seguimiento()

    # RC24: tres acciones conceptuales en vez de seis transiciones.
    # La operadora identifica el trabajo, lo marca y ejecuta lo sugerido; el
    # sistema decide cual es la transicion valida via next_action.
    resumen_accion = {"action": None, "ids": []}

    def seleccion_actual():
        return [w for w in estado_seguimiento["marcados"]
                if w in estado_seguimiento["filas"]]

    def ejecutar_accion_siguiente():
        ids = seleccion_actual()
        info = controller.tracking.next_action_for(ids)
        if info["action"] is None:
            messagebox.showinfo("Acción siguiente", info["reason"], parent=ventana)
            return
        accion = info["action"]
        if accion is NextAction.NONE:
            messagebox.showinfo("Acción siguiente", info["reason"], parent=ventana)
            return
        datos = {}
        if accion is NextAction.RESOLVE_CONFIRMATION:
            # Dos caminos validos y nada mas: confirmó o canceló. La operadora
            # no elige un estado de una lista, elige lo que pasó.
            resolver_confirmacion(ids)
            return
        if accion is NextAction.SEND_TO_LABORATORY:
            datos = pedir_destino_laboratorio(len(ids))
            if datos is None:
                return
        elif accion is NextAction.RESOLVE_RECEPTION:
            # Resolver la recepcion es recibir lo que no habia llegado. Se
            # confirma aparte porque cierra una discrepancia, no porque
            # bloquee: el trabajo nunca deja de poder avanzar.
            if not messagebox.askyesno(
                    "Resolver recepción",
                    f"Estos {len(ids)} trabajo(s) figuran como NO LLEGÓ.\n\n"
                    "¿Aparecieron y los recibís en Asunción?",
                    parent=ventana):
                return
        elif not messagebox.askyesno(
                "Confirmar", f"¿{info['label']}?", parent=ventana):
            return
        try:
            controller.tracking.apply_next_action(
                ids, responsible=responsable_actual(), **datos)
        except Exception as exc:
            mostrar_error(exc)
            return
        refrescar_seguimiento()

    def pedir_destino_laboratorio(cantidad):
        """Un solo dialogo para todo el grupo: laboratorio y plazo."""
        catalogo = controller.tracking.selectable_laboratories()
        if not catalogo:
            mostrar_error(InvalidCashDayError(
                "Cargá al menos un laboratorio antes de enviar trabajos."))
            return None
        elegido = simpledialog.askstring(
            "Enviar a laboratorio",
            f"{cantidad} trabajo(s).\nLaboratorio "
            f"({', '.join(l.name for l in catalogo)}):",
            parent=ventana)
        laboratorio = next(
            (l for l in catalogo
             if l.name.casefold() == str(elegido or "").strip().casefold()), None)
        if laboratorio is None:
            return None
        fecha = simpledialog.askstring(
            "Fecha esperada", "Fecha esperada (dd-mm-aaaa):", parent=ventana,
            initialvalue=date.today().strftime("%d-%m-%Y"))
        if not str(fecha or "").strip():
            return None
        hora = simpledialog.askstring(
            "Hora esperada", "Hora esperada (HH:MM):", parent=ventana,
            initialvalue=controller.tracking.default_expected_time().strftime("%H:%M"))
        return {"laboratory_id": laboratorio.id, "expected_date": fecha,
                "expected_time": hora or None}

    boton_accion_siguiente = ctk.CTkButton(
        acciones_seguimiento, text="Acción siguiente", width=250, height=34,
        fg_color=color_azul, hover_color="#0F5FC7",
        font=ctk.CTkFont(size=perfil["fuente_label"] + 1, weight="bold"),
        command=ejecutar_accion_siguiente,
    )
    boton_accion_siguiente.pack(side="left", padx=(0, 6))

    boton_novedad = ctk.CTkButton(
        acciones_seguimiento, text="Novedad", width=140, height=34,
        fg_color="#B45309", hover_color="#8F4207",
        font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
        command=lambda: abrir_novedad(seleccion_actual()),
    )
    boton_novedad.pack(side="left", padx=(0, 6))

    def abrir_menu_mas():
        """Acciones poco frecuentes, fuera de la vista normal."""
        menu = tk.Menu(ventana, tearoff=0)
        menu.add_command(label="Ver detalle", command=abrir_detalle_trabajo)
        menu.add_separator()
        menu.add_command(label="Seleccionar todo", command=seleccionar_todo)
        menu.add_command(label="Limpiar selección", command=limpiar_seleccion)
        menu.add_separator()
        menu.add_command(label="Queda a confirmar", command=marcar_queda_a_confirmar)
        menu.add_command(label="Corregir estado", command=corregir_estado)
        menu.add_command(label="Cerrar por excepción", command=cerrar_por_excepcion)
        menu.add_separator()
        menu.add_command(
            label=("Ver solo mi sucursal" if contexto_sucursal["todas"]
                   else "Ver todas las sucursales"),
            command=alternar_alcance_sucursal,
            state="normal" if contexto_sucursal["sucursal"] else "disabled")
        x = boton_mas.winfo_rootx()
        y = boton_mas.winfo_rooty() - menu.yposition(0) - 190
        menu.tk_popup(x, max(0, y))

    boton_mas = ctk.CTkButton(
        acciones_seguimiento, text="Más  ▾", width=100, height=34,
        fg_color="#FFFFFF", text_color=color_suave, border_width=1,
        border_color=color_borde_suave, hover_color="#EAF3FF",
        font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
        command=lambda: abrir_menu_mas(),
    )
    boton_mas.pack(side="left")

    etiqueta_seleccion = ctk.CTkLabel(
        acciones_seguimiento, text="", text_color=color_suave,
        font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
    )
    etiqueta_seleccion.pack(side="right", padx=(6, 0))

    def actualizar_acciones_seguimiento(_event=None):
        ids = seleccion_actual()
        info = controller.tracking.next_action_for(ids)
        resumen_accion.update(action=info["action"], ids=ids)
        hay = bool(ids)
        etiqueta_seleccion.configure(
            text=f"{len(ids)} seleccionado{'' if len(ids) == 1 else 's'}" if hay
            else "Marcá uno o varios trabajos")
        boton_accion_siguiente.configure(
            text=info["label"] or "Acción siguiente",
            state="normal" if info["action"] not in (None, NextAction.NONE) else "disabled")
        # La accion complementaria se sugiere en el boton que ya la ejecuta, en
        # vez de sumar un cuarto boton: sigue siendo Novedad, con el nombre de
        # lo que conviene hacer. Sugerir no es exigir — el boton principal
        # queda habilitado igual, porque contactar nunca bloquea el avance.
        sugerida = info.get("complementary")
        boton_novedad.configure(
            state="normal" if hay else "disabled",
            text=("Contactar laboratorio" if sugerida is NextAction.CONTACT_LABORATORY
                  else "Novedad"),
            width=200 if sugerida is NextAction.CONTACT_LABORATORY else 140,
            fg_color="#B42318" if sugerida is NextAction.CONTACT_LABORATORY else "#B45309")

    def ir_a_atrasados(_event=None):
        """Abre exactamente los trabajos que originaron la alerta.

        El grupo viaja con la alerta, asi que el filtro se aplica solo: la
        operadora no tiene que volver a tocar Atrasados ni buscar a mano.
        """
        grupo = contexto_alerta.get("grupo")
        contexto_grupo["foco"] = grupo
        # Un atraso cruza etapas: no es un grupo, se abre por su filtro.
        refrescar_seguimiento(
            "Activos" if grupo else (contexto_alerta.get("filtro") or "Atrasados"))

    alerta_seguimiento.configure(cursor="hand2")
    alerta_seguimiento.bind("<Button-1>", ir_a_atrasados, add="+")

    def abrir_detalle_trabajo(_event=None):
        """Ficha del trabajo: todo lo necesario sin cambiar de pantalla."""
        fila = trabajo_seleccionado()
        if fila is None:
            messagebox.showwarning(
                "Seleccioná un trabajo", "Elegí una fila.", parent=ventana)
            return
        try:
            detalle = controller.tracking.work_detail(fila.work.id)
        except Exception as exc:
            mostrar_error(exc)
            return
        dialogo = ctk.CTkToplevel(ventana)
        dialogo.title(f"Trabajo {detalle['envelope']}")
        dialogo.geometry("720x620")
        dialogo.transient(ventana)
        dialogo.grab_set()

        encabezado = ctk.CTkFrame(dialogo, fg_color=color_panel, corner_radius=0)
        encabezado.pack(fill="x")
        ctk.CTkLabel(
            encabezado, text=f"{detalle['envelope']}   ·   {detalle['customer_name']}",
            text_color=color_texto,
            font=ctk.CTkFont(size=perfil["fuente_seccion"], weight="bold"),
        ).pack(side="left", padx=12, pady=8)
        fondo, _b, texto_color = (
            COLOR_CHIP_ATRASADO if detalle["alert"] == "ATRASADO"
            else COLOR_CHIP_CONFIRMADO if detalle["alert"]
            else COLORES_ESTADO_SEGUIMIENTO.get(
                detalle["physical_status"], ("#EEF4FB", "#B9CDE5", "#132238"))
        )
        ctk.CTkLabel(
            encabezado, text=f"  {detalle['status']}  ", corner_radius=7,
            fg_color=fondo, text_color=texto_color,
            font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
        ).pack(side="right", padx=12, pady=8)

        ficha = ctk.CTkFrame(dialogo, fg_color="transparent")
        ficha.pack(fill="x", padx=14, pady=(10, 4))
        ficha.grid_columnconfigure(1, weight=1)
        for indice, (etiqueta, valor) in enumerate((
            ("Teléfono del cliente", detalle["customer_phone"] or "—"),
            ("Tipo de trabajo", detalle["work_type"] or "—"),
            ("Laboratorio", detalle["laboratory"]),
            ("Teléfono de línea", detalle["phone_line"] or "—"),
            ("WhatsApp", detalle["whatsapp"] or "—"),
            ("Fecha/hora esperada", detalle["expected"] or "—"),
            ("Última novedad", detalle["last_news"] or "—"),
            ("Registrado por", detalle["created_by"] or "—"),
        )):
            ctk.CTkLabel(
                ficha, text=etiqueta, anchor="w", text_color=color_suave,
                font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
            ).grid(row=indice, column=0, sticky="w", padx=(0, 12), pady=3)
            ctk.CTkLabel(
                ficha, text=str(valor), anchor="w", text_color=color_texto,
                font=ctk.CTkFont(size=perfil["fuente"]),
            ).grid(row=indice, column=1, sticky="w", pady=3)

        historial = ctk.CTkTextbox(
            dialogo, fg_color="#FFFFFF", border_width=1, border_color=color_borde_suave,
            wrap="word", font=ctk.CTkFont(size=perfil["fuente"]),
        )
        historial.pack(fill="both", expand=True, padx=14, pady=(8, 6))
        lineas = ["RECORRIDO DEL TRABAJO", ""]
        for transicion in detalle["transitions"]:
            momento = transicion.recorded_at.astimezone(BUSINESS_TIMEZONE)
            origen = transicion.from_status.value if transicion.from_status else "—"
            lineas.append(
                f"  {momento:%d-%m %H:%M}   {origen} → {transicion.to_status.value}"
                + (f"   ({transicion.responsible})" if transicion.responsible else "")
                + (f"   {transicion.note}" if transicion.note else "")
            )
        if not detalle["transitions"]:
            lineas.append("  Sin transiciones todavía.")
        lineas += ["", "NOVEDADES CON EL LABORATORIO", ""]
        for contacto in detalle["contacts"]:
            lineas.append("  " + contacto.summary())
        if not detalle["contacts"]:
            lineas.append("  Sin novedades registradas.")
        historial.insert("1.0", "\n".join(lineas))
        historial.configure(state="disabled")

        ctk.CTkButton(
            dialogo, text="Cerrar", width=120, height=32, fg_color="#52657D",
            command=dialogo.destroy,
        ).pack(pady=(0, 12))
        return dialogo

    def abrir_nuevo_envio_pilar():
        """Selector del lote: los trabajos ya existen, aca solo se eligen."""
        dialogo = ctk.CTkToplevel(ventana)
        dialogo.title("Nuevo envío desde Pilar")
        dialogo.geometry("980x620")
        dialogo.transient(ventana)
        dialogo.grab_set()

        encabezado = ctk.CTkFrame(dialogo, fg_color=color_panel, corner_radius=0)
        encabezado.pack(fill="x")
        ctk.CTkLabel(
            encabezado, text="Consulta de Pilar", text_color=color_texto,
            font=ctk.CTkFont(size=perfil["fuente_seccion"], weight="bold"),
        ).pack(side="left", padx=12, pady=8)
        ctk.CTkLabel(
            encabezado, text="Sucursal", text_color=color_suave,
            font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
        ).pack(side="left", padx=(12, 4))
        campo_sucursal = ctk.CTkEntry(encabezado, width=110, height=30)
        campo_sucursal.insert(0, "Pilar")
        campo_sucursal.pack(side="left", pady=8)
        ctk.CTkLabel(
            encabezado, text="Desde", text_color=color_suave,
            font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
        ).pack(side="left", padx=(12, 4))
        # Ventana por defecto de tres dias: la consulta del viernes sigue
        # visible el sabado sin que la operadora toque el selector, que queda
        # editable para cualquier otro rango.
        desde_defecto, hasta_defecto = controller.tracking.default_candidate_range()
        campo_desde = ctk.CTkEntry(encabezado, width=115, height=30)
        campo_desde.insert(0, desde_defecto.strftime("%d-%m-%Y"))
        campo_desde.pack(side="left", pady=8)
        ctk.CTkLabel(
            encabezado, text="Hasta", text_color=color_suave,
            font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
        ).pack(side="left", padx=(12, 4))
        campo_hasta = ctk.CTkEntry(encabezado, width=115, height=30)
        campo_hasta.insert(0, hasta_defecto.strftime("%d-%m-%Y"))
        campo_hasta.pack(side="left", pady=8)

        resumen_envio = ctk.CTkLabel(
            dialogo, text="", anchor="w", text_color=color_texto,
            font=ctk.CTkFont(size=perfil["fuente_label"] + 1, weight="bold"),
        )
        resumen_envio.pack(fill="x", padx=12, pady=(8, 4))

        marco_candidatos = ctk.CTkFrame(dialogo, fg_color="#FFFFFF", corner_radius=6)
        marco_candidatos.pack(fill="both", expand=True, padx=12)
        marco_candidatos.grid_rowconfigure(0, weight=1)
        marco_candidatos.grid_columnconfigure(0, weight=1)
        COLUMNAS_CANDIDATO = (
            ("marca", "✓", 40, "center", False),
            ("sobre", "Sobre", 90, "center", False),
            ("cliente", "Cliente", 240, "w", False),
            ("tipo", "Tipo de trabajo", 200, "w", False),
            ("entrega", "Entrega", 110, "center", False),
            ("vendedora", "Vendedora", 130, "center", True),
        )
        grilla_candidatos = ttk.Treeview(
            marco_candidatos, columns=[c for c, *_r in COLUMNAS_CANDIDATO],
            show="headings", style="Caja.Treeview", selectmode="none",
        )
        for clave, titulo, ancho, anclaje, expandible in COLUMNAS_CANDIDATO:
            grilla_candidatos.heading(clave, text=titulo, anchor=anclaje)
            grilla_candidatos.column(
                clave, width=ancho, minwidth=ancho, anchor=anclaje, stretch=expandible,
            )
        scroll_candidatos = ttk.Scrollbar(
            marco_candidatos, orient="vertical", command=grilla_candidatos.yview,
        )
        grilla_candidatos.configure(yscrollcommand=scroll_candidatos.set)
        grilla_candidatos.grid(row=0, column=0, sticky="nsew", padx=(5, 0), pady=5)
        scroll_candidatos.grid(row=0, column=1, sticky="ns", padx=(0, 5), pady=5)

        estado_envio = {"pedidos": {}, "elegidos": set()}

        def pintar_marca(order_id):
            grilla_candidatos.set(
                order_id, "marca", "✓" if order_id in estado_envio["elegidos"] else "",
            )

        def actualizar_resumen_envio():
            total = len(estado_envio["pedidos"])
            elegidos = len(estado_envio["elegidos"])
            resumen_envio.configure(
                text=f"{total} trabajos encontrados     ·     {elegidos} seleccionados"
            )
            boton_crear.configure(state="normal" if elegidos else "disabled")
            boton_crear.configure(text=f"Crear envío ({elegidos})" if elegidos else "Crear envío")

        def alternar(evento):
            fila = grilla_candidatos.identify_row(evento.y)
            if not fila:
                return
            if fila in estado_envio["elegidos"]:
                estado_envio["elegidos"].discard(fila)
            else:
                estado_envio["elegidos"].add(fila)
            pintar_marca(fila)
            actualizar_resumen_envio()

        grilla_candidatos.bind("<Button-1>", alternar, add="+")

        def buscar_candidatos():
            for item in grilla_candidatos.get_children():
                grilla_candidatos.delete(item)
            estado_envio["pedidos"].clear()
            estado_envio["elegidos"].clear()
            try:
                candidatos = controller.tracking.shipment_candidates(
                    branch=campo_sucursal.get().strip() or "Pilar",
                    consultation_date=campo_desde.get().strip(),
                    end_date=campo_hasta.get().strip(),
                )
            except Exception as exc:
                mostrar_error(exc)
                return
            for pedido in candidatos:
                grilla_candidatos.insert("", "end", iid=pedido.id, values=(
                    "", pedido.envelope, pedido.customer_name,
                    pedido.observations.splitlines()[0] if pedido.observations else "",
                    pedido.delivery_date.strftime("%d-%m-%Y"), pedido.saleswoman,
                ))
                estado_envio["pedidos"][pedido.id] = pedido
            actualizar_resumen_envio()

        def seleccionar_todos():
            estado_envio["elegidos"] = set(estado_envio["pedidos"])
            for order_id in estado_envio["pedidos"]:
                pintar_marca(order_id)
            actualizar_resumen_envio()

        def quitar_todos():
            estado_envio["elegidos"].clear()
            for order_id in estado_envio["pedidos"]:
                pintar_marca(order_id)
            actualizar_resumen_envio()

        def crear_envio():
            elegidos = sorted(estado_envio["elegidos"])
            if not elegidos:
                return
            if not messagebox.askyesno(
                "Confirmar envío",
                f"¿Crear el envío desde Pilar con {len(elegidos)} trabajo(s)?",
                parent=dialogo,
            ):
                return
            try:
                resultado = controller.tracking.create_pilar_shipment(
                    elegidos, operator=responsable_actual(),
                    consultation_date=campo_desde.get().strip(),
                )
            except Exception as exc:
                mostrar_error(exc)
                return
            messagebox.showinfo(
                "Envío creado",
                f"{resultado['count']} trabajo(s) marcados ENVIADO_DESDE_PILAR.",
                parent=dialogo,
            )
            dialogo.destroy()
            refrescar_seguimiento()

        acciones_envio = ctk.CTkFrame(dialogo, fg_color="transparent")
        acciones_envio.pack(fill="x", padx=12, pady=10)
        ctk.CTkButton(
            acciones_envio, text="Buscar", width=110, height=32, fg_color="#52657D",
            command=buscar_candidatos,
        ).pack(side="left", padx=(0, 3))
        ctk.CTkButton(
            acciones_envio, text="Seleccionar todos", width=150, height=32,
            fg_color="#FFFFFF", text_color=color_azul, border_width=1,
            border_color=color_borde_suave, hover_color="#EAF3FF",
            command=seleccionar_todos,
        ).pack(side="left", padx=3)
        ctk.CTkButton(
            acciones_envio, text="Quitar selección", width=150, height=32,
            fg_color="#FFFFFF", text_color=color_suave, border_width=1,
            border_color=color_borde_suave, hover_color="#EAF3FF",
            command=quitar_todos,
        ).pack(side="left", padx=3)
        ctk.CTkButton(
            acciones_envio, text="Cancelar", width=110, height=32, fg_color="#52657D",
            command=dialogo.destroy,
        ).pack(side="right", padx=(3, 0))
        boton_crear = ctk.CTkButton(
            acciones_envio, text="Crear envío", width=170, height=32,
            fg_color=color_verde, hover_color="#128A57", state="disabled",
            font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
            command=crear_envio,
        )
        boton_crear.pack(side="right", padx=3)

        buscar_candidatos()
        return dialogo

    def abrir_abm_laboratorios():
        """ABM compacto. Sin borrado fisico: solo alta, edicion y baja logica."""
        dialogo = ctk.CTkToplevel(ventana)
        dialogo.title("Laboratorios")
        # Ancho dimensionado a la fila del formulario: con 760 el campo
        # WhatsApp quedaba cortado contra el borde.
        dialogo.geometry("880x520")
        dialogo.transient(ventana)
        dialogo.grab_set()

        marco_labs = ctk.CTkFrame(dialogo, fg_color="#FFFFFF", corner_radius=6)
        marco_labs.pack(fill="both", expand=True, padx=12, pady=(12, 6))
        marco_labs.grid_rowconfigure(0, weight=1)
        marco_labs.grid_columnconfigure(0, weight=1)
        COLUMNAS_LAB = (
            ("nombre", "Nombre", 220, "w", True),
            ("linea", "Teléfono de línea", 165, "center", False),
            ("whatsapp", "WhatsApp", 165, "center", False),
            ("estado", "Estado", 110, "center", False),
        )
        grilla_labs = ttk.Treeview(
            marco_labs, columns=[c for c, *_r in COLUMNAS_LAB],
            show="headings", style="Caja.Treeview",
        )
        for clave, titulo, ancho, anclaje, expandible in COLUMNAS_LAB:
            grilla_labs.heading(clave, text=titulo, anchor=anclaje)
            grilla_labs.column(
                clave, width=ancho, minwidth=ancho, anchor=anclaje, stretch=expandible,
            )
        grilla_labs.tag_configure("inactivo", foreground="#8A94A6")
        scroll_labs = ttk.Scrollbar(marco_labs, orient="vertical", command=grilla_labs.yview)
        grilla_labs.configure(yscrollcommand=scroll_labs.set)
        grilla_labs.grid(row=0, column=0, sticky="nsew", padx=(5, 0), pady=5)
        scroll_labs.grid(row=0, column=1, sticky="ns", padx=(0, 5), pady=5)

        formulario_lab = ctk.CTkFrame(dialogo, fg_color=color_panel, corner_radius=6)
        formulario_lab.pack(fill="x", padx=12, pady=(0, 6))
        campos_lab = {}
        for indice, (clave, etiqueta, ancho) in enumerate((
            ("name", "Nombre", 200), ("phone_line", "Teléfono de línea", 150),
            ("whatsapp", "WhatsApp", 150),
        )):
            ctk.CTkLabel(
                formulario_lab, text=etiqueta, text_color=color_suave,
                font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
            ).grid(row=0, column=indice * 2, sticky="w", padx=(10, 4), pady=8)
            campo = ctk.CTkEntry(formulario_lab, width=ancho, height=30)
            campo.grid(row=0, column=indice * 2 + 1, padx=(0, 8), pady=8)
            campos_lab[clave] = campo

        seleccion_lab = {"id": None}

        def limpiar_formulario():
            seleccion_lab["id"] = None
            for campo in campos_lab.values():
                campo.delete(0, "end")
            boton_guardar_lab.configure(text="Agregar laboratorio")
            boton_estado_lab.configure(state="disabled", text="Activar / desactivar")

        def refrescar_labs():
            for item in grilla_labs.get_children():
                grilla_labs.delete(item)
            for laboratorio in controller.tracking.list_laboratories():
                grilla_labs.insert(
                    "", "end", iid=laboratorio.id,
                    tags=() if laboratorio.active else ("inactivo",),
                    values=(
                        laboratorio.name, laboratorio.phone_line, laboratorio.whatsapp,
                        "ACTIVO" if laboratorio.active else "INACTIVO",
                    ),
                )
            limpiar_formulario()

        def elegir_lab(_event=None):
            seleccion = grilla_labs.selection()
            if not seleccion:
                return
            laboratorio = controller.tracking.repository.get_laboratory(seleccion[0])
            if laboratorio is None:
                return
            seleccion_lab["id"] = laboratorio.id
            for clave, valor in (
                ("name", laboratorio.name), ("phone_line", laboratorio.phone_line),
                ("whatsapp", laboratorio.whatsapp),
            ):
                campos_lab[clave].delete(0, "end")
                campos_lab[clave].insert(0, valor)
            boton_guardar_lab.configure(text="Guardar cambios")
            boton_estado_lab.configure(
                state="normal",
                text="Desactivar" if laboratorio.active else "Activar",
            )

        def guardar_lab():
            datos = {clave: campo.get().strip() for clave, campo in campos_lab.items()}
            try:
                if seleccion_lab["id"]:
                    controller.tracking.update_laboratory(seleccion_lab["id"], **datos)
                else:
                    controller.tracking.save_laboratory(**datos)
            except Exception as exc:
                mostrar_error(exc)
                return
            refrescar_labs()
            refrescar_seguimiento()

        def alternar_estado_lab():
            laboratory_id = seleccion_lab["id"]
            if not laboratory_id:
                return
            laboratorio = controller.tracking.repository.get_laboratory(laboratory_id)
            if laboratorio.active and controller.tracking.laboratory_has_history(laboratory_id):
                if not messagebox.askyesno(
                    "Desactivar laboratorio",
                    f"{laboratorio.name} tiene trabajos asociados.\n\n"
                    "Se conserva en el historial y deja de ofrecerse para envíos nuevos. "
                    "¿Continuar?",
                    parent=dialogo,
                ):
                    return
            try:
                controller.tracking.set_laboratory_active(laboratory_id, not laboratorio.active)
            except Exception as exc:
                mostrar_error(exc)
                return
            refrescar_labs()
            refrescar_seguimiento()

        acciones_lab = ctk.CTkFrame(dialogo, fg_color="transparent")
        acciones_lab.pack(fill="x", padx=12, pady=(0, 12))
        boton_guardar_lab = ctk.CTkButton(
            acciones_lab, text="Agregar laboratorio", width=180, height=32,
            fg_color=color_verde, hover_color="#128A57",
            font=ctk.CTkFont(size=perfil["fuente_label"], weight="bold"),
            command=guardar_lab,
        )
        boton_guardar_lab.pack(side="left", padx=(0, 3))
        boton_estado_lab = ctk.CTkButton(
            acciones_lab, text="Activar / desactivar", width=170, height=32,
            fg_color="#B45309", state="disabled", command=alternar_estado_lab,
        )
        boton_estado_lab.pack(side="left", padx=3)
        ctk.CTkButton(
            acciones_lab, text="Nuevo", width=110, height=32, fg_color="#52657D",
            command=limpiar_formulario,
        ).pack(side="left", padx=3)
        ctk.CTkButton(
            acciones_lab, text="Cerrar", width=110, height=32, fg_color="#52657D",
            command=dialogo.destroy,
        ).pack(side="right")

        grilla_labs.bind("<<TreeviewSelect>>", elegir_lab, add="+")
        refrescar_labs()
        return dialogo

    refrescar_seguimiento()

    if not controller.admin.has_admin() and not os.environ.get("BC_CAJA_AUTOMATED"):
        ventana.after(450, lambda: abrir_acceso_administrador(configuracion_inicial=True))

    ventana.after(100, ventana.focus_set)
    return ventana
